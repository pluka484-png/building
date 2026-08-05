import io
import re
from difflib import SequenceMatcher
from datetime import date, datetime
from pathlib import Path
from typing import cast

import openpyxl
import pandas as pd
import streamlit as st


def is_blank(value) -> bool:
    return bool(pd.isna(value)) or str(value).strip().lower() in {"", "nan", "none"}


def cell_text(value) -> str:
    if is_blank(value):
        return ""
    return " ".join(str(value).replace("\n", " ").split())


def clean_column_name(value) -> str:
    text = str(value).strip().lower()
    text = "".join(ch if ch.isalnum() else "_" for ch in text)
    while "__" in text:
        text = text.replace("__", "_")
    text = text.strip("_") or "unnamed_column"
    if text[0].isdigit():
        text = f"col_{text}"
    return text


def dedupe_columns(columns):
    counts = {}
    result = []
    for col in columns:
        base = clean_column_name(col)
        counts[base] = counts.get(base, 0) + 1
        result.append(base if counts[base] == 1 else f"{base}_{counts[base]}")
    return result


def primary_excel_format_section(number_format: str) -> str:
    section = str(number_format or "General").split(";")[0]
    section = re.sub(r'"[^"]*"', "", section)
    section = re.sub(r"\[[^]]+]", "", section)
    return section


def decimal_places_from_excel_format(number_format: str) -> int | None:
    section = primary_excel_format_section(number_format)
    if "." not in section:
        return 0 if any(token in section for token in ("0", "#", "?")) else None
    decimals = section.split(".", 1)[1]
    decimals = decimals.split("%", 1)[0]
    placeholders = [char for char in decimals if char in {"0", "#", "?"}]
    return len(placeholders) if placeholders else 0


def formatted_excel_value(cell):
    value = cell.value
    if value is None:
        return ""

    if isinstance(value, (datetime, date)):
        return value.strftime("%d.%m.%Y")

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        number_format = str(cell.number_format or "General")
        decimals = decimal_places_from_excel_format(number_format)
        if decimals is None:
            return value

        is_percent = "%" in primary_excel_format_section(number_format)
        display_value = value * 100 if is_percent else value
        use_grouping = "," in number_format
        show_plus = "+" in number_format and display_value > 0
        sign = "+" if show_plus else ""
        grouped = "," if use_grouping else ""
        rendered = f"{sign}{display_value:{grouped}.{decimals}f}"
        if is_percent:
            rendered = f"{rendered}%"
        return rendered

    return value


def row_has_values(row, start_col=1) -> bool:
    return any(not is_blank(v) for v in row.iloc[start_col:].tolist())


def first_nonblank_after(row, start_col=1) -> str:
    for value in row.iloc[start_col:].tolist():
        if not is_blank(value):
            return cell_text(value)
    return ""


def is_unit_label(value) -> bool:
    text = cell_text(value).lower()
    return text == "%" or text.startswith(("eur ", "usd ", "gbp ", "in %"))


def nonblank_col_indexes(row, start_col=1):
    return [
        idx for idx in range(start_col, len(row))
        if not is_blank(row.iloc[idx])
    ]


def drop_all_blank_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame() if df is None else df
    keep_columns = [
        col for col in df.columns
        if not df[col].apply(is_blank).all()
    ]
    return pd.DataFrame(df.loc[:, keep_columns])


def drop_fully_duplicate_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Remove columns whose values are semantically identical to an earlier column."""
    if df is None or df.empty:
        return pd.DataFrame() if df is None else df
    seen_signatures: set = set()
    keep_cols = []
    for col in df.columns:
        sig = []
        for value in df[col].tolist():
            if is_blank(value):
                sig.append("")
            else:
                sig.append(str(value).strip())
        signature = tuple(sig)
        if signature in seen_signatures:
            continue
        seen_signatures.add(signature)
        keep_cols.append(col)
    return pd.DataFrame(df.loc[:, keep_cols])


def columns_semantically_equal(left: pd.Series, right: pd.Series) -> bool:
    if len(left) != len(right):
        return False
    for left_value, right_value in zip(left.tolist(), right.tolist()):
        if is_blank(left_value) and is_blank(right_value):
            continue
        if cell_text(left_value) != cell_text(right_value):
            return False
    return True


def looks_like_data_value(value) -> bool:
    text = cell_text(value)
    if not text or is_unit_label(text):
        return False
    if metric_context(text)["metric_type"]:
        return False
    cleaned = text.replace(",", "").replace("%", "").replace("+", "").replace("−", "-").strip()
    if cleaned.lower() in {"n.m.", "n.m", "nm"}:
        return True
    try:
        float(cleaned)
        return True
    except ValueError:
        return False


def _is_numeric_like(value) -> bool:
    """Quick check: can this value be parsed as numeric after removing formatting."""
    text = cell_text(value)
    if not text:
        return False
    # Strip common formatting and try to parse
    cleaned = (
        text.replace(" ", "")
        .replace(",", "")
        .replace("%", "")
        .replace("+", "")
        .replace("−", "-")
    )
    cleaned = re.sub(r"[()\[\]]", "", cleaned)
    try:
        float(cleaned)
        return True
    except ValueError:
        return False


def _looks_like_numeric_pipe_spillover(value) -> bool:
    """Detect numeric|text spillover tokens from matrix scratch calculations."""
    text = cell_text(value)
    if "|" not in text:
        return False
    parts = [part.strip() for part in re.split(r"\s*\|\s*", text) if part.strip()]
    if len(parts) < 2:
        return False
    if not _is_numeric_like(parts[0]):
        return False
    return any(bool(re.search(r"[A-Za-z]", part)) for part in parts[1:])


def _drop_numeric_pipe_spillover_rows(df: pd.DataFrame) -> pd.DataFrame:
    """Drop rows containing numeric|text pipe spillover in any text-like column."""
    if df is None or df.empty:
        return pd.DataFrame() if df is None else df

    text_cols = [
        col
        for col in df.columns
        if pd.api.types.is_object_dtype(df[col].dtype) or pd.api.types.is_string_dtype(df[col].dtype)
    ]
    if not text_cols:
        return df

    def _row_has_spillover(row: pd.Series) -> bool:
        for value in row.tolist():
            if _looks_like_numeric_pipe_spillover(value):
                return True
        return False

    keep_mask = ~df[text_cols].apply(_row_has_spillover, axis=1)
    return pd.DataFrame(df.loc[keep_mask])


def _drop_misaligned_matrix_rows(df: pd.DataFrame) -> pd.DataFrame:
    """Drop rows that look like misaligned scratch/calculation spillover.
    
    Red flag: numeric value in one field + alphabetic text in another,
    indicating a spillover of calculation scratch rather than a valid data row.
    """
    if df is None or df.empty:
        return pd.DataFrame() if df is None else df
    required = {"line_item", "metric", "value"}
    if not required.issubset(set(df.columns)):
        return df

    def _has_alpha(value) -> bool:
        return bool(re.search(r"[A-Za-z]", cell_text(value)))

    def _is_red_flag(row: pd.Series) -> bool:
        line_item = row.get("line_item", "")
        metric = row.get("metric", "")
        value = row.get("value", "")

        # Explicit user-requested red flag: numeric|text (or numeric|numeric|text) spillover tokens.
        pipe_fields = [
            "line_item",
            "line_item_path",
            "parent_line_item",
            "metric",
            "metric_detail",
        ]
        for field in pipe_fields:
            if field in row and _looks_like_numeric_pipe_spillover(row.get(field, "")):
                return True

        # Red flag: numeric value in one field + alphabetic text in another
        mixed_numeric_text = (
            (_is_numeric_like(line_item) and _has_alpha(metric))
            or (_is_numeric_like(metric) and _has_alpha(value))
        )
        if not mixed_numeric_text:
            return False

        # Keep rows where value still looks analytic; drop obvious spillover.
        return not looks_like_data_value(value)

    keep_mask = ~df.apply(_is_red_flag, axis=1)
    return pd.DataFrame(df.loc[keep_mask])


def extract_bottom_notes(raw_df: pd.DataFrame) -> list[str]:
    if raw_df is None or raw_df.empty:
        return []

    last_data_row = -1
    for row_idx in range(len(raw_df)):
        values = raw_df.iloc[row_idx].tolist()
        if sum(1 for value in values if looks_like_data_value(value)) >= 2:
            last_data_row = row_idx

    if last_data_row < 0:
        return []

    notes = []
    for row_idx in range(last_data_row + 1, len(raw_df)):
        parts = [cell_text(value) for value in raw_df.iloc[row_idx].tolist() if cell_text(value)]
        if not parts:
            continue
        note = " ".join(parts)
        if is_unit_label(note):
            continue
        notes.append(note)

    return notes


def find_unit_header_cells(raw_df: pd.DataFrame):
    cells = []
    if raw_df is None or raw_df.empty:
        return cells
    unit_prefixes = ("eur ", "usd ", "gbp ", "in %")
    for row_idx in range(len(raw_df)):
        row = raw_df.iloc[row_idx]
        for col_idx, value in enumerate(row.tolist()):
            text = cell_text(value).lower()
            if (is_unit_label(text) or text.startswith(unit_prefixes)) and first_nonblank_after(row, col_idx + 1):
                cells.append((row_idx, col_idx))
    return cells


def find_general_header_cells(
    raw_df: pd.DataFrame,
    max_rows: int | None = None,
    max_cols: int | None = None,
    stop_after_first: bool = False,
):
    """Find likely mini-table header cells without relying on sheet-specific names."""
    cells = []
    if raw_df is None or raw_df.empty:
        return cells
    header_prefixes = (
        "eur ", "usd ", "gbp ", "chf ", "aud ", "in %", "vs.", "vs ",
        "as of ", "period", "quarter", "year",
    )
    row_limit = len(raw_df) - 1
    if max_rows is not None:
        row_limit = min(row_limit, max_rows)

    col_limit = raw_df.shape[1]
    if max_cols is not None:
        col_limit = min(col_limit, max_cols)

    for row_idx in range(row_limit):
        row = raw_df.iloc[row_idx]
        for col_idx, value in enumerate(row.iloc[:col_limit].tolist()):
            label = cell_text(value)
            if not label:
                continue
            header_cols = [
                idx for idx in nonblank_col_indexes(row, start_col=col_idx + 1)
                if idx < col_limit
            ]
            if len(header_cols) < 2:
                continue

            data_rows = 0
            for next_idx in range(row_idx + 1, min(len(raw_df), row_idx + 8)):
                next_row = raw_df.iloc[next_idx]
                line_item = cell_text(next_row.iloc[col_idx])
                if not line_item:
                    continue
                has_values = any(
                    col < len(next_row) and not is_blank(next_row.iloc[col])
                    for col in header_cols[:12]
                )
                if has_values:
                    data_rows += 1

            looks_like_unit = label.lower().startswith(header_prefixes)
            if looks_like_unit or data_rows >= 2:
                cells.append((row_idx, col_idx))
                if stop_after_first:
                    return cells
    return cells


def sheet_looks_general_report_tables(raw_df: pd.DataFrame) -> bool:
    return len(
        find_general_header_cells(
            raw_df,
            max_rows=120,
            max_cols=30,
            stop_after_first=True,
        )
    ) >= 1


def section_name(value) -> str:
    text = cell_text(value)
    text = text.replace("¹", "").replace("²", "").replace("³", "")
    return text.strip()


def metric_parse_text(value) -> str:
    text = cell_text(value)
    text = re.sub(r"[¹²³⁴⁵⁶⁷⁸⁹⁰]+$", "", text).strip()
    text = re.sub(r"\s+\d+\)$", "", text).strip()
    return text


def carry_forward_header_values(values):
    carried = []
    current = ""
    for value in values:
        if not is_blank(value):
            current = cell_text(value)
        carried.append(current)
    return carried


def auto_flatten_sectioned_financial_sheet(raw_df: pd.DataFrame) -> pd.DataFrame:
    """Flatten stacked financial statement sections into a long analytical table.

    Handles patterns like:
    - Exchange rates / Spot or Average with periods across columns
    - Valuation rates with valuation dates spanning tenor columns
    """
    records = []
    i = 0
    while i < len(raw_df):
        row = raw_df.iloc[i]
        first = cell_text(row.iloc[0])
        first_lower = first.lower()
        rate_label = first_nonblank_after(row, start_col=1)

        if first_lower.startswith("exchange rates") and rate_label:
            rate_type = rate_label
            unit = cell_text(raw_df.iloc[i + 1, 0]) if i + 1 < len(raw_df) else ""
            headers = raw_df.iloc[i + 1].tolist() if i + 1 < len(raw_df) else []
            j = i + 2
            while j < len(raw_df):
                data_row = raw_df.iloc[j]
                label = cell_text(data_row.iloc[0])
                label_lower = label.lower()
                if (
                    not label
                    or label_lower.startswith("exchange rates")
                    or label_lower.startswith("valuation rates")
                    or label_lower.startswith("1)")
                    or label_lower.startswith("2)")
                    or label_lower.startswith("3)")
                ):
                    break
                for col_idx in range(1, len(data_row)):
                    value = data_row.iloc[col_idx]
                    period = cell_text(headers[col_idx]) if col_idx < len(headers) else ""
                    if period and not is_blank(value):
                        records.append({
                            "section": section_name(first),
                            "rate_type": rate_type,
                            "unit": unit,
                            "currency": label,
                            "period": period,
                            "valuation_date": "",
                            "tenor": "",
                            "contract_type": "",
                            "value": value,
                        })
                j += 1
            i = j
            continue

        if first_lower.startswith("valuation rates"):
            date_headers = carry_forward_header_values(row.tolist())
            tenor_row = raw_df.iloc[i + 1].tolist() if i + 1 < len(raw_df) else []
            unit = cell_text(tenor_row[0]) if tenor_row else ""
            current_contract = ""
            j = i + 2
            while j < len(raw_df):
                data_row = raw_df.iloc[j]
                label = cell_text(data_row.iloc[0])
                label_lower = label.lower()
                if (
                    not label
                    or label_lower.startswith("valuation rates")
                    or label_lower.startswith("exchange rates")
                    or label_lower.startswith("1)")
                    or label_lower.startswith("2)")
                    or label_lower.startswith("3)")
                ):
                    break

                if not row_has_values(data_row, start_col=1):
                    current_contract = label
                    j += 1
                    continue

                for col_idx in range(1, len(data_row)):
                    value = data_row.iloc[col_idx]
                    valuation_date = date_headers[col_idx] if col_idx < len(date_headers) else ""
                    tenor = cell_text(tenor_row[col_idx]) if col_idx < len(tenor_row) else ""
                    if valuation_date and tenor and not is_blank(value):
                        records.append({
                            "section": section_name(first),
                            "rate_type": "",
                            "unit": unit,
                            "currency": label,
                            "period": "",
                            "valuation_date": valuation_date,
                            "tenor": tenor,
                            "contract_type": current_contract,
                            "value": value,
                        })
                j += 1
            i = j
            continue

        i += 1

    return pd.DataFrame(records)


def nearest_title_above(raw_df: pd.DataFrame, row_idx: int) -> str:
    for idx in range(row_idx - 1, -1, -1):
        first = cell_text(raw_df.iloc[idx, 0])
        if not first:
            continue
        if first.lower().startswith(("eur ", "usd ", "gbp ", "in %")):
            continue
        if row_has_values(raw_df.iloc[idx], start_col=1):
            continue
        return first
    return ""


def nearest_group_above(raw_df: pd.DataFrame, row_idx: int) -> str:
    for idx in range(row_idx - 1, -1, -1):
        if cell_text(raw_df.iloc[idx, 0]):
            continue
        group = first_nonblank_after(raw_df.iloc[idx], start_col=1)
        if group:
            return group
    return ""


def nearest_left_label_above(raw_df: pd.DataFrame, row_idx: int, col_idx: int) -> str:
    for idx in range(row_idx - 1, -1, -1):
        label = cell_text(raw_df.iloc[idx, col_idx])
        if label and not label.replace(" ", "").isupper():
            return label
    return ""


def nearest_nonblank_right(raw_df: pd.DataFrame, row_idx: int, col_idx: int, block_end: int) -> str:
    row = raw_df.iloc[row_idx]
    for idx in range(col_idx + 1, min(block_end, len(row))):
        value = cell_text(row.iloc[idx])
        if value:
            return value
    return ""


def cell_indent(raw_df: pd.DataFrame, row_idx: int, col_idx: int) -> float:
    indents = raw_df.attrs.get("excel_indents") or []
    if row_idx >= len(indents) or col_idx >= len(indents[row_idx]):
        return 0
    return indents[row_idx][col_idx] or 0


def indented_row_context(raw_df: pd.DataFrame, row_idx: int, label_col: int) -> dict:
    current_indent = cell_indent(raw_df, row_idx, label_col)
    if current_indent <= 0:
        return {}

    parents = []
    next_indent = current_indent
    for idx in range(row_idx - 1, -1, -1):
        label = cell_text(raw_df.iloc[idx, label_col])
        if not label:
            continue
        indent = cell_indent(raw_df, idx, label_col)
        if indent < next_indent:
            parents.append(label)
            next_indent = indent
            if indent <= 0:
                break

    parents.reverse()
    if not parents:
        return {}

    line_item = cell_text(raw_df.iloc[row_idx, label_col])
    return {
        "parent_line_item": parents[-1],
        "line_item_path": " > ".join([*parents, line_item]),
    }


def metric_context(metric: str) -> dict:
    text = metric_parse_text(metric)
    empty_context = {
        "metric_type": "",
        "metric_date": "",
        "comparison_date": "",
        "metric_year": "",
        "metric_month": "",
        "metric_quarter": "",
        "comparison_year": "",
    }

    def normalize_year(year_text: str) -> int:
        year = int(year_text)
        return year + 2000 if len(year_text) == 2 else year

    def quarter_from_month(month: int) -> int:
        return ((month - 1) // 3) + 1

    date_pattern = r"(\d{1,2})\.(\d{1,2})\.(\d{2,4})"
    delta_date_match = re.fullmatch(rf"[∆Δ]\s*{date_pattern}\s*/\s*{date_pattern}", text)
    if delta_date_match:
        start_day, start_month, start_year, end_day, end_month, end_year = delta_date_match.groups()
        return {
            **empty_context,
            "metric_type": "date_delta",
            "metric_date": pd.Timestamp(normalize_year(start_year), int(start_month), int(start_day)),
            "comparison_date": pd.Timestamp(normalize_year(end_year), int(end_month), int(end_day)),
        }

    date_match = re.fullmatch(date_pattern, text)
    if date_match:
        day, month, year = date_match.groups()
        normalized_year = normalize_year(year)
        month_int = int(month)
        return {
            **empty_context,
            "metric_type": "date",
            "metric_date": pd.Timestamp(normalized_year, month_int, int(day)),
            "metric_year": normalized_year,
            "metric_month": month_int,
            "metric_quarter": quarter_from_month(month_int),
        }

    # ISO date/datetime (e.g. 2024-03-31 or 2024-03-31 00:00:00)
    iso_dt_match = re.fullmatch(
        r"(20\d{2})-(\d{1,2})-(\d{1,2})(?:[ T]\d{1,2}:\d{2}(?::\d{2})?)?",
        text,
    )
    if iso_dt_match:
        year_text, month_text, day_text = iso_dt_match.groups()
        year = int(year_text)
        month = int(month_text)
        day = int(day_text)
        return {
            **empty_context,
            "metric_type": "date",
            "metric_date": pd.Timestamp(year, month, day),
            "metric_year": year,
            "metric_month": month,
            "metric_quarter": quarter_from_month(month),
        }

    # Month period tokens (e.g. 12M24, 12M 2024, 2024 12M)
    month_period_patterns = [
        r"(0?[1-9]|1[0-2])M\s*(\d{2,4})",
        r"(\d{2,4})\s*(0?[1-9]|1[0-2])M",
    ]
    month_period_match = None
    year = None
    month = None
    for idx, pattern in enumerate(month_period_patterns):
        month_period_match = re.fullmatch(pattern, text, flags=re.IGNORECASE)
        if month_period_match:
            if idx == 0:
                month = int(month_period_match.group(1))
                year = normalize_year(month_period_match.group(2))
            else:
                year = normalize_year(month_period_match.group(1))
                month = int(month_period_match.group(2))
            break

    if month_period_match and year is not None and month is not None:
        metric_date = pd.Timestamp(year, month, 1) + pd.offsets.MonthEnd(0)  # type: ignore[operator]
        return {
            **empty_context,
            "metric_type": "month",
            "metric_date": metric_date,
            "metric_year": year,
            "metric_month": month,
        }

    quarter_match = re.fullmatch(r"([1-4])Q\s*(\d{2,4})", text)
    if quarter_match:
        year_text = quarter_match.group(2)
        year = int(year_text) + 2000 if len(year_text) == 2 else int(year_text)
        return {
            **empty_context,
            "metric_type": "quarter",
            "metric_date": pd.Timestamp(year, int(quarter_match.group(1)) * 3, 1) + pd.offsets.MonthEnd(0),  # type: ignore[operator]

        }

    year_match = re.fullmatch(r"(20\d{2})", text)
    if year_match:
        year = int(year_match.group(1))
        return {
            **empty_context,
            "metric_type": "year",
            "metric_date": pd.Timestamp(year, 12, 31),
            "metric_year": year,
            "metric_month": 12,
        }

    delta_match = re.fullmatch(r"[∆Δ]\s*(20\d{2})\s*/\s*(20\d{2})", text)
    if delta_match:
        year = int(delta_match.group(1))
        comparison_year = int(delta_match.group(2))
        return {
            **empty_context,
            "metric_type": "delta",
            "metric_date": pd.Timestamp(year, 12, 31),
            "comparison_date": pd.Timestamp(comparison_year, 12, 31),
            "metric_year": year,
            "comparison_year": comparison_year,
        }

    return empty_context


def header_row_score(row) -> int:
    return sum(
        1
        for value in row.tolist()[1:]
        if cell_text(value) and not metric_context(cell_text(value))["metric_type"] and not is_unit_label(value)
    )


def best_group_header_row(raw_df: pd.DataFrame, unit_row_idx: int) -> list[str]:
    best_idx = None
    best_score = 0
    for idx in range(max(0, unit_row_idx - 4), unit_row_idx):
        score = header_row_score(raw_df.iloc[idx])
        if score > best_score:
            best_idx = idx
            best_score = score
    if best_idx is None:
        return [""] * raw_df.shape[1]
    return carry_forward_header_values(raw_df.iloc[best_idx].tolist())


def best_label_column(raw_df: pd.DataFrame, unit_row_idx: int, unit_cols: list[int]) -> int:
    if not unit_cols:
        return 0

    best_col = 0
    best_score = -1
    for col_idx in range(min(unit_cols)):
        score = 0
        for row_idx in range(unit_row_idx + 1, len(raw_df)):
            row = raw_df.iloc[row_idx]
            label = cell_text(row.iloc[col_idx])
            if not label or is_unit_label(label):
                continue
            if any(not is_blank(row.iloc[value_col]) for value_col in unit_cols):
                score += 1
        if score > best_score:
            best_col = col_idx
            best_score = score

    return best_col


def auto_flatten_grouped_metric_blocks(raw_df: pd.DataFrame) -> pd.DataFrame:
    """Flatten wide report tables with row labels and repeated metric groups."""
    if raw_df is None or raw_df.empty:
        return pd.DataFrame()

    records = []
    for unit_row_idx in range(1, len(raw_df)):
        unit_row = raw_df.iloc[unit_row_idx]
        unit_cols = [
            col_idx for col_idx in range(1, raw_df.shape[1])
            if is_unit_label(unit_row.iloc[col_idx])
        ]
        if len(unit_cols) < 4:
            continue

        label_col = best_label_column(raw_df, unit_row_idx, unit_cols)
        period_row = raw_df.iloc[unit_row_idx - 1]
        period_headers = carry_forward_header_values(period_row.tolist())
        detail_headers = raw_df.iloc[unit_row_idx - 2].tolist() if unit_row_idx >= 2 else []
        group_headers = best_group_header_row(raw_df, unit_row_idx)
        table_name = nearest_title_above(raw_df, unit_row_idx)
        if not table_name:
            table_name = first_nonblank_after(raw_df.iloc[max(0, unit_row_idx - 4)], start_col=1)

        row_idx = unit_row_idx + 1
        blank_label_rows = 0
        while row_idx < len(raw_df):
            data_row = raw_df.iloc[row_idx]
            line_item = cell_text(data_row.iloc[label_col])
            if sum(1 for value in data_row.tolist() if is_unit_label(value)) >= 4:
                break
            if not line_item:
                blank_label_rows += 1
                if blank_label_rows >= 5:
                    break
                row_idx += 1
                continue
            blank_label_rows = 0
            if is_unit_label(line_item):
                break
            if all(is_blank(data_row.iloc[col_idx]) for col_idx in unit_cols):
                row_idx += 1
                continue

            row_context = indented_row_context(raw_df, row_idx, label_col)
            for col_idx in unit_cols:
                value = data_row.iloc[col_idx]
                if is_blank(value):
                    continue
                metric = cell_text(period_headers[col_idx])
                if not metric:
                    metric = cell_text(group_headers[col_idx])
                metric_detail = cell_text(detail_headers[col_idx]) if col_idx < len(detail_headers) else ""
                if metric_detail in {metric, cell_text(group_headers[col_idx])}:
                    metric_detail = ""
                record = {
                    "table_name": table_name,
                    "section": table_name,
                    "column_group": cell_text(group_headers[col_idx]),
                    "unit": cell_text(unit_row.iloc[col_idx]),
                    "line_item": line_item,
                    "metric": metric,
                    "metric_detail": metric_detail,
                    "value": value,
                    "block_key": f"{unit_row_idx}:{cell_text(group_headers[col_idx])}",
                    "block_start_column": col_idx,
                }
                record.update(row_context)
                record.update(metric_context(metric))
                records.append(record)
            row_idx += 1

    return pd.DataFrame(records)


def auto_flatten_report_blocks(raw_df: pd.DataFrame) -> pd.DataFrame:
    """Extract repeated side-by-side report blocks into long rows."""
    unit_cells = find_unit_header_cells(raw_df)
    if not unit_cells:
        return pd.DataFrame()

    records = []
    blocks_by_row: dict[int, list[int]] = {}
    for row_idx, col_idx in unit_cells:
        blocks_by_row.setdefault(row_idx, []).append(col_idx)

    for header_row_idx, starts in blocks_by_row.items():
        starts = sorted(starts)
        for pos, start_col in enumerate(starts):
            start_col = int(start_col)
            end_col = int(starts[pos + 1]) if pos + 1 < len(starts) else raw_df.shape[1]
            header_row = raw_df.iloc[header_row_idx]
            unit = cell_text(header_row.iloc[start_col])
            metric_cols = [
                col for col in range(start_col + 1, end_col)
                if not is_blank(header_row.iloc[col])
            ]
            if not metric_cols:
                continue

            table_name = nearest_left_label_above(raw_df, header_row_idx, start_col)
            section = cell_text(raw_df.iloc[header_row_idx - 1, start_col]) if header_row_idx > 0 else ""
            column_group = nearest_nonblank_right(
                raw_df,
                header_row_idx - 1,
                start_col,
                end_col,
            ) if header_row_idx > 0 else ""

            row_idx = header_row_idx + 1
            blank_label_rows = 0
            while row_idx < len(raw_df):
                data_row = raw_df.iloc[row_idx]
                line_item = cell_text(data_row.iloc[start_col])
                if not line_item:
                    blank_label_rows += 1
                    if blank_label_rows >= 5:
                        break
                    row_idx += 1
                    continue
                blank_label_rows = 0
                if line_item.lower().startswith(("eur ", "usd ", "gbp ", "in %")):
                    break
                if not row_has_values(data_row, start_col=start_col + 1):
                    row_idx += 1
                    continue

                row_context = indented_row_context(raw_df, row_idx, start_col)
                for metric_col in metric_cols:
                    metric = cell_text(header_row.iloc[metric_col])
                    value = data_row.iloc[metric_col]
                    if metric and not is_blank(value):
                        record = {
                            "table_name": table_name,
                            "section": section,
                            "column_group": column_group,
                            "unit": unit,
                            "line_item": line_item,
                            "metric": metric,
                            "value": value,
                            "block_key": f"{header_row_idx}:{start_col}",
                            "block_start_column": start_col,
                        }
                        record.update(row_context)
                        record.update(metric_context(metric))
                        records.append(record)
                row_idx += 1

    return pd.DataFrame(records)


def _fallback_matrix_flatten(raw_df: pd.DataFrame) -> pd.DataFrame:
    """Flatten matrix layouts with stacked headers and interior numeric grids."""
    if raw_df is None or raw_df.empty or raw_df.shape[0] < 2 or raw_df.shape[1] < 3:
        return pd.DataFrame()

    matrix_values = raw_df.to_numpy(copy=False)
    row_count, col_count = matrix_values.shape

    # Pre-cache cell_text and looks_like_data_value for all cells to avoid redundant calls
    cell_cache = {}
    data_value_cache = {}
    metric_meta_cache = {}

    def _get_cached_cell_text(row_idx: int, col_idx: int) -> str:
        key = (row_idx, col_idx)
        if key not in cell_cache:
            cell_cache[key] = cell_text(matrix_values[row_idx, col_idx])
        return cell_cache[key]

    def _is_cached_data_value(row_idx: int, col_idx: int) -> bool:
        key = (row_idx, col_idx)
        if key not in data_value_cache:
            value = matrix_values[row_idx, col_idx]
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                data_value_cache[key] = True
            else:
                data_value_cache[key] = looks_like_data_value(value)
        return data_value_cache[key]

    def _row_numeric_hits(row_idx: int, start_col: int = 0) -> int:
        return sum(1 for col in range(start_col, col_count) if _is_cached_data_value(row_idx, col))

    # Find rows that look like data lines (some text context + multiple values).
    # Scan only first 200 rows to avoid scanning entire sheet for huge files
    scan_limit = min(row_count, 200)
    candidate_rows = [
        row_idx for row_idx in range(scan_limit)
        if _row_numeric_hits(row_idx) >= 2
    ]
    if not candidate_rows:
        return pd.DataFrame()

    probe_rows = candidate_rows[: min(100, len(candidate_rows))]

    # Detect a likely business-line column (often col 1 in matrix files).
    label_candidates = range(min(6, col_count))
    label_col = 0
    label_score = -1
    for col_idx in label_candidates:
        score = 0
        for row_idx in probe_rows:
            label = _get_cached_cell_text(row_idx, col_idx)
            if label and not _is_cached_data_value(row_idx, col_idx) and not is_unit_label(label):
                score += 1
        if score > label_score:
            label_col = col_idx
            label_score = score

    # Detect metric/line-detail column near the label column.
    metric_col = min(label_col + 1, col_count - 1)
    metric_score = -1
    for col_idx in range(label_col + 1, min(col_count, label_col + 6)):
        score = 0
        for row_idx in probe_rows:
            metric_text = _get_cached_cell_text(row_idx, col_idx)
            if metric_text and not _is_cached_data_value(row_idx, col_idx):
                score += 1
        if score > metric_score:
            metric_col = col_idx
            metric_score = score

    # Detect where data rows begin for this label/metric pattern.
    data_start = candidate_rows[0]
    for row_idx in candidate_rows[:50]:  # Check only first 50 candidate rows
        line_item = _get_cached_cell_text(row_idx, label_col)
        metric = _get_cached_cell_text(row_idx, metric_col)
        if (line_item or metric) and _row_numeric_hits(row_idx, start_col=metric_col + 1) >= 2:
            data_start = row_idx
            break

    # Capture a wider header band so top hierarchy levels (period/country) are retained.
    header_start = max(0, data_start - 10)
    header_rows = list(range(header_start, data_start))

    # Keep columns that carry actual data in the matrix body.
    # Limit body_end to 300 rows for large files instead of 600
    body_end = min(row_count, data_start + 300)
    value_cols = []
    for col_idx in range(metric_col + 1, col_count):
        hits = 0
        for row_idx in range(data_start, body_end):
            if _is_cached_data_value(row_idx, col_idx):
                hits += 1
        if hits >= 2:
            value_cols.append(col_idx)

    if not value_cols:
        return pd.DataFrame()


    def _looks_like_period_token(token: str) -> bool:
        text = token.strip().lower()
        if not text:
            return False
        # Ignore descriptive titles; period tokens are usually compact labels.
        if len(text) > 35:
            return False
        if any(flag in text for flag in ("ytd", "qtd", "mtd", "fy", "year to date")):
            return True
        return bool(re.search(r"\b(jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec)\w*\s+\d{2,4}\b", text))

    global_period_token = ""
    for row_idx in header_rows:
        for col_idx in range(0, min(col_count, max(metric_col + 1, 6))):
            token = _get_cached_cell_text(row_idx, col_idx)
            if token and token.upper() not in {"#REF!", "#DIV/0!"} and _looks_like_period_token(token):
                global_period_token = token
                break
        if global_period_token:
            break

    header_noise = {"#REF!", "#DIV/0!", "#N/A", "#VALUE!"}
    subheader_tokens = {
        "actual", "previous", "plan", "budget", "target", "forecast", "ytd", "mtd", "qtd"
    }

    # Forward-fill sparse top headers row-wise (common in wide matrix exports where group
    # names appear once and subcolumns are blank). This avoids orphan headers like "Previous".
    carried_header_rows: dict[int, list[str]] = {}
    for row_idx in header_rows:
        carried: list[str] = []
        current = ""
        for col_idx in range(col_count):
            token = _get_cached_cell_text(row_idx, col_idx)
            if token and token.upper() not in header_noise:
                current = token
            carried.append(current)
        carried_header_rows[row_idx] = carried

    def _header_for_col(col_idx: int) -> str:
        tokens: list[str] = []
        if global_period_token:
            tokens.append(global_period_token)

        for row_idx in header_rows:
            token = carried_header_rows[row_idx][col_idx]
            if not token or token.upper() in header_noise:
                continue
            if not tokens or tokens[-1] != token:
                tokens.append(token)

        # If we still ended up with only a generic subheader, borrow nearest left group label.
        if len(tokens) == 1 and tokens[0].strip().lower() in subheader_tokens:
            base = ""
            for left_col in range(col_idx - 1, metric_col, -1):
                for row_idx in header_rows:
                    candidate = carried_header_rows[row_idx][left_col]
                    if candidate and candidate.upper() not in header_noise and candidate.strip().lower() not in subheader_tokens:
                        base = candidate
                        break
                if base:
                    break
            if base:
                tokens = [base, tokens[0]]

        return " | ".join(tokens)

    value_headers = {col_idx: _header_for_col(col_idx) for col_idx in value_cols}

    # Pick a broad section/title from the top-left area.
    section = ""
    for row_idx in range(min(8, row_count)):
        for col_idx in range(min(4, col_count)):
            token = _get_cached_cell_text(row_idx, col_idx)
            if token and token.upper() not in {"#REF!", "#DIV/0!"}:
                section = token
                break
        if section:
            break

    records = []
    current_line_item = ""
    blank_run = 0
    header_metric_meta = {
        col_idx: metric_context(value_headers.get(col_idx, "")) if value_headers.get(col_idx, "") else None
        for col_idx in value_cols
    }

    indent_grid = raw_df.attrs.get("excel_indents") or []
    label_texts = [_get_cached_cell_text(row_idx, label_col) for row_idx in range(row_count)]
    label_indents = [
        float(indent_grid[row_idx][label_col])
        if row_idx < len(indent_grid) and label_col < len(indent_grid[row_idx])
        else 0.0
        for row_idx in range(row_count)
    ]

    def _fast_row_context(row_idx: int) -> dict:
        current_indent = label_indents[row_idx]
        if current_indent <= 0:
            return {}

        parents = []
        next_indent = current_indent
        # Bound lookback depth for performance on large sheets.
        min_idx = max(row_idx - 120, 0)
        for idx in range(row_idx - 1, min_idx - 1, -1):
            label = label_texts[idx]
            if not label:
                continue
            indent = label_indents[idx]
            if indent < next_indent:
                parents.append(label)
                next_indent = indent
                if indent <= 0:
                    break

        parents.reverse()
        if not parents:
            return {}
        return {
            "parent_line_item": parents[-1],
            "line_item_path": " > ".join([*parents, label_texts[row_idx]]),
        }

    has_indent_hierarchy = any(label_indents[row_idx] > 0 for row_idx in range(data_start, min(data_start + 1000, row_count)))
    # Limit data extraction to first 1000 rows for huge files
    data_end = min(row_count, data_start + 1000)
    for row_idx in range(data_start, data_end):
        row = matrix_values[row_idx]
        line_item = _get_cached_cell_text(row_idx, label_col)
        metric = _get_cached_cell_text(row_idx, metric_col)
        row_value_hits = sum(1 for col_idx in value_cols if _is_cached_data_value(row_idx, col_idx))

        if not line_item and not metric and row_value_hits == 0:
            blank_run += 1
            if blank_run >= 25:
                break
            continue
        blank_run = 0

        if line_item and not is_unit_label(line_item):
            current_line_item = line_item

        active_line_item = current_line_item or line_item
        if not active_line_item:
            continue

        metric_name = metric or "value"
        row_context = _fast_row_context(row_idx) if has_indent_hierarchy else {}
        for col_idx in value_cols:
            value = row[col_idx]
            if is_blank(value) or not _is_cached_data_value(row_idx, col_idx):
                continue
            header_label = value_headers.get(col_idx, "")
            record = {
                "table_name": section,
                "section": section,
                "column_group": header_label,
                "line_item": active_line_item,
                "metric": metric_name,
                "value": value,
                "block_start_column": label_col,
            }
            if header_label:
                metric_meta = header_metric_meta.get(col_idx) or {}
            else:
                if metric_name not in metric_meta_cache:
                    metric_meta_cache[metric_name] = metric_context(metric_name)
                metric_meta = metric_meta_cache[metric_name]
            record.update(row_context)
            record.update(metric_meta)
            records.append(record)

    return pd.DataFrame(records)


def auto_flatten_report_tables(raw_df: pd.DataFrame, extraction_profile: str = "auto") -> pd.DataFrame:
     """General extractor for visually formatted report sheets.

     This is the single user-facing auto mode. Internally it tries a few
     structural strategies and chooses the richest useful output. That keeps the
     UI general without pretending every visual report has one physical shape.

     KEY FIX: If ALL strategies fail (return empty), fall back to detect and flatten
     matrix-style data where both rows and columns contain meaningful identifiers.
     """
     profile = (extraction_profile or "auto").strip().lower()
     if profile not in {"auto", "general", "matrix"}:
         profile = "auto"

     first_col_texts = [cell_text(value).lower() for value in raw_df.iloc[:, 0].tolist()]
     looks_like_rates_sheet = any(
         text.startswith("exchange rates") or text.startswith("valuation rates")
         for text in first_col_texts
     )

     top_values = [
         str(v).strip().upper()
         for row in raw_df.head(12).values
         for v in row
         if not is_blank(v)
     ]
     has_matrix_signals = any(v in {"#REF!", "#DIV/0!", "#N/A"} for v in top_values)

     strategy_order = []
     if profile == "matrix" or (profile == "auto" and has_matrix_signals and not looks_like_rates_sheet):
         strategy_order = [
             ("fallback_matrix", _fallback_matrix_flatten),
             ("grouped_metric_blocks", auto_flatten_grouped_metric_blocks),
             ("side_by_side_blocks", auto_flatten_report_blocks),
             ("stacked_tables", auto_flatten_stacked_tables),
             ("sectioned_tables", auto_flatten_sectioned_financial_sheet),
         ]
     elif looks_like_rates_sheet:
         strategy_order = [
             ("sectioned_tables", auto_flatten_sectioned_financial_sheet),
             ("grouped_metric_blocks", auto_flatten_grouped_metric_blocks),
             ("side_by_side_blocks", auto_flatten_report_blocks),
             ("stacked_tables", auto_flatten_stacked_tables),
             ("fallback_matrix", _fallback_matrix_flatten),
         ]
     else:
         strategy_order = [
             ("grouped_metric_blocks", auto_flatten_grouped_metric_blocks),
             ("side_by_side_blocks", auto_flatten_report_blocks),
             ("stacked_tables", auto_flatten_stacked_tables),
             ("sectioned_tables", auto_flatten_sectioned_financial_sheet),
             ("fallback_matrix", _fallback_matrix_flatten),
         ]

     best_name = ""
     best_df = pd.DataFrame()
     best_score = 0
     for name, extractor in strategy_order:
         df = extractor(raw_df)
         if df is None or df.empty:
             continue

         # If user explicitly selected matrix mode, trust matrix extractor output.
         if profile == "matrix" and name == "fallback_matrix":
             best_name = name
             best_df = df.copy()
             best_score = len(df)
             break

         context_cols = [
             col for col in df.columns
             if col not in {"value", "block_start_column"}
         ]
         score = len(df) * max(len(context_cols), 1)
         if looks_like_rates_sheet and name == "sectioned_tables":
             # Preserve Spot/Average/tenor context on market-data style sheets.
             score *= 10
         if score > best_score:
             best_name = name
             best_df = df.copy()
             best_score = score

         # Early exit for obvious wins to reduce work on large sheets.
         if name == "fallback_matrix" and len(df) >= 500:
             break
         if looks_like_rates_sheet and name == "sectioned_tables" and len(df) >= 80:
             break

     if best_df.empty:
         return best_df

     normalized = best_df.copy()

     if "block_id" not in normalized.columns:
         if "block_key" in normalized.columns:
             starts = normalized["block_key"].fillna("-1")
             normalized = normalized.drop(columns=["block_key"])
             normalized.insert(
                 1,
                 "block_id",
                 pd.factorize(starts.astype(str))[0] + 1,
             )
         elif "block_start_column" in normalized.columns:
             starts = normalized["block_start_column"].fillna(-1)
             normalized.insert(
                 1,
                 "block_id",
                 pd.factorize(starts.astype(str))[0] + 1,
             )
         else:
             normalized.insert(1, "block_id", 1)

     if "block_start_column" in normalized.columns:
         normalized = normalized.drop(columns=["block_start_column"])

     if "line_item" not in normalized.columns:
         if "currency" in normalized.columns:
             # currency IS the line item for rates sheets; don't create a duplicate column
             normalized["line_item"] = normalized["currency"]
             normalized = normalized.drop(columns=["currency"])
         else:
             normalized["line_item"] = ""

     # Drop currency if it's now a duplicate of line_item
     if (
         "currency" in normalized.columns
         and "line_item" in normalized.columns
         and columns_semantically_equal(normalized["currency"], normalized["line_item"])
     ):
         normalized = normalized.drop(columns=["currency"])

     # Drop table_name if it's a duplicate of section — checked after table_name is guaranteed to exist
     if (
         "table_name" in normalized.columns
         and "section" in normalized.columns
     ):
         pass  # checked below after table_name fallback runs

     if "metric" not in normalized.columns:
         metric_parts = []
         for _, row in normalized.iterrows():
             parts = [
                 cell_text(row.get("period", "")),
                 cell_text(row.get("valuation_date", "")),
                 cell_text(row.get("tenor", "")),
                 cell_text(row.get("rate_type", "")),
             ]
             metric_parts.append(" | ".join([part for part in parts if part]) or "value")
         normalized["metric"] = metric_parts

     # For rates sheets, populate metric_date/year/quarter from period or valuation_date
     # when those context columns are missing or empty.
     date_source_col = None
     if "period" in normalized.columns and normalized["period"].astype(str).str.strip().ne("").any():
         date_source_col = "period"
     elif "valuation_date" in normalized.columns and normalized["valuation_date"].astype(str).str.strip().ne("").any():
         date_source_col = "valuation_date"

     if date_source_col is not None:
         context_cols_to_fill = [
             "metric_type",
             "metric_date",
             "metric_year",
             "metric_month",
             "metric_quarter",
             "comparison_year",
             "comparison_date",
         ]
         # Only fill if all context columns are missing or entirely empty
         needs_fill = all(
             col not in normalized.columns or normalized[col].astype(str).str.strip().eq("").all()
             for col in context_cols_to_fill
         )
         if needs_fill:
             parsed = normalized[date_source_col].apply(lambda v: metric_context(str(v)))
             for ctx_col in context_cols_to_fill:
                 normalized[ctx_col] = parsed.apply(lambda d: d.get(ctx_col, ""))

     if "table_name" not in normalized.columns:
         normalized["table_name"] = normalized["section"] if "section" in normalized.columns else ""

     # Drop table_name if it's semantically identical to section
     if (
         "table_name" in normalized.columns
         and "section" in normalized.columns
         and columns_semantically_equal(normalized["table_name"], normalized["section"])
     ):
         normalized = normalized.drop(columns=["table_name"])

     if "unit" not in normalized.columns:
         normalized["unit"] = ""

     for col in normalized.columns:
         if col != "value":
             normalized[col] = normalized[col].fillna("")

     ordered_columns = [
         "table_name",
         "block_id",
         "section",
         "column_group",
         "unit",
         "rate_type",
         "period",
         "valuation_date",
         "tenor",
         "contract_type",
         "currency",
         "parent_line_item",
         "line_item_path",
         "line_item",
         "metric",
         "metric_detail",
         "metric_type",
         "metric_date",
         "comparison_date",
         "metric_year",
         "metric_month",
         "metric_quarter",
         "comparison_year",
         "value",
     ]
     normalized = _drop_misaligned_matrix_rows(normalized)

     existing_ordered = [col for col in ordered_columns if col in normalized.columns]
     remaining = [col for col in normalized.columns if col not in existing_ordered]
     return drop_all_blank_columns(normalized[existing_ordered + remaining])


def auto_flatten_stacked_tables(raw_df: pd.DataFrame) -> pd.DataFrame:
    """Flatten common stacked mini-table layouts into metric/value rows.

    Looks for header rows where column 0 is a unit label (for example EUR mn)
    and later nonblank cells are metrics/periods. Rows below become line items
    until a run of blank rows or the next header.
    """
    records = []
    if raw_df is None or raw_df.empty:
        return pd.DataFrame()

    unit_prefixes = ("eur ", "usd ", "gbp ", "in %")
    i = 0
    while i < len(raw_df):
        row = raw_df.iloc[i]
        first = cell_text(row.iloc[0])
        first_lower = first.lower()
        metric_cols = nonblank_col_indexes(row, start_col=1)

        if first_lower.startswith(unit_prefixes) and metric_cols:
            unit = first
            table_name = nearest_title_above(raw_df, i)
            column_group = nearest_group_above(raw_df, i)
            metric_headers = {
                col_idx: cell_text(row.iloc[col_idx])
                for col_idx in metric_cols
            }
            j = i + 1
            blank_label_rows = 0
            while j < len(raw_df):
                data_row = raw_df.iloc[j]
                item = cell_text(data_row.iloc[0])
                item_lower = item.lower()
                if not item:
                    blank_label_rows += 1
                    if blank_label_rows >= 5:
                        break
                    j += 1
                    continue
                blank_label_rows = 0
                if item_lower.startswith(unit_prefixes):
                    break
                if not row_has_values(data_row, start_col=1):
                    j += 1
                    continue

                row_context = indented_row_context(raw_df, j, 0)
                for col_idx, metric in metric_headers.items():
                    value = data_row.iloc[col_idx]
                    if not is_blank(value):
                        record = {
                            "table_name": table_name,
                            "column_group": column_group,
                            "unit": unit,
                            "line_item": item,
                            "metric": metric,
                            "value": value,
                        }
                        record.update(row_context)
                        record.update(metric_context(metric))
                        records.append(record)
                j += 1
            i = j
            continue

        i += 1

    return pd.DataFrame(records)


def read_sheet(uploaded_file, sheet_name, header=None):
    uploaded_file.seek(0)
    return pd.read_excel(uploaded_file, sheet_name=sheet_name, header=header)


# Function-level cache for when Streamlit session_state isn't available
_quick_read_memory_cache = {}

def _quick_read_cached(file_bytes: bytes, sheet_name: str) -> pd.DataFrame:
    """Fast pandas read for structure detection, cached when possible."""
    cache_key = f"_quick_{hash(file_bytes)}_{sheet_name}"
    
    # Try Streamlit session_state first (best caching in production)
    try:
        if cache_key not in st.session_state:
            st.session_state[cache_key] = pd.read_excel(
                io.BytesIO(file_bytes), sheet_name=sheet_name, header=None
            )
        return st.session_state[cache_key]
    except Exception:
        # Fallback: Use function-level memory cache when session_state unavailable
        # (e.g., when running outside Streamlit, in tests, or offline)
        if cache_key not in _quick_read_memory_cache:
            _quick_read_memory_cache[cache_key] = pd.read_excel(
                io.BytesIO(file_bytes), sheet_name=sheet_name, header=None
            )
        return _quick_read_memory_cache[cache_key]


def infer_report_layout_from_quick_df(quick_df: pd.DataFrame) -> tuple[str, str, bool, bool]:
    """Infer whether a sheet looks like a general report block layout or matrix layout."""
    if quick_df is None or quick_df.empty:
        return "general", "Sheet is mostly empty in quick preview.", False, False

    has_report_signals = (
        quick_df.shape[1] > 0
        and any(
            str(v).lower().startswith(
                ("exchange rates", "valuation rates", "eur ", "usd ", "gbp ", "in %")
            )
            for v in quick_df.iloc[:, 0].tolist()
            if not is_blank(v)
        )
    )

    top_values = [
        str(v).strip()
        for row in quick_df.head(12).values
        for v in row
        if not is_blank(v)
    ]
    has_matrix_signals = any(v.upper() in {"#REF!", "#DIV/0!", "#N/A"} for v in top_values)

    first_col_texts = [cell_text(v).lower() for v in quick_df.iloc[:, 0].head(40).tolist()]
    top_left_texts = [
        cell_text(v).lower()
        for row in quick_df.head(12).iloc[:, : min(12, quick_df.shape[1])].values
        for v in row
    ]
    if any("matrix" in text for text in first_col_texts if text) or any("matrix" in text for text in top_left_texts if text):
        has_matrix_signals = True

    # Matrix sheets are often very wide and have many value-heavy rows with sparse label columns.
    row_sample = quick_df.head(min(len(quick_df), 120))
    first_col_nonblank_top = sum(1 for value in quick_df.iloc[:, 0].head(20).tolist() if cell_text(value))
    wide_sheet_signal = quick_df.shape[1] >= 40
    value_heavy_rows = 0
    for _, row in row_sample.iterrows():
        values = row.tolist()
        label_left = cell_text(values[0]) if values else ""
        metric_left = cell_text(values[1]) if len(values) > 1 else ""
        numeric_hits = sum(1 for value in values[2:] if looks_like_data_value(value))
        if numeric_hits >= 8 and (label_left or metric_left):
            value_heavy_rows += 1
    structural_matrix_signal = (
        wide_sheet_signal
        and value_heavy_rows >= 5
        and first_col_nonblank_top <= 3
    )
    if structural_matrix_signal:
        has_matrix_signals = True

    # Run expensive general header detection only when matrix signals are absent.
    if not has_report_signals and not has_matrix_signals:
        has_report_signals = sheet_looks_general_report_tables(quick_df)

    if has_matrix_signals:
        return "matrix", "Detected matrix/error-token style header signals.", has_report_signals, has_matrix_signals
    if has_report_signals:
        return "general", "Detected general report block signals.", has_report_signals, has_matrix_signals
    return "general", "No strong report signals detected.", has_report_signals, has_matrix_signals


@st.cache_data(show_spinner=False)
def _extract_report_cached(file_bytes: bytes, sheet_name: str, extraction_profile: str = "auto") -> pd.DataFrame:
    # Always use openpyxl for proper merged cell handling; rely on extraction strategy optimization instead
    raw_df = _read_display_sheet_cached(file_bytes, sheet_name)
    return auto_flatten_report_tables(raw_df, extraction_profile=extraction_profile)


def read_display_sheet(uploaded_file, sheet_name) -> pd.DataFrame:
    """Read an Excel sheet using the values as they are displayed in Excel."""
    assert hasattr(uploaded_file, "name") and hasattr(uploaded_file, "seek"), "Expected a file-like object"
    suffix = Path(uploaded_file.name).suffix.lower()
    if suffix not in {".xlsx", ".xlsm"}:
        result = pd.read_excel(uploaded_file, sheet_name=sheet_name, header=None)
        return pd.DataFrame(result) if not isinstance(result, pd.DataFrame) else result

    uploaded_file.seek(0)
    file_bytes = uploaded_file.read()
    return _read_display_sheet_cached(file_bytes, sheet_name)


@st.cache_data(show_spinner=False)
def _read_display_sheet_cached(file_bytes: bytes, sheet_name: str) -> pd.DataFrame:
    """Cached inner implementation — keyed on file content + sheet name."""
    # Speed optimization: only read the active sheet and skip unused features
    workbook = openpyxl.load_workbook(
        io.BytesIO(file_bytes), 
        data_only=True,
        keep_vba=False,  # Don't load VBA/macros
        rich_text=False  # Don't parse rich text formatting
    )
    worksheet = workbook[sheet_name]
    
    # Optimize: only read the actual used range, not the entire worksheet
    if worksheet.dimensions:
        # Parse dimensions string like "A1:F100"
        dims = worksheet.dimensions
        if dims and ":" in dims:
            start_cell, end_cell = dims.split(":")
            # Extract row/col from end_cell
            import re
            match = re.match(r"([A-Z]+)(\d+)", end_cell)
            if match:
                max_col_letter, max_row = match.groups()
                max_row = int(max_row)
                # Convert column letter to number
                max_col = 0
                for char in max_col_letter:
                    max_col = max_col * 26 + (ord(char) - ord('A') + 1)
            else:
                max_row = worksheet.max_row
                max_col = worksheet.max_column
        else:
            max_row = worksheet.max_row
            max_col = worksheet.max_column
    else:
        max_row = worksheet.max_row
        max_col = worksheet.max_column
    
    # Build a row-indexed merged-range map (height-based), then resolve cells lazily.
    # This avoids materializing every cell inside wide merged ranges.
    row_merge_index: dict[int, list[tuple[int, int, object]]] = {}
    if worksheet.merged_cells.ranges:
        for merged_range in worksheet.merged_cells.ranges:
            source_cell = worksheet.cell(merged_range.min_row, merged_range.min_col)
            for row_idx in range(merged_range.min_row, merged_range.max_row + 1):
                row_merge_index.setdefault(row_idx, []).append(
                    (merged_range.min_col, merged_range.max_col, source_cell)
                )
        for row_idx in row_merge_index:
            row_merge_index[row_idx].sort(key=lambda item: item[0])

    resolved_cell_cache: dict[tuple[int, int], object] = {}

    def _resolve_actual_cell(cell):
        key = (cell.row, cell.column)
        if key in resolved_cell_cache:
            return resolved_cell_cache[key]

        intervals = row_merge_index.get(cell.row)
        if intervals:
            col_idx = cell.column
            for min_col, max_col, source_cell in intervals:
                if min_col <= col_idx <= max_col:
                    resolved_cell_cache[key] = source_cell
                    return source_cell
                if col_idx < min_col:
                    break

        resolved_cell_cache[key] = cell
        return cell

    rows = []
    indents = []
    # Only iterate the used range, not the entire worksheet
    for row_idx, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col),
        start=1,
    ):
        row_data = []
        row_indents = []
        for cell in row:
            actual_cell = _resolve_actual_cell(cell)
            row_data.append(formatted_excel_value(actual_cell))
            row_indents.append(
                float(actual_cell.alignment.indent or 0)
                if actual_cell.alignment else 0
            )
        rows.append(row_data)
        indents.append(row_indents)
    
    df = pd.DataFrame(rows)
    df.attrs["excel_indents"] = indents
    workbook.close()  # Explicitly close to free memory
    return df


def _column_match_alias(name: str) -> str:
    normalized = clean_column_name(name)
    aliases = {
        "lineitem": "line_item",
        "line_item_name": "line_item",
        "metric_dt": "metric_date",
        "metricdt": "metric_date",
        "metricdate": "metric_date",
        "metric_yr": "metric_year",
        "metricyr": "metric_year",
        "metricyear": "metric_year",
        "metric_qtr": "metric_quarter",
        "metricqtr": "metric_quarter",
        "metricquarter": "metric_quarter",
        "metric_mth": "metric_month",
        "metricmth": "metric_month",
        "metricmonth": "metric_month",
        "businessline": "business_line",
        "business_line_name": "business_line",
    }
    return aliases.get(normalized, normalized)


def _column_similarity(left: str, right: str) -> float:
    if left == right:
        return 1.0
    left_tokens = {token for token in left.split("_") if token}
    right_tokens = {token for token in right.split("_") if token}
    if not left_tokens or not right_tokens:
        return SequenceMatcher(None, left, right).ratio()
    overlap = len(left_tokens.intersection(right_tokens))
    union = len(left_tokens.union(right_tokens))
    token_score = overlap / union if union else 0.0
    text_score = SequenceMatcher(None, left, right).ratio()
    return max(text_score, (0.55 * text_score) + (0.45 * token_score))


def _best_canonical_column(name: str, canonical_columns: list[str]) -> str:
    normalized = _column_match_alias(name)
    if normalized in canonical_columns:
        return normalized

    best_name = ""
    best_score = 0.0
    for candidate in canonical_columns:
        score = _column_similarity(normalized, candidate)
        if score > best_score:
            best_name = candidate
            best_score = score

    # Strict threshold prevents unrelated columns from merging.
    return best_name if best_score >= 0.92 else normalized


def _coalesce_series(left: pd.Series, right: pd.Series) -> pd.Series:
    mask = left.apply(is_blank)
    return left.where(~mask, right)


def _drop_sparse_noise_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame() if df is None else df

    keep = []
    row_count = max(len(df), 1)
    for col in df.columns:
        series = df[col]
        non_blank_values = [v for v in series.tolist() if not is_blank(v)]
        non_blank = len(non_blank_values)
        density = non_blank / row_count
        normalized = clean_column_name(col)
        looks_noise = normalized.startswith("unnamed") or normalized.startswith("col_")

        if looks_noise:
            if non_blank == 0:
                continue

            data_like = 0
            for value in non_blank_values:
                text = cell_text(value)
                if looks_like_data_value(text) or metric_context(text).get("metric_type"):
                    data_like += 1
            data_ratio = data_like / max(non_blank, 1)

            unique_ratio = len({cell_text(v) for v in non_blank_values}) / max(non_blank, 1)
            sparse_small = non_blank <= max(3, int(0.03 * row_count))
            sparse_mid = density < 0.12

            # Drop unnamed spillover columns if they are sparse and mostly non-analytic text.
            if (sparse_small and data_ratio < 0.5) or (sparse_mid and data_ratio < 0.25 and unique_ratio > 0.7):
                continue

        keep.append(col)

    return pd.DataFrame(df.loc[:, keep])


def _is_note_like_text(value) -> bool:
    text = cell_text(value).lower()
    if not text:
        return False

    prefixes = (
        "note",
        "notes",
        "note:",
        "notes:",
        "source",
        "disclaimer",
        "see note",
        "refer to note",
        "for further information",
        "for more information",
        "of which",
        "thereof",
        "includes",
        "including",
        "excluding",
        "1)",
        "2)",
        "3)",
        "4)",
        "*",
    )
    if text.startswith(prefixes):
        return True

    if re.match(r"^\(?\d{1,3}\)?[).:\-]\s", text):
        return True

    if re.match(r"^\*+\s*", text):
        return True

    if "see note" in text or "refer to note" in text:
        return True

    if "for further information" in text or "for more information" in text:
        return True

    return False


def _row_data_signal_count(row: pd.Series) -> int:
    count = 0
    for value in row.tolist():
        if is_blank(value):
            continue
        text = cell_text(value)
        if looks_like_data_value(text) or metric_context(text).get("metric_type"):
            count += 1
    return count


def _drop_note_like_rows(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame() if df is None else df

    text_cols = [
        col for col in ["line_item", "metric", "metric_detail", "section", "table_name"]
        if col in df.columns
    ]
    if not text_cols:
        return df

    keep_mask = []
    for _, row in df.iterrows():
        has_note_signal = any(_is_note_like_text(row.get(col, "")) for col in text_cols)
        strong_data_signal = _row_data_signal_count(row) >= 2
        keep_mask.append(not (has_note_signal and not strong_data_signal))

    return pd.DataFrame(df.loc[keep_mask])


@st.cache_data(show_spinner=False)
def is_sheet_already_flat(file_bytes: bytes, sheet_name: str) -> bool:
    """
    Pre-flight flat-file detection — FAST check (no extraction).

    Returns True if sheet appears to be already flat (no report/matrix signals).
    Returns False if extraction might be beneficial.

    This function only reads a preview, making it ~100x faster than full extraction.
    """
    try:
        # Quick preview read only
        quick_df = _quick_read_cached(file_bytes, sheet_name)

        # Infer layout using the same detection logic as extract_and_flatten_sheet
        _, _, has_report_signals, has_matrix_signals = infer_report_layout_from_quick_df(quick_df)

        # If NO report or matrix signals detected, the sheet is already flat
        # Return True (skip extraction), False otherwise (do extraction)
        return not (has_report_signals or has_matrix_signals)
    except Exception as e:
        # On error, default to False (do extraction to be safe)
        print(f"[is_sheet_already_flat] Error checking sheet '{sheet_name}': {e}")
        return False


def finalize_extracted_sheet(df: pd.DataFrame, strip_text: bool = True, split_hierarchy: bool = True) -> pd.DataFrame:
    """Apply final post-processing transformations to an extracted/flattened sheet.
    
    Includes:
    - Trimming whitespace in text columns
    - Splitting hierarchy columns on | and dash delimiters (removes original pipe-delimited columns)
    - Cleaning up blank columns and duplicates
    """
    if df is None or df.empty:
        return df
    
    cleaned = df.copy()
    
    # Trim whitespace in text columns
    if strip_text:
        for col in cleaned.columns:
            if not (pd.api.types.is_object_dtype(cleaned[col].dtype) or pd.api.types.is_string_dtype(cleaned[col].dtype)):
                continue
            cleaned[col] = cleaned[col].astype(str).str.strip()

    # Apply spillover filters for both extracted and fallback/raw-read paths.
    cleaned = _drop_misaligned_matrix_rows(cleaned)
    cleaned = _drop_numeric_pipe_spillover_rows(cleaned)
    
    # Track columns that contain pipe delimiters BEFORE splitting
    cols_with_pipes = set()
    if split_hierarchy:
        for col in cleaned.columns:
            text_series = cleaned[col].map(cell_text)
            if bool(text_series.str.contains(r"\|", regex=True).any()):
                cols_with_pipes.add(str(col))
    
    # Apply hierarchy column splitting (| and dash delimiters)
    if split_hierarchy:
        cleaned = _auto_split_hierarchy_columns(cleaned, dash_split_mode="spaced")
        # Remove original columns only when split columns were successfully created.
        cols_to_remove = []
        for source_col in cols_with_pipes:
            split_prefix = clean_column_name(f"{source_col}_part_")
            has_split_cols = any(str(col).startswith(split_prefix) for col in cleaned.columns)
            if has_split_cols and source_col in cleaned.columns:
                cols_to_remove.append(source_col)
        if cols_to_remove:
            cleaned = cleaned.drop(columns=cols_to_remove)
    
    # Final cleanup: remove blank and duplicate columns
    cleaned = drop_fully_duplicate_columns(cleaned)
    return drop_all_blank_columns(cleaned).reset_index(drop=True)


def extract_and_flatten_sheet(file_bytes: bytes, sheet_name: str, preferred_profile: str = "auto") -> pd.DataFrame:
    """Extract and flatten complex Excel sheets, or return quickly for already-flat sheets.
    
    For flat sheets: Returns immediately with just a pd.read_excel() call (< 1 second).
    For complex sheets: Runs the full extraction pipeline with UI feedback.
    """
    # Pre-flight check: if sheet is already flat, skip ALL extraction and UI
    if is_sheet_already_flat(file_bytes, sheet_name):
        flat_df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name)
        return finalize_extracted_sheet(pd.DataFrame(flat_df), strip_text=True, split_hierarchy=True)
    
    # For complex sheets that need extraction, run the full pipeline
    quick_df = _quick_read_cached(file_bytes, sheet_name)
    hinted_profile = (preferred_profile or "auto").strip().lower()
    if hinted_profile not in {"auto", "general", "matrix"}:
        hinted_profile = "auto"

    auto_profile, _, has_report_signals, has_matrix_signals = infer_report_layout_from_quick_df(quick_df)
    extraction_profile = hinted_profile if hinted_profile != "auto" else auto_profile

    if hinted_profile == "auto" and not has_report_signals and not has_matrix_signals:
        flat_df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name)
        return finalize_extracted_sheet(pd.DataFrame(flat_df), strip_text=True, split_hierarchy=True)

    extracted_df = _extract_report_cached(file_bytes, sheet_name, extraction_profile=extraction_profile)
    if extracted_df is None or extracted_df.empty:
        fallback_df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name)
        return finalize_extracted_sheet(pd.DataFrame(fallback_df), strip_text=True, split_hierarchy=True)
    return finalize_extracted_sheet(pd.DataFrame(extracted_df), strip_text=True, split_hierarchy=True)


def _extract_like_one_sheet_mode(file_bytes: bytes, sheet_name: str) -> pd.DataFrame:
    """Mirror the default one-sheet extraction behavior for batch modes.

    This path intentionally avoids finalize_extracted_sheet() so multi-sheet modes
    retain the same detail level users see when processing one sheet in report mode.
    """
    quick_df = _quick_read_cached(file_bytes, sheet_name)
    suggested_layout, _, has_report_signals, has_matrix_signals = infer_report_layout_from_quick_df(quick_df)

    if not has_report_signals and not has_matrix_signals:
        return pd.DataFrame(pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name))

    extracted_df = _extract_report_cached(file_bytes, sheet_name, extraction_profile=suggested_layout)
    if extracted_df is None or extracted_df.empty:
        return pd.DataFrame(pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name))
    return pd.DataFrame(extracted_df)


def merge_sheets(sheet_dataframes: dict[str, pd.DataFrame], drop_note_rows: bool = True) -> pd.DataFrame:
    """Merge multiple sheet dataframes with minimal data loss.
    
    Preserves all columns and rows from all sheets. Only performs:
    1. Basic concatenation
    2. Optional note row filtering (if enabled)
    3. Column reordering for readability
    """
    if not sheet_dataframes:
        return pd.DataFrame()

    aligned_frames: list[pd.DataFrame] = []
    for sheet_name, df in sheet_dataframes.items():
        if df is None or df.empty:
            continue

        working = df.copy()
        # Add source tracking columns without aggressive column harmonization
        working["source_sheet"] = sheet_name
        aligned_frames.append(working)

    if not aligned_frames:
        return pd.DataFrame()

    # Simple concatenation - preserve all columns and rows from all sheets
    merged = pd.concat(aligned_frames, ignore_index=True, sort=False)
    
    # Only apply note filtering if explicitly requested
    if drop_note_rows:
        merged = _drop_note_like_rows(merged)

    # Reorder columns for readability (but keep all columns)
    preferred_order = [
        "source_sheet",
        "table_name",
        "section",
        "block_id",
        "column_group",
        "unit",
        "line_item",
        "business_line",
        "metric",
        "metric_detail",
        "metric_date",
        "metric_year",
        "metric_month",
        "metric_quarter",
        "value",
    ]
    
    # Put preferred columns first, then all remaining columns in original order
    ordered = [col for col in preferred_order if col in merged.columns]
    tail = [col for col in merged.columns if col not in ordered]
    
    return merged[ordered + tail].reset_index(drop=True)


def _auto_split_hierarchy_columns(
    df: pd.DataFrame,
    max_parts: int = 12,
    dash_split_mode: str = "spaced",
) -> pd.DataFrame:
    """Expand common hierarchy delimiters into dedicated `<column>_*_part_n` columns."""
    if df is None or df.empty or max_parts < 2:
        return pd.DataFrame() if df is None else df

    expanded = df.copy()

    def _unique_name(base_name: str) -> str:
        if base_name not in expanded.columns:
            return base_name
        idx = 2
        while f"{base_name}_{idx}" in expanded.columns:
            idx += 1
        return f"{base_name}_{idx}"

    delimiter_specs = [
        {
            "label": "part",
            "has_delim": lambda text: bool(re.search(r"\|", text)),
            "split": lambda text: re.split(r"\s*\|\s*", text),
        }
    ]
    mode = (dash_split_mode or "spaced").strip().lower()
    if mode == "spaced":
        delimiter_specs.append(
            {
                "label": "dash_part",
                "has_delim": lambda text: bool(re.search(r"\s+[-–—]\s+", text)),
                "split": lambda text: re.split(r"\s+[-–—]\s+", text),
            }
        )
    elif mode == "any":
        delimiter_specs.append(
            {
                "label": "dash_part",
                "has_delim": lambda text: bool(re.search(r"[-–—]", text)),
                "split": lambda text: re.split(r"\s*[-–—]\s*", text),
            }
        )

    for source_col in list(df.columns):
        text_values = expanded[source_col].map(cell_text)

        for spec in delimiter_specs:
            label = spec["label"]
            if not text_values.map(spec["has_delim"]).any():
                continue

            token_lists = text_values.map(
                lambda text: [part.strip() for part in spec["split"](text) if part.strip()]
                if spec["has_delim"](text)
                else []
            )
            observed_max = token_lists.map(len).max()
            if pd.isna(observed_max) or int(observed_max) < 2:
                continue

            part_count = min(int(observed_max), max_parts)
            for idx in range(part_count):
                base_name = clean_column_name(f"{source_col}_{label}_{idx + 1}")
                new_col = _unique_name(base_name)
                expanded[new_col] = token_lists.map(lambda parts: parts[idx] if idx < len(parts) else "")

    return expanded


def apply_pandas_cleanup(
    df: pd.DataFrame,
    drop_columns,
    rename_map,
    split_config,
    drop_blank_columns,
    type_conversions,
    strip_text,
    clean_names,
    dash_split_mode,
) -> pd.DataFrame:
    cleaned = df.copy()
    if cleaned.empty:
        return cleaned

    if drop_columns:
        cleaned = cleaned.drop(columns=[c for c in drop_columns if c in cleaned.columns])

    if rename_map:
        cleaned = cleaned.rename(columns={
            old: new for old, new in rename_map.items()
            if old in cleaned.columns and new
        })

    if strip_text:
        for col in cleaned.select_dtypes(include="object").columns:
            cleaned[col] = cleaned[col].astype(str).str.strip()

    cleaned = _auto_split_hierarchy_columns(cleaned, dash_split_mode=dash_split_mode)

    if split_config and split_config.get("column") in cleaned.columns:
        source_col = split_config["column"]
        delimiter = split_config.get("delimiter") or " "
        max_parts = int(split_config.get("max_parts") or 2)
        prefix = split_config.get("prefix") or source_col
        keep_original = split_config.get("keep_original", True)
        parts = cleaned[source_col].astype(str).str.split(
            delimiter,
            n=max_parts - 1,
            expand=True,
            regex=False,
        )
        for idx in range(max_parts):
            new_col = clean_column_name(f"{prefix}_{idx + 1}")
            cleaned[new_col] = parts[idx] if idx in parts.columns else ""
        if not keep_original:
            cleaned = cleaned.drop(columns=[source_col])

    if drop_blank_columns:
        present = [c for c in drop_blank_columns if c in cleaned.columns]
        if present:
            mask = cleaned[present].apply(
                lambda row: all(is_blank(value) for value in row),
                axis=1,
            )
            cleaned = cleaned.loc[~mask].copy()

    for col, target_type in (type_conversions or {}).items():
        if col not in cleaned.columns or target_type == "keep":
            continue
        if target_type == "numeric":
            cleaned[col] = pd.to_numeric(cleaned[col], errors="coerce")
        elif target_type == "datetime":
            cleaned[col] = pd.to_datetime(cleaned[col], errors="coerce")
        elif target_type == "text":
            cleaned[col] = cleaned[col].astype(str)

    if clean_names:
        cleaned.columns = dedupe_columns(cleaned.columns)

    cleaned = drop_fully_duplicate_columns(cleaned)
    return drop_all_blank_columns(cleaned).reset_index(drop=True)


def to_excel_bytes(df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="flat_table", index=False)
    return output.getvalue()


def _read_csv_bytes(data: bytes) -> pd.DataFrame:
    result = pd.read_csv(io.BytesIO(data))  # type: ignore[call-overload]
    return pd.DataFrame(result)


def excel_mime_type(file_name: str) -> str:
    suffix = Path(file_name).suffix.lower()
    if suffix == ".xls":
        return "application/vnd.ms-excel"
    return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def render_raw_excel_download(uploaded_file, key: str):
    st.download_button(
        "Download Raw Excel",
        data=uploaded_file.getvalue(),
        file_name=uploaded_file.name,
        mime=excel_mime_type(uploaded_file.name),
        use_container_width=True,
        key=key,
    )


def render_cleanup_and_preview(df: pd.DataFrame, file_stem: str, key_prefix: str):
    """Apply optional pandas cleanup, then preview and download the result."""
    import hashlib

    scope = hashlib.md5(f"{key_prefix}_{file_stem}".encode()).hexdigest()[:8]
    kp = f"{key_prefix}_{scope}"

    st.markdown("### Pandas Cleanup")
    with st.expander("Edit the table with pandas-style operations", expanded=False):
        cleanup_cols = list(df.columns)
        drop_columns = st.multiselect(
            "Delete columns",
            cleanup_cols,
            help="Remove columns from the final output.",
            key=f"{kp}_drop_columns",
        )

        remaining_for_rename = [c for c in cleanup_cols if c not in drop_columns]
        rename_df = pd.DataFrame({
            "Column": remaining_for_rename,
            "New Name": remaining_for_rename,
        })
        edited_rename_df = st.data_editor(
            rename_df,
            hide_index=True,
            use_container_width=True,
            num_rows="fixed",
            key=f"{kp}_rename_columns_editor",
        )
        rename_map = {
            row["Column"]: row["New Name"]
            for _, row in edited_rename_df.iterrows()
            if row["Column"] != row["New Name"]
        }

        st.markdown("**Split a column**")
        split_enabled = st.checkbox(
            "Split one column into multiple columns",
            key=f"{kp}_split_enabled",
        )
        split_config = {}
        if split_enabled and remaining_for_rename:
            col_split_a, col_split_b, col_split_c = st.columns(3)
            with col_split_a:
                split_column = st.selectbox(
                    "Column to split",
                    remaining_for_rename,
                    key=f"{kp}_split_column",
                )
            with col_split_b:
                delimiter = st.text_input(
                    "Delimiter",
                    value=" | ",
                    key=f"{kp}_split_delimiter",
                )
            with col_split_c:
                max_parts = st.number_input(
                    "Number of output columns",
                    min_value=2,
                    max_value=12,
                    value=2,
                    key=f"{kp}_split_max_parts",
                )
            col_prefix, col_keep = st.columns([2, 1])
            with col_prefix:
                split_prefix = st.text_input(
                    "Output column prefix",
                    value=clean_column_name(split_column),
                    key=f"{kp}_split_prefix",
                )
            with col_keep:
                keep_original = st.checkbox(
                    "Keep original",
                    value=True,
                    key=f"{kp}_split_keep_original",
                )
            split_config = {
                "column": split_column,
                "delimiter": delimiter,
                "max_parts": max_parts,
                "prefix": split_prefix,
                "keep_original": keep_original,
            }

        drop_blank_columns = []
        type_conversions = {}

        col_strip, col_clean, col_dash = st.columns(3)
        with col_strip:
            strip_text = st.checkbox(
                "Trim whitespace in text columns",
                value=True,
                key=f"{kp}_strip_text",
            )
        with col_clean:
            clean_names = st.checkbox(
                "Clean final column names",
                value=False,
                key=f"{kp}_clean_names",
            )
        with col_dash:
            dash_split_mode = st.selectbox(
                "Auto-split '-' mode",
                options=["Off", "Spaced only ( - )", "Any dash (-)"],
                index=1,
                key=f"{kp}_dash_split_mode",
                help="Controls automatic dash splitting in hierarchy columns.",
            )
        dash_split_mode_map = {
            "Off": "off",
            "Spaced only ( - )": "spaced",
            "Any dash (-)": "any",
        }
        dash_split_mode_value = dash_split_mode_map[dash_split_mode]

    import hashlib as _hl

    cleanup_cache_version = "v2_hierarchy_split_regex"
    cleanup_sig = _hl.md5(
        str((
            cleanup_cache_version,
            list(df.columns), len(df),
            sorted(drop_columns),
            sorted(rename_map.items()),
            str(split_config),
            strip_text,
            clean_names,
            dash_split_mode_value,
        )).encode()
    ).hexdigest()
    cleanup_cache_key = f"{kp}_final_df_{cleanup_sig}"

    if cleanup_cache_key not in st.session_state:
        st.session_state[cleanup_cache_key] = apply_pandas_cleanup(
            df,
            drop_columns=drop_columns,
            rename_map=rename_map,
            split_config=split_config,
            drop_blank_columns=drop_blank_columns,
            type_conversions=type_conversions,
            strip_text=strip_text,
            clean_names=clean_names,
            dash_split_mode=dash_split_mode_value,
        )
    final_df = st.session_state[cleanup_cache_key]

    st.markdown("### Final Table Preview")
    col_rows, col_cols = st.columns(2)
    col_rows.metric("Rows", f"{len(final_df):,}")
    col_cols.metric("Columns", f"{len(final_df.columns):,}")
    st.dataframe(final_df.head(200), use_container_width=True)

    dl_key = f"{kp}_download"
    dl_state_key = f"{kp}_excel_bytes_{cleanup_sig}"

    def _prepare_download():
        st.session_state[dl_state_key] = to_excel_bytes(final_df)

    if dl_state_key not in st.session_state:
        st.button(
            "Prepare Download",
            on_click=_prepare_download,
            use_container_width=True,
            key=f"{kp}_prepare_btn",
        )
    else:
        st.download_button(
            "⬇️ Download Excel",
            data=st.session_state[dl_state_key],
            file_name=f"{file_stem}_flat.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key=dl_key,
        )

def render_unified_file_tab():
    st.subheader("Upload File")
    st.caption("Upload one or more CSV/Excel files. Each file can be processed with auto-detect flattening or as already-flat data.")
    uploaded_files = st.file_uploader(
        "Upload CSV or Excel files",
        type=["csv", "xlsx", "xls", "xlsm"],
        accept_multiple_files=True,
        key="unified_file_upload",
    )
    if not uploaded_files:
        st.info("Upload at least one file to auto-detect its structure.")
        return

    if len(uploaded_files) == 1:
        uploaded_file = uploaded_files[0]
    else:
        name_to_file = {f.name: f for f in uploaded_files}
        selected_name = st.selectbox(
            "Choose file to process",
            options=list(name_to_file.keys()),
            key="unified_selected_file",
        )
        uploaded_file = name_to_file[selected_name]

    assert not isinstance(uploaded_file, list), "Expected a single file"
    suffix = Path(uploaded_file.name).suffix.lower()
    if suffix == ".csv":
        df: pd.DataFrame = _read_csv_bytes(uploaded_file.read())
        source_label = Path(uploaded_file.name).stem
        st.caption("Detected mode: flat table (CSV)")
        st.markdown("### Uploaded Table Preview")
        st.dataframe(df.head(100), use_container_width=True)
        render_cleanup_and_preview(df, source_label, "unified_flat")
        return

    xls = pd.ExcelFile(uploaded_file)
    uploaded_file.seek(0)
    file_bytes = uploaded_file.read()

    if len(xls.sheet_names) > 1:
        workbook_mode = st.radio(
            "Workbook mode",
            [
                "Process one sheet",
                "Process selected sheets separately (recommended)",
                "Merge selected sheets into one flat file",
            ],
            index=1,
            key="unified_workbook_mode",
        )
        if workbook_mode == "Process selected sheets separately (recommended)":
            selected_sheets = st.multiselect(
                "Sheets to flatten (kept as separate outputs)",
                options=xls.sheet_names,
                default=xls.sheet_names,
                key="unified_separate_sheets",
            )
            if not selected_sheets:
                st.warning("Select at least one sheet.")
                return

            import hashlib as _hl
            split_sig = _hl.md5(
                file_bytes + f"|{'|'.join(selected_sheets)}|mode=separate".encode()
            ).hexdigest()
            split_cache_key = f"_separate_tabs_{split_sig}"

            if st.button("Build separate flat tables", key="unified_separate_build", use_container_width=True):
                output_frames = {}
                progress = st.progress(0)
                status = st.empty()
                for idx, selected_sheet in enumerate(selected_sheets):
                    status.text(f"Extracting {idx + 1}/{len(selected_sheets)}: {selected_sheet}")
                    try:
                        output_frames[selected_sheet] = _extract_like_one_sheet_mode(
                            file_bytes,
                            selected_sheet,
                        )
                    except Exception as exc:
                        st.warning(f"Failed to process sheet '{selected_sheet}': {exc}")
                    progress.progress((idx + 1) / len(selected_sheets))

                st.session_state[split_cache_key] = output_frames
                progress.empty()
                status.empty()

            split_outputs = st.session_state.get(split_cache_key)
            if split_outputs is None:
                st.info("Click 'Build separate flat tables' to flatten selected sheets without merging them.")
                return
            if not split_outputs:
                st.warning("No rows found from selected sheets.")
                return

            available_sheets = [name for name, frame in split_outputs.items() if frame is not None and not frame.empty]
            if not available_sheets:
                st.warning("All selected sheets returned empty outputs.")
                return

            st.success(f"Flattened {len(available_sheets)} sheet(s) as separate outputs.")
            preview_sheet = st.selectbox(
                "Preview output sheet",
                options=available_sheets,
                key="unified_separate_preview_sheet",
            )
            preview_df = split_outputs[preview_sheet]
            st.markdown("### Flattened Table Preview")
            st.dataframe(preview_df.head(150), use_container_width=True)
            source_label = f"{Path(uploaded_file.name).stem}_{preview_sheet}"
            render_cleanup_and_preview(preview_df, source_label, "unified_separate")
            return

        if workbook_mode == "Merge selected sheets into one flat file":
            selected_sheets = st.multiselect(
                "Sheets to merge",
                options=xls.sheet_names,
                default=xls.sheet_names,
                key="unified_merge_sheets",
            )
            drop_note_rows = st.checkbox(
                "Exclude note-like rows in merged output",
                value=True,
                help="Removes footnotes/disclaimer-style rows that usually do not carry data values.",
                key="unified_merge_drop_notes",
            )
            if not selected_sheets:
                st.warning("Select at least one sheet to merge.")
                return

            import hashlib as _hl
            merge_sig = _hl.md5(
                file_bytes + f"|{'|'.join(selected_sheets)}|drop_notes={drop_note_rows}".encode()
            ).hexdigest()
            merge_cache_key = f"_merged_tabs_{merge_sig}"

            if st.button("Build merged flat file", key="unified_merge_build", use_container_width=True):
                sheet_dataframes = {}
                progress = st.progress(0)
                status = st.empty()
                for idx, selected_sheet in enumerate(selected_sheets):
                    status.text(f"Extracting {idx + 1}/{len(selected_sheets)}: {selected_sheet}")
                    try:
                        sheet_dataframes[selected_sheet] = _extract_like_one_sheet_mode(
                            file_bytes,
                            selected_sheet,
                        )
                    except Exception as exc:
                        st.warning(f"Failed to process sheet '{selected_sheet}': {exc}")
                    progress.progress((idx + 1) / len(selected_sheets))

                status.text("Merging columns by similarity...")
                st.session_state[merge_cache_key] = merge_sheets(
                    sheet_dataframes,
                    drop_note_rows=drop_note_rows,
                )
                progress.empty()
                status.empty()

            merged_df = st.session_state.get(merge_cache_key)
            if merged_df is None:
                st.info("Click 'Build merged flat file' to generate one combined flat output.")
                return
            if merged_df.empty:
                st.warning("No rows found from selected sheets after merge.")
                return

            st.success(f"Merged {len(selected_sheets)} sheet(s) into {len(merged_df):,} rows.")
            st.caption("Sheets are concatenated with all columns preserved; source tab is in 'source_sheet'.")
            st.markdown("### Merged Table Preview")
            st.dataframe(merged_df.head(150), use_container_width=True)
            source_label = f"{Path(uploaded_file.name).stem}_merged_tabs"
            render_cleanup_and_preview(merged_df, source_label, "unified_merged_tabs")
            return

    # --- Single sheet or individual sheet processing ---
    sheet_name = st.selectbox("Sheet", xls.sheet_names, key="unified_sheet")
    assert isinstance(sheet_name, str), "Expected sheet_name to be a string"

    # --- Step 1: light quick read to detect structure (no heavy extraction yet) ---
    # Read with pandas first for instant structural heuristic check. Cache by bytes.
    quick_df = _quick_read_cached(file_bytes, sheet_name)
    suggested_layout, detection_reason, has_report_signals, has_matrix_signals = infer_report_layout_from_quick_df(quick_df)
    extraction_profile = suggested_layout

    if not has_report_signals and not has_matrix_signals:
        # Looks flat — skip extraction entirely, no delay
        suggested_mode_default = "flat"
        reason = "No report block signals detected. Treating as flat table."
    else:
        suggested_mode_default = "report"
        reason = f"{detection_reason} Using {suggested_layout} extraction profile."

    mode_options = {
        f"Auto ({suggested_mode_default})": suggested_mode_default,
        "Treat as flat table": "flat",
        "Treat as report blocks": "report",
    }
    selection = st.selectbox(
        "Parsing mode",
        list(mode_options.keys()),
        index=0,
        help="Auto mode is recommended; override only if needed.",
        key="unified_parsing_mode",
    )
    active_mode = mode_options[selection]
    st.caption(f"Detection note: {reason}")

    source_label = f"{Path(uploaded_file.name).stem}_{sheet_name}"

    if active_mode == "flat":
        flat_df: pd.DataFrame = pd.DataFrame(pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name))

        st.markdown("### Uploaded Table Preview")
        st.dataframe(flat_df.head(100), use_container_width=True)
        render_cleanup_and_preview(flat_df, source_label, "unified_flat")
        return

    # --- Step 2: report mode — run heavy extraction only when needed ---
    raw_df = _read_display_sheet_cached(file_bytes, sheet_name)

    st.markdown("### Raw Sheet Preview")
    st.dataframe(raw_df.head(40), use_container_width=True)

    with st.spinner("Extracting report blocks…"):
        extracted_df = _extract_report_cached(file_bytes, sheet_name, extraction_profile=extraction_profile)

    if extracted_df.empty:
        # Fall back to a direct flat read so users are not blocked on complex/matrix layouts.
        fallback_df: pd.DataFrame = pd.DataFrame(pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name))
        if fallback_df.empty:
            st.info(
                "File structure could not be auto-flattened and no flat rows were detected for this sheet.\n\n"
                "Try a different tab (for example a summary tab) or export only the data range and re-upload."
            )
            render_raw_excel_download(uploaded_file, "unified_raw_fallback")
            return

        st.warning(
            "Auto report extraction returned no rows for this sheet. "
            "Loaded the sheet in flat mode as a fallback."
        )
        st.markdown("### Flat Fallback Preview")
        st.dataframe(fallback_df.head(100), use_container_width=True)
        render_cleanup_and_preview(fallback_df, source_label, "unified_flat_fallback")
        return

    st.markdown("### Extracted Blocks Preview")
    col_rows, col_cols, col_blocks = st.columns(3)
    col_rows.metric("Rows", f"{len(extracted_df):,}")
    col_cols.metric("Columns", f"{len(extracted_df.columns):,}")
    col_blocks.metric(
        "Blocks",
        f"{extracted_df['block_id'].nunique():,}" if "block_id" in extracted_df.columns else "n/a",
    )
    st.dataframe(extracted_df.head(200), use_container_width=True)

    notes = extract_bottom_notes(raw_df)
    if notes:
        st.markdown("### Notes")
        st.text_area(
            "Extracted sheet notes",
            value="\n".join(notes),
            height=140,
            disabled=True,
            label_visibility="collapsed",
            key="unified_extracted_notes",
        )

    render_cleanup_and_preview(extracted_df, source_label, "unified_report")





def main():
    st.set_page_config(
        page_title="Flat File Builder",
        page_icon="",
        layout="wide",
    )
    st.title("Flat File Builder")
    st.caption("Upload once, auto-detect structure, then preview and clean the resulting table.")
    render_unified_file_tab()


if __name__ == "__main__":
    main()
