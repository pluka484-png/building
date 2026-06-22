import io
import json
import re
from dataclasses import dataclass
from difflib import SequenceMatcher
from datetime import date, datetime
from functools import lru_cache
from pathlib import Path
from typing import cast

import numpy as np
import openpyxl
import pandas as pd
import streamlit as st


# Compile hot regex patterns once.
RE_QUOTED = re.compile(r'"[^"]*"')
RE_BRACKETED = re.compile(r"\[[^]]+]")
RE_NUMERIC_BRACKETS = re.compile(r"[()\[\]]")
RE_PIPE_SPLIT = re.compile(r"\s*\|\s*")
RE_ALPHA = re.compile(r"[A-Za-z]")
RE_DIMENSIONS = re.compile(r"([A-Z]+)(\d+)")
RE_METRIC_FOOTNOTE_SUFFIX = re.compile(r"[¹²³⁴⁵⁶⁷⁸⁹⁰]+$")
RE_METRIC_TRAILING_ENUM = re.compile(r"\s+\d+\)$")
RE_MONTH_YEAR_TOKEN = re.compile(r"\b(jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec)\w*\s+\d{2,4}\b")
RE_DURATION_TOKEN = re.compile(r"\d{1,3}[YyMmWwDd]|[Oo][Nn]|[Tt][Nn]|[Ss][Ww]|[Ss][Nn]")
RE_PERIOD_TOKEN = re.compile(r"\b(q[1-4]|fy\s*\d{2,4}|ytd|mtd|qtd)\b")
RE_DATE_DMY = re.compile(r"^\d{1,2}[./-]\d{1,2}[./-]\d{2,4}$")
RE_DATE_YEARISH = re.compile(r"^(19|20)\d{2}([./-]\d{1,2}){0,2}$")
RE_UPPER_CODE = re.compile(r"[A-Z]{2,6}")
RE_HEADER_PERIOD_QUALIFIER = re.compile(r"^(?:[AN]\s*\d{4}|(?:19|20)\d{2}|Q[1-4]\s*\d{2,4}|FY\s*\d{2,4})$", re.IGNORECASE)
RE_NOTE_ENUM = re.compile(r"^\(?\d{1,3}\)?[).:\-]\s")
RE_NOTE_STAR = re.compile(r"^\*+\s*")
RE_NUMERIC_TOKEN = re.compile(r"^-?(?:\d+\.?\d*|\.\d+)(?:[eE][+-]?\d+)?$")
RE_DELTA_PREFIX = re.compile(r"(?:[∆Δ]|(?i:delta))\s*(.+)")
RE_SPLIT_ON_SLASH = re.compile(r"\s*/\s*")
RE_DOT_DATE_CAPTURE = re.compile(r"(\d{1,2})\.(\d{1,2})\.(\d{2,4})")

ERROR_TOKENS = {"#REF!", "#DIV/0!", "#N/A", "#VALUE!", "#NULL!", "#NUM!", "#ERROR!", "#NAME?"}
BLANK_TEXT_TOKENS = {"", "nan", "none"}
ROW_ATTR_KEYS = {
    "excel_indents",
    "excel_row_outline_levels",
    "excel_row_bold_flags",
    "excel_row_merged_flags",
    "excel_row_indent_levels",
}


@dataclass
class SheetScanContext:
    text: np.ndarray
    blank_mask: np.ndarray
    numeric_mask: np.ndarray
    unit_mask: np.ndarray
    data_like_mask: np.ndarray
    non_blank_count: np.ndarray
    first_non_blank_col: np.ndarray
    single_non_blank_col: np.ndarray
    row_outline_levels: np.ndarray
    row_bold_flags: np.ndarray
    row_merged_flags: np.ndarray
    row_indent_levels: np.ndarray


def _slice_with_row_attrs(df: pd.DataFrame, start: int, end: int) -> pd.DataFrame:
    sliced = df.iloc[start:end].reset_index(drop=True)
    attrs = getattr(df, "attrs", {})
    if not attrs:
        return sliced

    new_attrs = {}
    for key, value in attrs.items():
        if key == "_scan_context":
            continue
        if key in ROW_ATTR_KEYS and isinstance(value, (list, tuple, np.ndarray)):
            new_attrs[key] = list(value[start:end])
        else:
            new_attrs[key] = value
    sliced.attrs = new_attrs
    return sliced


@lru_cache(maxsize=65536)
def _normalize_cell_text(raw_text: str) -> str:
    compact = " ".join(raw_text.replace("\n", " ").split())
    return "" if compact.upper() in ERROR_TOKENS else compact


@lru_cache(maxsize=65536)
def _is_blank_text(raw_text: str) -> bool:
    return raw_text.strip().lower() in BLANK_TEXT_TOKENS


def is_blank(value) -> bool:
    if value is None or value is pd.NA:
        return True
    if isinstance(value, (float, np.floating)):
        if np.isnan(value):
            return True
    else:
        try:
            # NaN is the only value that is not equal to itself.
            if value != value:
                return True
        except Exception:
            pass
    return _is_blank_text(str(value))


def cell_text(value) -> str:
    if is_blank(value):
        return ""
    return _normalize_cell_text(str(value))


def _series_cell_text(series: pd.Series) -> pd.Series:
    """Vectorized cell_text equivalent for pandas Series."""
    text = (
        series.astype("string")
        .str.replace("\n", " ", regex=False)
        .str.split()
        .str.join(" ")
        .fillna("")
    )
    text = text.mask(text.str.upper().isin(ERROR_TOKENS), "")
    text = text.mask(text.str.lower().isin(BLANK_TEXT_TOKENS), "")
    return text


@lru_cache(maxsize=65536)
def _is_numeric_like_text(text: str) -> bool:
    cleaned = (
        text.replace(" ", "")
        .replace(",", "")
        .replace("%", "")
        .replace("+", "")
        .replace("−", "-")
    )
    cleaned = RE_NUMERIC_BRACKETS.sub("", cleaned)
    if not cleaned or cleaned in {"-", ".", "-."}:
        return False
    return bool(RE_NUMERIC_TOKEN.fullmatch(cleaned))


def clean_column_name(value) -> str:
    text = str(value).strip().lower()
    text = "".join(ch if ch.isalnum() else "_" for ch in text)
    while "__" in text:
        text = text.replace("__", "_")
    text = text.strip("_") or "unnamed_column"
    if text[0].isdigit():
        text = f"col_{text}"
    return text


def dedupe_columns(columns):
    counts = {}
    result = []
    for col in columns:
        base = clean_column_name(col)
        counts[base] = counts.get(base, 0) + 1
        result.append(base if counts[base] == 1 else f"{base}_{counts[base]}")
    return result


def primary_excel_format_section(number_format: str) -> str:
    section = str(number_format or "General").split(";")[0]
    section = RE_QUOTED.sub("", section)
    section = RE_BRACKETED.sub("", section)
    return section


def decimal_places_from_excel_format(number_format: str) -> int | None:
    section = primary_excel_format_section(number_format)
    if "." not in section:
        return 0 if any(token in section for token in ("0", "#", "?")) else None
    decimals = section.split(".", 1)[1]
    decimals = decimals.split("%", 1)[0]
    placeholders = [char for char in decimals if char in {"0", "#", "?"}]
    return len(placeholders) if placeholders else 0


def formatted_excel_value(cell):
    value = getattr(cell, "value", None)
    if value is None:
        return ""

    if isinstance(value, (datetime, date)):
        return value.strftime("%d.%m.%Y")

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        number_format = str(getattr(cell, "number_format", "General") or "General")
        decimals = decimal_places_from_excel_format(number_format)
        if decimals is None:
            return value

        is_percent = "%" in primary_excel_format_section(number_format)
        display_value = value * 100 if is_percent else value
        use_grouping = "," in number_format
        show_plus = "+" in number_format and display_value > 0
        sign = "+" if show_plus else ""
        grouped = "," if use_grouping else ""
        rendered = f"{sign}{display_value:{grouped}.{decimals}f}"
        if is_percent:
            rendered = f"{rendered}%"
        return rendered

    return value


def row_has_values(row, start_col=1) -> bool:
    return any(not is_blank(v) for v in row.iloc[start_col:].tolist())


def first_nonblank_after(row, start_col=1) -> str:
    for value in row.iloc[start_col:].tolist():
        if not is_blank(value):
            return cell_text(value)
    return ""


def is_unit_label(value) -> bool:
    text = cell_text(value).lower()
    return text == "%" or text.startswith(("eur ", "usd ", "gbp ", "in %"))


def nonblank_col_indexes(row, start_col=1):
    return [
        idx for idx in range(start_col, len(row))
        if not is_blank(row.iloc[idx])
    ]


def _excel_col_letter(col_index_1based: int) -> str:
    if col_index_1based <= 0:
        return ""
    label = ""
    value = int(col_index_1based)
    while value > 0:
        value, remainder = divmod(value - 1, 26)
        label = chr(ord("A") + remainder) + label
    return label


def _infer_header_rows_for_debug(df: pd.DataFrame, max_scan: int = 20) -> list[int]:
    if df is None or df.empty:
        return []

    context = _sheet_scan_context(df)
    row_limit = min(int(max_scan), len(df))
    header_rows: list[int] = []
    for row_idx in range(row_limit):
        non_blank = int(context.non_blank_count[row_idx])
        if non_blank == 0:
            continue
        numeric_hits = int(context.numeric_mask[row_idx].sum())
        text_hits = non_blank - numeric_hits
        if numeric_hits > max(3, text_hits):
            break
        header_rows.append(row_idx)
    return header_rows


def build_column_mapping_debug_df(raw_df: pd.DataFrame, max_header_rows: int = 20) -> pd.DataFrame:
    """Build a per-column map: current index -> original excel index + composite header."""
    if raw_df is None or raw_df.empty:
        return pd.DataFrame(
            columns=[
                "current_col_index",
                "excel_col_index",
                "excel_col_letter",
                "non_blank_cells",
                "is_all_blank",
                "header_composite",
                "header_tokens",
            ]
        )

    context = _sheet_scan_context(raw_df)
    col_count = raw_df.shape[1]

    col_map_attr = raw_df.attrs.get("excel_col_map")
    if isinstance(col_map_attr, list) and len(col_map_attr) == col_count:
        col_map = [int(v) for v in col_map_attr]
    else:
        col_map = list(range(col_count))

    header_rows = raw_df.attrs.get("excel_header_rows")
    if not isinstance(header_rows, list) or not header_rows:
        header_rows = _infer_header_rows_for_debug(raw_df, max_scan=max_header_rows)

    valid_header_rows = [
        int(row_idx)
        for row_idx in header_rows
        if isinstance(row_idx, (int, np.integer)) and 0 <= int(row_idx) < len(raw_df)
    ][: max(1, int(max_header_rows))]

    rows = []
    for current_idx in range(col_count):
        excel_idx0 = col_map[current_idx] if current_idx < len(col_map) else current_idx
        token_parts: list[str] = []
        for row_idx in valid_header_rows:
            token = str(context.text[row_idx, current_idx]).strip()
            if not token:
                continue
            if token_parts and token_parts[-1] == token:
                continue
            token_parts.append(token)

        rows.append(
            {
                "current_col_index": int(current_idx),
                "excel_col_index": int(excel_idx0),
                "excel_col_letter": _excel_col_letter(int(excel_idx0) + 1),
                "non_blank_cells": int((~context.blank_mask[:, current_idx]).sum()),
                "is_all_blank": bool(context.blank_mask[:, current_idx].all()),
                "header_composite": " | ".join(token_parts) if token_parts else f"_col_{int(excel_idx0) + 1}",
                "header_tokens": token_parts,
            }
        )

    return pd.DataFrame(rows)


def export_column_mapping_debug_csv(raw_df: pd.DataFrame, output_path: str, max_header_rows: int = 20) -> pd.DataFrame:
    debug_df = build_column_mapping_debug_df(raw_df, max_header_rows=max_header_rows)
    debug_df.to_csv(output_path, index=False)
    return debug_df


def _build_sheet_scan_context(raw_df: pd.DataFrame) -> SheetScanContext:
    values = raw_df.to_numpy(dtype=object, copy=False)
    row_count = values.shape[0]

    row_outline_levels = np.asarray(raw_df.attrs.get("excel_row_outline_levels") or [0] * row_count, dtype=np.int32)
    row_bold_flags = np.asarray(raw_df.attrs.get("excel_row_bold_flags") or [False] * row_count, dtype=bool)
    row_merged_flags = np.asarray(raw_df.attrs.get("excel_row_merged_flags") or [False] * row_count, dtype=bool)
    row_indent_levels = np.asarray(raw_df.attrs.get("excel_row_indent_levels") or [0] * row_count, dtype=np.int32)

    if row_outline_levels.shape[0] != row_count:
        row_outline_levels = np.resize(row_outline_levels, row_count)
    if row_bold_flags.shape[0] != row_count:
        row_bold_flags = np.resize(row_bold_flags, row_count)
    if row_merged_flags.shape[0] != row_count:
        row_merged_flags = np.resize(row_merged_flags, row_count)
    if row_indent_levels.shape[0] != row_count:
        row_indent_levels = np.resize(row_indent_levels, row_count)

    if values.size == 0:
        empty = np.empty(values.shape, dtype=object)
        false_mask = np.zeros(values.shape, dtype=bool)
        return SheetScanContext(
            text=empty,
            blank_mask=false_mask,
            numeric_mask=false_mask,
            unit_mask=false_mask,
            data_like_mask=false_mask,
            non_blank_count=np.zeros((values.shape[0],), dtype=np.int32),
            first_non_blank_col=np.full((values.shape[0],), -1, dtype=np.int32),
            single_non_blank_col=np.full((values.shape[0],), -1, dtype=np.int32),
            row_outline_levels=row_outline_levels,
            row_bold_flags=row_bold_flags,
            row_merged_flags=row_merged_flags,
            row_indent_levels=row_indent_levels,
        )

    flat_values = values.ravel()
    flat_series = pd.Series(flat_values, dtype="object")

    na_mask = pd.isna(flat_values)
    text_flat = np.empty(flat_values.shape[0], dtype=object)
    text_flat[:] = ""

    non_na_idx = np.flatnonzero(~na_mask)
    if non_na_idx.size > 0:
        text_non_na = (
            flat_series.iloc[non_na_idx]
            .astype(str)
            .str.replace("\n", " ", regex=False)
            .str.split()
            .str.join(" ")
        )
        text_non_na = text_non_na.mask(text_non_na.str.upper().isin(ERROR_TOKENS), "")
        text_flat[non_na_idx] = text_non_na.to_numpy(dtype=object)

    text_series = pd.Series(text_flat, dtype="string")
    blank_flat = text_series.eq("").to_numpy(dtype=bool)

    cleaned_numeric = (
        text_series
        .str.replace(" ", "", regex=False)
        .str.replace(",", "", regex=False)
        .str.replace("%", "", regex=False)
        .str.replace("+", "", regex=False)
        .str.replace("−", "-", regex=False)
        .str.replace(r"[()\[\]]", "", regex=True)
    )
    numeric_flat = cleaned_numeric.str.fullmatch(RE_NUMERIC_TOKEN).fillna(False).to_numpy(dtype=bool)

    lower_text = text_series.str.lower()
    unit_flat = (
        lower_text.eq("%")
        | lower_text.str.startswith(("eur ", "usd ", "gbp ", "in %"), na=False)
    ).to_numpy(dtype=bool)

    nm_like = lower_text.isin(["n.m.", "n.m", "nm"]).to_numpy(dtype=bool)
    data_like_flat = (numeric_flat | nm_like) & ~unit_flat

    text = text_flat.reshape(values.shape)
    blank_mask = blank_flat.reshape(values.shape)
    numeric_mask = numeric_flat.reshape(values.shape)
    unit_mask = unit_flat.reshape(values.shape)
    data_like_mask = data_like_flat.reshape(values.shape)
    non_blank_count = (~blank_mask).sum(axis=1).astype(np.int32)
    has_non_blank = non_blank_count > 0
    first_non_blank_col = np.full(values.shape[0], -1, dtype=np.int32)
    if has_non_blank.any():
        first_non_blank_col[has_non_blank] = (~blank_mask[has_non_blank]).argmax(axis=1)
    has_single = non_blank_count == 1
    single_non_blank_col = np.full(values.shape[0], -1, dtype=np.int32)
    if has_single.any():
        single_non_blank_col[has_single] = (~blank_mask[has_single]).argmax(axis=1)

    return SheetScanContext(
        text=text,
        blank_mask=blank_mask,
        numeric_mask=numeric_mask,
        unit_mask=unit_mask,
        data_like_mask=data_like_mask,
        non_blank_count=non_blank_count,
        first_non_blank_col=first_non_blank_col,
        single_non_blank_col=single_non_blank_col,
        row_outline_levels=row_outline_levels,
        row_bold_flags=row_bold_flags,
        row_merged_flags=row_merged_flags,
        row_indent_levels=row_indent_levels,
    )


def _sheet_scan_context(raw_df: pd.DataFrame) -> SheetScanContext:
    cached = raw_df.attrs.get("_scan_context")
    if cached is not None and isinstance(cached, SheetScanContext):
        return cached
    context = _build_sheet_scan_context(raw_df)
    raw_df.attrs["_scan_context"] = context
    return context


def _skip_error_heavy_rows(raw_df: pd.DataFrame, max_check: int = 30) -> pd.DataFrame:
    """Skip leading rows that are mostly error tokens or blank.

    Benelux-style sheets often have rows of #REF! errors before actual headers.
    This function identifies and skips such leading junk rows.
    """
    if raw_df is None or raw_df.empty:
        return raw_df

    error_tokens = {"#REF!", "#DIV/0!", "#N/A", "#VALUE!", "#NULL!", "#NUM!", "#ERROR!", "#NAME?"}

    def _row_error_density(row_idx: int) -> float:
        """Return fraction of cells that are error tokens or blank."""
        row = raw_df.iloc[row_idx]
        total = len(row)
        if total == 0:
            return 1.0
        errors_and_blanks = sum(
            1 for v in row
            if is_blank(v) or str(v).strip().upper() in error_tokens
        )
        return errors_and_blanks / total

    # Find first meaningful row. Prefer explicit leading labels so sparse title rows
    # are preserved (for example, section/table names above unit rows).
    skip_until = 0
    for idx in range(min(max_check, len(raw_df))):
        lead_limit = min(3, raw_df.shape[1])
        leading_tokens = [cell_text(raw_df.iloc[idx, col_idx]) for col_idx in range(lead_limit)]
        has_leading_content = any(token and not is_unit_label(token) for token in leading_tokens)
        if has_leading_content:
            skip_until = idx
            break
        if _row_error_density(idx) < 0.80:
            skip_until = idx
            break

    if skip_until > 0:
        return _slice_with_row_attrs(raw_df, skip_until, len(raw_df))
    return raw_df


def drop_all_blank_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame() if df is None else df
    context = _sheet_scan_context(df)
    keep_mask = (~context.blank_mask).any(axis=0)
    return df.loc[:, keep_mask]


def trim_effective_row_window(df: pd.DataFrame, tail_buffer: int = 3) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame() if df is None else df
    non_blank_rows = df.ne("").any(axis=1).to_numpy(dtype=bool)
    if not bool(non_blank_rows.any()):
        return _slice_with_row_attrs(df, 0, 0)

    idx = np.flatnonzero(non_blank_rows)
    start = int(idx[0])
    end = min(int(idx[-1]) + 1 + max(int(tail_buffer), 0), len(df))
    if start == 0 and end == len(df):
        return df
    return _slice_with_row_attrs(df, start, end)


def drop_fully_duplicate_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Remove columns whose values are semantically identical to an earlier column."""
    if df is None or df.empty:
        return pd.DataFrame() if df is None else df
    # Hash-based dedupe runs at C-level and avoids O(n*m) tuple hashing in Python.
    flat = pd.Series(df.to_numpy(dtype=object, copy=False).ravel(), dtype="object")
    canonical_flat = _series_cell_text(flat)
    canonical = pd.DataFrame(
        canonical_flat.to_numpy(dtype=object).reshape(df.shape),
        index=df.index,
        columns=df.columns,
    )
    hashed_cols = [pd.util.hash_pandas_object(canonical.iloc[:, idx], index=False) for idx in range(canonical.shape[1])]
    hashed = pd.DataFrame(hashed_cols).T
    duplicated = hashed.T.duplicated(keep="first")
    return df.loc[:, ~duplicated.to_numpy()]


def columns_semantically_equal(left: pd.Series, right: pd.Series) -> bool:
    if len(left) != len(right):
        return False
    for left_value, right_value in zip(left.tolist(), right.tolist()):
        if is_blank(left_value) and is_blank(right_value):
            continue
        if cell_text(left_value) != cell_text(right_value):
            return False
    return True


def looks_like_data_value(value) -> bool:
    text = cell_text(value)
    if not text or is_unit_label(text):
        return False
    if metric_context(text)["metric_type"]:
        return False
    cleaned = text.replace(",", "").replace("%", "").replace("+", "").replace("−", "-").strip()
    if cleaned.lower() in {"n.m.", "n.m", "nm"}:
        return True
    try:
        float(cleaned)
        return True
    except ValueError:
        return False


def is_section_header_row(row, expected_col: int | None = None, allow_hybrid: bool = True) -> bool:
    """Determine if a row is a section header, including hybrid header+data rows."""
    non_blank_indexes = [idx for idx, value in enumerate(row.tolist()) if not is_blank(value)]
    if not non_blank_indexes:
        return False

    col_idx = non_blank_indexes[0]
    value = row.iloc[col_idx]

    if expected_col is not None and col_idx != expected_col:
        return False

    text = cell_text(value)
    if not text:
        return False
    if _is_numeric_like(value):
        return False
    if is_unit_label(value):
        return False
    if looks_like_data_value(value):
        return False

    if len(non_blank_indexes) == 1:
        return True

    if not allow_hybrid:
        return False

    remaining = [row.iloc[idx] for idx in non_blank_indexes[1:]]
    return all(_is_numeric_like(v) or looks_like_data_value(v) for v in remaining)


def detect_hierarchy_level(raw_df: pd.DataFrame) -> np.ndarray:
    """Vectorized hierarchy level detection using non-blank position + Excel metadata."""
    context = _sheet_scan_context(raw_df)
    row_count = context.blank_mask.shape[0]
    if row_count == 0:
        return np.asarray([], dtype=np.int32)

    non_blank_mask = ~context.blank_mask
    levels = context.first_non_blank_col.copy()
    blank_rows = ~non_blank_mask.any(axis=1)

    # Blend structural depth with explicit Excel grouping/indent metadata when present.
    levels = np.maximum(levels, context.row_indent_levels)
    levels = np.maximum(levels, context.row_outline_levels)
    levels[blank_rows] = -1
    return levels.astype(np.int32)


def _section_header_confidence(context: SheetScanContext, allow_hybrid: bool = True) -> np.ndarray:
    non_blank = ~context.blank_mask
    row_count, _ = non_blank.shape
    confidence = np.zeros(row_count, dtype=np.int8)

    has_any = context.first_non_blank_col >= 0
    row_idx = np.arange(row_count)
    safe_first_col = np.where(has_any, context.first_non_blank_col, 0)

    first_numeric = context.numeric_mask[row_idx, safe_first_col]
    first_unit = context.unit_mask[row_idx, safe_first_col]
    first_data_like = context.data_like_mask[row_idx, safe_first_col]
    first_text_like = has_any & ~first_numeric & ~first_unit & ~first_data_like

    confidence += first_text_like.astype(np.int8)

    tail_valid_matrix = context.blank_mask | context.numeric_mask | context.data_like_mask
    tail_valid_matrix[row_idx[has_any], safe_first_col[has_any]] = True
    tail_numeric_or_data = tail_valid_matrix.all(axis=1)
    strict_single = context.non_blank_count == 1
    hybrid_signal = strict_single | (allow_hybrid & (context.non_blank_count > 1) & tail_numeric_or_data)
    confidence += hybrid_signal.astype(np.int8)

    depth_signal = ((context.first_non_blank_col > 0) | (context.row_indent_levels > 0)).astype(np.int8)
    confidence += depth_signal
    confidence += context.row_bold_flags.astype(np.int8)
    confidence += (context.row_outline_levels > 0).astype(np.int8)
    confidence += context.row_merged_flags.astype(np.int8)
    return confidence


def propagate_section_context(raw_df: pd.DataFrame, section_col: int = 0) -> pd.DataFrame:
    """Propagate section labels by hierarchy depth using vectorized masks and ffill."""
    if raw_df is None or raw_df.empty:
        return pd.DataFrame() if raw_df is None else raw_df.copy()

    context = _sheet_scan_context(raw_df)
    levels = detect_hierarchy_level(raw_df)
    max_depth = int(levels.max()) if levels.size else -1
    max_depth = min(max_depth, 8)
    if max_depth < 0:
        result = raw_df.copy()
        result["_section_depth"] = levels
        return result

    confidence = _section_header_confidence(context, allow_hybrid=True)
    non_blank = ~context.blank_mask
    row_idx = np.arange(len(raw_df))
    has_any = context.first_non_blank_col >= 0
    safe_first_col = np.where(has_any, context.first_non_blank_col, 0)

    first_text_like = has_any & ~context.numeric_mask[row_idx, safe_first_col] & ~context.unit_mask[row_idx, safe_first_col]
    header_mask = first_text_like & (confidence >= 2)

    result = raw_df.copy()
    result["_section_depth"] = levels
    result["line_item"] = np.where(has_any, context.text[row_idx, safe_first_col], "")

    for depth in range(max_depth + 1):
        col_name = f"_section_L{depth}"
        is_header_at_depth = header_mask & (levels == depth)
        values = np.where(is_header_at_depth, context.text[row_idx, safe_first_col], np.nan)
        labels = pd.Series(values, index=raw_df.index, dtype="object")
        result[col_name] = labels.ffill()

        group_name = f"_section_gid_L{depth}"
        result[group_name] = pd.Series(is_header_at_depth.astype(np.int32), index=raw_df.index).cumsum()

    return result


def detect_section_column(raw_df: pd.DataFrame, allow_hybrid: bool = True) -> int | None:
    """Identify the dominant section column across standalone and hybrid headers."""
    if raw_df is None or raw_df.empty:
        return None

    context = _sheet_scan_context(raw_df)
    row_count = context.blank_mask.shape[0]
    if row_count == 0:
        return None

    row_idx = np.arange(row_count)
    has_any = context.first_non_blank_col >= 0
    if not bool(has_any.any()):
        return None

    safe_first_col = np.where(has_any, context.first_non_blank_col, 0)
    first_text_like = (
        has_any
        & ~context.numeric_mask[row_idx, safe_first_col]
        & ~context.unit_mask[row_idx, safe_first_col]
        & ~context.data_like_mask[row_idx, safe_first_col]
    )

    tail_valid_matrix = context.blank_mask | context.numeric_mask | context.data_like_mask
    tail_valid_matrix[row_idx[has_any], safe_first_col[has_any]] = True
    tail_ok = tail_valid_matrix.all(axis=1)

    strict_mask = first_text_like & (context.non_blank_count == 1)
    hybrid_mask = first_text_like & (context.non_blank_count > 1) & tail_ok
    structural_mask = strict_mask | (hybrid_mask if allow_hybrid else np.zeros_like(hybrid_mask, dtype=bool))

    confidence = _section_header_confidence(context, allow_hybrid=allow_hybrid)
    candidate_mask = structural_mask & (confidence >= 2)
    candidate_cols = context.first_non_blank_col[candidate_mask]

    if candidate_cols.size == 0:
        return None

    counts = np.bincount(candidate_cols.astype(np.int32))
    return int(counts.argmax()) if counts.size else None


def _is_numeric_like(value) -> bool:
    """Quick check: can this value be parsed as numeric after removing formatting."""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return True
    text = cell_text(value)
    if not text:
        return False
    # LBYL path avoids expensive exception-driven float parsing on text-heavy sheets.
    return _is_numeric_like_text(text)


def _parse_numeric_value(value) -> float | None:
    """Parse a numeric cell value while tolerating display formatting."""
    if value is None or is_blank(value):
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)

    text = cell_text(value)
    if not text:
        return None

    negative = text.strip().startswith("(") and text.strip().endswith(")")
    cleaned = (
        text.strip()
        .replace("(", "")
        .replace(")", "")
        .replace(",", "")
        .replace("%", "")
        .replace("+", "")
        .replace("−", "-")
        .replace(" ", "")
    )
    try:
        parsed = float(cleaned)
    except ValueError:
        return None
    return -abs(parsed) if negative else parsed


def _looks_like_numeric_pipe_spillover(value) -> bool:
    """Detect numeric|text spillover tokens from matrix scratch calculations."""
    text = cell_text(value)
    if "|" not in text:
        return False
    parts = [part.strip() for part in RE_PIPE_SPLIT.split(text) if part.strip()]
    if len(parts) < 2:
        return False
    if not _is_numeric_like(parts[0]):
        return False
    return any(bool(RE_ALPHA.search(part)) for part in parts[1:])


def _drop_numeric_pipe_spillover_rows(df: pd.DataFrame) -> pd.DataFrame:
    """Drop rows containing numeric|text pipe spillover in any text-like column."""
    if df is None or df.empty:
        return pd.DataFrame() if df is None else df

    text_cols = [
        col
        for col in df.columns
        if pd.api.types.is_object_dtype(df[col].dtype) or pd.api.types.is_string_dtype(df[col].dtype)
    ]
    if not text_cols:
        return df

    spillover_pattern = r"^\s*[-+()\d.,%\s]+\|.*[A-Za-z]"
    spillover_mask = np.zeros(len(df), dtype=bool)
    for col in text_cols:
        series = df[col].astype("string")
        spillover_mask |= series.str.contains(spillover_pattern, regex=True, na=False).to_numpy(dtype=bool)
    return df.loc[~spillover_mask]


def _drop_misaligned_matrix_rows(df: pd.DataFrame) -> pd.DataFrame:
    """Drop rows that look like misaligned scratch/calculation spillover.

    Red flag: numeric value in one field + alphabetic text in another,
    indicating a spillover of calculation scratch rather than a valid data row.
    """
    if df is None or df.empty:
        return pd.DataFrame() if df is None else df
    required = {"line_item", "metric", "value"}
    if not required.issubset(set(df.columns)):
        return df

    line_series = df["line_item"].astype("string") if "line_item" in df.columns else pd.Series("", index=df.index, dtype="string")
    metric_series = df["metric"].astype("string") if "metric" in df.columns else pd.Series("", index=df.index, dtype="string")
    value_series = df["value"].astype("string") if "value" in df.columns else pd.Series("", index=df.index, dtype="string")

    def _numeric_mask(series: pd.Series) -> pd.Series:
        cleaned = (
            series
            .str.replace(" ", "", regex=False)
            .str.replace(",", "", regex=False)
            .str.replace("%", "", regex=False)
            .str.replace("+", "", regex=False)
            .str.replace("−", "-", regex=False)
            .str.replace(r"[()\[\]]", "", regex=True)
        )
        return cleaned.str.fullmatch(RE_NUMERIC_TOKEN).fillna(False)

    line_numeric = _numeric_mask(line_series)
    metric_numeric = _numeric_mask(metric_series)
    metric_alpha = metric_series.str.contains(RE_ALPHA, na=False)
    value_alpha = value_series.str.contains(RE_ALPHA, na=False)

    pipe_fields = ["line_item", "line_item_path", "parent_line_item", "metric", "metric_detail"]
    spillover_pattern = r"^\s*[-+()\d.,%\s]+\|.*[A-Za-z]"
    spillover_mask = np.zeros(len(df), dtype=bool)
    for field in pipe_fields:
        if field in df.columns:
            spillover_mask |= df[field].astype("string").str.contains(spillover_pattern, regex=True, na=False).to_numpy(dtype=bool)

    mixed_numeric_text = (line_numeric & metric_alpha) | (metric_numeric & value_alpha)
    value_numeric = _numeric_mask(value_series)
    value_nm = value_series.str.lower().isin(["n.m.", "n.m", "nm"])
    analytic_value = value_numeric | value_nm
    red_flag = spillover_mask | ((mixed_numeric_text & ~analytic_value).to_numpy(dtype=bool))
    return df.loc[~red_flag]


def extract_bottom_notes(raw_df: pd.DataFrame) -> list[str]:
    if raw_df is None or raw_df.empty:
        return []

    context = _sheet_scan_context(raw_df)
    data_row_mask = context.data_like_mask.sum(axis=1) >= 2
    if not data_row_mask.any():
        return []

    last_data_row = int(np.flatnonzero(data_row_mask)[-1])

    notes: list[str] = []
    for row_idx in range(last_data_row + 1, len(raw_df)):
        row_text = context.text[row_idx]
        parts = [part for part in row_text.tolist() if part]
        if not parts:
            continue
        note = " ".join(parts)
        if is_unit_label(note):
            continue
        notes.append(note)

    return notes


def find_unit_header_cells(raw_df: pd.DataFrame):
    cells = []
    if raw_df is None or raw_df.empty:
        return cells
    context = _sheet_scan_context(raw_df)
    unit_prefixes = ("eur ", "usd ", "gbp ", "in %")
    row_count, col_count = context.text.shape
    for row_idx in range(row_count):
        for col_idx in range(col_count):
            text = str(context.text[row_idx, col_idx]).lower()
            if not text:
                continue
            if not (context.unit_mask[row_idx, col_idx] or text.startswith(unit_prefixes)):
                continue
            if (~context.blank_mask[row_idx, col_idx + 1:]).any():
                cells.append((row_idx, col_idx))
    return cells


def find_general_header_cells(
        raw_df: pd.DataFrame,
        max_rows: int | None = None,
        max_cols: int | None = None,
        stop_after_first: bool = False,
):
    """Find likely mini-table header cells without relying on sheet-specific names."""
    cells = []
    if raw_df is None or raw_df.empty:
        return cells
    header_prefixes = (
        "eur ", "usd ", "gbp ", "chf ", "aud ", "in %", "vs.", "vs ",
        "as of ", "period", "quarter", "year",
    )
    row_limit = len(raw_df) - 1
    if max_rows is not None:
        row_limit = min(row_limit, max_rows)

    col_limit = raw_df.shape[1]
    if max_cols is not None:
        col_limit = min(col_limit, max_cols)

    for row_idx in range(row_limit):
        row = raw_df.iloc[row_idx]
        for col_idx, value in enumerate(row.iloc[:col_limit].tolist()):
            label = cell_text(value)
            if not label:
                continue
            header_cols = [
                idx for idx in nonblank_col_indexes(row, start_col=col_idx + 1)
                if idx < col_limit
            ]
            if len(header_cols) < 2:
                continue

            data_rows = 0
            for next_idx in range(row_idx + 1, min(len(raw_df), row_idx + 8)):
                next_row = raw_df.iloc[next_idx]
                line_item = cell_text(next_row.iloc[col_idx])
                if not line_item:
                    continue
                has_values = any(
                    col < len(next_row) and not is_blank(next_row.iloc[col])
                    for col in header_cols[:12]
                )
                if has_values:
                    data_rows += 1

            looks_like_unit = label.lower().startswith(header_prefixes)
            if looks_like_unit or data_rows >= 2:
                cells.append((row_idx, col_idx))
                if stop_after_first:
                    return cells
    return cells


def sheet_looks_general_report_tables(raw_df: pd.DataFrame) -> bool:
    return len(
        find_general_header_cells(
            raw_df,
            max_rows=120,
            max_cols=30,
            stop_after_first=True,
        )
    ) >= 1


def section_name(value) -> str:
    text = cell_text(value)
    text = text.replace("¹", "").replace("²", "").replace("³", "")
    return text.strip()


def section_name_validated(row, expected_col: int | None = None, allow_hybrid: bool = True) -> str:
    """Extract section text only when the row matches standalone header structure."""
    if not is_section_header_row(row, expected_col=expected_col, allow_hybrid=allow_hybrid):
        return ""

    for value in row.tolist():
        if not is_blank(value):
            text = cell_text(value)
            text = text.replace("¹", "").replace("²", "").replace("³", "")
            return text.strip()

    return ""


def metric_parse_text(value) -> str:
    text = cell_text(value)
    text = RE_METRIC_FOOTNOTE_SUFFIX.sub("", text).strip()
    text = RE_METRIC_TRAILING_ENUM.sub("", text).strip()
    return text


def carry_forward_header_values(values):
    carried = []
    current = ""
    for value in values:
        if not is_blank(value):
            current = cell_text(value)
        carried.append(current)
    return carried


def _detect_summary_header_rows(raw_df: pd.DataFrame, max_scan: int = 20) -> list[int]:
    if raw_df is None or raw_df.empty:
        return []
    context = _sheet_scan_context(raw_df)
    row_count, col_count = raw_df.shape
    header_rows: list[int] = []
    blank_seen = False

    for row_idx in range(min(max_scan, row_count)):
        non_blank = int(context.non_blank_count[row_idx])
        if non_blank == 0:
            if header_rows:
                blank_seen = True
            continue

        period_hits = 0
        financial_hits = 0
        for col_idx in range(col_count):
            token = str(context.text[row_idx, col_idx]).strip()
            if token and RE_HEADER_PERIOD_QUALIFIER.fullmatch(token):
                period_hits += 1
            if col_idx >= 2:
                parsed = _parse_numeric_value(raw_df.iat[row_idx, col_idx])
                if parsed is not None and abs(parsed) >= 100:
                    financial_hits += 1

        if financial_hits >= 2 and header_rows:
            break
        if blank_seen and financial_hits >= 1:
            break
        header_rows.append(row_idx)

        # Summary sheets usually have a compact 1-3 row header band.
        if period_hits >= 2 and len(header_rows) >= 2:
            continue

    return header_rows


def _detect_summary_data_region(raw_df: pd.DataFrame, max_scan: int = 250) -> tuple[int, int]:
    if raw_df is None or raw_df.empty:
        return -1, -1
    row_count, col_count = raw_df.shape
    candidates: list[tuple[int, int, int]] = []

    for row_idx in range(min(max_scan, row_count)):
        first_numeric = -1
        numeric_hits = 0
        for col_idx in range(col_count):
            if _parse_numeric_value(raw_df.iat[row_idx, col_idx]) is None:
                continue
            numeric_hits += 1
            if first_numeric < 0:
                first_numeric = col_idx

        if numeric_hits < 2 or first_numeric < 0:
            continue
        label_hits = sum(1 for col_idx in range(min(first_numeric, 10)) if cell_text(raw_df.iat[row_idx, col_idx]))
        if label_hits == 0:
            continue
        candidates.append((row_idx, first_numeric, numeric_hits))

    if not candidates:
        return -1, -1

    best_row = -1
    best_score = -1
    best_value_start = -1
    for row_idx, first_col, _ in candidates:
        window_end = row_idx + 20
        cluster = [
            cand_first
            for cand_row, cand_first, cand_hits in candidates
            if row_idx <= cand_row <= window_end and abs(cand_first - first_col) <= 2 and cand_hits >= 2
        ]
        score = len(cluster)
        if score > best_score:
            best_score = score
            best_row = row_idx
            best_value_start = int(np.median(cluster)) if cluster else first_col

    if best_row < 0 or best_value_start < 0:
        return -1, -1
    return best_row, best_value_start


def _detect_summary_header_band(raw_df: pd.DataFrame, data_start: int, data_col_start: int) -> list[int]:
    if raw_df is None or raw_df.empty or data_start <= 0:
        return []
    context = _sheet_scan_context(raw_df)
    col_count = raw_df.shape[1]

    # Prefer a top-anchored header band and stop at first separator blank row.
    top_rows: list[int] = []
    saw_non_blank = False
    for row_idx in range(min(data_start, len(raw_df))):
        non_blank = int(context.non_blank_count[row_idx])
        if non_blank == 0:
            if saw_non_blank and len(top_rows) >= 2:
                break
            continue

        saw_non_blank = True
        label_hits = sum(1 for col_idx in range(min(data_col_start, 10)) if cell_text(raw_df.iat[row_idx, col_idx]))
        numeric_hits = sum(
            1
            for col_idx in range(data_col_start, col_count)
            if _parse_numeric_value(raw_df.iat[row_idx, col_idx]) is not None
        )
        if label_hits >= 1 and numeric_hits >= 3:
            break
        top_rows.append(row_idx)

    if len(top_rows) >= 2:
        return top_rows

    rows: list[int] = []
    band_start = max(0, data_start - 15)
    for row_idx in range(band_start, data_start):
        if int(context.non_blank_count[row_idx]) == 0:
            continue
        label_hits = sum(1 for col_idx in range(min(data_col_start, 10)) if cell_text(raw_df.iat[row_idx, col_idx]))
        numeric_hits = sum(
            1
            for col_idx in range(data_col_start, col_count)
            if _parse_numeric_value(raw_df.iat[row_idx, col_idx]) is not None
        )
        # Exclude likely data rows from the header band.
        if label_hits >= 1 and numeric_hits >= 3:
            continue
        non_blank_data = sum(1 for col_idx in range(data_col_start, col_count) if cell_text(raw_df.iat[row_idx, col_idx]))
        if non_blank_data == 0:
            continue
        rows.append(row_idx)
    return rows


def _is_header_noise_token(token: str) -> bool:
    text = str(token or "").strip()
    if not text:
        return True
    lower = text.lower()
    if lower.startswith("as per "):
        return True
    if "review update" in lower or "bom update" in lower:
        return True
    if len(text) > 40 and RE_DOT_DATE_CAPTURE.search(text):
        return True
    return False


def _compact_summary_header_tokens(tokens: list[str], max_tokens: int = 4) -> list[str]:
    compact: list[str] = []
    for token in tokens:
        text = str(token or "").strip()
        if not text or _is_header_noise_token(text):
            continue

        # Drop measure-like numerics from header chains while keeping year tokens.
        if text.isdigit():
            number = int(text)
            if not (1900 <= number <= 2100):
                continue

        if compact and compact[-1] == text:
            continue
        compact.append(text)

    # Remove immediate repeated suffix patterns, e.g. A|B|C|X|B|C -> A|B|C|X
    if len(compact) >= 6:
        for size in range(2, min(4, len(compact) // 2 + 1)):
            left = compact[-2 * size: -size]
            right = compact[-size:]
            if left == right:
                compact = compact[:-size]
                break

    if len(compact) <= max_tokens:
        return compact
    return [*compact[: max_tokens - 1], compact[-1]]


def _build_summary_composite_headers(raw_df: pd.DataFrame, header_rows: list[int]) -> dict[int, str]:
    row_count, col_count = raw_df.shape
    if row_count == 0 or col_count == 0:
        return {}

    context = _sheet_scan_context(raw_df)
    carried_rows: list[list[str]] = []
    for row_idx in header_rows:
        if 0 <= row_idx < row_count:
            carried_rows.append(carry_forward_header_values(context.text[row_idx].tolist()))

    excel_col_map = raw_df.attrs.get("excel_col_map")
    if not isinstance(excel_col_map, list) or len(excel_col_map) != col_count:
        excel_col_map = list(range(col_count))

    headers: dict[int, str] = {}
    for col_idx in range(col_count):
        parts: list[str] = []
        for carried in carried_rows:
            token = str(carried[col_idx]).strip() if col_idx < len(carried) else ""
            if not token:
                continue
            if parts and parts[-1] == token:
                continue
            parts.append(token)

        parts = _compact_summary_header_tokens(parts, max_tokens=4)

        if parts:
            headers[col_idx] = " | ".join(parts)
        else:
            excel_idx0 = int(excel_col_map[col_idx]) if col_idx < len(excel_col_map) else col_idx
            headers[col_idx] = f"Column {_excel_col_letter(excel_idx0 + 1)}"
    return headers


def _looks_like_financial_summary_layout(raw_df: pd.DataFrame) -> bool:
    if raw_df is None or raw_df.empty or raw_df.shape[1] < 6 or raw_df.shape[0] < 6:
        return False

    data_start, data_col_start = _detect_summary_data_region(raw_df, max_scan=180)
    if data_start < 0 or data_col_start < 2:
        return False

    header_rows = _detect_summary_header_band(raw_df, data_start, data_col_start)
    if len(header_rows) < 2:
        header_rows = _detect_summary_header_rows(raw_df, max_scan=15)
    if len(header_rows) < 2:
        return False

    headers = _build_summary_composite_headers(raw_df, header_rows)
    period_header_count = 0
    for label in headers.values():
        tokens = [part.strip() for part in str(label).split("|") if part.strip()]
        if any(RE_HEADER_PERIOD_QUALIFIER.fullmatch(token) for token in tokens):
            period_header_count += 1
    if period_header_count < 2:
        return False

    # Require at least one row with a descriptive label and numeric payload in C+.
    sample_end = min(len(raw_df), data_start + 40)
    for row_idx in range(data_start, sample_end):
        row = raw_df.iloc[row_idx]
        label_hits = sum(1 for col_idx in range(min(data_col_start, 10)) if cell_text(row.iloc[col_idx]))
        numeric_hits = sum(1 for col_idx in range(data_col_start, raw_df.shape[1]) if _parse_numeric_value(row.iloc[col_idx]) is not None)
        if label_hits >= 1 and numeric_hits >= 1:
            return True
    return False


def auto_flatten_financial_summary_layout(raw_df: pd.DataFrame) -> pd.DataFrame:
    """Extract sparse financial summary matrices while preserving true column identity."""
    if raw_df is None or raw_df.empty:
        return pd.DataFrame()
    if not _looks_like_financial_summary_layout(raw_df):
        return pd.DataFrame()

    row_count, col_count = raw_df.shape
    if col_count < 3:
        return pd.DataFrame()

    data_start, data_col_start = _detect_summary_data_region(raw_df, max_scan=250)
    if data_start < 0 or data_col_start < 2:
        return pd.DataFrame()

    header_rows = _detect_summary_header_band(raw_df, data_start, data_col_start)
    if len(header_rows) < 2:
        header_rows = _detect_summary_header_rows(raw_df, max_scan=20)
    if len(header_rows) < 2:
        return pd.DataFrame()
    headers = _build_summary_composite_headers(raw_df, header_rows)

    if data_col_start >= col_count:
        return pd.DataFrame()

    value_cols: list[int] = []
    body_end = min(row_count, data_start + 300)
    for col_idx in range(data_col_start, col_count):
        numeric_hits = 0
        text_hits = 0
        for row_idx in range(data_start, body_end):
            value = raw_df.iat[row_idx, col_idx]
            if is_blank(value):
                continue
            if _parse_numeric_value(value) is not None:
                numeric_hits += 1
            else:
                text_hits += 1
        if numeric_hits >= 1 and numeric_hits >= text_hits:
            value_cols.append(col_idx)

    if not value_cols:
        return pd.DataFrame()

    records: list[dict] = []
    processed_rows: set[int] = set()
    current_parent = ""
    metric_meta_cache: dict[str, dict] = {}
    indents = raw_df.attrs.get("excel_indents") or []

    values = raw_df.to_numpy(copy=False)
    for row_idx in range(data_start, row_count):
        if row_idx in processed_rows:
            continue
        processed_rows.add(row_idx)

        row = values[row_idx]
        has_data = any(not is_blank(row[col_idx]) for col_idx in value_cols)
        if not has_data:
            continue

        left_tokens = [cell_text(row[col_idx]) for col_idx in range(min(data_col_start, col_count))]
        col_a = left_tokens[0] if left_tokens else ""
        col_b = left_tokens[1] if len(left_tokens) > 1 else ""
        primary_label = ""
        for token in left_tokens:
            if token:
                primary_label = token
                break

        if col_a and not is_unit_label(col_a):
            current_parent = col_a

        line_item = col_b or primary_label
        if not line_item:
            continue

        indent_level = 0
        if row_idx < len(indents) and indents[row_idx] and len(indents[row_idx]) > 0:
            indent_level = int(float(indents[row_idx][0] or 0))

        for col_idx in value_cols:
            value = row[col_idx]
            if is_blank(value):
                continue

            metric_name = headers.get(col_idx, f"Column {col_idx + 1}")
            if metric_name not in metric_meta_cache:
                metric_meta_cache[metric_name] = metric_context(metric_name)

            record = {
                "table_name": "Summary Grid",
                "section": current_parent,
                "column_group": metric_name,
                "unit": "",
                "line_item": line_item,
                "parent_line_item": current_parent if current_parent and current_parent != line_item else "",
                "line_item_path": f"{current_parent} > {line_item}" if current_parent and current_parent != line_item else line_item,
                "metric": metric_name,
                "metric_detail": col_b,
                "indent_level": indent_level,
                "value": value,
                "block_key": f"summary_layout:{data_start}",
                "block_start_column": data_col_start,
            }
            record.update(metric_meta_cache[metric_name])
            records.append(record)

    return pd.DataFrame(records)


def auto_flatten_sectioned_financial_sheet(raw_df: pd.DataFrame) -> pd.DataFrame:
    """Flatten stacked financial statement sections into a long analytical table.

    Handles patterns like:
    - Exchange rates / Spot or Average with periods across columns
    - Valuation rates with valuation dates spanning tenor columns
    """
    records = []
    i = 0
    while i < len(raw_df):
        row = raw_df.iloc[i]
        first = cell_text(row.iloc[0])
        first_lower = first.lower()
        rate_label = first_nonblank_after(row, start_col=1)

        if first_lower.startswith("exchange rates") and rate_label:
            rate_type = rate_label
            unit = cell_text(raw_df.iloc[i + 1, 0]) if i + 1 < len(raw_df) else ""
            headers = raw_df.iloc[i + 1].tolist() if i + 1 < len(raw_df) else []
            j = i + 2
            while j < len(raw_df):
                data_row = raw_df.iloc[j]
                label = cell_text(data_row.iloc[0])
                label_lower = label.lower()
                if (
                        not label
                        or label_lower.startswith("exchange rates")
                        or label_lower.startswith("valuation rates")
                        or label_lower.startswith("1)")
                        or label_lower.startswith("2)")
                        or label_lower.startswith("3)")
                ):
                    break
                for col_idx in range(1, len(data_row)):
                    value = data_row.iloc[col_idx]
                    period = cell_text(headers[col_idx]) if col_idx < len(headers) else ""
                    if period and not is_blank(value):
                        records.append({
                            "section": section_name(first),
                            "rate_type": rate_type,
                            "unit": unit,
                            "currency": label,
                            "period": period,
                            "valuation_date": "",
                            "tenor": "",
                            "contract_type": "",
                            "value": value,
                        })
                j += 1
            i = j
            continue

        if first_lower.startswith("valuation rates"):
            date_headers = carry_forward_header_values(row.tolist())
            tenor_row = raw_df.iloc[i + 1].tolist() if i + 1 < len(raw_df) else []
            unit = cell_text(tenor_row[0]) if tenor_row else ""
            current_contract = ""
            j = i + 2
            while j < len(raw_df):
                data_row = raw_df.iloc[j]
                label = cell_text(data_row.iloc[0])
                label_lower = label.lower()
                if (
                        not label
                        or label_lower.startswith("valuation rates")
                        or label_lower.startswith("exchange rates")
                        or label_lower.startswith("1)")
                        or label_lower.startswith("2)")
                        or label_lower.startswith("3)")
                ):
                    break

                if not row_has_values(data_row, start_col=1):
                    current_contract = label
                    j += 1
                    continue

                for col_idx in range(1, len(data_row)):
                    value = data_row.iloc[col_idx]
                    valuation_date = date_headers[col_idx] if col_idx < len(date_headers) else ""
                    tenor = cell_text(tenor_row[col_idx]) if col_idx < len(tenor_row) else ""
                    if valuation_date and tenor and not is_blank(value):
                        records.append({
                            "section": section_name(first),
                            "rate_type": "",
                            "unit": unit,
                            "currency": label,
                            "period": "",
                            "valuation_date": valuation_date,
                            "tenor": tenor,
                            "contract_type": current_contract,
                            "value": value,
                        })
                j += 1
            i = j
            continue

        i += 1

    return pd.DataFrame(records)


def nearest_title_above(raw_df: pd.DataFrame, row_idx: int) -> str:
    for idx in range(row_idx - 1, -1, -1):
        first = cell_text(raw_df.iloc[idx, 0])
        if not first:
            continue
        if first.lower().startswith(("eur ", "usd ", "gbp ", "in %")):
            continue
        if row_has_values(raw_df.iloc[idx], start_col=1):
            continue
        return first
    return ""


def nearest_group_above(raw_df: pd.DataFrame, row_idx: int) -> str:
    for idx in range(row_idx - 1, -1, -1):
        if cell_text(raw_df.iloc[idx, 0]):
            continue
        group = first_nonblank_after(raw_df.iloc[idx], start_col=1)
        if group:
            return group
    return ""


def _is_col0_unit_anchor(value) -> bool:
    text = cell_text(value).lower()
    return text in {"eur", "eur mn", "eur bn", "usd", "usd mn", "usd bn", "gbp", "gbp mn", "gbp bn"} or text.startswith(("eur ", "usd ", "gbp ", "in %"))


def _resolve_block_header_from_unit_anchor(raw_df: pd.DataFrame, unit_row_idx: int) -> tuple[str, str]:
    """Resolve (table_name, column_group) by scanning upward from a unit anchor row.

    Expected pattern:
    - unit row at col 0 (e.g., EUR mn)
    - qualifier row immediately above with text in col 2
    - next non-empty col 0 row above qualifier is section/table name
    """
    if raw_df is None or raw_df.empty or unit_row_idx <= 0:
        return "", ""

    qualifier_row_idx = unit_row_idx - 1
    column_group = ""
    if qualifier_row_idx >= 0 and raw_df.shape[1] > 2:
        qualifier_row = raw_df.iloc[qualifier_row_idx]
        col2_text = cell_text(qualifier_row.iloc[2])
        non_blank = [idx for idx, value in enumerate(qualifier_row.tolist()) if not is_blank(value)]
        if col2_text and (non_blank == [2] or (0 not in non_blank and 2 in non_blank)):
            column_group = col2_text

    scan_idx = qualifier_row_idx - 1 if column_group else qualifier_row_idx
    while scan_idx >= 0:
        candidate = cell_text(raw_df.iloc[scan_idx, 0])
        if candidate:
            return candidate, column_group
        scan_idx -= 1

    return "", column_group


def nearest_left_label_above(raw_df: pd.DataFrame, row_idx: int, col_idx: int) -> str:
    for idx in range(row_idx - 1, -1, -1):
        label = cell_text(raw_df.iloc[idx, col_idx])
        if label and not label.replace(" ", "").isupper():
            return label
    return ""


def nearest_nonblank_right(raw_df: pd.DataFrame, row_idx: int, col_idx: int, block_end: int) -> str:
    row = raw_df.iloc[row_idx]
    for idx in range(col_idx + 1, min(block_end, len(row))):
        value = cell_text(row.iloc[idx])
        if value:
            return value
    return ""


def cell_indent(raw_df: pd.DataFrame, row_idx: int, col_idx: int) -> float:
    indents = raw_df.attrs.get("excel_indents") or []
    if row_idx >= len(indents) or col_idx >= len(indents[row_idx]):
        return 0
    return indents[row_idx][col_idx] or 0


def indented_row_context(raw_df: pd.DataFrame, row_idx: int, label_col: int) -> dict:
    current_indent = cell_indent(raw_df, row_idx, label_col)
    if current_indent <= 0:
        return {}

    parents = []
    next_indent = current_indent
    for idx in range(row_idx - 1, -1, -1):
        label = cell_text(raw_df.iloc[idx, label_col])
        if not label:
            continue
        indent = cell_indent(raw_df, idx, label_col)
        if indent < next_indent:
            parents.append(label)
            next_indent = indent
            if indent <= 0:
                break

    parents.reverse()
    if not parents:
        return {}

    line_item = cell_text(raw_df.iloc[row_idx, label_col])
    return {
        "parent_line_item": parents[-1],
        "line_item_path": " > ".join([*parents, line_item]),
    }


def metric_context(metric: str) -> dict:
    text = metric_parse_text(metric)
    text = text.strip(" ,;:()[]")
    empty_context = {
        "metric_type": "",
        "metric_date": "",
        "comparison_date": "",
        "metric_quarter": "",
        "comparison_year": "",
    }

    def normalize_year(year_text: str) -> int:
        year = int(year_text)
        return year + 2000 if len(year_text) == 2 else year

    def quarter_from_month(month: int) -> int:
        return ((month - 1) // 3) + 1

    def parse_dot_date_token(date_text: str) -> tuple[int, int, int] | None:
        token = date_text.strip(" ,;:()[]")
        date_match_local = RE_DOT_DATE_CAPTURE.fullmatch(token)
        if not date_match_local:
            date_match_local = RE_DOT_DATE_CAPTURE.search(token)
        if not date_match_local:
            return None
        day_text, month_text, year_text = date_match_local.groups()
        return normalize_year(year_text), int(month_text), int(day_text)

    delta_parts: list[str] = []
    delta_prefix_match = RE_DELTA_PREFIX.match(text)
    if delta_prefix_match:
        delta_body = delta_prefix_match.group(1).strip()
        delta_parts = [part.strip() for part in RE_SPLIT_ON_SLASH.split(delta_body, maxsplit=1)]
        start_date_parts = parse_dot_date_token(delta_parts[0]) if delta_parts else None
        end_date_parts = parse_dot_date_token(delta_parts[1]) if len(delta_parts) == 2 else None
        if (not start_date_parts or not end_date_parts) and len(delta_parts) == 2:
            # Fallback for noisy labels like "31.12.25, / (31.12.24)".
            date_tokens = RE_DOT_DATE_CAPTURE.findall(delta_body)
            if len(date_tokens) >= 2:
                start_date_parts = (
                    normalize_year(date_tokens[0][2]),
                    int(date_tokens[0][1]),
                    int(date_tokens[0][0]),
                )
                end_date_parts = (
                    normalize_year(date_tokens[1][2]),
                    int(date_tokens[1][1]),
                    int(date_tokens[1][0]),
                )

        if start_date_parts and end_date_parts:
            start_year, start_month, start_day = start_date_parts
            end_year, end_month, end_day = end_date_parts
            return {
                **empty_context,
                "metric_type": "date_delta",
                "metric_date": pd.Timestamp(start_year, start_month, start_day),
                "comparison_date": pd.Timestamp(end_year, end_month, end_day),
            }

    date_match = RE_DOT_DATE_CAPTURE.fullmatch(text)
    if date_match:
        day, month, year = date_match.groups()
        normalized_year = normalize_year(year)
        month_int = int(month)
        return {
            **empty_context,
            "metric_type": "date",
            "metric_date": pd.Timestamp(normalized_year, month_int, int(day)),
            "metric_quarter": quarter_from_month(month_int),
        }

    # ISO date/datetime (e.g. 2024-03-31 or 2024-03-31 00:00:00)
    iso_dt_match = re.fullmatch(
        r"(20\d{2})-(\d{1,2})-(\d{1,2})(?:[ T]\d{1,2}:\d{2}(?::\d{2})?)?",
        text,
    )
    if iso_dt_match:
        year_text, month_text, day_text = iso_dt_match.groups()
        year = int(year_text)
        month = int(month_text)
        day = int(day_text)
        return {
            **empty_context,
            "metric_type": "date",
            "metric_date": pd.Timestamp(year, month, day),
            "metric_quarter": quarter_from_month(month),
        }

    # Month period tokens (e.g. 12M24, 12M 2024, 2024 12M)
    month_period_patterns = [
        r"(0?[1-9]|1[0-2])M\s*(\d{2,4})",
        r"(\d{2,4})\s*(0?[1-9]|1[0-2])M",
    ]
    month_period_match = None
    year = None
    month = None
    for idx, pattern in enumerate(month_period_patterns):
        month_period_match = re.fullmatch(pattern, text, flags=re.IGNORECASE)
        if month_period_match:
            if idx == 0:
                month = int(month_period_match.group(1))
                year = normalize_year(month_period_match.group(2))
            else:
                year = normalize_year(month_period_match.group(1))
                month = int(month_period_match.group(2))
            break

    if month_period_match and year is not None and month is not None:
        metric_date = pd.Timestamp(year, month, 1) + pd.offsets.MonthEnd(0)  # type: ignore[operator]
        return {
            **empty_context,
            "metric_type": "month",
            "metric_date": metric_date,
        }

    quarter_match = re.fullmatch(r"([1-4])Q\s*(\d{2,4})", text)
    if quarter_match:
        year_text = quarter_match.group(2)
        year = int(year_text) + 2000 if len(year_text) == 2 else int(year_text)
        return {
            **empty_context,
            "metric_type": "quarter",
            "metric_date": pd.Timestamp(year, int(quarter_match.group(1)) * 3, 1) + pd.offsets.MonthEnd(0),
            # type: ignore[operator]

        }

    year_match = re.fullmatch(r"(20\d{2})", text)
    if year_match:
        year = int(year_match.group(1))
        return {
            **empty_context,
            "metric_type": "year",
            "metric_date": pd.Timestamp(year, 12, 31),
        }

    if len(delta_parts) == 2:
        year_match = re.fullmatch(r"20\d{2}", delta_parts[0])
        comparison_year_match = re.fullmatch(r"20\d{2}", delta_parts[1])
        if year_match and comparison_year_match:
            year = int(delta_parts[0])
            comparison_year = int(delta_parts[1])
            return {
                **empty_context,
                "metric_type": "delta",
                "metric_date": pd.Timestamp(year, 12, 31),
                "comparison_date": pd.Timestamp(comparison_year, 12, 31),
                "comparison_year": comparison_year,
            }

    return empty_context


def header_row_score(row) -> int:
    return sum(
        1
        for value in row.tolist()[1:]
        if cell_text(value) and not metric_context(cell_text(value))["metric_type"] and not is_unit_label(value)
    )


def best_group_header_row(raw_df: pd.DataFrame, unit_row_idx: int) -> list[str]:
    best_idx = None
    best_score = 0
    for idx in range(max(0, unit_row_idx - 4), unit_row_idx):
        score = header_row_score(raw_df.iloc[idx])
        if score > best_score:
            best_idx = idx
            best_score = score
    if best_idx is None:
        return [""] * raw_df.shape[1]
    return carry_forward_header_values(raw_df.iloc[best_idx].tolist())


def best_label_column(raw_df: pd.DataFrame, unit_row_idx: int, unit_cols: list[int]) -> int:
    if not unit_cols:
        return 0

    best_col = 0
    best_score = -1
    for col_idx in range(min(unit_cols)):
        score = 0
        for row_idx in range(unit_row_idx + 1, len(raw_df)):
            row = raw_df.iloc[row_idx]
            label = cell_text(row.iloc[col_idx])
            if not label or is_unit_label(label):
                continue
            if any(not is_blank(row.iloc[value_col]) for value_col in unit_cols):
                score += 1
        if score > best_score:
            best_col = col_idx
            best_score = score

    return best_col


def auto_flatten_grouped_metric_blocks(raw_df: pd.DataFrame) -> pd.DataFrame:
    """Flatten wide report tables with row labels and repeated metric groups."""
    if raw_df is None or raw_df.empty:
        return pd.DataFrame()

    records = []
    for unit_row_idx in range(1, len(raw_df)):
        unit_row = raw_df.iloc[unit_row_idx]
        unit_cols = [
            col_idx for col_idx in range(1, raw_df.shape[1])
            if is_unit_label(unit_row.iloc[col_idx])
        ]
        if len(unit_cols) < 4:
            continue

        label_col = best_label_column(raw_df, unit_row_idx, unit_cols)
        period_row = raw_df.iloc[unit_row_idx - 1]
        period_headers = carry_forward_header_values(period_row.tolist())
        detail_headers = raw_df.iloc[unit_row_idx - 2].tolist() if unit_row_idx >= 2 else []
        group_headers = best_group_header_row(raw_df, unit_row_idx)
        table_name = nearest_title_above(raw_df, unit_row_idx)
        if not table_name:
            table_name = first_nonblank_after(raw_df.iloc[max(0, unit_row_idx - 4)], start_col=1)
        if _is_col0_unit_anchor(unit_row.iloc[0]):
            anchored_table, _ = _resolve_block_header_from_unit_anchor(raw_df, unit_row_idx)
            if anchored_table:
                table_name = anchored_table

        row_idx = unit_row_idx + 1
        blank_label_rows = 0
        while row_idx < len(raw_df):
            data_row = raw_df.iloc[row_idx]
            line_item = cell_text(data_row.iloc[label_col])
            if sum(1 for value in data_row.tolist() if is_unit_label(value)) >= 4:
                break
            if not line_item:
                blank_label_rows += 1
                if blank_label_rows >= 5:
                    break
                row_idx += 1
                continue
            blank_label_rows = 0
            if is_unit_label(line_item):
                break
            if all(is_blank(data_row.iloc[col_idx]) for col_idx in unit_cols):
                row_idx += 1
                continue

            row_context = indented_row_context(raw_df, row_idx, label_col)
            for col_idx in unit_cols:
                value = data_row.iloc[col_idx]
                if is_blank(value):
                    continue
                metric = cell_text(period_headers[col_idx])
                if not metric:
                    metric = cell_text(group_headers[col_idx])
                metric_detail = cell_text(detail_headers[col_idx]) if col_idx < len(detail_headers) else ""
                if metric_detail in {metric, cell_text(group_headers[col_idx])}:
                    metric_detail = ""
                record = {
                    "table_name": table_name,
                    "section": table_name,
                    "column_group": cell_text(group_headers[col_idx]),
                    "unit": cell_text(unit_row.iloc[col_idx]),
                    "line_item": line_item,
                    "metric": metric,
                    "metric_detail": metric_detail,
                    "value": value,
                    "block_key": f"{unit_row_idx}:{cell_text(group_headers[col_idx])}",
                    "block_start_column": col_idx,
                }
                record.update(row_context)
                record.update(metric_context(metric))
                records.append(record)
            row_idx += 1

    return pd.DataFrame(records)


def auto_flatten_report_blocks(raw_df: pd.DataFrame) -> pd.DataFrame:
    """Extract repeated side-by-side report blocks into long rows."""
    unit_cells = find_unit_header_cells(raw_df)
    if not unit_cells:
        return pd.DataFrame()

    records = []
    blocks_by_row: dict[int, list[int]] = {}
    for row_idx, col_idx in unit_cells:
        blocks_by_row.setdefault(row_idx, []).append(col_idx)

    for header_row_idx, starts in blocks_by_row.items():
        starts = sorted(starts)
        for pos, start_col in enumerate(starts):
            start_col = int(start_col)
            end_col = int(starts[pos + 1]) if pos + 1 < len(starts) else raw_df.shape[1]
            header_row = raw_df.iloc[header_row_idx]
            unit = cell_text(header_row.iloc[start_col])
            metric_cols = [
                col for col in range(start_col + 1, end_col)
                if not is_blank(header_row.iloc[col])
            ]
            if not metric_cols:
                continue

            table_name = nearest_left_label_above(raw_df, header_row_idx, start_col)
            section = cell_text(raw_df.iloc[header_row_idx - 1, start_col]) if header_row_idx > 0 else ""
            column_group = nearest_nonblank_right(
                raw_df,
                header_row_idx - 1,
                start_col,
                end_col,
            ) if header_row_idx > 0 else ""

            if start_col == 0 and _is_col0_unit_anchor(header_row.iloc[0]):
                anchored_table, anchored_group = _resolve_block_header_from_unit_anchor(raw_df, header_row_idx)
                if anchored_table:
                    table_name = anchored_table
                if anchored_group:
                    column_group = anchored_group

            row_idx = header_row_idx + 1
            blank_label_rows = 0
            while row_idx < len(raw_df):
                data_row = raw_df.iloc[row_idx]
                line_item = cell_text(data_row.iloc[start_col])
                if not line_item:
                    blank_label_rows += 1
                    if blank_label_rows >= 5:
                        break
                    row_idx += 1
                    continue
                blank_label_rows = 0
                if line_item.lower().startswith(("eur ", "usd ", "gbp ", "in %")):
                    break
                if not row_has_values(data_row, start_col=start_col + 1):
                    row_idx += 1
                    continue

                row_context = indented_row_context(raw_df, row_idx, start_col)
                for metric_col in metric_cols:
                    metric = cell_text(header_row.iloc[metric_col])
                    value = data_row.iloc[metric_col]
                    if metric and not is_blank(value):
                        record = {
                            "table_name": table_name,
                            "section": section,
                            "column_group": column_group,
                            "unit": unit,
                            "line_item": line_item,
                            "metric": metric,
                            "value": value,
                            "block_key": f"{header_row_idx}:{start_col}",
                            "block_start_column": start_col,
                        }
                        record.update(row_context)
                        record.update(metric_context(metric))
                        records.append(record)
                row_idx += 1

    return pd.DataFrame(records)


def auto_flatten_horizontal_balance_sheet(raw_df: pd.DataFrame) -> pd.DataFrame:
    """Flatten balance-sheet tabs where business segments are side-by-side in 20-col blocks."""
    if raw_df is None or raw_df.empty or raw_df.shape[1] < 20:
        return pd.DataFrame()

    date_columns = [
        "31.03.2024",
        "30.06.2024",
        "30.09.2024",
        "31.12.2024",
        "31.03.2025",
        "30.06.2025",
        "30.09.2025",
        "31.12.2025",
    ]
    value_offsets = [2, 4, 6, 8, 10, 12, 14, 16]
    delta_offset = 18
    block_width = 20

    if raw_df.shape[1] < (min(value_offsets) + 1):
        return pd.DataFrame()
    block_starts = [
        start_col
        for start_col in range(0, raw_df.shape[1], block_width)
        if start_col + max(value_offsets) < raw_df.shape[1]
    ]

    records = []
    valid_block_count = 0
    for start_col in block_starts:
        # Signature checks for this very specific layout.
        title = cell_text(raw_df.iloc[0, start_col]) if len(raw_df) > 0 else ""
        section_seed = cell_text(raw_df.iloc[3, start_col]) if len(raw_df) > 3 else ""
        segment_name = cell_text(raw_df.iloc[3, start_col + 2]) if len(raw_df) > 3 and start_col + 2 < raw_df.shape[1] else ""
        unit = cell_text(raw_df.iloc[4, start_col]) if len(raw_df) > 4 else ""

        has_balance_title = "balance sheet" in title.lower()
        section_seed_compact = re.sub(r"\s+", "", section_seed.lower())
        has_assets_seed = "assets" in section_seed_compact
        has_unit_seed = unit.lower().startswith(("eur ", "usd ", "gbp ", "in %"))

        # Require a plausible segment block signature so we don't hijack unrelated sheets.
        if not ((has_balance_title or has_assets_seed) and (segment_name or has_unit_seed)):
            continue

        valid_block_count += 1
        current_section = "ASSETS"
        if "liabilities" in section_seed_compact:
            current_section = "LIABILITIES AND EQUITY"

        for row_idx in range(5, len(raw_df)):
            row = raw_df.iloc[row_idx]
            line_item = cell_text(row.iloc[start_col])
            if not line_item:
                continue

            compact = re.sub(r"\s+", "", line_item.lower())
            if "assets" in compact and len(compact) <= 24:
                current_section = "ASSETS"
                continue
            if "liabilities" in compact and "equity" in compact:
                current_section = "LIABILITIES AND EQUITY"
                continue

            if is_unit_label(line_item):
                continue
            if line_item.lower().startswith("consolidated balance sheet"):
                continue
            if line_item.lower().startswith("by business segments"):
                continue

            row_has_value = False
            for metric, offset in zip(date_columns, value_offsets):
                col_idx = start_col + offset
                if col_idx >= raw_df.shape[1]:
                    continue
                value = row.iloc[col_idx]
                if is_blank(value):
                    continue
                row_has_value = True
                record = {
                    "table_name": segment_name or "Balance Sheet",
                    "section": current_section,
                    "column_group": segment_name,
                    "unit": unit,
                    "line_item": line_item,
                    "metric": metric,
                    "value": value,
                    "block_key": f"horizontal_bs:{start_col}",
                    "block_start_column": start_col,
                }
                record.update(metric_context(metric))
                records.append(record)

            delta_col = start_col + delta_offset
            if delta_col < raw_df.shape[1]:
                delta_value = row.iloc[delta_col]
                if not is_blank(delta_value):
                    row_has_value = True
                    delta_metric = "Delta 31.12.25 / 31.12.24"
                    record = {
                        "table_name": segment_name or "Balance Sheet",
                        "section": current_section,
                        "column_group": segment_name,
                        "unit": unit,
                        "line_item": line_item,
                        "metric": delta_metric,
                        "value": delta_value,
                        "block_key": f"horizontal_bs:{start_col}",
                        "block_start_column": start_col,
                    }
                    record.update(metric_context(delta_metric))
                    records.append(record)

            # Skip decorative label-only rows that don't carry values.
            if not row_has_value:
                continue

    # Need multiple matching blocks; otherwise treat as non-match and let other extractors run.
    if valid_block_count < 2 or not records:
        return pd.DataFrame()
    return pd.DataFrame(records)


def _fallback_matrix_flatten(raw_df: pd.DataFrame) -> pd.DataFrame:
    """Flatten matrix layouts with stacked headers and interior numeric grids."""
    if raw_df is None or raw_df.empty or raw_df.shape[0] < 2 or raw_df.shape[1] < 3:
        return pd.DataFrame()

    # Skip error-heavy leading rows (benelux sheets have many #REF! errors at top)
    raw_df = _skip_error_heavy_rows(raw_df, max_check=50)
    raw_df = drop_all_blank_columns(raw_df).reset_index(drop=True)

    # Column pruning changes positional indexes; drop stale indent metadata.
    raw_df.attrs.pop("excel_indents", None)

    section_col = detect_section_column(raw_df)

    matrix_values = raw_df.to_numpy(copy=False)
    row_count, col_count = matrix_values.shape

    # Pre-cache cell_text and looks_like_data_value for all cells to avoid redundant calls
    cell_cache = {}
    data_value_cache = {}
    metric_meta_cache = {}

    def _get_cached_cell_text(row_idx: int, col_idx: int) -> str:
        key = (row_idx, col_idx)
        if key not in cell_cache:
            cell_cache[key] = cell_text(matrix_values[row_idx, col_idx])
        return cell_cache[key]

    def _is_cached_data_value(row_idx: int, col_idx: int) -> bool:
        key = (row_idx, col_idx)
        if key not in data_value_cache:
            value = matrix_values[row_idx, col_idx]
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                data_value_cache[key] = True
            else:
                data_value_cache[key] = looks_like_data_value(value)
        return data_value_cache[key]

    def _row_numeric_hits(row_idx: int, start_col: int = 0) -> int:
        return sum(1 for col in range(start_col, col_count) if _is_cached_data_value(row_idx, col))

    # Find rows that look like data lines (some text context + multiple values).
    # Scan only first 200 rows to avoid scanning entire sheet for huge files
    scan_limit = min(row_count, 200)
    candidate_rows = [
        row_idx for row_idx in range(scan_limit)
        if _row_numeric_hits(row_idx) >= 2
    ]
    if not candidate_rows:
        return pd.DataFrame()

    probe_rows = candidate_rows[: min(100, len(candidate_rows))]

    # Detect a likely business-line column (often col 1 in matrix files).
    label_candidates = range(min(6, col_count))
    label_col = 0
    label_score = -1
    for col_idx in label_candidates:
        score = 0
        for row_idx in probe_rows:
            label = _get_cached_cell_text(row_idx, col_idx)
            if label and not _is_cached_data_value(row_idx, col_idx) and not is_unit_label(label):
                score += 1
        if score > label_score:
            label_col = col_idx
            label_score = score

    # Detect metric/line-detail column near the label column.
    metric_col = min(label_col + 1, col_count - 1)
    metric_score = -1
    for col_idx in range(label_col + 1, min(col_count, label_col + 6)):
        score = 0
        for row_idx in probe_rows:
            metric_text = _get_cached_cell_text(row_idx, col_idx)
            if metric_text and not _is_cached_data_value(row_idx, col_idx):
                score += 1
        if score > metric_score:
            metric_col = col_idx
            metric_score = score

    # Detect where data rows begin for this label/metric pattern.
    data_start = candidate_rows[0]
    for row_idx in candidate_rows[:50]:  # Check only first 50 candidate rows
        line_item = _get_cached_cell_text(row_idx, label_col)
        metric = _get_cached_cell_text(row_idx, metric_col)
        if (line_item or metric) and _row_numeric_hits(row_idx, start_col=metric_col + 1) >= 2:
            data_start = row_idx
            break

    # Capture a wider header band so top hierarchy levels (period/country) are retained.
    header_start = max(0, data_start - 10)
    header_rows = list(range(header_start, data_start))

    # Keep columns that carry actual data in the matrix body.
    # Limit body_end to 300 rows for large files instead of 600
    body_end = min(row_count, data_start + 300)
    value_cols = []
    for col_idx in range(metric_col + 1, col_count):
        hits = 0
        for row_idx in range(data_start, body_end):
            if _is_cached_data_value(row_idx, col_idx):
                hits += 1
        if hits >= 2:
            value_cols.append(col_idx)

    if not value_cols:
        return pd.DataFrame()

    def _looks_like_period_token(token: str) -> bool:
        text = token.strip().lower()
        if not text:
            return False
        # Ignore descriptive titles; period tokens are usually compact labels.
        if len(text) > 35:
            return False
        if any(flag in text for flag in ("ytd", "qtd", "mtd", "fy", "year to date")):
            return True
        return bool(RE_MONTH_YEAR_TOKEN.search(text))

    global_period_token = ""
    for row_idx in header_rows:
        for col_idx in range(0, min(col_count, max(metric_col + 1, 6))):
            token = _get_cached_cell_text(row_idx, col_idx)
            if token and token.upper() not in {"#REF!", "#DIV/0!"} and _looks_like_period_token(token):
                global_period_token = token
                break
        if global_period_token:
            break

    header_noise = {"#REF!", "#DIV/0!", "#N/A", "#VALUE!"}
    subheader_tokens = {
        "actual", "previous", "plan", "budget", "target", "forecast", "ytd", "mtd", "qtd"
    }

    # Forward-fill sparse top headers row-wise (common in wide matrix exports where group
    # names appear once and subcolumns are blank). This avoids orphan headers like "Previous".
    carried_header_rows: dict[int, list[str]] = {}
    for row_idx in header_rows:
        carried: list[str] = []
        current = ""
        for col_idx in range(col_count):
            token = _get_cached_cell_text(row_idx, col_idx)
            if token and token.upper() not in header_noise:
                current = token
            carried.append(current)
        carried_header_rows[row_idx] = carried

    def _header_for_col(col_idx: int) -> str:
        tokens: list[str] = []
        if global_period_token:
            tokens.append(global_period_token)

        for row_idx in header_rows:
            token = carried_header_rows[row_idx][col_idx]
            if not token or token.upper() in header_noise:
                continue
            if not tokens or tokens[-1] != token:
                tokens.append(token)

        # If we still ended up with only a generic subheader, borrow nearest left group label.
        if len(tokens) == 1 and tokens[0].strip().lower() in subheader_tokens:
            base = ""
            for left_col in range(col_idx - 1, metric_col, -1):
                for row_idx in header_rows:
                    candidate = carried_header_rows[row_idx][left_col]
                    if candidate and candidate.upper() not in header_noise and candidate.strip().lower() not in subheader_tokens:
                        base = candidate
                        break
                if base:
                    break
            if base:
                tokens = [base, tokens[0]]

        return " | ".join(tokens)

    value_headers = {col_idx: _header_for_col(col_idx) for col_idx in value_cols}

    # Pick a broad section/title from the top-left area.
    section = ""
    for row_idx in range(min(8, row_count)):
        for col_idx in range(min(4, col_count)):
            token = _get_cached_cell_text(row_idx, col_idx)
            if token and token.upper() not in {"#REF!", "#DIV/0!"}:
                section = token
                break
        if section:
            break

    records = []
    current_line_item = ""
    blank_run = 0
    current_section = section
    header_metric_meta = {
        col_idx: metric_context(value_headers.get(col_idx, "")) if value_headers.get(col_idx, "") else None
        for col_idx in value_cols
    }

    indent_grid = raw_df.attrs.get("excel_indents") or []
    label_texts = [_get_cached_cell_text(row_idx, label_col) for row_idx in range(row_count)]
    label_indents = [
        float(indent_grid[row_idx][label_col])
        if row_idx < len(indent_grid) and label_col < len(indent_grid[row_idx])
        else 0.0
        for row_idx in range(row_count)
    ]

    def _fast_row_context(row_idx: int) -> dict:
        current_indent = label_indents[row_idx]
        if current_indent <= 0:
            return {}

        parents = []
        next_indent = current_indent
        # Bound lookback depth for performance on large sheets.
        min_idx = max(row_idx - 120, 0)
        for idx in range(row_idx - 1, min_idx - 1, -1):
            label = label_texts[idx]
            if not label:
                continue
            indent = label_indents[idx]
            if indent < next_indent:
                parents.append(label)
                next_indent = indent
                if indent <= 0:
                    break

        parents.reverse()
        if not parents:
            return {}
        return {
            "parent_line_item": parents[-1],
            "line_item_path": " > ".join([*parents, label_texts[row_idx]]),
        }

    has_indent_hierarchy = any(
        label_indents[row_idx] > 0 for row_idx in range(data_start, min(data_start + 1000, row_count)))
    # Limit data extraction to first 1000 rows for huge files
    data_end = min(row_count, data_start + 1000)
    for row_idx in range(data_start, data_end):
        row = matrix_values[row_idx]
        row_series = raw_df.iloc[row_idx]
        header = section_name_validated(row_series, expected_col=section_col)
        if header:
            current_section = header
            current_line_item = ""
            blank_run = 0
            continue

        line_item = _get_cached_cell_text(row_idx, label_col)
        metric = _get_cached_cell_text(row_idx, metric_col)
        row_value_hits = sum(1 for col_idx in value_cols if _is_cached_data_value(row_idx, col_idx))

        if not line_item and not metric and row_value_hits == 0:
            blank_run += 1
            if blank_run >= 25:
                break
            continue
        blank_run = 0

        if line_item and not is_unit_label(line_item):
            current_line_item = line_item

        active_line_item = current_line_item or line_item
        if not active_line_item:
            continue

        metric_name = metric or "value"
        row_context = _fast_row_context(row_idx) if has_indent_hierarchy else {}
        for col_idx in value_cols:
            value = row[col_idx]
            if is_blank(value) or not _is_cached_data_value(row_idx, col_idx):
                continue
            header_label = value_headers.get(col_idx, "")
            record = {
                "table_name": current_section,
                "section": current_section,
                "column_group": header_label,
                "line_item": active_line_item,
                "metric": metric_name,
                "value": value,
                "block_start_column": label_col,
            }
            if header_label:
                metric_meta = header_metric_meta.get(col_idx) or {}
            else:
                if metric_name not in metric_meta_cache:
                    metric_meta_cache[metric_name] = metric_context(metric_name)
                metric_meta = metric_meta_cache[metric_name]
            record.update(row_context)
            record.update(metric_meta)
            records.append(record)

    return pd.DataFrame(records)


def auto_flatten_heavy_summary_grid(raw_df: pd.DataFrame) -> pd.DataFrame:
    if raw_df is None or raw_df.empty:
        return pd.DataFrame()

    row_count, col_count = raw_df.shape
    if row_count == 0 or col_count < 3:
        return pd.DataFrame()

    top_window = min(row_count, 40)
    metric_row_idx = 0
    best_score = -1
    for row_idx in range(top_window):
        score = 0
        row = raw_df.iloc[row_idx]
        for value in row.tolist():
            text = cell_text(value)
            if text and not _is_numeric_like(value):
                score += 1
        if score > best_score:
            metric_row_idx = row_idx
            best_score = score

    metric_headers = [cell_text(v) for v in raw_df.iloc[metric_row_idx].tolist()]

    records = []
    for row_idx in range(metric_row_idx + 1, row_count):
        row = raw_df.iloc[row_idx]
        label_col = -1
        line_item = ""
        for col_idx, value in enumerate(row.tolist()):
            text = cell_text(value)
            if text:
                label_col = col_idx
                line_item = text
                break
        if label_col < 0:
            continue
        if not line_item or is_unit_label(line_item):
            continue

        row_has_values = False
        for col_idx in range(label_col + 1, col_count):
            if col_idx <= label_col:
                continue
            value = row.iloc[col_idx]
            if is_blank(value):
                continue
            if not _is_numeric_like(value):
                continue
            row_has_values = True
            metric = metric_headers[col_idx] if col_idx < len(metric_headers) else ""
            metric_text = metric or f"col_{col_idx}"
            records.append({
                "table_name": "Summary Grid",
                "section": "",
                "column_group": "",
                "unit": "",
                "line_item": line_item,
                "metric": metric_text,
                "metric_type": "",
                "metric_date": "",
                "comparison_date": "",
                "metric_quarter": "",
                "comparison_year": "",
                "value": value,
                "block_key": f"heavy_summary:{metric_row_idx}",
                "block_start_column": int(label_col),
            })

        if not row_has_values:
            continue

    return pd.DataFrame(records)


def auto_flatten_report_tables(raw_df: pd.DataFrame, extraction_profile: str = "auto") -> pd.DataFrame:
    """General extractor for visually formatted report sheets.

    This is the single user-facing auto mode. Internally it tries a few
    structural strategies and chooses the richest useful output. That keeps the
    UI general without pretending every visual report has one physical shape.

    KEY FIX: If ALL strategies fail (return empty), fall back to detect and flatten
    matrix-style data where both rows and columns contain meaningful identifiers.
    """
    profile = (extraction_profile or "auto").strip().lower()
    if profile not in {"auto", "general", "matrix"}:
        profile = "auto"

    raw_df = trim_effective_row_window(raw_df, tail_buffer=3)

    # Keep the original grid for sparse wide layouts (e.g., horizontal balance sheet)
    # where top rows are intentionally sparse and would be removed by error-density skipping.
    original_df = raw_df
    raw_df = _skip_error_heavy_rows(raw_df, max_check=50)
    raw_df = trim_effective_row_window(raw_df, tail_buffer=2)

    first_col_texts = [cell_text(value).lower() for value in raw_df.iloc[:, 0].tolist()]
    looks_like_rates_sheet = any(
        text.startswith("exchange rates") or text.startswith("valuation rates")
        for text in first_col_texts
    )

    top_values = [
        str(v).strip().upper()
        for row in raw_df.head(12).values
        for v in row
        if not is_blank(v)
    ]
    has_matrix_signals = any(v in {"#REF!", "#DIV/0!", "#N/A"} for v in top_values)
    is_heavy_layout = raw_df.shape[0] >= 500 and raw_df.shape[1] >= 80

    if is_heavy_layout:
        fast_df = auto_flatten_financial_summary_layout(raw_df)
        if fast_df is None or fast_df.empty:
            fast_df = auto_flatten_heavy_summary_grid(raw_df)
        if fast_df is not None and not fast_df.empty:
            best_df = fast_df.copy()
            normalized = best_df
            # Continue with the shared normalization path below.
        else:
            normalized = pd.DataFrame()
    else:
        normalized = pd.DataFrame()

    strategy_order = []
    if normalized.empty and (profile == "matrix" or (profile == "auto" and has_matrix_signals and not looks_like_rates_sheet)):
        strategy_order = [
            ("horizontal_balance_sheet", auto_flatten_horizontal_balance_sheet),
            ("financial_summary_layout", auto_flatten_financial_summary_layout),
            ("fallback_matrix", _fallback_matrix_flatten),
            ("grouped_metric_blocks", auto_flatten_grouped_metric_blocks),
            ("side_by_side_blocks", auto_flatten_report_blocks),
            ("stacked_tables", auto_flatten_stacked_tables),
            ("sectioned_tables", auto_flatten_sectioned_financial_sheet),
        ]
    elif normalized.empty and looks_like_rates_sheet:
        strategy_order = [
            ("financial_summary_layout", auto_flatten_financial_summary_layout),
            ("sectioned_tables", auto_flatten_sectioned_financial_sheet),
            ("grouped_metric_blocks", auto_flatten_grouped_metric_blocks),
            ("side_by_side_blocks", auto_flatten_report_blocks),
            ("stacked_tables", auto_flatten_stacked_tables),
            ("fallback_matrix", _fallback_matrix_flatten),
            ("horizontal_balance_sheet", auto_flatten_horizontal_balance_sheet),
        ]
    elif normalized.empty:
        strategy_order = [
            ("horizontal_balance_sheet", auto_flatten_horizontal_balance_sheet),
            ("financial_summary_layout", auto_flatten_financial_summary_layout),
            ("heavy_summary_grid", auto_flatten_heavy_summary_grid),
            ("side_by_side_blocks", auto_flatten_report_blocks),
            ("stacked_tables", auto_flatten_stacked_tables),
            ("sectioned_tables", auto_flatten_sectioned_financial_sheet),
            ("fallback_matrix", _fallback_matrix_flatten),
        ]
        if is_heavy_layout:
            strategy_order = [
                ("horizontal_balance_sheet", auto_flatten_horizontal_balance_sheet),
                ("financial_summary_layout", auto_flatten_financial_summary_layout),
                ("heavy_summary_grid", auto_flatten_heavy_summary_grid),
                ("side_by_side_blocks", auto_flatten_report_blocks),
            ]

    best_name = ""
    best_df = normalized.copy()
    best_score = len(best_df) if not best_df.empty else 0
    for name, extractor in strategy_order:
        source_df = original_df if name == "horizontal_balance_sheet" else raw_df
        df = extractor(source_df)
        if df is None or df.empty:
            continue

        if name == "horizontal_balance_sheet" and len(df) >= 40:
            best_name = name
            best_df = df.copy()
            best_score = len(df)
            break

        # If user explicitly selected matrix mode, trust matrix extractor output.
        if profile == "matrix" and name == "fallback_matrix":
            best_name = name
            best_df = df.copy()
            best_score = len(df)
            break

        context_cols = [
            col for col in df.columns
            if col not in {"value", "block_start_column"}
        ]
        score = len(df) * max(len(context_cols), 1)
        if looks_like_rates_sheet and name == "sectioned_tables":
            # Preserve Spot/Average/tenor context on market-data style sheets.
            score *= 10
        if score > best_score:
            best_name = name
            best_df = df.copy()
            best_score = score

        # Early exit for obvious wins to reduce work on large sheets.
        if name == "fallback_matrix" and len(df) >= 500:
            break
        if looks_like_rates_sheet and name == "sectioned_tables" and len(df) >= 80:
            break

    if best_df.empty:
        return best_df

    normalized = best_df.copy()

    hierarchy_df = propagate_section_context(raw_df)
    if "line_item" in normalized.columns and not hierarchy_df.empty and "line_item" in hierarchy_df.columns:
        lineage_cols = [col for col in hierarchy_df.columns if isinstance(col, str) and col.startswith("_section_L")]
        if lineage_cols:
            hierarchy_map = hierarchy_df.loc[:, ["line_item", *lineage_cols]].copy()
            hierarchy_map["line_item"] = _series_cell_text(hierarchy_map["line_item"])
            hierarchy_map = hierarchy_map[hierarchy_map["line_item"].ne("")]
            hierarchy_map = hierarchy_map.drop_duplicates(subset=["line_item"], keep="first")
            normalized["line_item"] = _series_cell_text(normalized["line_item"])
            normalized = normalized.merge(hierarchy_map, on="line_item", how="left")

    if "block_id" not in normalized.columns:
        if "block_key" in normalized.columns:
            starts = normalized["block_key"].fillna("-1")
            normalized = normalized.drop(columns=["block_key"])
            normalized.insert(
                1,
                "block_id",
                pd.factorize(starts.astype(str))[0] + 1,
            )
        elif "block_start_column" in normalized.columns:
            starts = normalized["block_start_column"].fillna(-1)
            normalized.insert(
                1,
                "block_id",
                pd.factorize(starts.astype(str))[0] + 1,
            )
        else:
            normalized.insert(1, "block_id", 1)

    if "block_start_column" in normalized.columns:
        normalized = normalized.drop(columns=["block_start_column"])

    if "line_item" not in normalized.columns:
        if "currency" in normalized.columns:
            # currency IS the line item for rates sheets; don't create a duplicate column
            normalized["line_item"] = normalized["currency"]
            normalized = normalized.drop(columns=["currency"])
        else:
            normalized["line_item"] = ""

    # Drop currency if it's now a duplicate of line_item
    if (
            "currency" in normalized.columns
            and "line_item" in normalized.columns
            and columns_semantically_equal(normalized["currency"], normalized["line_item"])
    ):
        normalized = normalized.drop(columns=["currency"])

    # Drop table_name if it's a duplicate of section — checked after table_name is guaranteed to exist
    if (
            "table_name" in normalized.columns
            and "section" in normalized.columns
    ):
        pass  # checked below after table_name fallback runs

    if "metric" not in normalized.columns:
        metric_parts = []
        for _, row in normalized.iterrows():
            parts = [
                cell_text(row.get("period", "")),
                cell_text(row.get("valuation_date", "")),
                cell_text(row.get("tenor", "")),
                cell_text(row.get("rate_type", "")),
            ]
            metric_parts.append(" | ".join([part for part in parts if part]) or "value")
        normalized["metric"] = metric_parts

    # For rates sheets, populate metric_date/year/quarter from period or valuation_date
    # when those context columns are missing or empty.
    date_source_col = None
    if "period" in normalized.columns and normalized["period"].astype(str).str.strip().ne("").any():
        date_source_col = "period"
    elif "valuation_date" in normalized.columns and normalized["valuation_date"].astype(str).str.strip().ne("").any():
        date_source_col = "valuation_date"

    if date_source_col is not None:
        context_cols_to_fill = [
            "metric_type",
            "metric_date",
            "metric_quarter",
            "comparison_year",
            "comparison_date",
        ]
        # Only fill if all context columns are missing or entirely empty
        needs_fill = all(
            col not in normalized.columns or normalized[col].astype(str).str.strip().eq("").all()
            for col in context_cols_to_fill
        )
        if needs_fill:
            parsed = [metric_context(str(v)) for v in normalized[date_source_col].tolist()]
            for ctx_col in context_cols_to_fill:
                normalized[ctx_col] = [item.get(ctx_col, "") for item in parsed]

    if "table_name" not in normalized.columns:
        normalized["table_name"] = normalized["section"] if "section" in normalized.columns else ""

    # Drop table_name if it's semantically identical to section
    if (
            "table_name" in normalized.columns
            and "section" in normalized.columns
            and columns_semantically_equal(normalized["table_name"], normalized["section"])
    ):
        normalized = normalized.drop(columns=["table_name"])

    if "unit" not in normalized.columns:
        normalized["unit"] = ""

    for col in normalized.columns:
        if col != "value":
            normalized[col] = normalized[col].fillna("")

    normalized = _assign_schema_group_id(normalized)
    normalized = _reassign_block_id_from_schema_group(normalized)

    normalized = normalized.drop(
        columns=[
            col for col in (
                "block_id",
                "schema_group_id",
                "source_block_id",
                "source_schema_group_id",
                "logical_block_key",
                "block_key",
                "block_start_column",
            )
            if col in normalized.columns
        ],
        errors="ignore",
    )

    ordered_columns = [
        "tab_name",
        "table_name",
        "schema_signature",
        "section",
        "column_group",
        "unit",
        "rate_type",
        "period",
        "valuation_date",
        "tenor",
        "contract_type",
        "currency",
        "parent_line_item",
        "line_item_path",
        "line_item",
        "metric",
        "metric_detail",
        "metric_type",
        "metric_date",
        "comparison_date",
        "metric_quarter",
        "comparison_year",
        "value",
    ]
    normalized = _drop_misaligned_matrix_rows(normalized)
    normalized = _drop_numeric_pipe_spillover_rows(normalized)

    existing_ordered = [col for col in ordered_columns if col in normalized.columns]
    remaining = [col for col in normalized.columns if col not in existing_ordered]
    return drop_all_blank_columns(normalized[existing_ordered + remaining])


def auto_flatten_stacked_tables(raw_df: pd.DataFrame) -> pd.DataFrame:
    """Flatten common stacked mini-table layouts into metric/value rows.

    Looks for header rows where column 0 is a unit label (for example EUR mn)
    and later nonblank cells are metrics/periods. Rows below become line items
    until a run of blank rows or the next header.
    """
    records = []
    if raw_df is None or raw_df.empty:
        return pd.DataFrame()

    unit_prefixes = ("eur ", "usd ", "gbp ", "in %")
    i = 0
    while i < len(raw_df):
        row = raw_df.iloc[i]
        first = cell_text(row.iloc[0])
        first_lower = first.lower()
        metric_cols = nonblank_col_indexes(row, start_col=1)

        if first_lower.startswith(unit_prefixes) and metric_cols:
            unit = first
            table_name, column_group = "", ""
            if _is_col0_unit_anchor(row.iloc[0]):
                table_name, column_group = _resolve_block_header_from_unit_anchor(raw_df, i)
            if not table_name:
                table_name = nearest_title_above(raw_df, i)
            if not column_group:
                column_group = nearest_group_above(raw_df, i)
            metric_headers = {
                col_idx: cell_text(row.iloc[col_idx])
                for col_idx in metric_cols
            }
            j = i + 1
            blank_label_rows = 0
            while j < len(raw_df):
                data_row = raw_df.iloc[j]
                item = cell_text(data_row.iloc[0])
                item_lower = item.lower()
                if not item:
                    blank_label_rows += 1
                    if blank_label_rows >= 5:
                        break
                    j += 1
                    continue
                blank_label_rows = 0
                if item_lower.startswith(unit_prefixes):
                    break
                if not row_has_values(data_row, start_col=1):
                    j += 1
                    continue

                row_context = indented_row_context(raw_df, j, 0)
                for col_idx, metric in metric_headers.items():
                    value = data_row.iloc[col_idx]
                    if not is_blank(value):
                        record = {
                            "table_name": table_name,
                            "column_group": column_group,
                            "unit": unit,
                            "line_item": item,
                            "metric": metric,
                            "value": value,
                        }
                        record.update(row_context)
                        record.update(metric_context(metric))
                        records.append(record)
                j += 1
            i = j
            continue

        i += 1

    return pd.DataFrame(records)


def read_sheet(uploaded_file, sheet_name, header=None):
    uploaded_file.seek(0)
    return pd.read_excel(uploaded_file, sheet_name=sheet_name, header=header)


# Function-level cache for when Streamlit session_state isn't available
_quick_read_memory_cache = {}


def _quick_read_cached(file_bytes: bytes, sheet_name: str) -> pd.DataFrame:
    """Fast pandas read for structure detection, cached when possible."""
    cache_key = f"_quick_{hash(file_bytes)}_{sheet_name}"

    # Try Streamlit session_state first (best caching in production)
    try:
        if cache_key not in st.session_state:
            st.session_state[cache_key] = pd.read_excel(
                io.BytesIO(file_bytes), sheet_name=sheet_name, header=None
            )
        return st.session_state[cache_key]
    except Exception:
        # Fallback: Use function-level memory cache when session_state unavailable
        # (e.g., when running outside Streamlit, in tests, or offline)
        if cache_key not in _quick_read_memory_cache:
            _quick_read_memory_cache[cache_key] = pd.read_excel(
                io.BytesIO(file_bytes), sheet_name=sheet_name, header=None
            )
        return _quick_read_memory_cache[cache_key]


@st.cache_data(show_spinner=False)
def infer_report_layout_from_quick_df(quick_df: pd.DataFrame) -> tuple[str, str, bool, bool]:
    """Infer whether a sheet looks like a general report block layout or matrix layout."""
    if quick_df is None or quick_df.empty:
        return "general", "Sheet is mostly empty in quick preview.", False, False

    has_report_signals = (
            quick_df.shape[1] > 0
            and any(
        str(v).lower().startswith(
            ("exchange rates", "valuation rates", "eur ", "usd ", "gbp ", "in %")
        )
        for v in quick_df.iloc[:, 0].tolist()
        if not is_blank(v)
    )
    )

    top_values = [
        str(v).strip()
        for row in quick_df.head(20).values
        for v in row
        if not is_blank(v)
    ]
    # Check for error tokens more thoroughly (up to 20 rows, not just 12)
    error_tokens_list = {"#REF!", "#DIV/0!", "#N/A", "#VALUE!", "#NULL!", "#NUM!", "#ERROR!", "#NAME?"}
    has_matrix_signals = any(v.upper() in error_tokens_list for v in top_values)

    first_col_texts = [cell_text(v).lower() for v in quick_df.iloc[:, 0].head(40).tolist()]
    top_left_texts = [
        cell_text(v).lower()
        for row in quick_df.head(12).iloc[:, : min(12, quick_df.shape[1])].values
        for v in row
    ]
    if any("matrix" in text for text in first_col_texts if text) or any(
            "matrix" in text for text in top_left_texts if text):
        has_matrix_signals = True

    # Matrix sheets are often very wide and have many value-heavy rows with sparse label columns.
    row_sample = quick_df.head(min(len(quick_df), 120))
    first_col_nonblank_top = sum(1 for value in quick_df.iloc[:, 0].head(20).tolist() if cell_text(value))
    wide_sheet_signal = quick_df.shape[1] >= 40
    value_heavy_rows = 0
    for _, row in row_sample.iterrows():
        values = row.tolist()
        label_left = cell_text(values[0]) if values else ""
        metric_left = cell_text(values[1]) if len(values) > 1 else ""
        numeric_hits = sum(1 for value in values[2:] if looks_like_data_value(value))
        if numeric_hits >= 8 and (label_left or metric_left):
            value_heavy_rows += 1
    structural_matrix_signal = (
            wide_sheet_signal
            and value_heavy_rows >= 5
            and first_col_nonblank_top <= 5  # Relaxed from 3 to account for error tokens being skipped
    )
    if structural_matrix_signal:
        has_matrix_signals = True

    # Run expensive general header detection only when matrix signals are absent.
    if not has_report_signals and not has_matrix_signals:
        has_report_signals = sheet_looks_general_report_tables(quick_df)

    if has_matrix_signals:
        return "matrix", "Detected matrix/error-token style header signals.", has_report_signals, has_matrix_signals
    if has_report_signals:
        return "general", "Detected general report block signals.", has_report_signals, has_matrix_signals
    return "general", "No strong report signals detected.", has_report_signals, has_matrix_signals


@st.cache_data(show_spinner=False)
def _extract_report_cached(file_bytes: bytes, sheet_name: str, extraction_profile: str = "auto") -> pd.DataFrame:
    # Always use openpyxl for proper merged cell handling; rely on extraction strategy optimization instead
    raw_df = _read_display_sheet_cached(file_bytes, sheet_name)
    return auto_flatten_report_tables(raw_df, extraction_profile=extraction_profile)


def read_display_sheet(uploaded_file, sheet_name) -> pd.DataFrame:
    """Read an Excel sheet using the values as they are displayed in Excel."""
    assert hasattr(uploaded_file, "name") and hasattr(uploaded_file, "seek"), "Expected a file-like object"
    suffix = Path(uploaded_file.name).suffix.lower()
    if suffix not in {".xlsx", ".xlsm"}:
        result = pd.read_excel(uploaded_file, sheet_name=sheet_name, header=None)
        return pd.DataFrame(result) if not isinstance(result, pd.DataFrame) else result

    uploaded_file.seek(0)
    file_bytes = uploaded_file.read()
    return _read_display_sheet_cached(file_bytes, sheet_name)


@st.cache_data(show_spinner=False)
def _read_display_sheet_cached(file_bytes: bytes, sheet_name: str) -> pd.DataFrame:
    """Cached inner implementation — keyed on file content + sheet name."""
    workbook = openpyxl.load_workbook(
        io.BytesIO(file_bytes),
        read_only=True,
        data_only=True,
        keep_vba=False,
        rich_text=False,
    )
    worksheet = workbook[sheet_name]

    # Read-only path keeps memory usage bounded for large matrix sheets.
    needs_merge_fallback = bool(getattr(getattr(worksheet, "merged_cells", None), "ranges", []))
    if needs_merge_fallback:
        workbook.close()
        workbook = openpyxl.load_workbook(
            io.BytesIO(file_bytes),
            read_only=False,
            data_only=True,
            keep_vba=False,
            rich_text=False,
        )
        worksheet = workbook[sheet_name]

    # Optimize: only read the actual used range, not the entire worksheet
    dims = ""
    try:
        dims = worksheet.calculate_dimension() or ""
    except Exception:
        dims = getattr(worksheet, "dimensions", "") or ""
    if dims:
        # Parse dimensions string like "A1:F100"
        if dims and ":" in dims:
            start_cell, end_cell = dims.split(":")
            # Extract row/col from end_cell
            match = RE_DIMENSIONS.match(end_cell)
            if match:
                max_col_letter, max_row = match.groups()
                max_row = int(max_row)
                # Convert column letter to number
                max_col = 0
                for char in max_col_letter:
                    max_col = max_col * 26 + (ord(char) - ord('A') + 1)
            else:
                max_row = worksheet.max_row
                max_col = worksheet.max_column
        else:
            max_row = worksheet.max_row
            max_col = worksheet.max_column
    else:
        max_row = worksheet.max_row
        max_col = worksheet.max_column

    row_merge_index: dict[int, list[tuple[int, int, object]]] = {}
    worksheet_merged = getattr(getattr(worksheet, "merged_cells", None), "ranges", None)
    if worksheet_merged and not getattr(workbook, "read_only", False):
        for merged_range in worksheet_merged:
            source_cell = worksheet.cell(merged_range.min_row, merged_range.min_col)
            for row_idx in range(merged_range.min_row, merged_range.max_row + 1):
                row_merge_index.setdefault(row_idx, []).append(
                    (merged_range.min_col, merged_range.max_col, source_cell)
                )
        for row_idx in row_merge_index:
            row_merge_index[row_idx].sort(key=lambda item: item[0])

    resolved_cell_cache: dict[tuple[int, int], object] = {}

    def _resolve_actual_cell(cell):
        if not hasattr(cell, "row") or not hasattr(cell, "column"):
            return cell
        key = (cell.row, cell.column)
        if key in resolved_cell_cache:
            return resolved_cell_cache[key]

        intervals = row_merge_index.get(cell.row)
        if intervals:
            col_idx = cell.column
            for min_col, max_col, source_cell in intervals:
                if min_col <= col_idx <= max_col:
                    resolved_cell_cache[key] = source_cell
                    return source_cell
                if col_idx < min_col:
                    break

        resolved_cell_cache[key] = cell
        return cell

    rows = []
    indents = []
    row_outline_levels: list[int] = []
    row_bold_flags: list[bool] = []
    row_merged_flags: list[bool] = []
    row_indent_levels: list[int] = []
    # Bounded iteration avoids parsing full XML rows outside the used range.
    for row_idx, row in enumerate(
            worksheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col),
            start=1,
    ):
        row_data = []
        row_indents = []
        row_has_bold = False
        row_has_merged = False
        for cell in row:
            actual_cell = _resolve_actual_cell(cell)
            row_data.append(formatted_excel_value(actual_cell))
            alignment = getattr(actual_cell, "alignment", None)
            row_indents.append(
                float(getattr(alignment, "indent", 0) or 0)
                if alignment else 0
            )
            font = getattr(actual_cell, "font", None)
            if bool(getattr(font, "bold", False)):
                row_has_bold = True
            if actual_cell is not cell:
                row_has_merged = True
        rows.append(row_data)
        indents.append(row_indents)
        row_bold_flags.append(row_has_bold)
        row_merged_flags.append(row_has_merged)
        row_indent_levels.append(int(max(row_indents)) if row_indents else 0)

        outline_level = 0
        try:
            outline_level = int(getattr(worksheet.row_dimensions.get(row_idx), "outlineLevel", 0) or 0)
        except Exception:
            outline_level = 0
        row_outline_levels.append(outline_level)

    df = pd.DataFrame(rows)
    df.attrs["excel_indents"] = indents
    df.attrs["excel_row_outline_levels"] = row_outline_levels
    df.attrs["excel_row_bold_flags"] = row_bold_flags
    df.attrs["excel_row_merged_flags"] = row_merged_flags
    df.attrs["excel_row_indent_levels"] = row_indent_levels
    df.attrs["excel_col_map"] = list(range(df.shape[1]))
    workbook.close()  # Explicitly close to free memory
    return df


def _column_match_alias(name: str) -> str:
    normalized = clean_column_name(name)
    aliases = {
        "lineitem": "line_item",
        "line_item_name": "line_item",
        "metric_dt": "metric_date",
        "metricdt": "metric_date",
        "metricdate": "metric_date",
        "metric_yr": "metric_year",
        "metricyr": "metric_year",
        "metricyear": "metric_year",
        "metric_qtr": "metric_quarter",
        "metricqtr": "metric_quarter",
        "metricquarter": "metric_quarter",
        "metric_mth": "metric_month",
        "metricmth": "metric_month",
        "metricmonth": "metric_month",
        "businessline": "business_line",
        "business_line_name": "business_line",
    }
    return aliases.get(normalized, normalized)


def _column_similarity(left: str, right: str) -> float:
    if left == right:
        return 1.0
    left_tokens = {token for token in left.split("_") if token}
    right_tokens = {token for token in right.split("_") if token}
    if not left_tokens or not right_tokens:
        return SequenceMatcher(None, left, right).ratio()
    overlap = len(left_tokens.intersection(right_tokens))
    union = len(left_tokens.union(right_tokens))
    token_score = overlap / union if union else 0.0
    text_score = SequenceMatcher(None, left, right).ratio()
    return max(text_score, (0.55 * text_score) + (0.45 * token_score))


def _best_canonical_column(name: str, canonical_columns: list[str]) -> str:
    normalized = _column_match_alias(name)
    if normalized in canonical_columns:
        return normalized

    best_name = ""
    best_score = 0.0
    for candidate in canonical_columns:
        score = _column_similarity(normalized, candidate)
        if score > best_score:
            best_name = candidate
            best_score = score

    # Strict threshold prevents unrelated columns from merging.
    return best_name if best_score >= 0.92 else normalized


def _coalesce_series(left: pd.Series, right: pd.Series) -> pd.Series:
    mask = _series_cell_text(left).eq("")
    return left.where(~mask, right)


def _drop_sparse_noise_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame() if df is None else df

    keep = []
    row_count = max(len(df), 1)
    for col in df.columns:
        series = df[col]
        non_blank_values = [v for v in series.tolist() if not is_blank(v)]
        non_blank = len(non_blank_values)
        density = non_blank / row_count
        normalized = clean_column_name(col)
        looks_noise = normalized.startswith("unnamed") or normalized.startswith("col_")

        if looks_noise:
            if non_blank == 0:
                continue

            data_like = 0
            for value in non_blank_values:
                text = cell_text(value)
                if looks_like_data_value(text) or metric_context(text).get("metric_type"):
                    data_like += 1
            data_ratio = data_like / max(non_blank, 1)

            unique_ratio = len({cell_text(v) for v in non_blank_values}) / max(non_blank, 1)
            sparse_small = non_blank <= max(3, int(0.03 * row_count))
            sparse_mid = density < 0.12

            # Drop unnamed spillover columns if they are sparse and mostly non-analytic text.
            if (sparse_small and data_ratio < 0.5) or (sparse_mid and data_ratio < 0.25 and unique_ratio > 0.7):
                continue

        keep.append(col)

    return df.loc[:, keep]


def _is_note_like_text(value) -> bool:
    text = cell_text(value).lower()
    if not text:
        return False

    prefixes = (
        "note",
        "notes",
        "note:",
        "notes:",
        "source",
        "disclaimer",
        "see note",
        "refer to note",
        "for further information",
        "for more information",
        # NOT "thereof", "of which", "includes", "including", "excluding" -
        # these are financial sub-item labels (hierarchy children), not footnotes.
        "1)",
        "2)",
        "3)",
        "4)",
        "*",
    )
    if text.startswith(prefixes):
        return True

    if RE_NOTE_ENUM.match(text):
        return True

    if RE_NOTE_STAR.match(text):
        return True

    if "see note" in text or "refer to note" in text:
        return True

    if "for further information" in text or "for more information" in text:
        return True

    return False


def _row_data_signal_count(row: pd.Series) -> int:
    count = 0
    for value in row.tolist():
        if is_blank(value):
            continue
        text = cell_text(value)
        if looks_like_data_value(text) or metric_context(text).get("metric_type"):
            count += 1
    return count


def _drop_note_like_rows(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame() if df is None else df

    text_cols = [
        col for col in ["line_item", "metric", "metric_detail", "section", "table_name"]
        if col in df.columns
    ]
    if not text_cols:
        return df

    keep_mask = []
    for _, row in df.iterrows():
        has_note_signal = any(_is_note_like_text(row.get(col, "")) for col in text_cols)
        strong_data_signal = _row_data_signal_count(row) >= 2
        keep_mask.append(not (has_note_signal and not strong_data_signal))

    return df.loc[keep_mask]


# Keywords that mark a row as a sub-item (child) of the nearest preceding root row.
_HIERARCHY_CHILD_PREFIXES: tuple[str, ...] = (
    "thereof",
    "of which",
    "whereof",
    "includes",
    "including",
    "excl.",
    "excluding",
    "hereof",
)


def _enrich_hierarchy_from_text_patterns(df: pd.DataFrame) -> pd.DataFrame:
    """Fill parent_line_item, line_item_path, and indent_level using text patterns.

    Handles melted/matrix data by resolving hierarchy from ordered unique labels per
    block and applying the result back to all rows.
    """
    if df is None or df.empty:
        return pd.DataFrame() if df is None else df

    label_candidates = [c for c in ("line_item", "metric", "metric_detail") if c in df.columns]
    if not label_candidates:
        return df

    working = df.copy()

    for col in ("parent_line_item", "line_item_path"):
        if col not in working.columns:
            working[col] = ""
    if "indent_level" not in working.columns:
        working["indent_level"] = 0

    def _is_child_keyword(text: str) -> bool:
        return text.strip().lower().startswith(_HIERARCHY_CHILD_PREFIXES)

    def _resolve_hierarchy_for_sequence(unique_items: list[str]) -> dict[str, dict]:
        result: dict[str, dict] = {}
        stack: list[tuple[int, str]] = []

        for item in unique_items:
            if not item:
                continue

            if _is_child_keyword(item):
                if stack and stack[-1][0] >= 1:
                    depth = stack[-1][0]
                else:
                    depth = 1

                while stack and stack[-1][0] >= depth:
                    stack.pop()

                parent_text = stack[-1][1] if stack else ""
                ancestors = [entry[1] for entry in stack]
                path = " > ".join([*ancestors, item]) if ancestors else item
                result[item] = {"parent": parent_text, "path": path, "depth": depth}
                stack.append((depth, item))

            else:
                if item in result and result[item]["depth"] == 0:
                    while stack and stack[-1][0] > 0:
                        stack.pop()
                    if not stack or stack[-1][1] != item:
                        stack = [(0, item)]
                else:
                    stack = [(0, item)]
                result[item] = {"parent": "", "path": "", "depth": 0}

        return result

    group_col = "block_id" if "block_id" in working.columns else None

    def _process_group(group_df: pd.DataFrame):
        # Select the label column that best represents hierarchical line labels.
        label_col = label_candidates[0]
        best_score = (-1, -1)
        for cand in label_candidates:
            values = [cell_text(v) for v in group_df[cand].tolist()]
            non_blank = [v for v in values if v]
            child_hits = sum(1 for v in non_blank if _is_child_keyword(v))
            score = (child_hits, len(set(non_blank)))
            if score > best_score:
                best_score = score
                label_col = cand

        seen: dict[str, int] = {}
        for i, val in enumerate(group_df[label_col].tolist()):
            text = cell_text(val)
            if text and text not in seen:
                seen[text] = i
        ordered_unique = sorted(seen.keys(), key=lambda k: seen[k])

        excel_hierarchy: dict[str, dict] = {}
        for idx in group_df.index:
            existing_path = cell_text(working.at[idx, "line_item_path"])
            if existing_path and ">" in existing_path:
                item = cell_text(working.at[idx, label_col])
                depth = existing_path.count(">")
                parent = cell_text(working.at[idx, "parent_line_item"])
                excel_hierarchy[item] = {"parent": parent, "path": existing_path, "depth": depth}
                working.at[idx, "indent_level"] = depth

        hierarchy = _resolve_hierarchy_for_sequence(ordered_unique)

        for idx in group_df.index:
            item = cell_text(working.at[idx, label_col])
            if not item:
                working.at[idx, "indent_level"] = 0
                continue

            if item in excel_hierarchy:
                info = excel_hierarchy[item]
            elif item in hierarchy:
                info = hierarchy[item]
            else:
                info = {"parent": "", "path": "", "depth": 0}

            if info["depth"] > 0:
                if is_blank(working.at[idx, "parent_line_item"]):
                    working.at[idx, "parent_line_item"] = info["parent"]
                if is_blank(working.at[idx, "line_item_path"]):
                    working.at[idx, "line_item_path"] = info["path"]

            current_level = int(working.at[idx, "indent_level"]) if not is_blank(working.at[idx, "indent_level"]) else 0
            working.at[idx, "indent_level"] = max(current_level, info["depth"])

    if group_col and group_col in working.columns:
        for _, grp in working.groupby(group_col, dropna=False, sort=False):
            _process_group(grp)
    else:
        _process_group(working)

    working["indent_level"] = pd.to_numeric(working["indent_level"], errors="coerce").fillna(0).astype(int)

    hier_cols = ["parent_line_item", "line_item_path", "indent_level"]
    present_hier = [c for c in hier_cols if c in working.columns]
    if present_hier:
        other_cols = [c for c in working.columns if c not in present_hier]
        anchor = next((c for c in ("line_item", "metric", "value") if c in other_cols), None)
        if anchor:
            pos = other_cols.index(anchor)
            new_order = other_cols[:pos] + present_hier + other_cols[pos:]
        else:
            new_order = other_cols + present_hier
        working = working[new_order]

    return working


def _assign_schema_group_id(df: pd.DataFrame) -> pd.DataFrame:
    """Split mixed schemas within the same block into stable schema_group_id values."""
    if df is None or df.empty:
        return pd.DataFrame() if df is None else df

    working = df.copy()
    if "schema_group_id" not in working.columns:
        working["schema_group_id"] = ""
    if "schema_signature" not in working.columns:
        working["schema_signature"] = ""

    essential_presence_cols = [
        "column_group",
        "unit",
        "rate_type",
        "period",
        "valuation_date",
        "tenor",
        "contract_type",
        "metric_type",
        "metric_detail",
    ]
    valuation_cols = ["valuation_date", "tenor", "contract_type"]
    period_cols = ["period", "rate_type"]

    _AXIS_INHERITANCE: dict[str, list[str]] = {
        "period_like": [],
        "valuation_like": ["period_like"],
        "hybrid": ["period_like", "valuation_like"],
        "generic": [],
    }

    def _value_shape(value) -> str:
        text = cell_text(value)
        if not text:
            return "blank"

        if _parse_numeric_value(text) is not None:
            return "numeric"

        lowered = text.lower()

        if RE_DURATION_TOKEN.fullmatch(text.strip()):
            return "duration"

        if RE_PERIOD_TOKEN.search(lowered):
            return "period_token"
        if RE_DATE_DMY.search(text):
            return "date"
        if RE_DATE_YEARISH.search(text):
            return "date"
        try:
            parsed = pd.to_datetime(text, errors="coerce", dayfirst=True)
            if pd.notna(parsed):
                return "date"
        except Exception:
            pass

        if RE_UPPER_CODE.fullmatch(text):
            return "code"
        if "|" in text:
            return "composite"
        if RE_ALPHA.search(text):
            return "text"
        return "other"

    def _row_signature(row: pd.Series) -> tuple:
        # Build structural signature from presence + stable value shapes.
        presence = tuple(bool(cell_text(row.get(col, ""))) for col in essential_presence_cols)
        shape_profile = tuple(
            _value_shape(row.get(col, "")) if present else "blank"
            for col, present in zip(essential_presence_cols, presence)
        )

        valuation_hits = sum(1 for col in valuation_cols if bool(cell_text(row.get(col, ""))))
        period_hits = sum(1 for col in period_cols if bool(cell_text(row.get(col, ""))))

        if valuation_hits >= 1 and period_hits == 0:
            axis_kind = "valuation_like"
        elif period_hits >= 1 and valuation_hits == 0:
            axis_kind = "period_like"
        elif valuation_hits >= 1 and period_hits >= 1:
            axis_kind = "hybrid"
        else:
            axis_kind = "generic"

        essential_count = sum(1 for bit in presence if bit)
        return (axis_kind, essential_count, presence, shape_profile)

    def _context_anchor(row: pd.Series | None) -> dict:
        if row is None:
            return {}
        result: dict = {}
        for col in ["section", "table_name", "column_group", "unit", "rate_type", "contract_type"]:
            raw = cell_text(row.get(col, ""))
            if raw:
                result[col] = re.sub(r"\s+", " ", raw).strip()[:64]
        return result

    def _signature_label(signature: tuple, row: pd.Series | None = None) -> str:
        axis_kind = signature[0]
        presence_bits = signature[2]
        shape_profile = signature[3]

        bases = _AXIS_INHERITANCE.get(axis_kind, [])

        ctx = _context_anchor(row)
        section = ctx.get("section") or ctx.get("table_name") or ""
        qualifier = ctx.get("rate_type") or ctx.get("contract_type") or ctx.get("column_group") or ""
        unit = ctx.get("unit") or ""

        present_cols = [
            col for col, present in zip(essential_presence_cols, presence_bits)
            if bool(present)
        ]

        note_shapes = {"date", "period_token", "duration", "code"}
        shape_hints = {
            col: shape
            for col, present, shape in zip(essential_presence_cols, presence_bits, shape_profile)
            if bool(present) and shape in note_shapes
        }

        payload = {
            "schema_version": "v2",
            "axis": {
                "kind": axis_kind,
                "inherits": bases,
            },
            "context": {
                "section": section,
                "qualifier": qualifier,
                "unit": unit,
            },
            "fields_present": present_cols,
            "shape_hints": shape_hints,
        }
        return json.dumps(payload, ensure_ascii=True, sort_keys=True, separators=(",", ":"))

    if "block_id" in working.columns:
        block_groups = working.groupby("block_id", dropna=False, sort=False)
    else:
        block_groups = [("default", working)]

    for block_key, block_df in block_groups:
        if block_df.empty:
            continue

        base = "default" if pd.isna(block_key) else str(block_key)
        signatures = pd.Series(
            [_row_signature(block_df.loc[idx]) for idx in block_df.index],
            index=block_df.index,
            dtype="object",
        )
        if signatures.empty:
            working.loc[block_df.index, "schema_group_id"] = base
            continue

        counts = signatures.value_counts()
        if len(counts) <= 1:
            working.loc[block_df.index, "schema_group_id"] = base
            signature = counts.index[0]
            sample_row = working.loc[block_df.index[0]] if len(block_df.index) > 0 else None
            working.loc[block_df.index, "schema_signature"] = _signature_label(signature, sample_row)
            continue

        # Trend logic: split only when alternate row-types have enough mass, otherwise smooth as noise.
        dominant_signature = counts.index[0]
        dominant_count = int(counts.iloc[0])
        total_count = len(signatures)
        dominant_ratio = dominant_count / max(total_count, 1)
        outlier_threshold = max(1, int(total_count * 0.10))
        significance_threshold = max(2, int(total_count * 0.15))

        signature_map: dict[tuple, tuple] = {}
        if dominant_ratio >= 0.70:
            dominant_axis = dominant_signature[0]
            for signature, count in counts.items():
                axis = signature[0]
                count_int = int(count)
                if count_int <= outlier_threshold:
                    signature_map[signature] = dominant_signature
                elif axis != dominant_axis and count_int < significance_threshold:
                    signature_map[signature] = dominant_signature
                else:
                    signature_map[signature] = signature
        else:
            for signature in counts.index.tolist():
                signature_map[signature] = signature

        normalized_signatures = pd.Series(
            [signature_map[sig] for sig in signatures.tolist()],
            index=signatures.index,
            dtype="object",
        )
        unique_signatures = list(dict.fromkeys(normalized_signatures.tolist()))

        if len(unique_signatures) <= 1:
            working.loc[block_df.index, "schema_group_id"] = base
            sample_row = working.loc[block_df.index[0]] if len(block_df.index) > 0 else None
            label = _signature_label(unique_signatures[0], sample_row) if unique_signatures else "[generic] unknown"
            working.loc[block_df.index, "schema_signature"] = label
            continue

        signature_to_suffix = {
            signature: idx + 1
            for idx, signature in enumerate(sorted(unique_signatures, key=lambda sig: str(sig)))
        }
        for row_idx, signature in normalized_signatures.items():
            working.at[row_idx, "schema_group_id"] = f"{base}:{signature_to_suffix[signature]}"
            row_obj = working.loc[row_idx] if row_idx in working.index else None
            working.at[row_idx, "schema_signature"] = _signature_label(signature, row_obj)

    return working


def _reassign_block_id_from_schema_group(df: pd.DataFrame) -> pd.DataFrame:
    """Assign distinct numeric block_id values when one source block splits into multiple schema groups."""
    if df is None or df.empty:
        return pd.DataFrame() if df is None else df
    if "schema_group_id" not in df.columns:
        return df

    working = df.copy()
    # Use schema_signature as canonical grouping key without exposing helper columns.
    if "schema_signature" in working.columns:
        canonical_key = working["schema_signature"].astype(str)
    else:
        canonical_key = working["schema_group_id"].astype(str)

    logical_keys = sorted(canonical_key.unique().tolist())
    key_to_id = {key: idx + 1 for idx, key in enumerate(logical_keys)}
    working["block_id"] = canonical_key.replace(key_to_id).astype(int)

    # Canonical schema groups are logical (cross-source), not extractor-order based.
    key_to_schema_group = {key: f"schema_{idx + 1}" for idx, key in enumerate(logical_keys)}
    working["schema_group_id"] = canonical_key.replace(key_to_schema_group)

    # Remove intermediate helper columns if they already exist in older dataframes.
    drop_cols = [c for c in ("source_block_id", "source_schema_group_id", "logical_block_key") if c in working.columns]
    if drop_cols:
        working = working.drop(columns=drop_cols)
    return working


@st.cache_data(show_spinner=False)
def is_sheet_already_flat(file_bytes: bytes, sheet_name: str) -> bool:
    """
    Pre-flight flat-file detection — FAST check (no extraction).

    Returns True if sheet appears to be already flat (no report/matrix signals).
    Returns False if extraction might be beneficial.

    This function only reads a preview, making it ~100x faster than full extraction.
    """
    try:
        # Quick preview read only
        quick_df = _quick_read_cached(file_bytes, sheet_name)

        # Infer layout using the same detection logic as extract_and_flatten_sheet
        _, _, has_report_signals, has_matrix_signals = infer_report_layout_from_quick_df(quick_df)

        # If NO report or matrix signals detected, the sheet is already flat
        # Return True (skip extraction), False otherwise (do extraction)
        return not (has_report_signals or has_matrix_signals)
    except Exception as e:
        # On error, default to False (do extraction to be safe)
        print(f"[is_sheet_already_flat] Error checking sheet '{sheet_name}': {e}")
        return False


def finalize_extracted_sheet(df: pd.DataFrame, strip_text: bool = True, split_hierarchy: bool = True) -> pd.DataFrame:
    """Apply final post-processing transformations to an extracted/flattened sheet.

    Includes:
    - Trimming whitespace in text columns
    - Splitting hierarchy columns on | and dash delimiters (removes original pipe-delimited columns)
    - Cleaning up blank columns and duplicates
    """
    if df is None or df.empty:
        return df

    cleaned = df.copy()

    # Trim whitespace in text columns
    if strip_text:
        for col in cleaned.columns:
            if not (pd.api.types.is_object_dtype(cleaned[col].dtype) or pd.api.types.is_string_dtype(
                    cleaned[col].dtype)):
                continue
            cleaned[col] = cleaned[col].astype(str).str.strip()

    # Apply spillover filters for both extracted and fallback/raw-read paths.
    cleaned = _drop_misaligned_matrix_rows(cleaned)
    cleaned = _drop_numeric_pipe_spillover_rows(cleaned)

    # Track columns that contain pipe delimiters BEFORE splitting
    cols_with_pipes = set()
    if split_hierarchy:
        for col in cleaned.columns:
            text_series = _series_cell_text(cleaned[col]).astype(str)
            if bool(text_series.str.contains(r"\|", regex=True).any()):
                cols_with_pipes.add(str(col))

    # Apply hierarchy column splitting (| and dash delimiters)
    if split_hierarchy:
        cleaned = _auto_split_hierarchy_columns(cleaned, dash_split_mode="spaced")
        # Remove original columns only when split columns were successfully created.
        cols_to_remove = []
        for source_col in cols_with_pipes:
            split_prefix = clean_column_name(f"{source_col}_part_")
            has_split_cols = any(str(col).startswith(split_prefix) for col in cleaned.columns)
            if has_split_cols and source_col in cleaned.columns:
                cols_to_remove.append(source_col)
        if cols_to_remove:
            cleaned = cleaned.drop(columns=cols_to_remove)

    cleaned = _assign_schema_group_id(cleaned)
    cleaned = _reassign_block_id_from_schema_group(cleaned)

    cleaned = _enrich_hierarchy_from_text_patterns(cleaned)

    # Final cleanup: only remove columns that are entirely blank.
    # We intentionally do NOT call drop_fully_duplicate_columns here because
    # two columns with identical content are still semantically distinct and
    # silently dropping one of them causes confusing data loss.
    cleaned = drop_all_blank_columns(cleaned).reset_index(drop=True)

    hierarchy_defaults: list[tuple[str, object]] = [
        ("parent_line_item", ""),
        ("line_item_path", ""),
        ("indent_level", 0),
    ]
    missing_hier = [(c, d) for c, d in hierarchy_defaults if c not in cleaned.columns]
    if missing_hier:
        for hcol, hdefault in missing_hier:
            cleaned[hcol] = hdefault
        hier_present = [c for c, _ in hierarchy_defaults if c in cleaned.columns]
        other_cols = [c for c in cleaned.columns if c not in hier_present]
        anchor = next((c for c in ("line_item", "metric", "value") if c in other_cols), None)
        if anchor:
            pos = other_cols.index(anchor)
            cleaned = cleaned[other_cols[:pos] + hier_present + other_cols[pos:]]

    cleaned = cleaned.drop(
        columns=[
            col for col in (
                "block_id",
                "schema_group_id",
                "source_block_id",
                "source_schema_group_id",
                "logical_block_key",
                "block_key",
                "block_start_column",
            )
            if col in cleaned.columns
        ],
        errors="ignore",
    )

    return cleaned


def extract_and_flatten_sheet(file_bytes: bytes, sheet_name: str, preferred_profile: str = "auto") -> pd.DataFrame:
    """Extract and flatten complex Excel sheets, or return quickly for already-flat sheets.

    For flat sheets: Returns immediately with just a pd.read_excel() call (< 1 second).
    For complex sheets: Runs the full extraction pipeline with UI feedback.
    """
    # Always run extraction here; callers control "already flat" behavior explicitly.
    quick_df = _quick_read_cached(file_bytes, sheet_name)
    hinted_profile = (preferred_profile or "auto").strip().lower()
    if hinted_profile not in {"auto", "general", "matrix"}:
        hinted_profile = "auto"

    auto_profile, _, has_report_signals, has_matrix_signals = infer_report_layout_from_quick_df(quick_df)
    extraction_profile = hinted_profile if hinted_profile != "auto" else auto_profile

    extracted_df = _extract_report_cached(file_bytes, sheet_name, extraction_profile=extraction_profile)
    if extracted_df is None or extracted_df.empty:
        fallback_df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name)
        result = finalize_extracted_sheet(pd.DataFrame(fallback_df), strip_text=True, split_hierarchy=True)
    else:
        result = finalize_extracted_sheet(pd.DataFrame(extracted_df), strip_text=True, split_hierarchy=True)
    if result is not None and not result.empty and "tab_name" not in result.columns:
        result.insert(0, "tab_name", sheet_name)
    return result


def _extract_like_one_sheet_mode(file_bytes: bytes, sheet_name: str) -> pd.DataFrame:
    """Mirror the default one-sheet extraction behavior for batch modes.

    This path intentionally avoids finalize_extracted_sheet() so multi-sheet modes
    retain the same detail level users see when processing one sheet in report mode.
    """
    quick_df = _quick_read_cached(file_bytes, sheet_name)
    suggested_layout, _, has_report_signals, has_matrix_signals = infer_report_layout_from_quick_df(quick_df)

    if not has_report_signals and not has_matrix_signals:
        result = pd.DataFrame(pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name))
        if result is not None and not result.empty and "tab_name" not in result.columns:
            result.insert(0, "tab_name", sheet_name)
        return result

    extracted_df = _extract_report_cached(file_bytes, sheet_name, extraction_profile=suggested_layout)
    if extracted_df is None or extracted_df.empty:
        result = pd.DataFrame(pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name))
    else:
        result = pd.DataFrame(extracted_df)
    if result is not None and not result.empty and "tab_name" not in result.columns:
        result.insert(0, "tab_name", sheet_name)
    return result


def merge_sheets(sheet_dataframes: dict[str, pd.DataFrame], drop_note_rows: bool = True) -> pd.DataFrame:
    """Merge multiple sheet dataframes with minimal data loss.

    Preserves all columns and rows from all sheets. Only performs:
    1. Basic concatenation
    2. Optional note row filtering (if enabled)
    3. Column reordering for readability
    """
    if not sheet_dataframes:
        return pd.DataFrame()

    aligned_frames: list[pd.DataFrame] = []
    for sheet_name, df in sheet_dataframes.items():
        if df is None or df.empty:
            continue

        working = df.copy()
        # Add source tracking columns without aggressive column harmonization
        if "tab_name" not in working.columns:
            working.insert(0, "tab_name", sheet_name)
        aligned_frames.append(working)

    if not aligned_frames:
        return pd.DataFrame()

    # Simple concatenation - preserve all columns and rows from all sheets
    merged = pd.concat(aligned_frames, ignore_index=True, sort=False)

    # Only apply note filtering if explicitly requested
    if drop_note_rows:
        merged = _drop_note_like_rows(merged)

    # Reorder columns for readability (but keep all columns)
    preferred_order = [
        "tab_name",
        "table_name",
        "section",
        "schema_signature",
        "column_group",
        "unit",
        "line_item",
        "business_line",
        "metric",
        "metric_detail",
        "metric_date",
        "metric_quarter",
        "value",
    ]

    # Put preferred columns first, then all remaining columns in original order
    ordered = [col for col in preferred_order if col in merged.columns]
    tail = [col for col in merged.columns if col not in ordered]

    return merged[ordered + tail].reset_index(drop=True)


def _auto_split_hierarchy_columns(
        df: pd.DataFrame,
        max_parts: int = 12,
        dash_split_mode: str = "spaced",
) -> pd.DataFrame:
    """Expand common hierarchy delimiters into dedicated `<column>_*_part_n` columns."""
    if df is None or df.empty or max_parts < 2:
        return pd.DataFrame() if df is None else df

    expanded = df.copy()

    def _unique_name(base_name: str) -> str:
        if base_name not in expanded.columns:
            return base_name
        idx = 2
        while f"{base_name}_{idx}" in expanded.columns:
            idx += 1
        return f"{base_name}_{idx}"

    delimiter_specs = [
        {
            "label": "part",
            "has_delim": lambda text: bool(re.search(r"\|", text)),
            "split": lambda text: re.split(r"\s*\|\s*", text),
        }
    ]
    mode = (dash_split_mode or "spaced").strip().lower()
    if mode == "spaced":
        delimiter_specs.append(
            {
                "label": "dash_part",
                "has_delim": lambda text: bool(re.search(r"\s+[-–—]\s+", text)),
                "split": lambda text: re.split(r"\s+[-–—]\s+", text),
            }
        )
    elif mode == "any":
        delimiter_specs.append(
            {
                "label": "dash_part",
                "has_delim": lambda text: bool(re.search(r"[-–—]", text)),
                "split": lambda text: re.split(r"\s*[-–—]\s*", text),
            }
        )

    for source_col in list(df.columns):
        text_values = _series_cell_text(expanded[source_col]).astype(str)

        for spec in delimiter_specs:
            label = spec["label"]
            has_delim = text_values.str.contains(r"\s+[-–—]\s+", regex=True, na=False) if label == "dash_part" and mode == "spaced" else (
                text_values.str.contains(r"[-–—]", regex=True, na=False) if label == "dash_part" and mode == "any" else text_values.str.contains(r"\|", regex=True, na=False)
            )
            if not has_delim.any():
                continue

            if label == "part":
                split_df = text_values.where(has_delim, "").str.split(r"\s*\|\s*", expand=True, regex=True)
            elif mode == "spaced":
                split_df = text_values.where(has_delim, "").str.split(r"\s+[-–—]\s+", expand=True, regex=True)
            else:
                split_df = text_values.where(has_delim, "").str.split(r"\s*[-–—]\s*", expand=True, regex=True)

            split_df = split_df.fillna("").replace(r"^\s+|\s+$", "", regex=True)
            col_has_values = split_df.ne("").any(axis=0)
            observed_max = int(col_has_values.sum())
            if observed_max < 2:
                continue

            part_count = min(int(observed_max), max_parts)
            for idx in range(part_count):
                base_name = clean_column_name(f"{source_col}_{label}_{idx + 1}")
                new_col = _unique_name(base_name)
                expanded[new_col] = split_df.iloc[:, idx] if idx < split_df.shape[1] else ""

    return expanded


def apply_pandas_cleanup(
        df: pd.DataFrame,
        drop_columns,
        rename_map,
        split_config,
        drop_blank_columns,
        type_conversions,
        strip_text,
        clean_names,
        dash_split_mode,
) -> pd.DataFrame:
    cleaned = df.copy()
    if cleaned.empty:
        return cleaned

    if drop_columns:
        cleaned = cleaned.drop(columns=[c for c in drop_columns if c in cleaned.columns])

    if rename_map:
        cleaned = cleaned.rename(columns={
            old: new for old, new in rename_map.items()
            if old in cleaned.columns and new
        })

    if strip_text:
        for col in cleaned.select_dtypes(include="object").columns:
            cleaned[col] = cleaned[col].astype(str).str.strip()

    # Remove matrix scratch/comment spillover rows before and after splitting to
    # catch both source pipe fields and derived split columns.
    cleaned = _drop_misaligned_matrix_rows(cleaned)
    cleaned = _drop_numeric_pipe_spillover_rows(cleaned)

    cleaned = _auto_split_hierarchy_columns(cleaned, dash_split_mode=dash_split_mode)
    cleaned = _drop_misaligned_matrix_rows(cleaned)
    cleaned = _drop_numeric_pipe_spillover_rows(cleaned)

    if split_config and split_config.get("column") in cleaned.columns:
        source_col = split_config["column"]
        delimiter = split_config.get("delimiter") or " "
        max_parts = int(split_config.get("max_parts") or 2)
        prefix = split_config.get("prefix") or source_col
        keep_original = split_config.get("keep_original", True)
        parts = cleaned[source_col].astype(str).str.split(
            delimiter,
            n=max_parts - 1,
            expand=True,
            regex=False,
        )
        for idx in range(max_parts):
            new_col = clean_column_name(f"{prefix}_{idx + 1}")
            cleaned[new_col] = parts[idx] if idx in parts.columns else ""
        if not keep_original:
            cleaned = cleaned.drop(columns=[source_col])

    if drop_blank_columns:
        present = [c for c in drop_blank_columns if c in cleaned.columns]
        if present:
            candidate = cleaned[present]
            flat = pd.Series(candidate.to_numpy(dtype=object, copy=False).ravel(), dtype="object")
            blank_flat = _series_cell_text(flat).eq("").to_numpy(dtype=bool)
            blank_matrix = blank_flat.reshape(candidate.shape)
            mask = pd.Series(blank_matrix.all(axis=1), index=candidate.index)
            cleaned = cleaned.loc[~mask].copy()

    for col, target_type in (type_conversions or {}).items():
        if col not in cleaned.columns or target_type == "keep":
            continue
        if target_type == "numeric":
            cleaned[col] = pd.to_numeric(cleaned[col], errors="coerce")
        elif target_type == "datetime":
            cleaned[col] = pd.to_datetime(cleaned[col], errors="coerce")
        elif target_type == "text":
            cleaned[col] = cleaned[col].astype(str)

    if clean_names:
        cleaned.columns = dedupe_columns(cleaned.columns)

    cleaned = drop_fully_duplicate_columns(cleaned)
    return drop_all_blank_columns(cleaned).reset_index(drop=True)


def to_excel_bytes(df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="flat_table", index=False)
    return output.getvalue()


def to_multisheet_excel_bytes(frames: dict[str, pd.DataFrame]) -> bytes:
    """Write multiple DataFrames into one Excel workbook, one sheet per entry.

    Sheet names are truncated to 31 characters (Excel limit) and de-duplicated.
    """
    output = io.BytesIO()
    used_names: list[str] = []

    def _safe_sheet_name(name: str) -> str:
        # Excel sheet names: max 31 chars, no [ ] : * ? / \
        cleaned = re.sub(r"[\\/:*?\[\]]", "_", str(name))[:31]
        base = cleaned
        counter = 2
        while cleaned in used_names:
            suffix = f"_{counter}"
            cleaned = base[: 31 - len(suffix)] + suffix
            counter += 1
        used_names.append(cleaned)
        return cleaned

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, df in frames.items():
            if df is None or df.empty:
                continue
            df.to_excel(writer, sheet_name=_safe_sheet_name(sheet_name), index=False)
    return output.getvalue()


def _read_csv_bytes(data: bytes) -> pd.DataFrame:
    result = pd.read_csv(io.BytesIO(data))  # type: ignore[call-overload]
    return pd.DataFrame(result)


def excel_mime_type(file_name: str) -> str:
    suffix = Path(file_name).suffix.lower()
    if suffix == ".xls":
        return "application/vnd.ms-excel"
    return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def render_raw_excel_download(uploaded_file, key: str):
    st.download_button(
        "Download Raw Excel",
        data=uploaded_file.getvalue(),
        file_name=uploaded_file.name,
        mime=excel_mime_type(uploaded_file.name),
        use_container_width=True,
        key=key,
    )


def render_preview_excel_download(df: pd.DataFrame, file_name: str, key: str, label: str = "Download Preview as Excel"):
    st.download_button(
        label,
        data=to_excel_bytes(df),
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        key=key,
    )


def render_cleanup_and_preview(df: pd.DataFrame, file_stem: str, key_prefix: str):
    """Apply optional pandas cleanup, then preview and download the result."""
    import hashlib

    scope = hashlib.md5(f"{key_prefix}_{file_stem}".encode()).hexdigest()[:8]
    kp = f"{key_prefix}_{scope}"

    st.markdown("### Pandas Cleanup")
    with st.expander("Edit the table with pandas-style operations", expanded=False):
        cleanup_cols = list(df.columns)
        drop_columns = st.multiselect(
            "Delete columns",
            cleanup_cols,
            help="Remove columns from the final output.",
            key=f"{kp}_drop_columns",
        )

        remaining_for_rename = [c for c in cleanup_cols if c not in drop_columns]
        rename_df = pd.DataFrame({
            "Column": remaining_for_rename,
            "New Name": remaining_for_rename,
        })
        edited_rename_df = st.data_editor(
            rename_df,
            hide_index=True,
            use_container_width=True,
            num_rows="fixed",
            key=f"{kp}_rename_columns_editor",
        )
        rename_map = {
            row["Column"]: row["New Name"]
            for _, row in edited_rename_df.iterrows()
            if row["Column"] != row["New Name"]
        }

        st.markdown("**Split a column**")
        split_enabled = st.checkbox(
            "Split one column into multiple columns",
            key=f"{kp}_split_enabled",
        )
        split_config = {}
        if split_enabled and remaining_for_rename:
            col_split_a, col_split_b, col_split_c = st.columns(3)
            with col_split_a:
                split_column = st.selectbox(
                    "Column to split",
                    remaining_for_rename,
                    key=f"{kp}_split_column",
                )
            with col_split_b:
                delimiter = st.text_input(
                    "Delimiter",
                    value=" | ",
                    key=f"{kp}_split_delimiter",
                )
            with col_split_c:
                max_parts = st.number_input(
                    "Number of output columns",
                    min_value=2,
                    max_value=12,
                    value=2,
                    key=f"{kp}_split_max_parts",
                )
            col_prefix, col_keep = st.columns([2, 1])
            with col_prefix:
                split_prefix = st.text_input(
                    "Output column prefix",
                    value=clean_column_name(split_column),
                    key=f"{kp}_split_prefix",
                )
            with col_keep:
                keep_original = st.checkbox(
                    "Keep original",
                    value=True,
                    key=f"{kp}_split_keep_original",
                )
            split_config = {
                "column": split_column,
                "delimiter": delimiter,
                "max_parts": max_parts,
                "prefix": split_prefix,
                "keep_original": keep_original,
            }

        drop_blank_columns = []
        type_conversions = {}

        col_strip, col_clean, col_dash = st.columns(3)
        with col_strip:
            strip_text = st.checkbox(
                "Trim whitespace in text columns",
                value=True,
                key=f"{kp}_strip_text",
            )
        with col_clean:
            clean_names = st.checkbox(
                "Clean final column names",
                value=False,
                key=f"{kp}_clean_names",
            )
        with col_dash:
            dash_split_mode = st.selectbox(
                "Auto-split '-' mode",
                options=["Off", "Spaced only ( - )", "Any dash (-)"],
                index=1,
                key=f"{kp}_dash_split_mode",
                help="Controls automatic dash splitting in hierarchy columns.",
            )
        dash_split_mode_map = {
            "Off": "off",
            "Spaced only ( - )": "spaced",
            "Any dash (-)": "any",
        }
        dash_split_mode_value = dash_split_mode_map[dash_split_mode]

    import hashlib as _hl

    cleanup_cache_version = "v2_hierarchy_split_regex"
    cleanup_sig = _hl.md5(
        str((
            cleanup_cache_version,
            list(df.columns), len(df),
            sorted(drop_columns),
            sorted(rename_map.items()),
            str(split_config),
            strip_text,
            clean_names,
            dash_split_mode_value,
        )).encode()
    ).hexdigest()
    cleanup_cache_key = f"{kp}_final_df_{cleanup_sig}"

    if cleanup_cache_key not in st.session_state:
        st.session_state[cleanup_cache_key] = apply_pandas_cleanup(
            df,
            drop_columns=drop_columns,
            rename_map=rename_map,
            split_config=split_config,
            drop_blank_columns=drop_blank_columns,
            type_conversions=type_conversions,
            strip_text=strip_text,
            clean_names=clean_names,
            dash_split_mode=dash_split_mode_value,
        )
    final_df = st.session_state[cleanup_cache_key]

    st.markdown("### Final Table Preview")
    col_rows, col_cols = st.columns(2)
    col_rows.metric("Rows", f"{len(final_df):,}")
    col_cols.metric("Columns", f"{len(final_df.columns):,}")
    st.dataframe(final_df.head(200), use_container_width=True)

    dl_key = f"{kp}_download"
    dl_state_key = f"{kp}_excel_bytes_{cleanup_sig}"

    def _prepare_download():
        st.session_state[dl_state_key] = to_excel_bytes(final_df)

    if dl_state_key not in st.session_state:
        st.button(
            "Prepare Download",
            on_click=_prepare_download,
            use_container_width=True,
            key=f"{kp}_prepare_btn",
        )
    else:
        st.download_button(
            "⬇️ Download Excel",
            data=st.session_state[dl_state_key],
            file_name=f"{file_stem}_flat.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key=dl_key,
        )


def render_unified_file_tab():
    st.subheader("Upload File")
    st.caption(
        "Upload one or more CSV/Excel files. Each file can be processed with auto-detect flattening or as already-flat data.")
    uploaded_files = st.file_uploader(
        "Upload CSV or Excel files",
        type=["csv", "xlsx", "xls", "xlsm"],
        accept_multiple_files=True,
        key="unified_file_upload",
    )
    if not uploaded_files:
        st.info("Upload at least one file to auto-detect its structure.")
        return

    if len(uploaded_files) == 1:
        uploaded_file = uploaded_files[0]
    else:
        name_to_file = {f.name: f for f in uploaded_files}
        selected_name = st.selectbox(
            "Choose file to process",
            options=list(name_to_file.keys()),
            key="unified_selected_file",
        )
        uploaded_file = name_to_file[selected_name]

    assert not isinstance(uploaded_file, list), "Expected a single file"
    suffix = Path(uploaded_file.name).suffix.lower()
    if suffix == ".csv":
        df: pd.DataFrame = _read_csv_bytes(uploaded_file.read())
        source_label = Path(uploaded_file.name).stem
        st.caption("Detected mode: flat table (CSV)")
        st.markdown("### Uploaded Table Preview")
        st.dataframe(df.head(100), use_container_width=True)
        render_preview_excel_download(df, f"{source_label}_raw_preview.xlsx", "unified_csv_raw_preview_download")
        render_cleanup_and_preview(df, source_label, "unified_flat")
        return

    xls = pd.ExcelFile(uploaded_file)
    uploaded_file.seek(0)
    file_bytes = uploaded_file.read()

    if len(xls.sheet_names) > 1:
        workbook_mode = st.radio(
            "Workbook mode",
            [
                "Process one sheet",
                "Process selected sheets separately (recommended)",
            ],
            index=1,
            key="unified_workbook_mode",
        )
        if workbook_mode == "Process selected sheets separately (recommended)":
            selected_sheets = st.multiselect(
                "Sheets to flatten (kept as separate outputs)",
                options=xls.sheet_names,
                default=xls.sheet_names,
                key="unified_separate_sheets",
            )
            if not selected_sheets:
                st.warning("Select at least one sheet.")
                return

            import hashlib as _hl
            split_sig = _hl.md5(
                file_bytes + f"|{'|'.join(selected_sheets)}|mode=separate".encode()
            ).hexdigest()
            split_cache_key = f"_separate_tabs_{split_sig}"

            if st.button("Build separate flat tables", key="unified_separate_build", use_container_width=True):
                output_frames = {}
                progress = st.progress(0)
                status = st.empty()
                for idx, selected_sheet in enumerate(selected_sheets):
                    status.text(f"Extracting {idx + 1}/{len(selected_sheets)}: {selected_sheet}")
                    try:
                        output_frames[selected_sheet] = _extract_like_one_sheet_mode(
                            file_bytes,
                            selected_sheet,
                        )
                    except Exception as exc:
                        st.warning(f"Failed to process sheet '{selected_sheet}': {exc}")
                    progress.progress((idx + 1) / len(selected_sheets))

                st.session_state[split_cache_key] = output_frames
                progress.empty()
                status.empty()

            split_outputs = st.session_state.get(split_cache_key)
            if split_outputs is None:
                st.info("Click 'Build separate flat tables' to flatten selected sheets without merging them.")
                return
            if not split_outputs:
                st.warning("No rows found from selected sheets.")
                return

            available_sheets = [name for name, frame in split_outputs.items() if frame is not None and not frame.empty]
            if not available_sheets:
                st.warning("All selected sheets returned empty outputs.")
                return

            st.success(f"Flattened {len(available_sheets)} sheet(s) as separate outputs.")

            st.success(f"Flattened {len(available_sheets)} sheet(s).")

            # --- Select which sheets to include ---
            sheets_to_include = st.multiselect(
                "Sheets to include in final download",
                options=available_sheets,
                default=available_sheets,
                key="unified_separate_sheets_to_include",
            )
            if not sheets_to_include:
                st.warning("Select at least one sheet to download.")
                return

            # --- Unified cleanup settings (apply to all selected sheets) ---
            st.markdown("### Cleanup Settings (apply to all selected sheets)")
            with st.expander("Optional: configure cleanup rules", expanded=False):
                col_strip, col_clean = st.columns(2)
                with col_strip:
                    strip_text = st.checkbox("Trim whitespace in text columns", value=True,
                                             key="unified_separate_strip")
                with col_clean:
                    clean_names = st.checkbox("Clean final column names", value=False,
                                              key="unified_separate_clean_names")

            # --- Preview first selected sheet ---
            preview_sheet = st.selectbox(
                "Preview a sheet before download",
                options=sheets_to_include,
                key="unified_separate_preview_sheet",
            )
            preview_df = split_outputs[preview_sheet]
            st.markdown(f"### {preview_sheet} Preview")
            st.dataframe(preview_df.head(150), use_container_width=True)

            # --- One-click download all ---
            def _download_all_batch():
                batch = {}
                for sname in sheets_to_include:
                    sdf = finalize_extracted_sheet(
                        split_outputs[sname].copy(),
                        strip_text=strip_text,
                        split_hierarchy=True,
                    )
                    batch[sname] = sdf
                return to_multisheet_excel_bytes(batch)

            import hashlib as _hl_batch
            batch_sig = _hl_batch.md5(
                str((sheets_to_include, strip_text, clean_names)).encode()
            ).hexdigest()
            batch_cache_key = f"_batch_download_{batch_sig}"

            if batch_cache_key not in st.session_state:
                st.session_state[batch_cache_key] = _download_all_batch()

            st.download_button(
                f"⬇️ Download {len(sheets_to_include)} sheet(s) as Excel workbook",
                data=st.session_state[batch_cache_key],
                file_name=f"{Path(uploaded_file.name).stem}_flat_all_sheets.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="unified_separate_download_batch",
            )
            return

    # --- Single sheet or individual sheet processing ---
    sheet_name = st.selectbox("Sheet", xls.sheet_names, key="unified_sheet")
    assert isinstance(sheet_name, str), "Expected sheet_name to be a string"

    # --- Step 1: light quick read to detect structure (no heavy extraction yet) ---
    # Read with pandas first for instant structural heuristic check. Cache by bytes.
    quick_df = _quick_read_cached(file_bytes, sheet_name)
    suggested_layout, detection_reason, has_report_signals, has_matrix_signals = infer_report_layout_from_quick_df(
        quick_df)
    extraction_profile = suggested_layout

    if not has_report_signals and not has_matrix_signals:
        # Looks flat — skip extraction entirely, no delay
        suggested_mode_default = "flat"
        reason = "No report block signals detected. Treating as flat table."
    else:
        suggested_mode_default = "report"
        reason = f"{detection_reason} Using {suggested_layout} extraction profile."

    mode_options = {
        f"Auto ({suggested_mode_default})": suggested_mode_default,
        "Treat as flat table": "flat",
        "Treat as report blocks": "report",
    }
    selection = st.selectbox(
        "Parsing mode",
        list(mode_options.keys()),
        index=0,
        help="Auto mode is recommended; override only if needed.",
        key="unified_parsing_mode",
    )
    active_mode = mode_options[selection]
    st.caption(f"Detection note: {reason}")

    source_label = f"{Path(uploaded_file.name).stem}_{sheet_name}"

    if active_mode == "flat":
        flat_df: pd.DataFrame = pd.DataFrame(pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name))

        st.markdown("### Uploaded Table Preview")
        st.dataframe(flat_df.head(100), use_container_width=True)
        render_preview_excel_download(
            flat_df,
            f"{source_label}_raw_preview.xlsx",
            "unified_flat_raw_preview_download",
        )
        render_cleanup_and_preview(flat_df, source_label, "unified_flat")
        return

    # --- Step 2: report mode — run heavy extraction only when needed ---
    raw_df = _read_display_sheet_cached(file_bytes, sheet_name)

    st.markdown("### Raw Sheet Preview")
    st.dataframe(raw_df.head(40), use_container_width=True)
    render_preview_excel_download(
        raw_df,
        f"{source_label}_raw_preview.xlsx",
        "unified_report_raw_preview_download",
    )

    with st.spinner("Extracting report blocks…"):
        extracted_df = _extract_report_cached(file_bytes, sheet_name, extraction_profile=extraction_profile)

    if extracted_df.empty:
        # Fall back to a direct flat read so users are not blocked on complex/matrix layouts.
        fallback_df: pd.DataFrame = pd.DataFrame(pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name))
        if fallback_df.empty:
            st.info(
                "File structure could not be auto-flattened and no flat rows were detected for this sheet.\n\n"
                "Try a different tab (for example a summary tab) or export only the data range and re-upload."
            )
            render_raw_excel_download(uploaded_file, "unified_raw_fallback")
            return

        st.warning(
            "Auto report extraction returned no rows for this sheet. "
            "Loaded the sheet in flat mode as a fallback."
        )
        st.markdown("### Flat Fallback Preview")
        st.dataframe(fallback_df.head(100), use_container_width=True)
        render_cleanup_and_preview(fallback_df, source_label, "unified_flat_fallback")
        return

    st.markdown("### Extracted Blocks Preview")
    col_rows, col_cols, col_blocks = st.columns(3)
    col_rows.metric("Rows", f"{len(extracted_df):,}")
    col_cols.metric("Columns", f"{len(extracted_df.columns):,}")
    col_blocks.metric(
        "Blocks",
        f"{extracted_df['block_id'].nunique():,}" if "block_id" in extracted_df.columns else "n/a",
    )
    st.dataframe(extracted_df.head(200), use_container_width=True)

    id_cols = [
        col for col in [
            "block_id",
            "schema_group_id",
            "schema_signature",
        ]
        if col in extracted_df.columns
    ]
    if id_cols:
        mapping_df = extracted_df[id_cols].drop_duplicates().sort_values(id_cols).reset_index(drop=True)
        st.markdown("### Block Mapping")
        st.dataframe(mapping_df.head(200), use_container_width=True)

    notes = extract_bottom_notes(raw_df)
    if notes:
        st.markdown("### Notes")
        st.text_area(
            "Extracted sheet notes",
            value="\n".join(notes),
            height=140,
            disabled=True,
            label_visibility="collapsed",
            key="unified_extracted_notes",
        )

    render_cleanup_and_preview(extracted_df, source_label, "unified_report")



def generate_structural_schema(
        df: pd.DataFrame,
        max_unique_values: int = 100,
        group_by: tuple[str, ...] = ("schema_group_id", "section", "table_name", "block_id"),
) -> dict:
    """Layer 1: Generate structural schema from flat-file output.
    
    Groups by a structural context (default: block_id + section + table_name) and extracts:
    - Column names, types, null counts
    - Unique values per dimension column
    - Fact column statistics
    - Row counts
    
    Returns dict keyed by group identifier with schema metadata.
    """
    if df is None or df.empty:
        return {}

    working_df = _assign_schema_group_id(df) if "schema_group_id" not in df.columns else df.copy()
    schema = {}

    grouping_cols = [col for col in group_by if col in working_df.columns]
    grouped_frames: list[tuple[dict, pd.DataFrame]] = []

    if grouping_cols:
        grouped = working_df.groupby(grouping_cols, dropna=False, sort=False)
        for keys, group_df in grouped:
            if not isinstance(keys, tuple):
                keys = (keys,)
            context = {
                col: "" if pd.isna(value) else str(value)
                for col, value in zip(grouping_cols, keys)
            }
            grouped_frames.append((context, group_df))
    else:
        grouped_frames.append(({"group": "default"}, working_df))

    for context, block_df in grouped_frames:
        key_parts = [f"{col}={context.get(col, '')}" for col in grouping_cols] or ["default"]
        block_key = " | ".join(key_parts)
        
        if block_df.empty:
            continue
        
        dimension_candidates = [
            "line_item", "metric", "period", "valuation_date", "tenor",
            "contract_type", "rate_type", "currency", "section",
            "column_group", "unit", "parent_line_item", "business_line",
        ]
        dimension_cols = [col for col in dimension_candidates if col in block_df.columns]
        
        fact_col = "value" if "value" in block_df.columns else None
        
        schema[block_key] = {
            "group_context": context,
            "row_count": len(block_df),
            "columns": {},
            "dimensions": {},
            "fact_column": fact_col,
            "fact_stats": {},
        }
        
        for col in block_df.columns:
            if col in {
                "block_id",
                "schema_group_id",
                "schema_signature",
                "tab_name",
                "table_name",
                "section",
                "block_start_column",
            }:
                continue
            
            col_data = block_df[col]
            null_count = int(_series_cell_text(col_data).eq("").sum())
            
            schema[block_key]["columns"][col] = {
                "type": str(col_data.dtype),
                "null_count": int(null_count),
                "fill_rate": round(1.0 - (null_count / len(block_df)), 3) if len(block_df) > 0 else 0.0,
            }
        
        for dim_col in dimension_cols:
            if dim_col not in block_df.columns:
                continue
            
            non_blank = [v for v in block_df[dim_col].tolist() if not is_blank(v)]
            unique_vals = sorted({str(v) for v in non_blank})
            unique_count = len(unique_vals)
            
            sampled_unique = unique_vals[:max_unique_values]
            
            schema[block_key]["dimensions"][dim_col] = {
                "unique_count": unique_count,
                "unique_values": sampled_unique,
                "truncated": unique_count > max_unique_values,
            }
        
        if fact_col and fact_col in block_df.columns:
            fact_values = block_df[fact_col].dropna()
            numeric_values = []
            for v in fact_values:
                parsed = _parse_numeric_value(v)
                if parsed is not None:
                    numeric_values.append(parsed)
            
            if numeric_values:
                schema[block_key]["fact_stats"] = {
                    "count": len(numeric_values),
                    "mean": round(sum(numeric_values) / len(numeric_values), 2),
                    "min": round(min(numeric_values), 2),
                    "max": round(max(numeric_values), 2),
                    "sample_values": numeric_values[:5],
                }
    
    return schema

def main():
    st.set_page_config(
        page_title="Flat File Builder",
        page_icon="",
        layout="wide",
    )
    st.title("Flat File Builder")
    st.caption("Upload once, auto-detect structure, then preview and clean the resulting table.")
    render_unified_file_tab()


if __name__ == "__main__":
    main()
