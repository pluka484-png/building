import io
import hashlib
import json
import re
import weakref
import zipfile
from collections import OrderedDict
from dataclasses import dataclass
from datetime import date, datetime
from functools import lru_cache
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
import streamlit as st


# Compile hot regex patterns once.
RE_QUOTED = re.compile(r'"[^"]*"')
RE_BRACKETED = re.compile(r"\[[^]]+]")
RE_ALPHA = re.compile(r"[A-Za-z]")
RE_DIMENSIONS = re.compile(r"([A-Z]+)(\d+)")
RE_METRIC_FOOTNOTE_SUFFIX = re.compile(r"[¹²³⁴⁵⁶⁷⁸⁹⁰]+$")
RE_METRIC_TRAILING_ENUM = re.compile(r"\s+\d+\)$")
RE_MONTH_YEAR_TOKEN = re.compile(r"\b(jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec)\w*\s+\d{2,4}\b")
RE_DURATION_TOKEN = re.compile(r"\d{1,3}[YyMmWwDd]|[Oo][Nn]|[Tt][Nn]|[Ss][Ww]|[Ss][Nn]")
RE_PERIOD_TOKEN = re.compile(r"\b(q[1-4]|fy\s*\d{2,4}|ytd|mtd|qtd)\b")
RE_DATE_DMY = re.compile(r"^\d{1,2}[./-]\d{1,2}[./-]\d{2,4}$")
RE_DATE_YEARISH = re.compile(r"^(19|20)\d{2}([./-]\d{1,2}){0,2}$")
RE_UPPER_CODE = re.compile(r"[A-Z]{2,6}")
RE_HEADER_PERIOD_QUALIFIER = re.compile(
    r"^(?:"
    r"[AN]\s*\d{4}|"
    r"(?:19|20)\d{2}[AE]?|"
    r"Q[1-4]\s*\d{2,4}|"
    r"FY\s*\d{2,4}|"
    r"H[12]\s*\d{2,4}|"
    r"(?:3|6|9|12)M\s*\d{2,4}|"
    r"YTD|MTD|QTD"
    r")$",
    re.IGNORECASE,
)
RE_NUMERIC_TOKEN = re.compile(r"^[+-]?(?:\d+\.?\d*|\.\d+)(?:[eE][+-]?\d+)?$")
RE_DELTA_PREFIX = re.compile(r"(?:[∆Δ]|(?i:delta))\s*(.+)")
RE_SPLIT_ON_SLASH = re.compile(r"\s*/\s*")
RE_DOT_DATE_CAPTURE = re.compile(r"(\d{1,2})\.(\d{1,2})\.(\d{2,4})")
RE_CURRENCY_PREFIX = re.compile(
    r"^(?:[$€£¥]|(?:EUR|USD|GBP|CHF|JPY|CNY|CAD|AUD)\s+)",
    re.IGNORECASE,
)
RE_MAGNITUDE_SUFFIX = re.compile(r"([KMBT])$", re.IGNORECASE)
RE_SCIENTIFIC_EXPONENT = re.compile(r"[eE]([+-]?\d+)$")
RE_UNIT_LABEL = re.compile(
    r"^(?:%|in\s+%|(?:EUR|USD|GBP|CHF|JPY|CNY|CAD|AUD)"
    r"(?:\s+(?:mn|m|bn|b|thousand|million|billion|per\s+unit))?)$",
    re.IGNORECASE,
)

ERROR_TOKENS = {"#REF!", "#DIV/0!", "#N/A", "#VALUE!", "#NULL!", "#NUM!", "#ERROR!", "#NAME?"}
BLANK_TEXT_TOKENS = {"", "nan", "none"}
ROW_ATTR_KEYS = {
    "excel_indents",
    "excel_row_outline_levels",
    "excel_row_bold_flags",
    "excel_row_merged_flags",
    "excel_row_indent_levels",
    "excel_row_top_rule_flags",
    "excel_row_bottom_rule_flags",
}


class _SharedAttrValue(list):
    """List subclass for df.attrs payloads that derived frames share.

    pandas' __finalize__ deep-copies attrs into every derived frame (each
    slice, filter, or copy), and the extractor's style/merge metadata is
    O(cells), so that deepcopy dominated large-file profiles. Extraction code
    treats attrs values as immutable — it always builds replacement lists,
    never mutates them in place — so sharing one instance is safe and turns
    the per-operation copy into O(1).
    """
    __slots__ = ()

    def __deepcopy__(self, memo):
        return self


def _freeze_attrs(df: pd.DataFrame) -> pd.DataFrame:
    """Wrap list-valued attrs so pandas shares them instead of deep-copying."""
    attrs = getattr(df, "attrs", None)
    if attrs:
        for key, value in attrs.items():
            if isinstance(value, list) and not isinstance(value, _SharedAttrValue):
                attrs[key] = _SharedAttrValue(value)
    return df


@dataclass
class SheetScanContext:
    text: np.ndarray
    blank_mask: np.ndarray
    numeric_mask: np.ndarray
    unit_mask: np.ndarray
    data_like_mask: np.ndarray
    non_blank_count: np.ndarray
    first_non_blank_col: np.ndarray
    row_outline_levels: np.ndarray
    row_bold_flags: np.ndarray
    row_merged_flags: np.ndarray
    row_indent_levels: np.ndarray


@dataclass(frozen=True)
class TableRegion:
    start_row: int
    end_row: int
    start_col: int
    end_col: int

    @property
    def source_range(self) -> str:
        start = f"{_excel_col_letter(self.start_col + 1)}{self.start_row + 1}"
        end = f"{_excel_col_letter(self.end_col)}{self.end_row}"
        return f"{start}:{end}"


_SHEET_SCAN_CONTEXT_CACHE: dict[
    int,
    tuple[weakref.ReferenceType[pd.DataFrame], tuple[int, int], SheetScanContext],
] = {}


def _slice_with_row_attrs(df: pd.DataFrame, start: int, end: int) -> pd.DataFrame:
    sliced = df.iloc[start:end].reset_index(drop=True)
    attrs = getattr(df, "attrs", {})
    if not attrs:
        return sliced

    new_attrs = {}
    for key, value in attrs.items():
        if key == "_scan_context":
            continue
        if key in ROW_ATTR_KEYS and isinstance(value, (list, tuple, np.ndarray)):
            new_attrs[key] = list(value[start:end])
        else:
            new_attrs[key] = value
    sliced.attrs = new_attrs
    return _freeze_attrs(sliced)


@lru_cache(maxsize=65536)
def _normalize_cell_text(raw_text: str) -> str:
    compact = " ".join(raw_text.replace("\n", " ").split())
    return "" if compact.upper() in ERROR_TOKENS else compact


def _is_blank_text(raw_text: str) -> bool:
    return raw_text.strip().lower() in BLANK_TEXT_TOKENS


def is_blank(value) -> bool:
    if value is None or value is pd.NA:
        return True
    if isinstance(value, (float, np.floating)):
        if np.isnan(value):
            return True
    else:
        try:
            # NaN is the only value that is not equal to itself.
            if value != value:
                return True
        except Exception:
            pass
    return _is_blank_text(str(value))


def cell_text(value) -> str:
    if is_blank(value):
        return ""
    return _normalize_cell_text(str(value))


def _series_cell_text(series: pd.Series) -> pd.Series:
    """Vectorized cell_text equivalent for pandas Series."""
    text = (
        series.astype("string")
        .str.replace("\n", " ", regex=False)
        .str.split()
        .str.join(" ")
        .fillna("")
    )
    text = text.mask(text.str.upper().isin(ERROR_TOKENS), "")
    text = text.mask(text.str.lower().isin(BLANK_TEXT_TOKENS), "")
    return text


def _looks_like_date_or_period_text(text: str) -> bool:
    return _looks_like_date_or_period_token(str(text or "").strip())


@lru_cache(maxsize=65536)
def _looks_like_date_or_period_token(token: str) -> bool:
    if not token:
        return False
    if RE_DATE_DMY.fullmatch(token) or RE_DATE_YEARISH.fullmatch(token):
        return True
    if RE_HEADER_PERIOD_QUALIFIER.fullmatch(token):
        return True
    if RE_DURATION_TOKEN.fullmatch(token):
        return True
    if RE_MONTH_YEAR_TOKEN.search(token) or RE_PERIOD_TOKEN.fullmatch(token.lower()):
        return True
    if re.fullmatch(r"(?:[1-4]Q|Q[1-4])\s*\d{2,4}", token, flags=re.IGNORECASE):
        return True
    if re.fullmatch(r"(?:0?[1-9]|1[0-2])M\s*\d{2,4}", token, flags=re.IGNORECASE):
        return True
    return False


def _parse_numeric_text_for_detection(text: str) -> float | None:
    """Parse numeric-looking display text without destroying date/code semantics."""
    return _parse_numeric_token_for_detection(str(text or "").strip())


@lru_cache(maxsize=65536)
def _parse_numeric_token_for_detection(token: str) -> float | None:
    if not token or _looks_like_date_or_period_token(token):
        return None

    negative_parentheses = token.startswith("(") and token.endswith(")")
    if negative_parentheses:
        token = token[1:-1].strip()

    token = RE_CURRENCY_PREFIX.sub("", token).strip()
    token = token.replace("−", "-").replace(",", "").replace(" ", "")
    if token.endswith("%"):
        token = token[:-1]

    multiplier = 1.0
    magnitude_match = RE_MAGNITUDE_SUFFIX.search(token)
    if magnitude_match:
        multiplier = {
            "K": 1e3,
            "M": 1e6,
            "B": 1e9,
            "T": 1e12,
        }[magnitude_match.group(1).upper()]
        token = token[:-1]

    if not token or token in {"-", "+", ".", "-.", "+."}:
        return None
    if not RE_NUMERIC_TOKEN.fullmatch(token):
        return None

    exponent_match = RE_SCIENTIFIC_EXPONENT.search(token)
    if exponent_match and abs(int(exponent_match.group(1))) > 100:
        return None

    try:
        parsed = float(token) * multiplier
    except (ValueError, OverflowError):
        return None
    return -abs(parsed) if negative_parentheses else parsed


def _is_numeric_like_text(text: str) -> bool:
    return _parse_numeric_text_for_detection(text) is not None


def clean_column_name(value) -> str:
    text = str(value).strip().lower()
    text = "".join(ch if ch.isalnum() else "_" for ch in text)
    while "__" in text:
        text = text.replace("__", "_")
    text = text.strip("_") or "unnamed_column"
    if text[0].isdigit():
        text = f"col_{text}"
    return text


def dedupe_columns(columns):
    counts = {}
    result = []
    for col in columns:
        base = clean_column_name(col)
        counts[base] = counts.get(base, 0) + 1
        result.append(base if counts[base] == 1 else f"{base}_{counts[base]}")
    return result


def primary_excel_format_section(number_format: str) -> str:
    section = str(number_format or "General").split(";")[0]
    section = RE_QUOTED.sub("", section)
    section = RE_BRACKETED.sub("", section)
    return section


def decimal_places_from_excel_format(number_format: str) -> int | None:
    section = primary_excel_format_section(number_format)
    if "." not in section:
        return 0 if any(token in section for token in ("0", "#", "?")) else None
    decimals = section.split(".", 1)[1]
    decimals = decimals.split("%", 1)[0]
    placeholders = [char for char in decimals if char in {"0", "#", "?"}]
    return len(placeholders) if placeholders else 0


def formatted_excel_value(cell):
    value = getattr(cell, "value", None)
    if value is None:
        return ""

    if isinstance(value, (datetime, date)):
        return value.strftime("%d.%m.%Y")

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        number_format = str(getattr(cell, "number_format", "General") or "General")
        decimals = decimal_places_from_excel_format(number_format)
        if decimals is None:
            return value

        is_percent = "%" in primary_excel_format_section(number_format)
        display_value = value * 100 if is_percent else value
        use_grouping = "," in number_format
        show_plus = "+" in number_format and display_value > 0
        sign = "+" if show_plus else ""
        grouped = "," if use_grouping else ""
        rendered = f"{sign}{display_value:{grouped}.{decimals}f}"
        if is_percent:
            rendered = f"{rendered}%"
        return rendered

    return value


def row_has_values(row, start_col=1) -> bool:
    return any(not is_blank(v) for v in row.iloc[start_col:].tolist())


def first_nonblank_after(row, start_col=1) -> str:
    for value in row.iloc[start_col:].tolist():
        if not is_blank(value):
            return cell_text(value)
    return ""


def is_unit_label(value) -> bool:
    return _is_unit_label_text(cell_text(value))


@lru_cache(maxsize=8192)
def _is_unit_label_text(text: str) -> bool:
    return bool(RE_UNIT_LABEL.fullmatch(text))


def nonblank_col_indexes(row, start_col=1):
    return [
        idx for idx in range(start_col, len(row))
        if not is_blank(row.iloc[idx])
    ]


def _excel_col_letter(col_index_1based: int) -> str:
    if col_index_1based <= 0:
        return ""
    label = ""
    value = int(col_index_1based)
    while value > 0:
        value, remainder = divmod(value - 1, 26)
        label = chr(ord("A") + remainder) + label
    return label


def _infer_header_rows_for_debug(df: pd.DataFrame, max_scan: int = 20) -> list[int]:
    """Return likely header rows for diagnostics and regression checks."""
    if df is None or df.empty:
        return []

    context = _sheet_scan_context(df)
    row_limit = min(int(max_scan), len(df))
    header_rows: list[int] = []
    for row_idx in range(row_limit):
        non_blank = int(context.non_blank_count[row_idx])
        if non_blank == 0:
            continue
        period_hits = sum(
            1
            for token in context.text[row_idx].tolist()
            if token and _looks_like_date_or_period_text(str(token))
        )
        numeric_hits = max(int(context.numeric_mask[row_idx].sum()) - period_hits, 0)
        text_hits = non_blank - numeric_hits
        if numeric_hits > max(3, text_hits):
            break
        header_rows.append(row_idx)
    return header_rows



def _build_sheet_scan_context(raw_df: pd.DataFrame) -> SheetScanContext:
    values = raw_df.to_numpy(dtype=object, copy=False)
    row_count = values.shape[0]

    row_outline_levels = np.asarray(raw_df.attrs.get("excel_row_outline_levels") or [0] * row_count, dtype=np.int32)
    row_bold_flags = np.asarray(raw_df.attrs.get("excel_row_bold_flags") or [False] * row_count, dtype=bool)
    row_merged_flags = np.asarray(raw_df.attrs.get("excel_row_merged_flags") or [False] * row_count, dtype=bool)
    row_indent_levels = np.asarray(raw_df.attrs.get("excel_row_indent_levels") or [0] * row_count, dtype=np.int32)

    if row_outline_levels.shape[0] != row_count:
        row_outline_levels = np.resize(row_outline_levels, row_count)
    if row_bold_flags.shape[0] != row_count:
        row_bold_flags = np.resize(row_bold_flags, row_count)
    if row_merged_flags.shape[0] != row_count:
        row_merged_flags = np.resize(row_merged_flags, row_count)
    if row_indent_levels.shape[0] != row_count:
        row_indent_levels = np.resize(row_indent_levels, row_count)

    if values.size == 0:
        empty = np.empty(values.shape, dtype=object)
        false_mask = np.zeros(values.shape, dtype=bool)
        return SheetScanContext(
            text=empty,
            blank_mask=false_mask,
            numeric_mask=false_mask,
            unit_mask=false_mask,
            data_like_mask=false_mask,
            non_blank_count=np.zeros((values.shape[0],), dtype=np.int32),
            first_non_blank_col=np.full((values.shape[0],), -1, dtype=np.int32),
            row_outline_levels=row_outline_levels,
            row_bold_flags=row_bold_flags,
            row_merged_flags=row_merged_flags,
            row_indent_levels=row_indent_levels,
        )

    flat_values = values.ravel()
    cell_count = flat_values.shape[0]
    text_flat = np.empty(cell_count, dtype=object)
    blank_flat = np.zeros(cell_count, dtype=bool)
    numeric_flat = np.zeros(cell_count, dtype=bool)
    unit_flat = np.zeros(cell_count, dtype=bool)
    data_like_flat = np.zeros(cell_count, dtype=bool)

    # Sheets repeat the same cell values heavily (blank markers, units, section
    # labels, common numbers), so classify each distinct value once.
    blank_entry = ("", True, False, False, False)
    nm_tokens = {"n.m.", "n.m", "nm"}
    memo: dict = {}
    for idx, value in enumerate(flat_values):
        # Key on (type, value): 1, 1.0, and True are equal/hash-equal dict
        # keys but stringify differently, so they must not share an entry.
        entry = None
        hashable = True
        try:
            key = (value.__class__, value)
            entry = memo.get(key)
        except TypeError:
            hashable = False

        if entry is None:
            if is_blank(value):
                entry = blank_entry
            else:
                text = _normalize_cell_text(str(value))
                if not text:
                    entry = blank_entry
                else:
                    unit_like = _is_unit_label_text(text)
                    if isinstance(value, (int, float, np.integer, np.floating)) and not isinstance(value, bool):
                        numeric_like = not bool(np.isnan(value)) if isinstance(value, (float, np.floating)) else True
                    else:
                        numeric_like = _parse_numeric_token_for_detection(text) is not None
                    nm_like = text.lower() in nm_tokens
                    data_like = (numeric_like or nm_like) and not unit_like
                    entry = (text, False, numeric_like, unit_like, data_like)
            if hashable:
                memo[key] = entry

        text_flat[idx] = entry[0]
        blank_flat[idx] = entry[1]
        numeric_flat[idx] = entry[2]
        unit_flat[idx] = entry[3]
        data_like_flat[idx] = entry[4]

    text = text_flat.reshape(values.shape)
    blank_mask = blank_flat.reshape(values.shape)
    numeric_mask = numeric_flat.reshape(values.shape)
    unit_mask = unit_flat.reshape(values.shape)
    data_like_mask = data_like_flat.reshape(values.shape)
    non_blank_count = (~blank_mask).sum(axis=1).astype(np.int32)
    has_non_blank = non_blank_count > 0
    first_non_blank_col = np.full(values.shape[0], -1, dtype=np.int32)
    if has_non_blank.any():
        first_non_blank_col[has_non_blank] = (~blank_mask[has_non_blank]).argmax(axis=1)
    return SheetScanContext(
        text=text,
        blank_mask=blank_mask,
        numeric_mask=numeric_mask,
        unit_mask=unit_mask,
        data_like_mask=data_like_mask,
        non_blank_count=non_blank_count,
        first_non_blank_col=first_non_blank_col,
        row_outline_levels=row_outline_levels,
        row_bold_flags=row_bold_flags,
        row_merged_flags=row_merged_flags,
        row_indent_levels=row_indent_levels,
    )


def _sheet_scan_context(raw_df: pd.DataFrame) -> SheetScanContext:
    cache_key = id(raw_df)
    cached = _SHEET_SCAN_CONTEXT_CACHE.get(cache_key)
    if cached is not None:
        frame_ref, cached_shape, context = cached
        if frame_ref() is raw_df and cached_shape == raw_df.shape:
            return context

    context = _build_sheet_scan_context(raw_df)

    def _discard_context(_):
        _SHEET_SCAN_CONTEXT_CACHE.pop(cache_key, None)

    _SHEET_SCAN_CONTEXT_CACHE[cache_key] = (
        weakref.ref(raw_df, _discard_context),
        raw_df.shape,
        context,
    )
    return context


def _skip_error_heavy_rows(raw_df: pd.DataFrame, max_check: int = 30) -> pd.DataFrame:
    """Skip leading rows that are mostly error tokens or blank.

    Benelux-style sheets often have rows of #REF! errors before actual headers.
    This function identifies and skips such leading junk rows.
    """
    if raw_df is None or raw_df.empty:
        return raw_df

    error_tokens = {"#REF!", "#DIV/0!", "#N/A", "#VALUE!", "#NULL!", "#NUM!", "#ERROR!", "#NAME?"}

    def _row_error_density(row_idx: int) -> float:
        """Return fraction of cells that are error tokens or blank."""
        row = raw_df.iloc[row_idx]
        total = len(row)
        if total == 0:
            return 1.0
        errors_and_blanks = sum(
            1 for v in row
            if is_blank(v) or str(v).strip().upper() in error_tokens
        )
        return errors_and_blanks / total

    # Find first meaningful row. Prefer explicit leading labels so sparse title rows
    # are preserved (for example, section/table names above unit rows).
    skip_until = 0
    for idx in range(min(max_check, len(raw_df))):
        lead_limit = min(3, raw_df.shape[1])
        leading_tokens = [cell_text(raw_df.iloc[idx, col_idx]) for col_idx in range(lead_limit)]
        has_leading_content = any(token and not is_unit_label(token) for token in leading_tokens)
        if has_leading_content:
            skip_until = idx
            break
        if _row_error_density(idx) < 0.80:
            skip_until = idx
            break

    if skip_until > 0:
        return _slice_with_row_attrs(raw_df, skip_until, len(raw_df))
    return raw_df


def drop_all_blank_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame() if df is None else df
    context = _sheet_scan_context(df)
    keep_mask = (~context.blank_mask).any(axis=0)
    return df.loc[:, keep_mask]


def drop_empty_or_zero_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Remove columns containing only blanks and numeric zero values.

    Text labels, dates, errors, and any non-zero value keep the column. Excel
    provenance and merged-range metadata are remapped to the reduced grid.
    Columns participating in merged ranges are retained because their position
    carries header structure even when their visible cells are blank or zero.
    """
    if df is None or df.empty or df.shape[1] <= 1:
        return pd.DataFrame() if df is None else df

    values = df.to_numpy(dtype=object, copy=False)
    structural_columns: set[int] = set()
    merged_ranges = df.attrs.get("excel_merged_ranges")
    if isinstance(merged_ranges, list):
        for merged in merged_ranges:
            if not isinstance(merged, dict):
                continue
            min_col = max(int(merged.get("min_col", 0)), 0)
            max_col = min(int(merged.get("max_col", 0)), values.shape[1])
            structural_columns.update(range(min_col, max_col))

    keep_indexes: list[int] = []
    for col_idx in range(values.shape[1]):
        if col_idx in structural_columns:
            keep_indexes.append(col_idx)
            continue
        meaningful = False
        for value in values[:, col_idx]:
            if is_blank(value):
                continue
            parsed = _parse_numeric_value(value)
            if parsed is not None and parsed == 0:
                continue
            meaningful = True
            break
        if meaningful:
            keep_indexes.append(col_idx)

    if not keep_indexes:
        return df.iloc[:, :1].copy()
    if len(keep_indexes) == df.shape[1]:
        return df

    result = df.iloc[:, keep_indexes].copy()
    attrs = dict(getattr(df, "attrs", {}))

    indents = attrs.get("excel_indents")
    if isinstance(indents, (list, tuple, np.ndarray)):
        attrs["excel_indents"] = [
            [row[col_idx] for col_idx in keep_indexes if col_idx < len(row)]
            for row in indents
        ]

    old_col_map = attrs.get("excel_col_map")
    if isinstance(old_col_map, list) and len(old_col_map) == df.shape[1]:
        attrs["excel_col_map"] = [old_col_map[idx] for idx in keep_indexes]
    else:
        attrs["excel_col_map"] = keep_indexes

    old_to_new = {
        old_idx: new_idx
        for new_idx, old_idx in enumerate(keep_indexes)
    }
    merged_ranges = attrs.get("excel_merged_ranges")
    if isinstance(merged_ranges, list):
        remapped_ranges = []
        for merged in merged_ranges:
            if not isinstance(merged, dict):
                continue
            retained = [
                old_to_new[col_idx]
                for col_idx in range(
                    int(merged.get("min_col", 0)),
                    int(merged.get("max_col", 0)),
                )
                if col_idx in old_to_new
            ]
            if not retained:
                continue
            remapped_ranges.append(
                {
                    **merged,
                    "min_col": min(retained),
                    "max_col": max(retained) + 1,
                }
            )
        attrs["excel_merged_ranges"] = remapped_ranges

    attrs.pop("_scan_context", None)
    result.attrs = attrs
    return _freeze_attrs(result)


def trim_effective_row_window(df: pd.DataFrame, tail_buffer: int = 3) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame() if df is None else df
    non_blank_rows = df.ne("").any(axis=1).to_numpy(dtype=bool)
    if not bool(non_blank_rows.any()):
        return _slice_with_row_attrs(df, 0, 0)

    idx = np.flatnonzero(non_blank_rows)
    start = int(idx[0])
    end = min(int(idx[-1]) + 1 + max(int(tail_buffer), 0), len(df))
    if start == 0 and end == len(df):
        return df
    return _slice_with_row_attrs(df, start, end)


def _true_runs(mask: np.ndarray) -> list[tuple[int, int]]:
    """Return half-open runs where a one-dimensional boolean mask is true."""
    if mask.size == 0 or not bool(mask.any()):
        return []
    padded = np.concatenate(([False], mask.astype(bool), [False]))
    transitions = np.flatnonzero(padded[1:] != padded[:-1])
    return [
        (int(transitions[idx]), int(transitions[idx + 1]))
        for idx in range(0, len(transitions), 2)
    ]


def _row_bands_with_blank_gaps(
        occupied_rows: np.ndarray,
        min_blank_gap: int = 2,
) -> list[tuple[int, int]]:
    """Split used rows when there is a meaningful vertical whitespace gap."""
    used_rows = np.flatnonzero(occupied_rows)
    if used_rows.size == 0:
        return []

    bands: list[tuple[int, int]] = []
    start = int(used_rows[0])
    previous = int(used_rows[0])
    for row_idx in used_rows[1:]:
        row_idx = int(row_idx)
        blank_gap = row_idx - previous - 1
        if blank_gap >= max(int(min_blank_gap), 1):
            bands.append((start, previous + 1))
            start = row_idx
        previous = row_idx
    bands.append((start, previous + 1))
    return bands


def _region_has_table_signal(region_df: pd.DataFrame) -> bool:
    if region_df is None or region_df.empty:
        return False
    if region_df.shape[1] < 2:
        return False

    values = region_df.to_numpy(dtype=object, copy=False)
    if region_df.shape[0] < 3:
        has_header_row = False
        for row in values:
            tokens = [value for value in row if not is_blank(value)]
            text_tokens = [
                cell_text(value)
                for value in tokens
                if _parse_numeric_value(value) is None
            ]
            if len(text_tokens) >= 2 and len(text_tokens) == len(tokens):
                has_header_row = True
                break
        if not has_header_row:
            return False

    for row in values:
        non_blank = [idx for idx, value in enumerate(row) if not is_blank(value)]
        if len(non_blank) < 2:
            continue
        first_col = non_blank[0]
        label = cell_text(row[first_col])
        if not label or _parse_numeric_value(row[first_col]) is not None:
            continue
        payload = [row[idx] for idx in non_blank[1:]]
        if any(_parse_numeric_value(value) is not None for value in payload):
            return True
    return False


def _lightweight_blank_mask(raw_df: pd.DataFrame) -> np.ndarray:
    """Blank-only first pass used before any full cell classification."""
    values = raw_df.to_numpy(dtype=object, copy=False)
    flat_values = values.ravel()
    blank_flat = np.zeros(flat_values.shape[0], dtype=bool)
    for idx, value in enumerate(flat_values):
        if is_blank(value):
            blank_flat[idx] = True
            continue
        blank_flat[idx] = not bool(_normalize_cell_text(str(value)))
    blank_mask = blank_flat.reshape(values.shape)

    # Values copied across merged cells are useful for header interpretation,
    # but they must not bridge independent tables during whitespace-based
    # region detection. Treat every merged cell except its top-left anchor as
    # structurally blank in this first pass.
    merged_ranges = raw_df.attrs.get("excel_merged_ranges")
    if isinstance(merged_ranges, list):
        for merged in merged_ranges:
            if not isinstance(merged, dict):
                continue
            min_row = max(int(merged.get("min_row", 0)), 0)
            max_row = min(int(merged.get("max_row", 0)), values.shape[0])
            min_col = max(int(merged.get("min_col", 0)), 0)
            max_col = min(int(merged.get("max_col", 0)), values.shape[1])
            if min_row >= max_row or min_col >= max_col:
                continue
            blank_mask[min_row:max_row, min_col:max_col] = True
            blank_mask[min_row, min_col] = False
    return blank_mask


def detect_table_regions(
        raw_df: pd.DataFrame,
        min_blank_row_gap: int = 2,
) -> list[TableRegion]:
    """Detect independent table rectangles separated by worksheet whitespace."""
    if raw_df is None or raw_df.empty:
        return []

    occupied = ~_lightweight_blank_mask(raw_df)
    row_bands = _row_bands_with_blank_gaps(
        occupied.any(axis=1),
        min_blank_gap=min_blank_row_gap,
    )

    regions: list[TableRegion] = []
    for start_row, end_row in row_bands:
        band_occupied = occupied[start_row:end_row]
        active_columns = band_occupied.any(axis=0)
        for start_col, end_col in _true_runs(active_columns):
            region_occupied = band_occupied[:, start_col:end_col]
            active_rows = region_occupied.any(axis=1)
            if not bool(active_rows.any()):
                continue

            local_rows = np.flatnonzero(active_rows)
            tight_start_row = start_row + int(local_rows[0])
            tight_end_row = start_row + int(local_rows[-1]) + 1
            region = TableRegion(
                start_row=tight_start_row,
                end_row=tight_end_row,
                start_col=start_col,
                end_col=end_col,
            )
            region_df = raw_df.iloc[
                region.start_row:region.end_row,
                region.start_col:region.end_col,
            ]
            if not _region_has_table_signal(region_df):
                continue
            regions.extend(_split_table_region_vertically(raw_df, region))

    regions.sort(key=lambda item: (item.start_row, item.start_col))
    return regions


def _parent_report_titles_for_regions(
        raw_df: pd.DataFrame,
        regions: list[TableRegion],
) -> dict[int, str]:
    """Map child regions to the nearest merged title spanning multiple regions."""
    merged_ranges = raw_df.attrs.get("excel_merged_ranges")
    if not isinstance(merged_ranges, list) or not regions:
        return {}

    # Sibling merged ranges on one row are column-group headers, not titles.
    merged_rows_seen: dict[int, int] = {}
    for merged in merged_ranges:
        if isinstance(merged, dict):
            row_key = int(merged.get("min_row", 0))
            merged_rows_seen[row_key] = merged_rows_seen.get(row_key, 0) + 1

    candidates: list[tuple[int, int, str, list[int]]] = []
    for merged in merged_ranges:
        if not isinstance(merged, dict):
            continue
        title = cell_text(merged.get("text", ""))
        if not title or is_unit_label(title) or _looks_like_date_or_period_text(title):
            continue
        if merged_rows_seen.get(int(merged.get("min_row", 0)), 0) >= 2:
            continue
        min_row = int(merged.get("min_row", 0))
        max_row = int(merged.get("max_row", 0))
        min_col = int(merged.get("min_col", 0))
        max_col = int(merged.get("max_col", 0))
        if max_col - min_col < 2:
            continue

        child_indexes: list[int] = []
        for region_idx, region in enumerate(regions, start=1):
            horizontal_overlap = min(max_col, region.end_col) - max(min_col, region.start_col)
            row_distance = max(region.start_row - max_row, 0)
            title_is_near_region = (
                min_row <= region.start_row + 3
                and row_distance <= 8
            )
            if horizontal_overlap > 0 and title_is_near_region:
                child_indexes.append(region_idx)

        if len(child_indexes) >= 2:
            candidates.append((max_row, min_row, title, child_indexes))

    # Process higher titles first so a closer/lower merged title replaces a
    # broad workbook-level heading for the affected children.
    mapping: dict[int, str] = {}
    for _, _, title, child_indexes in sorted(candidates, key=lambda item: (item[0], item[1])):
        for region_idx in child_indexes:
            mapping[region_idx] = title
    return mapping



def _apply_sheet_level_context(
        extracted: pd.DataFrame,
        raw_df: pd.DataFrame,
) -> pd.DataFrame:
    """Apply shared headers and merged parent titles after all rescue paths."""
    if extracted is None or extracted.empty:
        return extracted
    working = extracted.copy()

    if {"source_row", "source_column_index"}.issubset(working.columns):
        context = _sheet_scan_context(raw_df)
        header_candidates: list[tuple[int, dict[int, str]]] = []
        for row_idx in range(len(raw_df)):
            period_count = sum(
                1
                for token in context.text[row_idx].tolist()
                if str(token).strip()
                and _looks_like_date_or_period_text(str(token).strip())
            )
            if period_count < 2:
                continue
            headers: dict[int, str] = {}
            for col_idx in range(raw_df.shape[1]):
                token = str(context.text[row_idx, col_idx]).strip()
                if (
                    not token
                    or is_unit_label(token)
                    or token.lower() in {
                        "account",
                        "description",
                        "item",
                        "line item",
                        "metric",
                        "risk metric",
                    }
                ):
                    continue
                headers[col_idx + 1] = token
            header_candidates.append((row_idx + 1, headers))

        source_rows = pd.to_numeric(working["source_row"], errors="coerce")
        source_cols = pd.to_numeric(
            working["source_column_index"], errors="coerce"
        )
        inherited_values: list[str] = []
        for source_row, source_col in zip(
                source_rows.tolist(),
                source_cols.tolist(),
        ):
            inherited = ""
            if not pd.isna(source_row) and not pd.isna(source_col):
                applicable = [
                    (header_row, headers.get(int(source_col), ""))
                    for header_row, headers in header_candidates
                    if header_row < int(source_row)
                    and headers.get(int(source_col), "")
                ]
                if applicable:
                    inherited = max(applicable, key=lambda item: item[0])[1]
            inherited_values.append(inherited)

        existing_shared = (
            _series_cell_text(working["shared_header"])
            if "shared_header" in working.columns
            else pd.Series("", index=working.index, dtype="string")
        )
        working["shared_header"] = [
            existing or inherited
            for existing, inherited in zip(
                existing_shared.tolist(),
                inherited_values,
            )
        ]
        # Fold row-specific inherited headers into metrics.
        updated_metrics: list[str] = []
        for metric_value, inherited in zip(
                working.get("metric", pd.Series("", index=working.index)).tolist(),
                working["shared_header"].tolist(),
        ):
            metric = cell_text(metric_value)
            inherited_text = cell_text(inherited)
            parts = [part.strip() for part in metric.split("|") if part.strip()]
            if not inherited_text or inherited_text.lower() in {
                part.lower() for part in parts
            }:
                updated_metrics.append(metric)
            elif not metric or re.fullmatch(
                    r"Column\s+[A-Z]+", metric, flags=re.IGNORECASE
            ):
                updated_metrics.append(inherited_text)
            else:
                updated_metrics.append(f"{inherited_text} | {metric}")
        working["metric"] = updated_metrics
        for context_column in (
            "metric_type",
            "metric_date",
            "comparison_date",
            "metric_quarter",
            "comparison_year",
        ):
            if context_column not in working.columns:
                continue
            existing = _series_cell_text(working[context_column])
            parsed_values = [
                metric_context(metric).get(context_column, "")
                for metric in updated_metrics
            ]
            working[context_column] = working[context_column].where(
                existing.ne(""),
                parsed_values,
            )

    merged_ranges = raw_df.attrs.get("excel_merged_ranges")
    if (
        isinstance(merged_ranges, list)
        and {"table_id", "source_row", "source_column_index"}.issubset(working.columns)
    ):
        # Several merged ranges sharing one row are column-group headers
        # (segment / region / scenario bands over value columns), not report
        # titles — a real parent title is the only merged range on its row.
        merged_rows_seen: dict[int, int] = {}
        for merged in merged_ranges:
            if isinstance(merged, dict):
                row_key = int(merged.get("min_row", 0))
                merged_rows_seen[row_key] = merged_rows_seen.get(row_key, 0) + 1

        assignments: dict[str, tuple[int, str]] = {}
        for merged in merged_ranges:
            if not isinstance(merged, dict):
                continue
            title = cell_text(merged.get("text", ""))
            if not title or is_unit_label(title) or _looks_like_date_or_period_text(title):
                continue
            if merged_rows_seen.get(int(merged.get("min_row", 0)), 0) >= 2:
                continue
            min_row = int(merged.get("min_row", 0)) + 1
            max_row = int(merged.get("max_row", 0))
            min_col = int(merged.get("min_col", 0)) + 1
            max_col = int(merged.get("max_col", 0))
            row_numbers = pd.to_numeric(working["source_row"], errors="coerce")
            col_numbers = pd.to_numeric(
                working["source_column_index"], errors="coerce"
            )
            governed = working.loc[
                row_numbers.gt(max_row)
                & col_numbers.ge(min_col)
                & col_numbers.le(max_col)
            ]
            # A title governs tables that start near it or that chain onto an
            # already-governed sibling (stacked dashboard mini-tables).
            # Without this bound a wide heading would claim every table
            # further down the sheet in its column span.
            governed_rows_numeric = pd.to_numeric(governed["source_row"], errors="coerce")
            governed_ids = governed["table_id"].map(cell_text)
            child_spans = sorted(
                (
                    (float(rows.min()), float(rows.max()), child_id)
                    for child_id, rows in governed_rows_numeric.groupby(governed_ids)
                    if child_id and not rows.dropna().empty
                ),
                key=lambda span: span[0],
            )
            child_ids: list[str] = []
            frontier = float(max_row)
            for child_min, child_max, child_id in child_spans:
                if child_min - frontier > 8:
                    continue
                child_ids.append(child_id)
                frontier = max(frontier, child_max)
            if len(child_ids) < 2:
                continue
            # A governing title must span (most of) its children horizontally;
            # a narrow band over a slice of the value columns is a column
            # header even when it is the only merged range on its row.
            governed_cols = pd.to_numeric(
                governed["source_column_index"], errors="coerce"
            ).dropna()
            child_ids_all = working["table_id"].map(cell_text)
            children_cols = pd.to_numeric(
                working.loc[child_ids_all.isin(child_ids), "source_column_index"],
                errors="coerce",
            ).dropna()
            if not children_cols.empty:
                child_extent = float(children_cols.max() - children_cols.min()) + 1.0
                merged_width = float(max_col - min_col) + 1.0
                if merged_width < 0.5 * child_extent:
                    continue
            for child_id in child_ids:
                previous = assignments.get(child_id)
                if previous is None or min_row >= previous[0]:
                    assignments[child_id] = (min_row, title)
        if assignments:
            existing_parent = (
                _series_cell_text(working["parent_report_name"])
                if "parent_report_name" in working.columns
                else pd.Series("", index=working.index, dtype="string")
            )
            working["parent_report_name"] = [
                existing or assignments.get(cell_text(table_id), (0, ""))[1]
                for existing, table_id in zip(
                    existing_parent.tolist(),
                    working["table_id"].tolist(),
                )
            ]

    # Ruled titles: a single-text row carrying a horizontal rule (like the
    # "A S S E T S" line above a balance-sheet block) groups the rows beneath
    # it. Fill the section dimension from the nearest ruled title above each
    # observation, but never overwrite sections detected by other means.
    if "source_row" in working.columns:
        top_rules = raw_df.attrs.get("excel_row_top_rule_flags") or []
        bottom_rules = raw_df.attrs.get("excel_row_bottom_rule_flags") or []
        if top_rules or bottom_rules:
            context = _sheet_scan_context(raw_df)
            ruled_titles: list[tuple[int, str]] = []
            for row_idx in range(len(raw_df)):
                has_rule = (
                    (row_idx < len(top_rules) and bool(top_rules[row_idx]))
                    or (row_idx < len(bottom_rules) and bool(bottom_rules[row_idx]))
                )
                if not has_rule:
                    continue
                non_blank = int(context.non_blank_count[row_idx])
                if non_blank == 0 or non_blank > 2:
                    continue
                if bool(context.data_like_mask[row_idx].any()):
                    continue
                first_col = int(context.first_non_blank_col[row_idx])
                token = str(context.text[row_idx, first_col]).strip()
                if not token or is_unit_label(token) or _looks_like_date_or_period_text(token):
                    continue
                ruled_titles.append((row_idx + 1, token))

            if ruled_titles:
                source_rows_numeric = pd.to_numeric(working["source_row"], errors="coerce")
                existing_section = (
                    _series_cell_text(working["section"])
                    if "section" in working.columns
                    else pd.Series("", index=working.index, dtype="string")
                )
                filled_sections: list[str] = []
                for existing, source_row in zip(
                        existing_section.tolist(),
                        source_rows_numeric.tolist(),
                ):
                    if existing or pd.isna(source_row):
                        filled_sections.append(existing)
                        continue
                    inherited = ""
                    for title_row, title in ruled_titles:
                        if title_row < source_row:
                            inherited = title
                        else:
                            break
                    filled_sections.append(inherited)
                if any(filled_sections):
                    working["section"] = filled_sections
    return working


def _slice_table_region(raw_df: pd.DataFrame, region: TableRegion) -> pd.DataFrame:
    """Slice a table region while retaining only metadata relevant to that rectangle."""
    sliced = raw_df.iloc[
        region.start_row:region.end_row,
        region.start_col:region.end_col,
    ].copy()
    sliced = sliced.reset_index(drop=True)
    sliced.columns = range(sliced.shape[1])

    attrs = getattr(raw_df, "attrs", {})
    new_attrs: dict = {
        "_source_row_offset": int(region.start_row),
        "_source_col_offset": int(region.start_col),
        "_source_range": region.source_range,
    }

    for key in (
        "excel_row_outline_levels",
        "excel_row_bold_flags",
        "excel_row_merged_flags",
        "excel_row_indent_levels",
        "excel_row_top_rule_flags",
        "excel_row_bottom_rule_flags",
    ):
        value = attrs.get(key)
        if isinstance(value, (list, tuple, np.ndarray)):
            new_attrs[key] = list(value[region.start_row:region.end_row])

    merged_ranges = attrs.get("excel_merged_ranges")
    if isinstance(merged_ranges, list):
        relevant_merged_ranges = []
        for merged in merged_ranges:
            if not isinstance(merged, dict):
                continue
            min_row = int(merged.get("min_row", 0))
            max_row = int(merged.get("max_row", 0))
            min_col = int(merged.get("min_col", 0))
            max_col = int(merged.get("max_col", 0))
            if (
                max_row <= region.start_row
                or min_row >= region.end_row
                or max_col <= region.start_col
                or min_col >= region.end_col
            ):
                continue
            relevant_merged_ranges.append(
                {
                    **merged,
                    "min_row": max(min_row, region.start_row) - region.start_row,
                    "max_row": min(max_row, region.end_row) - region.start_row,
                    "min_col": max(min_col, region.start_col) - region.start_col,
                    "max_col": min(max_col, region.end_col) - region.start_col,
                }
            )
        new_attrs["excel_merged_ranges"] = relevant_merged_ranges

    indents = attrs.get("excel_indents")
    if isinstance(indents, (list, tuple, np.ndarray)):
        new_attrs["excel_indents"] = [
            list(row[region.start_col:region.end_col])
            for row in indents[region.start_row:region.end_row]
        ]

    col_map = attrs.get("excel_col_map")
    if isinstance(col_map, list) and len(col_map) >= region.end_col:
        new_attrs["excel_col_map"] = list(col_map[region.start_col:region.end_col])
    else:
        new_attrs["excel_col_map"] = list(range(region.start_col, region.end_col))

    sliced.attrs = new_attrs
    return _freeze_attrs(sliced)


def columns_semantically_equal(left: pd.Series, right: pd.Series) -> bool:
    if len(left) != len(right):
        return False
    for left_value, right_value in zip(left.tolist(), right.tolist()):
        if is_blank(left_value) and is_blank(right_value):
            continue
        if cell_text(left_value) != cell_text(right_value):
            return False
    return True


def looks_like_data_value(value) -> bool:
    text = cell_text(value)
    if not text or is_unit_label(text):
        return False
    if metric_context(text)["metric_type"]:
        return False
    cleaned = text.replace(",", "").replace("%", "").replace("+", "").replace("−", "-").strip()
    if cleaned.lower() in {"n.m.", "n.m", "nm"}:
        return True
    try:
        float(cleaned)
        return True
    except ValueError:
        return False


def is_section_header_row(row, expected_col: int | None = None, allow_hybrid: bool = True) -> bool:
    """Determine if a row is a section header, including hybrid header+data rows."""
    non_blank_indexes = [idx for idx, value in enumerate(row.tolist()) if not is_blank(value)]
    if not non_blank_indexes:
        return False

    col_idx = non_blank_indexes[0]
    value = row.iloc[col_idx]

    if expected_col is not None and col_idx != expected_col:
        return False

    text = cell_text(value)
    if not text:
        return False
    if _is_numeric_like(value):
        return False
    if is_unit_label(value):
        return False
    if looks_like_data_value(value):
        return False

    if len(non_blank_indexes) == 1:
        return True

    if not allow_hybrid:
        return False

    remaining = [row.iloc[idx] for idx in non_blank_indexes[1:]]
    return all(_is_numeric_like(v) or looks_like_data_value(v) for v in remaining)


def detect_hierarchy_level(raw_df: pd.DataFrame) -> np.ndarray:
    """Vectorized hierarchy level detection using non-blank position + Excel metadata."""
    context = _sheet_scan_context(raw_df)
    row_count = context.blank_mask.shape[0]
    if row_count == 0:
        return np.asarray([], dtype=np.int32)

    non_blank_mask = ~context.blank_mask
    levels = context.first_non_blank_col.copy()
    blank_rows = ~non_blank_mask.any(axis=1)

    # Blend structural depth with explicit Excel grouping/indent metadata when present.
    levels = np.maximum(levels, context.row_indent_levels)
    levels = np.maximum(levels, context.row_outline_levels)
    levels[blank_rows] = -1
    return levels.astype(np.int32)


def _section_header_confidence(context: SheetScanContext, allow_hybrid: bool = True) -> np.ndarray:
    non_blank = ~context.blank_mask
    row_count, _ = non_blank.shape
    confidence = np.zeros(row_count, dtype=np.int8)

    has_any = context.first_non_blank_col >= 0
    row_idx = np.arange(row_count)
    safe_first_col = np.where(has_any, context.first_non_blank_col, 0)

    first_numeric = context.numeric_mask[row_idx, safe_first_col]
    first_unit = context.unit_mask[row_idx, safe_first_col]
    first_data_like = context.data_like_mask[row_idx, safe_first_col]
    first_text_like = has_any & ~first_numeric & ~first_unit & ~first_data_like

    confidence += first_text_like.astype(np.int8)

    tail_valid_matrix = context.blank_mask | context.numeric_mask | context.data_like_mask
    tail_valid_matrix[row_idx[has_any], safe_first_col[has_any]] = True
    tail_numeric_or_data = tail_valid_matrix.all(axis=1)
    strict_single = context.non_blank_count == 1
    hybrid_signal = strict_single | (allow_hybrid & (context.non_blank_count > 1) & tail_numeric_or_data)
    confidence += hybrid_signal.astype(np.int8)

    depth_signal = ((context.first_non_blank_col > 0) | (context.row_indent_levels > 0)).astype(np.int8)
    confidence += depth_signal
    confidence += context.row_bold_flags.astype(np.int8)
    confidence += (context.row_outline_levels > 0).astype(np.int8)
    confidence += context.row_merged_flags.astype(np.int8)
    return confidence


def propagate_section_context(raw_df: pd.DataFrame, section_col: int = 0) -> pd.DataFrame:
    """Propagate section labels by hierarchy depth using vectorized masks and ffill."""
    if raw_df is None or raw_df.empty:
        return pd.DataFrame() if raw_df is None else raw_df.copy()

    context = _sheet_scan_context(raw_df)
    levels = detect_hierarchy_level(raw_df)
    max_depth = int(levels.max()) if levels.size else -1
    max_depth = min(max_depth, 8)
    if max_depth < 0:
        result = raw_df.copy()
        result["_section_depth"] = levels
        return result

    confidence = _section_header_confidence(context, allow_hybrid=True)
    row_idx = np.arange(len(raw_df))
    has_any = context.first_non_blank_col >= 0
    safe_first_col = np.where(has_any, context.first_non_blank_col, 0)

    first_text_like = has_any & ~context.numeric_mask[row_idx, safe_first_col] & ~context.unit_mask[row_idx, safe_first_col]
    header_mask = first_text_like & (confidence >= 2)

    result = raw_df.copy()
    result["_section_depth"] = levels
    result["line_item"] = np.where(has_any, context.text[row_idx, safe_first_col], "")

    for depth in range(max_depth + 1):
        col_name = f"_section_L{depth}"
        is_header_at_depth = header_mask & (levels == depth)
        # Forward-fill header text via a numpy cummax-of-index trick rather than
        # pd.Series.ffill(): an object column with zero non-null values (no header
        # at this depth) is common here and triggers pandas' ffill/fillna
        # downcasting FutureWarning on real-sized data. The "" placeholder for
        # rows before the first header is harmless — every non-value column gets
        # a blanket .fillna("") a few steps downstream regardless.
        header_values = np.where(is_header_at_depth, context.text[row_idx, safe_first_col], "")
        fill_idx = np.where(is_header_at_depth, row_idx, 0)
        np.maximum.accumulate(fill_idx, out=fill_idx)
        result[col_name] = header_values[fill_idx]

        group_name = f"_section_gid_L{depth}"
        result[group_name] = pd.Series(is_header_at_depth.astype(np.int32), index=raw_df.index).cumsum()

    return result


def detect_section_column(raw_df: pd.DataFrame, allow_hybrid: bool = True) -> int | None:
    """Identify the dominant section column across standalone and hybrid headers."""
    if raw_df is None or raw_df.empty:
        return None

    context = _sheet_scan_context(raw_df)
    row_count = context.blank_mask.shape[0]
    if row_count == 0:
        return None

    row_idx = np.arange(row_count)
    has_any = context.first_non_blank_col >= 0
    if not bool(has_any.any()):
        return None

    safe_first_col = np.where(has_any, context.first_non_blank_col, 0)
    first_text_like = (
        has_any
        & ~context.numeric_mask[row_idx, safe_first_col]
        & ~context.unit_mask[row_idx, safe_first_col]
        & ~context.data_like_mask[row_idx, safe_first_col]
    )

    tail_valid_matrix = context.blank_mask | context.numeric_mask | context.data_like_mask
    tail_valid_matrix[row_idx[has_any], safe_first_col[has_any]] = True
    tail_ok = tail_valid_matrix.all(axis=1)

    strict_mask = first_text_like & (context.non_blank_count == 1)
    hybrid_mask = first_text_like & (context.non_blank_count > 1) & tail_ok
    structural_mask = strict_mask | (hybrid_mask if allow_hybrid else np.zeros_like(hybrid_mask, dtype=bool))

    confidence = _section_header_confidence(context, allow_hybrid=allow_hybrid)
    candidate_mask = structural_mask & (confidence >= 2)
    candidate_cols = context.first_non_blank_col[candidate_mask]

    if candidate_cols.size == 0:
        return None

    counts = np.bincount(candidate_cols.astype(np.int32))
    return int(counts.argmax()) if counts.size else None


def _is_numeric_like(value) -> bool:
    """Quick check: can this value be parsed as numeric after removing formatting."""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return True
    text = cell_text(value)
    if not text:
        return False
    # LBYL path avoids expensive exception-driven float parsing on text-heavy sheets.
    return _is_numeric_like_text(text)


def _parse_numeric_value(value) -> float | None:
    """Parse a numeric cell value while tolerating display formatting."""
    if value is None or is_blank(value):
        return None
    if isinstance(value, (int, float, np.integer, np.floating)) and not isinstance(value, bool):
        return float(value)

    text = cell_text(value)
    if not text:
        return None
    return _parse_numeric_text_for_detection(text)



def find_unit_header_cells(raw_df: pd.DataFrame):
    cells = []
    if raw_df is None or raw_df.empty:
        return cells
    context = _sheet_scan_context(raw_df)
    unit_prefixes = ("eur ", "usd ", "gbp ", "in %")
    row_count, col_count = context.text.shape
    for row_idx in range(row_count):
        for col_idx in range(col_count):
            text = str(context.text[row_idx, col_idx]).lower()
            if not text:
                continue
            if not (context.unit_mask[row_idx, col_idx] or text.startswith(unit_prefixes)):
                continue
            if (~context.blank_mask[row_idx, col_idx + 1:]).any():
                cells.append((row_idx, col_idx))
    return cells


def find_general_header_cells(
        raw_df: pd.DataFrame,
        max_rows: int | None = None,
        max_cols: int | None = None,
        stop_after_first: bool = False,
):
    """Find likely mini-table header cells without relying on sheet-specific names."""
    cells = []
    if raw_df is None or raw_df.empty:
        return cells
    header_prefixes = (
        "eur ", "usd ", "gbp ", "chf ", "aud ", "in %", "vs.", "vs ",
        "as of ", "period", "quarter", "year",
    )
    row_limit = len(raw_df) - 1
    if max_rows is not None:
        row_limit = min(row_limit, max_rows)

    col_limit = raw_df.shape[1]
    if max_cols is not None:
        col_limit = min(col_limit, max_cols)

    for row_idx in range(row_limit):
        row = raw_df.iloc[row_idx]
        for col_idx, value in enumerate(row.iloc[:col_limit].tolist()):
            label = cell_text(value)
            if not label:
                continue
            header_cols = [
                idx for idx in nonblank_col_indexes(row, start_col=col_idx + 1)
                if idx < col_limit
            ]
            if len(header_cols) < 2:
                continue

            data_rows = 0
            for next_idx in range(row_idx + 1, min(len(raw_df), row_idx + 8)):
                next_row = raw_df.iloc[next_idx]
                line_item = cell_text(next_row.iloc[col_idx])
                if not line_item:
                    continue
                has_values = any(
                    col < len(next_row) and not is_blank(next_row.iloc[col])
                    for col in header_cols[:12]
                )
                if has_values:
                    data_rows += 1

            looks_like_unit = label.lower().startswith(header_prefixes)
            if looks_like_unit or data_rows >= 2:
                cells.append((row_idx, col_idx))
                if stop_after_first:
                    return cells
    return cells


def sheet_looks_general_report_tables(raw_df: pd.DataFrame) -> bool:
    return len(
        find_general_header_cells(
            raw_df,
            max_rows=120,
            max_cols=30,
            stop_after_first=True,
        )
    ) >= 1


def section_name(value) -> str:
    text = cell_text(value)
    text = text.replace("¹", "").replace("²", "").replace("³", "")
    return text.strip()


def section_name_validated(row, expected_col: int | None = None, allow_hybrid: bool = True) -> str:
    """Extract section text only when the row matches standalone header structure."""
    if not is_section_header_row(row, expected_col=expected_col, allow_hybrid=allow_hybrid):
        return ""

    for value in row.tolist():
        if not is_blank(value):
            text = cell_text(value)
            text = text.replace("¹", "").replace("²", "").replace("³", "")
            return text.strip()

    return ""


def metric_parse_text(value) -> str:
    text = cell_text(value)
    text = RE_METRIC_FOOTNOTE_SUFFIX.sub("", text).strip()
    text = RE_METRIC_TRAILING_ENUM.sub("", text).strip()
    return text


def carry_forward_header_values(values):
    carried = []
    current = ""
    for value in values:
        if not is_blank(value):
            current = cell_text(value)
        carried.append(current)
    return carried


def _detect_summary_header_rows(raw_df: pd.DataFrame, max_scan: int = 20) -> list[int]:
    if raw_df is None or raw_df.empty:
        return []
    context = _sheet_scan_context(raw_df)
    values = raw_df.to_numpy(dtype=object, copy=False)
    row_count, col_count = raw_df.shape
    header_rows: list[int] = []
    blank_seen = False

    for row_idx in range(min(max_scan, row_count)):
        non_blank = int(context.non_blank_count[row_idx])
        if non_blank == 0:
            if header_rows:
                blank_seen = True
            continue

        period_hits = 0
        financial_hits = 0
        for col_idx in range(col_count):
            token = str(context.text[row_idx, col_idx]).strip()
            if token and RE_HEADER_PERIOD_QUALIFIER.fullmatch(token):
                period_hits += 1
            if col_idx >= 2:
                parsed = _parse_numeric_value(values[row_idx, col_idx])
                if parsed is not None and abs(parsed) >= 100:
                    financial_hits += 1

        if financial_hits >= 2 and header_rows:
            break
        if blank_seen and financial_hits >= 1:
            break
        header_rows.append(row_idx)

        # Summary sheets usually have a compact 1-3 row header band.
        if period_hits >= 2 and len(header_rows) >= 2:
            continue

    return header_rows


def _detect_summary_data_region(raw_df: pd.DataFrame, max_scan: int = 250) -> tuple[int, int]:
    if raw_df is None or raw_df.empty:
        return -1, -1
    row_count, col_count = raw_df.shape
    values = raw_df.to_numpy(dtype=object, copy=False)
    candidates: list[tuple[int, int, int]] = []

    for row_idx in range(min(max_scan, row_count)):
        first_numeric = -1
        numeric_hits = 0
        for col_idx in range(col_count):
            if _parse_numeric_value(values[row_idx, col_idx]) is None:
                continue
            numeric_hits += 1
            if first_numeric < 0:
                first_numeric = col_idx

        if numeric_hits < 2 or first_numeric < 0:
            continue
        label_hits = sum(
            1
            for col_idx in range(min(first_numeric, 10))
            if cell_text(values[row_idx, col_idx])
        )
        if label_hits == 0:
            continue
        candidates.append((row_idx, first_numeric, numeric_hits))

    if not candidates:
        return -1, -1

    best_row = -1
    best_score = -1
    best_value_start = -1
    for row_idx, first_col, _ in candidates:
        window_end = row_idx + 20
        cluster = [
            cand_first
            for cand_row, cand_first, cand_hits in candidates
            if row_idx <= cand_row <= window_end and abs(cand_first - first_col) <= 2 and cand_hits >= 2
        ]
        score = len(cluster)
        if score > best_score:
            best_score = score
            best_row = row_idx
            best_value_start = int(np.median(cluster)) if cluster else first_col

    if best_row < 0 or best_value_start < 0:
        return -1, -1
    return best_row, best_value_start


def _detect_summary_header_band(raw_df: pd.DataFrame, data_start: int, data_col_start: int) -> list[int]:
    if raw_df is None or raw_df.empty or data_start <= 0:
        return []
    context = _sheet_scan_context(raw_df)
    col_count = raw_df.shape[1]
    values = raw_df.to_numpy(dtype=object, copy=False)

    # Prefer a top-anchored header band and stop at first separator blank row.
    top_rows: list[int] = []
    saw_non_blank = False
    for row_idx in range(min(data_start, len(raw_df))):
        non_blank = int(context.non_blank_count[row_idx])
        if non_blank == 0:
            if saw_non_blank and len(top_rows) >= 2:
                break
            continue

        saw_non_blank = True
        label_hits = sum(
            1
            for col_idx in range(min(data_col_start, 10))
            if cell_text(values[row_idx, col_idx])
        )
        numeric_hits = sum(
            1
            for col_idx in range(data_col_start, col_count)
            if _parse_numeric_value(values[row_idx, col_idx]) is not None
        )
        if label_hits >= 1 and numeric_hits >= 3:
            break
        top_rows.append(row_idx)

    if len(top_rows) >= 2:
        return top_rows

    rows: list[int] = []
    band_start = max(0, data_start - 15)
    for row_idx in range(band_start, data_start):
        if int(context.non_blank_count[row_idx]) == 0:
            continue
        label_hits = sum(
            1
            for col_idx in range(min(data_col_start, 10))
            if cell_text(values[row_idx, col_idx])
        )
        numeric_hits = sum(
            1
            for col_idx in range(data_col_start, col_count)
            if _parse_numeric_value(values[row_idx, col_idx]) is not None
        )
        # Exclude likely data rows from the header band.
        if label_hits >= 1 and numeric_hits >= 3:
            continue
        non_blank_data = sum(
            1
            for col_idx in range(data_col_start, col_count)
            if cell_text(values[row_idx, col_idx])
        )
        if non_blank_data == 0:
            continue
        rows.append(row_idx)
    return rows


def _is_header_noise_token(token: str) -> bool:
    text = str(token or "").strip()
    if not text:
        return True
    lower = text.lower()
    if lower.startswith("as per "):
        return True
    if "review update" in lower or "bom update" in lower:
        return True
    if len(text) > 40 and RE_DOT_DATE_CAPTURE.search(text):
        return True
    return False


def _compact_summary_header_tokens(tokens: list[str], max_tokens: int = 4) -> list[str]:
    compact: list[str] = []
    for token in tokens:
        text = str(token or "").strip()
        if not text or _is_header_noise_token(text):
            continue

        # Drop measure-like numerics from header chains while keeping year tokens.
        if text.isdigit():
            number = int(text)
            if not (1900 <= number <= 2100):
                continue

        if compact and compact[-1] == text:
            continue
        compact.append(text)

    # Remove immediate repeated suffix patterns, e.g. A|B|C|X|B|C -> A|B|C|X
    if len(compact) >= 6:
        for size in range(2, min(4, len(compact) // 2 + 1)):
            left = compact[-2 * size: -size]
            right = compact[-size:]
            if left == right:
                compact = compact[:-size]
                break

    if len(compact) <= max_tokens:
        return compact
    return [*compact[: max_tokens - 1], compact[-1]]


def _build_summary_composite_headers(raw_df: pd.DataFrame, header_rows: list[int]) -> dict[int, str]:
    row_count, col_count = raw_df.shape
    if row_count == 0 or col_count == 0:
        return {}

    context = _sheet_scan_context(raw_df)
    carried_rows: list[list[str]] = []
    for row_idx in header_rows:
        if 0 <= row_idx < row_count:
            carried_rows.append(carry_forward_header_values(context.text[row_idx].tolist()))

    excel_col_map = raw_df.attrs.get("excel_col_map")
    if not isinstance(excel_col_map, list) or len(excel_col_map) != col_count:
        excel_col_map = list(range(col_count))

    headers: dict[int, str] = {}
    for col_idx in range(col_count):
        parts: list[str] = []
        for carried in carried_rows:
            token = str(carried[col_idx]).strip() if col_idx < len(carried) else ""
            if not token:
                continue
            if parts and parts[-1] == token:
                continue
            parts.append(token)

        parts = _compact_summary_header_tokens(parts, max_tokens=4)

        if parts:
            headers[col_idx] = " | ".join(parts)
        else:
            excel_idx0 = int(excel_col_map[col_idx]) if col_idx < len(excel_col_map) else col_idx
            headers[col_idx] = f"Column {_excel_col_letter(excel_idx0 + 1)}"
    return headers


def _looks_like_financial_summary_layout(raw_df: pd.DataFrame) -> bool:
    if raw_df is None or raw_df.empty or raw_df.shape[1] < 6 or raw_df.shape[0] < 6:
        return False

    data_start, data_col_start = _detect_summary_data_region(raw_df, max_scan=180)
    if data_start < 0 or data_col_start < 2:
        return False

    header_rows = _detect_summary_header_band(raw_df, data_start, data_col_start)
    if len(header_rows) < 2:
        header_rows = _detect_summary_header_rows(raw_df, max_scan=15)
    if len(header_rows) < 2:
        return False

    headers = _build_summary_composite_headers(raw_df, header_rows)
    period_header_count = 0
    for label in headers.values():
        tokens = [part.strip() for part in str(label).split("|") if part.strip()]
        if any(RE_HEADER_PERIOD_QUALIFIER.fullmatch(token) for token in tokens):
            period_header_count += 1
    if period_header_count < 2:
        return False

    # Require at least one row with a descriptive label and numeric payload in C+.
    sample_end = min(len(raw_df), data_start + 40)
    for row_idx in range(data_start, sample_end):
        row = raw_df.iloc[row_idx]
        label_hits = sum(1 for col_idx in range(min(data_col_start, 10)) if cell_text(row.iloc[col_idx]))
        numeric_hits = sum(1 for col_idx in range(data_col_start, raw_df.shape[1]) if _parse_numeric_value(row.iloc[col_idx]) is not None)
        if label_hits >= 1 and numeric_hits >= 1:
            return True
    return False


def auto_flatten_financial_summary_layout(raw_df: pd.DataFrame) -> pd.DataFrame:
    """Extract sparse financial summary matrices while preserving true column identity."""
    if raw_df is None or raw_df.empty:
        return pd.DataFrame()
    if not _looks_like_financial_summary_layout(raw_df):
        return pd.DataFrame()

    row_count, col_count = raw_df.shape
    values = raw_df.to_numpy(dtype=object, copy=False)
    if col_count < 3:
        return pd.DataFrame()

    data_start, data_col_start = _detect_summary_data_region(raw_df, max_scan=250)
    if data_start < 0 or data_col_start < 2:
        return pd.DataFrame()

    header_rows = _detect_summary_header_band(raw_df, data_start, data_col_start)
    if len(header_rows) < 2:
        header_rows = _detect_summary_header_rows(raw_df, max_scan=20)
    if len(header_rows) < 2:
        return pd.DataFrame()
    headers = _build_summary_composite_headers(raw_df, header_rows)

    if data_col_start >= col_count:
        return pd.DataFrame()

    value_cols: list[int] = []
    body_end = row_count
    for col_idx in range(data_col_start, col_count):
        numeric_hits = 0
        text_hits = 0
        for row_idx in range(data_start, body_end):
            value = values[row_idx, col_idx]
            if is_blank(value):
                continue
            if _parse_numeric_value(value) is not None:
                numeric_hits += 1
            else:
                text_hits += 1
        if numeric_hits >= 1 and numeric_hits >= text_hits:
            value_cols.append(col_idx)

    if not value_cols:
        return pd.DataFrame()

    records: list[dict] = []
    processed_rows: set[int] = set()
    current_parent = ""
    metric_meta_cache: dict[str, dict] = {}
    indents = raw_df.attrs.get("excel_indents") or []

    values = raw_df.to_numpy(copy=False)
    for row_idx in range(data_start, row_count):
        if row_idx in processed_rows:
            continue
        processed_rows.add(row_idx)

        row = values[row_idx]
        has_data = any(not is_blank(row[col_idx]) for col_idx in value_cols)
        if not has_data:
            continue

        left_tokens = [cell_text(row[col_idx]) for col_idx in range(min(data_col_start, col_count))]
        col_a = left_tokens[0] if left_tokens else ""
        col_b = left_tokens[1] if len(left_tokens) > 1 else ""
        primary_label = ""
        for token in left_tokens:
            if token:
                primary_label = token
                break

        if col_a and not is_unit_label(col_a):
            current_parent = col_a

        line_item = col_b or primary_label
        if not line_item:
            continue

        indent_level = 0
        if row_idx < len(indents) and indents[row_idx] and len(indents[row_idx]) > 0:
            indent_level = int(float(indents[row_idx][0] or 0))

        for col_idx in value_cols:
            value = row[col_idx]
            if is_blank(value):
                continue

            metric_name = headers.get(col_idx, f"Column {col_idx + 1}")
            if metric_name not in metric_meta_cache:
                metric_meta_cache[metric_name] = metric_context(metric_name)

            record = {
                "table_name": "Summary Grid",
                "section": current_parent,
                "column_group": metric_name,
                "unit": "",
                "line_item": line_item,
                "parent_line_item": current_parent if current_parent and current_parent != line_item else "",
                "line_item_path": f"{current_parent} > {line_item}" if current_parent and current_parent != line_item else line_item,
                "metric": metric_name,
                "metric_detail": col_b,
                "indent_level": indent_level,
                "value": value,
                "block_key": f"summary_layout:{data_start}",
                "block_start_column": data_col_start,
            }
            record.update(metric_meta_cache[metric_name])
            records.append(record)

    return pd.DataFrame(records)


def auto_flatten_sectioned_financial_sheet(raw_df: pd.DataFrame) -> pd.DataFrame:
    """Flatten stacked financial statement sections into a long analytical table.

    Handles patterns like:
    - Exchange rates / Spot or Average with periods across columns
    - Valuation rates with valuation dates spanning tenor columns
    """
    records = []
    i = 0
    while i < len(raw_df):
        row = raw_df.iloc[i]
        first = cell_text(row.iloc[0])
        first_lower = first.lower()
        rate_label = first_nonblank_after(row, start_col=1)

        if first_lower.startswith("exchange rates") and rate_label:
            rate_type = rate_label
            unit = cell_text(raw_df.iloc[i + 1, 0]) if i + 1 < len(raw_df) else ""
            headers = raw_df.iloc[i + 1].tolist() if i + 1 < len(raw_df) else []
            j = i + 2
            while j < len(raw_df):
                data_row = raw_df.iloc[j]
                label = cell_text(data_row.iloc[0])
                label_lower = label.lower()
                if (
                        not label
                        or label_lower.startswith("exchange rates")
                        or label_lower.startswith("valuation rates")
                        or label_lower.startswith("1)")
                        or label_lower.startswith("2)")
                        or label_lower.startswith("3)")
                ):
                    break
                for col_idx in range(1, len(data_row)):
                    value = data_row.iloc[col_idx]
                    period = cell_text(headers[col_idx]) if col_idx < len(headers) else ""
                    if period and not is_blank(value):
                        records.append({
                            "section": section_name(first),
                            "rate_type": rate_type,
                            "unit": unit,
                            "currency": label,
                            "period": period,
                            "valuation_date": "",
                            "tenor": "",
                            "contract_type": "",
                            "value": value,
                        })
                j += 1
            i = j
            continue

        if first_lower.startswith("valuation rates"):
            date_headers = carry_forward_header_values(row.tolist())
            tenor_row = raw_df.iloc[i + 1].tolist() if i + 1 < len(raw_df) else []
            unit = cell_text(tenor_row[0]) if tenor_row else ""
            current_contract = ""
            j = i + 2
            while j < len(raw_df):
                data_row = raw_df.iloc[j]
                label = cell_text(data_row.iloc[0])
                label_lower = label.lower()
                if (
                        not label
                        or label_lower.startswith("valuation rates")
                        or label_lower.startswith("exchange rates")
                        or label_lower.startswith("1)")
                        or label_lower.startswith("2)")
                        or label_lower.startswith("3)")
                ):
                    break

                if not row_has_values(data_row, start_col=1):
                    current_contract = label
                    j += 1
                    continue

                for col_idx in range(1, len(data_row)):
                    value = data_row.iloc[col_idx]
                    valuation_date = date_headers[col_idx] if col_idx < len(date_headers) else ""
                    tenor = cell_text(tenor_row[col_idx]) if col_idx < len(tenor_row) else ""
                    if valuation_date and tenor and not is_blank(value):
                        records.append({
                            "section": section_name(first),
                            "rate_type": "",
                            "unit": unit,
                            "currency": label,
                            "period": "",
                            "valuation_date": valuation_date,
                            "tenor": tenor,
                            "contract_type": current_contract,
                            "value": value,
                        })
                j += 1
            i = j
            continue

        i += 1

    return pd.DataFrame(records)


def nearest_title_above(raw_df: pd.DataFrame, row_idx: int) -> str:
    for idx in range(row_idx - 1, -1, -1):
        first = cell_text(raw_df.iloc[idx, 0])
        if not first:
            continue
        if first.lower().startswith(("eur ", "usd ", "gbp ", "in %")):
            continue
        if row_has_values(raw_df.iloc[idx], start_col=1):
            continue
        return first
    return ""


def nearest_group_above(raw_df: pd.DataFrame, row_idx: int) -> str:
    for idx in range(row_idx - 1, -1, -1):
        if cell_text(raw_df.iloc[idx, 0]):
            continue
        group = first_nonblank_after(raw_df.iloc[idx], start_col=1)
        if group:
            return group
    return ""


def _is_col0_unit_anchor(value) -> bool:
    text = cell_text(value).lower()
    return text in {"eur", "eur mn", "eur bn", "usd", "usd mn", "usd bn", "gbp", "gbp mn", "gbp bn"} or text.startswith(("eur ", "usd ", "gbp ", "in %"))


def _resolve_block_header_from_unit_anchor(raw_df: pd.DataFrame, unit_row_idx: int) -> tuple[str, str]:
    """Resolve (table_name, column_group) by scanning upward from a unit anchor row.

    Expected pattern:
    - unit row at col 0 (e.g., EUR mn)
    - qualifier row immediately above with text in col 2
    - next non-empty col 0 row above qualifier is section/table name
    """
    if raw_df is None or raw_df.empty or unit_row_idx <= 0:
        return "", ""

    qualifier_row_idx = unit_row_idx - 1
    column_group = ""
    if qualifier_row_idx >= 0 and raw_df.shape[1] > 2:
        qualifier_row = raw_df.iloc[qualifier_row_idx]
        col2_text = cell_text(qualifier_row.iloc[2])
        non_blank = [idx for idx, value in enumerate(qualifier_row.tolist()) if not is_blank(value)]
        if col2_text and (non_blank == [2] or (0 not in non_blank and 2 in non_blank)):
            column_group = col2_text

    scan_idx = qualifier_row_idx - 1 if column_group else qualifier_row_idx
    while scan_idx >= 0:
        candidate = cell_text(raw_df.iloc[scan_idx, 0])
        if candidate:
            return candidate, column_group
        scan_idx -= 1

    return "", column_group


def nearest_left_label_above(raw_df: pd.DataFrame, row_idx: int, col_idx: int) -> str:
    for idx in range(row_idx - 1, -1, -1):
        label = cell_text(raw_df.iloc[idx, col_idx])
        if label and not label.replace(" ", "").isupper():
            return label
    return ""


def nearest_nonblank_right(raw_df: pd.DataFrame, row_idx: int, col_idx: int, block_end: int) -> str:
    row = raw_df.iloc[row_idx]
    for idx in range(col_idx + 1, min(block_end, len(row))):
        value = cell_text(row.iloc[idx])
        if value:
            return value
    return ""


def cell_indent(raw_df: pd.DataFrame, row_idx: int, col_idx: int) -> float:
    indents = raw_df.attrs.get("excel_indents") or []
    value = 0.0
    if row_idx < len(indents) and col_idx < len(indents[row_idx]):
        value = float(indents[row_idx][col_idx] or 0)
    # Row outline levels (the +/- grouping feature) are author-declared
    # hierarchy; treat them as an indent floor for label cells.
    outline_levels = raw_df.attrs.get("excel_row_outline_levels") or []
    if row_idx < len(outline_levels) and outline_levels[row_idx]:
        value = max(value, float(outline_levels[row_idx]))
    return value


def indented_row_context(raw_df: pd.DataFrame, row_idx: int, label_col: int) -> dict:
    current_indent = cell_indent(raw_df, row_idx, label_col)
    if current_indent <= 0:
        return {}

    parents = []
    next_indent = current_indent
    for idx in range(row_idx - 1, -1, -1):
        label = cell_text(raw_df.iloc[idx, label_col])
        if not label:
            continue
        indent = cell_indent(raw_df, idx, label_col)
        if indent < next_indent:
            parents.append(label)
            next_indent = indent
            if indent <= 0:
                break

    parents.reverse()
    if not parents:
        return {}

    line_item = cell_text(raw_df.iloc[row_idx, label_col])
    return {
        "parent_line_item": parents[-1],
        "line_item_path": " > ".join([*parents, line_item]),
    }


def metric_context(metric: str) -> dict:
    # Header labels repeat heavily across rows/blocks; parsing is regex-heavy.
    # Return a shallow copy so callers can never mutate the cached entry
    # (values are flat strings/ints).
    return dict(_metric_context_cached(str(metric)))


@lru_cache(maxsize=8192)
def _metric_context_cached(metric: str) -> dict:
    text = metric_parse_text(metric)
    text = text.strip(" ,;:()[]")
    empty_context = {
        "metric_type": "",
        "metric_date": "",
        "comparison_date": "",
        "metric_quarter": "",
        "comparison_year": "",
    }

    def normalize_year(year_text: str) -> int:
        year = int(year_text)
        return year + 2000 if len(year_text) == 2 else year

    def quarter_from_month(month: int) -> int:
        return ((month - 1) // 3) + 1

    def parse_dot_date_token(date_text: str) -> tuple[int, int, int] | None:
        token = date_text.strip(" ,;:()[]")
        date_match_local = RE_DOT_DATE_CAPTURE.fullmatch(token)
        if not date_match_local:
            date_match_local = RE_DOT_DATE_CAPTURE.search(token)
        if not date_match_local:
            return None
        day_text, month_text, year_text = date_match_local.groups()
        return normalize_year(year_text), int(month_text), int(day_text)

    delta_parts: list[str] = []
    delta_prefix_match = RE_DELTA_PREFIX.match(text)
    if delta_prefix_match:
        delta_body = delta_prefix_match.group(1).strip()
        delta_parts = [part.strip() for part in RE_SPLIT_ON_SLASH.split(delta_body, maxsplit=1)]
        start_date_parts = parse_dot_date_token(delta_parts[0]) if delta_parts else None
        end_date_parts = parse_dot_date_token(delta_parts[1]) if len(delta_parts) == 2 else None
        if (not start_date_parts or not end_date_parts) and len(delta_parts) == 2:
            # Fallback for noisy labels like "31.12.25, / (31.12.24)".
            date_tokens = RE_DOT_DATE_CAPTURE.findall(delta_body)
            if len(date_tokens) >= 2:
                start_date_parts = (
                    normalize_year(date_tokens[0][2]),
                    int(date_tokens[0][1]),
                    int(date_tokens[0][0]),
                )
                end_date_parts = (
                    normalize_year(date_tokens[1][2]),
                    int(date_tokens[1][1]),
                    int(date_tokens[1][0]),
                )

        if start_date_parts and end_date_parts:
            start_year, start_month, start_day = start_date_parts
            end_year, end_month, end_day = end_date_parts
            return {
                **empty_context,
                "metric_type": "date_delta",
                "metric_date": pd.Timestamp(start_year, start_month, start_day),
                "comparison_date": pd.Timestamp(end_year, end_month, end_day),
            }

    date_match = RE_DOT_DATE_CAPTURE.fullmatch(text)
    if date_match:
        day, month, year = date_match.groups()
        normalized_year = normalize_year(year)
        month_int = int(month)
        return {
            **empty_context,
            "metric_type": "date",
            "metric_date": pd.Timestamp(normalized_year, month_int, int(day)),
            "metric_quarter": quarter_from_month(month_int),
        }

    # ISO date/datetime (e.g. 2024-03-31 or 2024-03-31 00:00:00)
    iso_dt_match = re.fullmatch(
        r"(20\d{2})-(\d{1,2})-(\d{1,2})(?:[ T]\d{1,2}:\d{2}(?::\d{2})?)?",
        text,
    )
    if iso_dt_match:
        year_text, month_text, day_text = iso_dt_match.groups()
        year = int(year_text)
        month = int(month_text)
        day = int(day_text)
        return {
            **empty_context,
            "metric_type": "date",
            "metric_date": pd.Timestamp(year, month, day),
            "metric_quarter": quarter_from_month(month),
        }

    # Month period tokens (e.g. 12M24, 12M 2024, 2024 12M)
    month_period_patterns = [
        r"(0?[1-9]|1[0-2])M\s*(\d{2,4})",
        r"(\d{2,4})\s*(0?[1-9]|1[0-2])M",
    ]
    month_period_match = None
    year = None
    month = None
    for idx, pattern in enumerate(month_period_patterns):
        month_period_match = re.fullmatch(pattern, text, flags=re.IGNORECASE)
        if month_period_match:
            if idx == 0:
                month = int(month_period_match.group(1))
                year = normalize_year(month_period_match.group(2))
            else:
                year = normalize_year(month_period_match.group(1))
                month = int(month_period_match.group(2))
            break

    if month_period_match and year is not None and month is not None:
        metric_date = pd.Timestamp(year, month, 1) + pd.offsets.MonthEnd(0)  # type: ignore[operator]
        return {
            **empty_context,
            "metric_type": "month",
            "metric_date": metric_date,
        }

    quarter_match = re.fullmatch(r"([1-4])Q\s*(\d{2,4})", text)
    if quarter_match:
        year_text = quarter_match.group(2)
        year = int(year_text) + 2000 if len(year_text) == 2 else int(year_text)
        return {
            **empty_context,
            "metric_type": "quarter",
            "metric_date": pd.Timestamp(year, int(quarter_match.group(1)) * 3, 1) + pd.offsets.MonthEnd(0),
            # type: ignore[operator]

        }

    year_match = re.fullmatch(r"(20\d{2})", text)
    if year_match:
        year = int(year_match.group(1))
        return {
            **empty_context,
            "metric_type": "year",
            "metric_date": pd.Timestamp(year, 12, 31),
        }

    if len(delta_parts) == 2:
        year_match = re.fullmatch(r"20\d{2}", delta_parts[0])
        comparison_year_match = re.fullmatch(r"20\d{2}", delta_parts[1])
        if year_match and comparison_year_match:
            year = int(delta_parts[0])
            comparison_year = int(delta_parts[1])
            return {
                **empty_context,
                "metric_type": "delta",
                "metric_date": pd.Timestamp(year, 12, 31),
                "comparison_date": pd.Timestamp(comparison_year, 12, 31),
                "comparison_year": comparison_year,
            }

    return empty_context


def header_row_score(row) -> int:
    return sum(
        1
        for value in row.tolist()[1:]
        if cell_text(value) and not metric_context(cell_text(value))["metric_type"] and not is_unit_label(value)
    )


def best_group_header_row(raw_df: pd.DataFrame, unit_row_idx: int) -> list[str]:
    best_idx = None
    best_score = 0
    for idx in range(max(0, unit_row_idx - 4), unit_row_idx):
        score = header_row_score(raw_df.iloc[idx])
        if score > best_score:
            best_idx = idx
            best_score = score
    if best_idx is None:
        return [""] * raw_df.shape[1]
    return carry_forward_header_values(raw_df.iloc[best_idx].tolist())


def best_label_column(raw_df: pd.DataFrame, unit_row_idx: int, unit_cols: list[int]) -> int:
    if not unit_cols:
        return 0

    best_col = 0
    best_score = -1
    for col_idx in range(min(unit_cols)):
        score = 0
        for row_idx in range(unit_row_idx + 1, len(raw_df)):
            row = raw_df.iloc[row_idx]
            label = cell_text(row.iloc[col_idx])
            if not label or is_unit_label(label):
                continue
            if any(not is_blank(row.iloc[value_col]) for value_col in unit_cols):
                score += 1
        if score > best_score:
            best_col = col_idx
            best_score = score

    return best_col


def auto_flatten_grouped_metric_blocks(raw_df: pd.DataFrame) -> pd.DataFrame:
    """Flatten wide report tables with row labels and repeated metric groups."""
    if raw_df is None or raw_df.empty:
        return pd.DataFrame()

    records = []
    for unit_row_idx in range(1, len(raw_df)):
        unit_row = raw_df.iloc[unit_row_idx]
        unit_cols = [
            col_idx for col_idx in range(1, raw_df.shape[1])
            if is_unit_label(unit_row.iloc[col_idx])
        ]
        if len(unit_cols) < 4:
            continue

        label_col = best_label_column(raw_df, unit_row_idx, unit_cols)
        period_row = raw_df.iloc[unit_row_idx - 1]
        period_headers = carry_forward_header_values(period_row.tolist())
        detail_headers = raw_df.iloc[unit_row_idx - 2].tolist() if unit_row_idx >= 2 else []
        group_headers = best_group_header_row(raw_df, unit_row_idx)
        table_name = nearest_title_above(raw_df, unit_row_idx)
        if not table_name:
            table_name = first_nonblank_after(raw_df.iloc[max(0, unit_row_idx - 4)], start_col=1)
        if _is_col0_unit_anchor(unit_row.iloc[0]):
            anchored_table, _ = _resolve_block_header_from_unit_anchor(raw_df, unit_row_idx)
            if anchored_table:
                table_name = anchored_table

        row_idx = unit_row_idx + 1
        blank_label_rows = 0
        while row_idx < len(raw_df):
            data_row = raw_df.iloc[row_idx]
            line_item = cell_text(data_row.iloc[label_col])
            if sum(1 for value in data_row.tolist() if is_unit_label(value)) >= 4:
                break
            if not line_item:
                blank_label_rows += 1
                if blank_label_rows >= 5:
                    break
                row_idx += 1
                continue
            blank_label_rows = 0
            if is_unit_label(line_item):
                break
            if all(is_blank(data_row.iloc[col_idx]) for col_idx in unit_cols):
                row_idx += 1
                continue

            row_context = indented_row_context(raw_df, row_idx, label_col)
            for col_idx in unit_cols:
                value = data_row.iloc[col_idx]
                if is_blank(value):
                    continue
                metric = cell_text(period_headers[col_idx])
                if not metric:
                    metric = cell_text(group_headers[col_idx])
                metric_detail = cell_text(detail_headers[col_idx]) if col_idx < len(detail_headers) else ""
                if metric_detail in {metric, cell_text(group_headers[col_idx])}:
                    metric_detail = ""
                record = {
                    "table_name": table_name,
                    "section": table_name,
                    "column_group": cell_text(group_headers[col_idx]),
                    "unit": cell_text(unit_row.iloc[col_idx]),
                    "line_item": line_item,
                    "metric": metric,
                    "metric_detail": metric_detail,
                    "value": value,
                    "block_key": f"{unit_row_idx}:{cell_text(group_headers[col_idx])}",
                    "block_start_column": col_idx,
                }
                record.update(row_context)
                record.update(metric_context(metric))
                records.append(record)
            row_idx += 1

    return pd.DataFrame(records)


def auto_flatten_report_blocks(raw_df: pd.DataFrame) -> pd.DataFrame:
    """Extract repeated side-by-side report blocks into long rows."""
    unit_cells = find_unit_header_cells(raw_df)
    if not unit_cells:
        return pd.DataFrame()

    records = []
    blocks_by_row: dict[int, list[int]] = {}
    for row_idx, col_idx in unit_cells:
        blocks_by_row.setdefault(row_idx, []).append(col_idx)

    for header_row_idx, starts in blocks_by_row.items():
        starts = sorted(starts)
        for pos, start_col in enumerate(starts):
            start_col = int(start_col)
            end_col = int(starts[pos + 1]) if pos + 1 < len(starts) else raw_df.shape[1]
            header_row = raw_df.iloc[header_row_idx]
            unit = cell_text(header_row.iloc[start_col])
            metric_cols = [
                col for col in range(start_col + 1, end_col)
                if not is_blank(header_row.iloc[col])
            ]
            if not metric_cols:
                continue

            table_name = nearest_left_label_above(raw_df, header_row_idx, start_col)
            section = cell_text(raw_df.iloc[header_row_idx - 1, start_col]) if header_row_idx > 0 else ""
            column_group = nearest_nonblank_right(
                raw_df,
                header_row_idx - 1,
                start_col,
                end_col,
            ) if header_row_idx > 0 else ""

            if start_col == 0 and _is_col0_unit_anchor(header_row.iloc[0]):
                anchored_table, anchored_group = _resolve_block_header_from_unit_anchor(raw_df, header_row_idx)
                if anchored_table:
                    table_name = anchored_table
                if anchored_group:
                    column_group = anchored_group

            row_idx = header_row_idx + 1
            blank_label_rows = 0
            while row_idx < len(raw_df):
                data_row = raw_df.iloc[row_idx]
                line_item = cell_text(data_row.iloc[start_col])
                if not line_item:
                    blank_label_rows += 1
                    if blank_label_rows >= 5:
                        break
                    row_idx += 1
                    continue
                blank_label_rows = 0
                if line_item.lower().startswith(("eur ", "usd ", "gbp ", "in %")):
                    break
                if not row_has_values(data_row, start_col=start_col + 1):
                    row_idx += 1
                    continue

                row_context = indented_row_context(raw_df, row_idx, start_col)
                for metric_col in metric_cols:
                    metric = cell_text(header_row.iloc[metric_col])
                    value = data_row.iloc[metric_col]
                    if metric and not is_blank(value):
                        record = {
                            "table_name": table_name,
                            "section": section,
                            "column_group": column_group,
                            "unit": unit,
                            "line_item": line_item,
                            "metric": metric,
                            "value": value,
                            "block_key": f"{header_row_idx}:{start_col}",
                            "block_start_column": start_col,
                        }
                        record.update(row_context)
                        record.update(metric_context(metric))
                        records.append(record)
                row_idx += 1

    return pd.DataFrame(records)


def auto_flatten_horizontal_balance_sheet(raw_df: pd.DataFrame) -> pd.DataFrame:
    """Flatten balance-sheet tabs where business segments are side-by-side in 20-col blocks."""
    if raw_df is None or raw_df.empty or raw_df.shape[1] < 20:
        return pd.DataFrame()

    date_columns = [
        "31.03.2024",
        "30.06.2024",
        "30.09.2024",
        "31.12.2024",
        "31.03.2025",
        "30.06.2025",
        "30.09.2025",
        "31.12.2025",
    ]
    value_offsets = [2, 4, 6, 8, 10, 12, 14, 16]
    delta_offset = 18
    block_width = 20

    if raw_df.shape[1] < (min(value_offsets) + 1):
        return pd.DataFrame()
    block_starts = [
        start_col
        for start_col in range(0, raw_df.shape[1], block_width)
        if start_col + max(value_offsets) < raw_df.shape[1]
    ]

    records = []
    valid_block_count = 0
    for start_col in block_starts:
        # Signature checks for this very specific layout.
        title = cell_text(raw_df.iloc[0, start_col]) if len(raw_df) > 0 else ""
        section_seed = cell_text(raw_df.iloc[3, start_col]) if len(raw_df) > 3 else ""
        segment_name = cell_text(raw_df.iloc[3, start_col + 2]) if len(raw_df) > 3 and start_col + 2 < raw_df.shape[1] else ""
        unit = cell_text(raw_df.iloc[4, start_col]) if len(raw_df) > 4 else ""

        has_balance_title = "balance sheet" in title.lower()
        section_seed_compact = re.sub(r"\s+", "", section_seed.lower())
        has_assets_seed = "assets" in section_seed_compact
        has_unit_seed = unit.lower().startswith(("eur ", "usd ", "gbp ", "in %"))

        # Require a plausible segment block signature so we don't hijack unrelated sheets.
        if not ((has_balance_title or has_assets_seed) and (segment_name or has_unit_seed)):
            continue

        valid_block_count += 1
        current_section = "ASSETS"
        if "liabilities" in section_seed_compact:
            current_section = "LIABILITIES AND EQUITY"

        for row_idx in range(5, len(raw_df)):
            row = raw_df.iloc[row_idx]
            line_item = cell_text(row.iloc[start_col])
            if not line_item:
                continue

            compact = re.sub(r"\s+", "", line_item.lower())
            if "assets" in compact and len(compact) <= 24:
                current_section = "ASSETS"
                continue
            if "liabilities" in compact and "equity" in compact:
                current_section = "LIABILITIES AND EQUITY"
                continue

            if is_unit_label(line_item):
                continue
            if line_item.lower().startswith("consolidated balance sheet"):
                continue
            if line_item.lower().startswith("by business segments"):
                continue

            row_has_value = False
            for metric, offset in zip(date_columns, value_offsets):
                col_idx = start_col + offset
                if col_idx >= raw_df.shape[1]:
                    continue
                value = row.iloc[col_idx]
                if is_blank(value):
                    continue
                row_has_value = True
                record = {
                    "table_name": segment_name or "Balance Sheet",
                    "section": current_section,
                    "column_group": segment_name,
                    "unit": unit,
                    "line_item": line_item,
                    "metric": metric,
                    "value": value,
                    "block_key": f"horizontal_bs:{start_col}",
                    "block_start_column": start_col,
                }
                record.update(metric_context(metric))
                records.append(record)

            delta_col = start_col + delta_offset
            if delta_col < raw_df.shape[1]:
                delta_value = row.iloc[delta_col]
                if not is_blank(delta_value):
                    row_has_value = True
                    delta_metric = "Delta 31.12.25 / 31.12.24"
                    record = {
                        "table_name": segment_name or "Balance Sheet",
                        "section": current_section,
                        "column_group": segment_name,
                        "unit": unit,
                        "line_item": line_item,
                        "metric": delta_metric,
                        "value": delta_value,
                        "block_key": f"horizontal_bs:{start_col}",
                        "block_start_column": start_col,
                    }
                    record.update(metric_context(delta_metric))
                    records.append(record)

            # Skip decorative label-only rows that don't carry values.
            if not row_has_value:
                continue

    # Need multiple matching blocks; otherwise treat as non-match and let other extractors run.
    if valid_block_count < 2 or not records:
        return pd.DataFrame()
    return pd.DataFrame(records)


def _fallback_matrix_flatten(raw_df: pd.DataFrame) -> pd.DataFrame:
    """Flatten matrix layouts with stacked headers and interior numeric grids."""
    if raw_df is None or raw_df.empty or raw_df.shape[0] < 2 or raw_df.shape[1] < 3:
        return pd.DataFrame()

    # Skip error-heavy leading rows (benelux sheets have many #REF! errors at top)
    raw_df = _skip_error_heavy_rows(raw_df, max_check=50)
    raw_df = drop_all_blank_columns(raw_df).reset_index(drop=True)

    # Column pruning changes positional indexes; drop stale indent metadata.
    raw_df.attrs.pop("excel_indents", None)

    section_col = detect_section_column(raw_df)

    matrix_values = raw_df.to_numpy(copy=False)
    row_count, col_count = matrix_values.shape

    # Pre-cache cell_text and looks_like_data_value for all cells to avoid redundant calls
    cell_cache = {}
    data_value_cache = {}
    metric_meta_cache = {}

    def _get_cached_cell_text(row_idx: int, col_idx: int) -> str:
        key = (row_idx, col_idx)
        if key not in cell_cache:
            cell_cache[key] = cell_text(matrix_values[row_idx, col_idx])
        return cell_cache[key]

    def _is_cached_data_value(row_idx: int, col_idx: int) -> bool:
        key = (row_idx, col_idx)
        if key not in data_value_cache:
            value = matrix_values[row_idx, col_idx]
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                data_value_cache[key] = True
            else:
                data_value_cache[key] = looks_like_data_value(value)
        return data_value_cache[key]

    def _row_numeric_hits(row_idx: int, start_col: int = 0) -> int:
        return sum(1 for col in range(start_col, col_count) if _is_cached_data_value(row_idx, col))

    # Find rows that look like data lines (some text context + multiple values).
    # Scan the full sheet. A large title/notes band must not hide a matrix that
    # starts after an arbitrary preview cutoff.
    scan_limit = row_count
    candidate_rows = [
        row_idx for row_idx in range(scan_limit)
        if _row_numeric_hits(row_idx) >= 2
    ]
    if not candidate_rows:
        return pd.DataFrame()

    probe_rows = candidate_rows[: min(100, len(candidate_rows))]

    # Detect a likely business-line column (often col 1 in matrix files).
    label_candidates = range(min(6, col_count))
    label_col = 0
    label_score = -1
    for col_idx in label_candidates:
        score = 0
        for row_idx in probe_rows:
            label = _get_cached_cell_text(row_idx, col_idx)
            if label and not _is_cached_data_value(row_idx, col_idx) and not is_unit_label(label):
                score += 1
        if score > label_score:
            label_col = col_idx
            label_score = score

    # Detect metric/line-detail column near the label column.
    metric_col = min(label_col + 1, col_count - 1)
    metric_score = -1
    for col_idx in range(label_col + 1, min(col_count, label_col + 6)):
        score = 0
        for row_idx in probe_rows:
            metric_text = _get_cached_cell_text(row_idx, col_idx)
            if metric_text and not _is_cached_data_value(row_idx, col_idx):
                score += 1
        if score > metric_score:
            metric_col = col_idx
            metric_score = score

    # Detect where data rows begin for this label/metric pattern.
    data_start = candidate_rows[0]
    for row_idx in candidate_rows[:50]:  # Check only first 50 candidate rows
        line_item = _get_cached_cell_text(row_idx, label_col)
        metric = _get_cached_cell_text(row_idx, metric_col)
        if (line_item or metric) and _row_numeric_hits(row_idx, start_col=metric_col + 1) >= 2:
            data_start = row_idx
            break

    # Capture a wider header band so top hierarchy levels (period/country) are retained.
    header_start = max(0, data_start - 10)
    header_rows = list(range(header_start, data_start))

    # Keep columns that carry actual data in the matrix body.
    # Inspect the full body so columns populated only in later sections remain
    # discoverable.
    body_end = row_count
    value_cols = []
    for col_idx in range(metric_col + 1, col_count):
        hits = 0
        for row_idx in range(data_start, body_end):
            if _is_cached_data_value(row_idx, col_idx):
                hits += 1
        if hits >= 2:
            value_cols.append(col_idx)

    if not value_cols:
        return pd.DataFrame()

    def _looks_like_period_token(token: str) -> bool:
        text = token.strip().lower()
        if not text:
            return False
        # Ignore descriptive titles; period tokens are usually compact labels.
        if len(text) > 35:
            return False
        if any(flag in text for flag in ("ytd", "qtd", "mtd", "fy", "year to date")):
            return True
        return bool(RE_MONTH_YEAR_TOKEN.search(text))

    global_period_token = ""
    for row_idx in header_rows:
        for col_idx in range(0, min(col_count, max(metric_col + 1, 6))):
            token = _get_cached_cell_text(row_idx, col_idx)
            if token and token.upper() not in {"#REF!", "#DIV/0!"} and _looks_like_period_token(token):
                global_period_token = token
                break
        if global_period_token:
            break

    header_noise = {"#REF!", "#DIV/0!", "#N/A", "#VALUE!"}
    subheader_tokens = {
        "actual", "previous", "plan", "budget", "target", "forecast", "ytd", "mtd", "qtd"
    }

    # Forward-fill sparse top headers row-wise (common in wide matrix exports where group
    # names appear once and subcolumns are blank). This avoids orphan headers like "Previous".
    carried_header_rows: dict[int, list[str]] = {}
    for row_idx in header_rows:
        carried: list[str] = []
        current = ""
        for col_idx in range(col_count):
            token = _get_cached_cell_text(row_idx, col_idx)
            if token and token.upper() not in header_noise:
                current = token
            carried.append(current)
        carried_header_rows[row_idx] = carried

    def _header_for_col(col_idx: int) -> str:
        tokens: list[str] = []
        if global_period_token:
            tokens.append(global_period_token)

        for row_idx in header_rows:
            token = carried_header_rows[row_idx][col_idx]
            if not token or token.upper() in header_noise:
                continue
            if not tokens or tokens[-1] != token:
                tokens.append(token)

        # If we still ended up with only a generic subheader, borrow nearest left group label.
        if len(tokens) == 1 and tokens[0].strip().lower() in subheader_tokens:
            base = ""
            for left_col in range(col_idx - 1, metric_col, -1):
                for row_idx in header_rows:
                    candidate = carried_header_rows[row_idx][left_col]
                    if candidate and candidate.upper() not in header_noise and candidate.strip().lower() not in subheader_tokens:
                        base = candidate
                        break
                if base:
                    break
            if base:
                tokens = [base, tokens[0]]

        return " | ".join(tokens)

    value_headers = {col_idx: _header_for_col(col_idx) for col_idx in value_cols}

    # Pick a broad section/title from the top-left area.
    section = ""
    for row_idx in range(min(8, row_count)):
        for col_idx in range(min(4, col_count)):
            token = _get_cached_cell_text(row_idx, col_idx)
            if token and token.upper() not in {"#REF!", "#DIV/0!"}:
                section = token
                break
        if section:
            break

    records = []
    current_line_item = ""
    blank_run = 0
    current_section = section
    header_metric_meta = {
        col_idx: metric_context(value_headers.get(col_idx, "")) if value_headers.get(col_idx, "") else None
        for col_idx in value_cols
    }

    indent_grid = raw_df.attrs.get("excel_indents") or []
    label_texts = [_get_cached_cell_text(row_idx, label_col) for row_idx in range(row_count)]
    label_indents = [
        float(indent_grid[row_idx][label_col])
        if row_idx < len(indent_grid) and label_col < len(indent_grid[row_idx])
        else 0.0
        for row_idx in range(row_count)
    ]

    def _fast_row_context(row_idx: int) -> dict:
        current_indent = label_indents[row_idx]
        if current_indent <= 0:
            return {}

        parents = []
        next_indent = current_indent
        # Bound lookback depth for performance on large sheets.
        min_idx = max(row_idx - 120, 0)
        for idx in range(row_idx - 1, min_idx - 1, -1):
            label = label_texts[idx]
            if not label:
                continue
            indent = label_indents[idx]
            if indent < next_indent:
                parents.append(label)
                next_indent = indent
                if indent <= 0:
                    break

        parents.reverse()
        if not parents:
            return {}
        return {
            "parent_line_item": parents[-1],
            "line_item_path": " > ".join([*parents, label_texts[row_idx]]),
        }

    has_indent_hierarchy = any(
        label_indents[row_idx] > 0 for row_idx in range(data_start, row_count)
    )
    for row_idx in range(data_start, row_count):
        row = matrix_values[row_idx]
        row_series = raw_df.iloc[row_idx]
        header = section_name_validated(row_series, expected_col=section_col)
        if header:
            current_section = header
            current_line_item = ""
            blank_run = 0
            continue

        line_item = _get_cached_cell_text(row_idx, label_col)
        metric = _get_cached_cell_text(row_idx, metric_col)
        row_value_hits = sum(1 for col_idx in value_cols if _is_cached_data_value(row_idx, col_idx))

        if not line_item and not metric and row_value_hits == 0:
            blank_run += 1
            continue
        blank_run = 0

        if line_item and not is_unit_label(line_item):
            current_line_item = line_item

        active_line_item = current_line_item or line_item
        if not active_line_item:
            continue

        metric_name = metric or "value"
        row_context = _fast_row_context(row_idx) if has_indent_hierarchy else {}
        for col_idx in value_cols:
            value = row[col_idx]
            if is_blank(value) or not _is_cached_data_value(row_idx, col_idx):
                continue
            header_label = value_headers.get(col_idx, "")
            record = {
                "table_name": current_section,
                "section": current_section,
                "column_group": header_label,
                "line_item": active_line_item,
                "metric": metric_name,
                "value": value,
                "block_start_column": label_col,
            }
            if header_label:
                metric_meta = header_metric_meta.get(col_idx) or {}
            else:
                if metric_name not in metric_meta_cache:
                    metric_meta_cache[metric_name] = metric_context(metric_name)
                metric_meta = metric_meta_cache[metric_name]
            record.update(row_context)
            record.update(metric_meta)
            records.append(record)

    return pd.DataFrame(records)


def auto_flatten_heavy_summary_grid(raw_df: pd.DataFrame) -> pd.DataFrame:
    if raw_df is None or raw_df.empty:
        return pd.DataFrame()

    row_count, col_count = raw_df.shape
    if row_count == 0 or col_count < 3:
        return pd.DataFrame()

    values = raw_df.to_numpy(dtype=object, copy=False)

    top_window = min(row_count, 40)
    metric_row_idx = 0
    best_score = -1
    for row_idx in range(top_window):
        score = 0
        for value in values[row_idx]:
            text = cell_text(value)
            if text and not _is_numeric_like(value):
                score += 1
        if score > best_score:
            metric_row_idx = row_idx
            best_score = score

    def _row_has_numeric_payload(row_idx: int) -> bool:
        return any(_parse_numeric_value(value) is not None for value in values[row_idx])

    def _row_is_nonblank(row_idx: int) -> bool:
        return any(not is_blank(value) for value in values[row_idx])

    # The header is frequently a BAND, not a single row: a merged group row
    # (e.g. segment names) with a period row below it ("12M 24" / "12M 25"),
    # and sometimes a sparse qualifier row above. Fold adjacent non-data rows
    # into the band so each value column gets a composite header such as
    # "Retail lines | 12M 24" instead of one ambiguous label per group.
    band_start = metric_row_idx
    while (
        band_start > 0
        and metric_row_idx - band_start < 2
        and _row_is_nonblank(band_start - 1)
        and not _row_has_numeric_payload(band_start - 1)
    ):
        band_start -= 1
    # An underline transition (this row carries a bottom rule, the next row is
    # unruled) is the visual end of the header band — stop folding rows below
    # it even when they hold no numbers (e.g. text-valued data like ratings).
    # Fully gridded sheets have rules on every row, so they never transition.
    top_rules = raw_df.attrs.get("excel_row_top_rule_flags") or []
    bottom_rules = raw_df.attrs.get("excel_row_bottom_rule_flags") or []

    def _underline_closes_band(row_idx: int) -> bool:
        if row_idx >= len(bottom_rules) or not bottom_rules[row_idx]:
            return False
        next_idx = row_idx + 1
        next_top = bool(top_rules[next_idx]) if next_idx < len(top_rules) else False
        next_bottom = bool(bottom_rules[next_idx]) if next_idx < len(bottom_rules) else False
        return not next_top and not next_bottom

    band_end = metric_row_idx
    while (
        band_end + 1 < row_count
        and band_end - metric_row_idx < 3
        and _row_is_nonblank(band_end + 1)
        and not _row_has_numeric_payload(band_end + 1)
        and not _underline_closes_band(band_end)
    ):
        band_end += 1
    header_band = list(range(band_start, band_end + 1))
    data_start = band_end + 1

    # Value columns: any column carrying numeric payload in the data area.
    value_col_set: set[int] = set()
    for row_idx in range(data_start, row_count):
        first_text_col = -1
        for col_idx in range(col_count):
            if cell_text(values[row_idx, col_idx]):
                first_text_col = col_idx
                break
        if first_text_col < 0:
            continue
        for col_idx in range(first_text_col + 1, col_count):
            if _is_numeric_like(values[row_idx, col_idx]):
                value_col_set.add(col_idx)
    value_cols = sorted(value_col_set)

    headers: dict[int, str] = {}
    column_groups: dict[int, str] = {}
    if value_cols:
        headers, column_groups = _composite_headers_for_region(
            raw_df,
            header_band,
            value_cols,
        )

    source_row_offset = int(raw_df.attrs.get("_source_row_offset", 0) or 0)
    source_col_offset = int(raw_df.attrs.get("_source_col_offset", 0) or 0)
    source_range = str(raw_df.attrs.get("_source_range", ""))

    def _unit_anchor_text(row_idx: int) -> str:
        """Return the unit token when this row (re)declares the block unit."""
        anchor = ""
        for col_idx in range(col_count):
            text = cell_text(values[row_idx, col_idx])
            if not text:
                continue
            if anchor:
                return ""  # more than one meaningful token → not a pure anchor
            if is_unit_label(text) or _is_col0_unit_anchor(text):
                anchor = text
            else:
                return ""
        return anchor

    # Seed the unit from the header band ("EUR mn" frequently sits at the far
    # left of the period row).
    current_unit = ""
    for row_idx in header_band:
        for col_idx in range(col_count):
            text = cell_text(values[row_idx, col_idx])
            if text and (is_unit_label(text) or _is_col0_unit_anchor(text)):
                current_unit = text
                break

    metric_meta_cache: dict[str, dict] = {}
    records = []
    for row_idx in range(data_start, row_count):
        anchor = _unit_anchor_text(row_idx)
        if anchor:
            current_unit = anchor
            continue

        label_col = -1
        line_item = ""
        for col_idx in range(col_count):
            text = cell_text(values[row_idx, col_idx])
            if text:
                label_col = col_idx
                line_item = text
                break
        if label_col < 0:
            continue
        if not line_item or is_unit_label(line_item):
            continue

        for col_idx in range(label_col + 1, col_count):
            value = values[row_idx, col_idx]
            if is_blank(value):
                continue
            if not _is_numeric_like(value):
                continue
            metric_text = headers.get(col_idx) or f"col_{col_idx}"
            source_row = source_row_offset + row_idx + 1
            source_col_idx = source_col_offset + col_idx + 1
            source_column = _excel_col_letter(source_col_idx)
            if metric_text not in metric_meta_cache:
                metric_meta_cache[metric_text] = metric_context(
                    metric_text.split("|")[-1].strip()
                )
            record = {
                "table_name": "Summary Grid",
                "source_range": source_range,
                "source_cell": f"{source_column}{source_row}",
                "source_row": source_row,
                "source_column": source_column,
                "source_column_index": source_col_idx,
                "section": "",
                "column_group": column_groups.get(col_idx, ""),
                "unit": current_unit,
                "line_item": line_item,
                "metric": metric_text,
                "value": value,
                "block_key": f"heavy_summary:{metric_row_idx}",
                "block_start_column": int(label_col),
            }
            record.update(metric_meta_cache[metric_text])
            records.append(record)

    return pd.DataFrame(records)


def _candidate_data_rows_for_region(raw_df: pd.DataFrame) -> list[tuple[int, int]]:
    """Return (row index, first value column) candidates for a rectangular table."""
    if raw_df is None or raw_df.empty:
        return []

    values = raw_df.to_numpy(dtype=object, copy=False)
    candidates: list[tuple[int, int]] = []
    for row_idx, row in enumerate(values):
        numeric_cols = []
        numeric_values = []
        for col_idx, value in enumerate(row):
            parsed = _parse_numeric_value(value)
            if parsed is None:
                continue
            numeric_cols.append(col_idx)
            numeric_values.append(parsed)
        if not numeric_cols:
            continue

        first_value_col = int(numeric_cols[0])
        if first_value_col <= 0:
            continue
        label_tokens = [
            cell_text(row[col_idx])
            for col_idx in range(first_value_col)
            if cell_text(row[col_idx])
        ]
        if not label_tokens:
            continue
        if all(_parse_numeric_value(token) is not None for token in label_tokens):
            continue
        # A period-header row such as "in % | 2021 | 2022 | 2023" is not
        # data. Do not use the year range alone: a large numeric matrix can
        # legitimately contain an entire row whose values happen to fall
        # between 1900 and 2100.
        year_only_payload = numeric_values and all(
            float(value).is_integer() and 1900 <= int(value) <= 2100
            for value in numeric_values
        )
        normalized_labels = {
            re.sub(r"[^a-z0-9]+", " ", token.lower()).strip()
            for token in label_tokens
        }
        explicit_header_labels = {
            "year",
            "period",
            "date",
            "fiscal year",
            "reporting period",
            "line item",
            "description",
        }
        if year_only_payload and (
            any(is_unit_label(token) for token in label_tokens)
            or bool(normalized_labels & explicit_header_labels)
            or (row_idx < 10 and len(label_tokens) == 1)
        ):
            continue
        candidates.append((row_idx, first_value_col))
    return candidates


def _candidate_row_clusters(
        candidates: list[tuple[int, int]],
        max_row_gap: int = 2,
) -> list[list[tuple[int, int]]]:
    if not candidates:
        return []
    clusters: list[list[tuple[int, int]]] = [[candidates[0]]]
    for candidate in candidates[1:]:
        if candidate[0] - clusters[-1][-1][0] <= max(int(max_row_gap), 1):
            clusters[-1].append(candidate)
        else:
            clusters.append([candidate])
    return clusters


def _largest_candidate_row_cluster(
        candidates: list[tuple[int, int]],
        max_row_gap: int = 2,
) -> list[tuple[int, int]]:
    clusters = _candidate_row_clusters(candidates, max_row_gap=max_row_gap)
    if not clusters:
        return []
    return max(clusters, key=lambda cluster: (len(cluster), cluster[-1][0] - cluster[0][0]))


def _ruled_boundary_rows(
        raw_df: pd.DataFrame,
        region: TableRegion,
        context: SheetScanContext,
) -> set[int]:
    """Region-local rows where a ruled title starts a new stacked table.

    A boundary needs a border TRANSITION — a top rule whose preceding row has
    no bottom rule — so fully gridded sheets (every row ruled) never split,
    which keeps in-table section headers intact. The row itself must look like
    a title: one or two text cells, no numeric payload, not a unit or period.
    """
    top_rules = raw_df.attrs.get("excel_row_top_rule_flags") or []
    bottom_rules = raw_df.attrs.get("excel_row_bottom_rule_flags") or []
    if not top_rules:
        return set()

    boundaries: set[int] = set()
    region_length = region.end_row - region.start_row
    for local_row in range(1, region_length):
        sheet_row = region.start_row + local_row
        if sheet_row >= len(top_rules) or not top_rules[sheet_row]:
            continue
        previous_has_bottom = (
            bool(bottom_rules[sheet_row - 1])
            if sheet_row - 1 < len(bottom_rules)
            else False
        )
        if previous_has_bottom:
            continue
        non_blank = int(context.non_blank_count[local_row])
        if non_blank == 0 or non_blank > 2:
            continue
        if bool(context.data_like_mask[local_row].any()):
            continue
        first_col = int(context.first_non_blank_col[local_row])
        token = str(context.text[local_row, first_col]).strip()
        if not token or is_unit_label(token) or _looks_like_date_or_period_text(token):
            continue
        boundaries.add(local_row)
    return boundaries


def _split_table_region_vertically(
        raw_df: pd.DataFrame,
        region: TableRegion,
) -> list[TableRegion]:
    """Split stacked tables sharing the same columns, including one-row gaps."""
    region_df = raw_df.iloc[
        region.start_row:region.end_row,
        region.start_col:region.end_col,
    ]
    candidates = _candidate_data_rows_for_region(region_df)
    context = _sheet_scan_context(region_df)
    boundaries = _ruled_boundary_rows(raw_df, region, context)

    if boundaries:
        # A ruled title between two payload runs marks a new stacked table
        # even without blank rows; cluster each side independently.
        segments: list[list[tuple[int, int]]] = []
        current_segment: list[tuple[int, int]] = []
        pending = sorted(boundaries)
        for candidate_row in candidates:
            while pending and int(candidate_row[0]) >= pending[0]:
                if current_segment:
                    segments.append(current_segment)
                    current_segment = []
                pending = pending[1:]
            current_segment.append(candidate_row)
        if current_segment:
            segments.append(current_segment)
        clusters = [
            cluster
            for segment in segments
            for cluster in _candidate_row_clusters(segment, max_row_gap=2)
        ]
    else:
        clusters = _candidate_row_clusters(candidates, max_row_gap=2)
    if len(clusters) <= 1:
        return [region]

    split_starts = [0]
    for previous, current in zip(clusters, clusters[1:]):
        previous_end = int(previous[-1][0])
        current_start = int(current[0][0])
        gap_rows = list(range(previous_end + 1, current_start))
        boundary_in_gap = [row_idx for row_idx in gap_rows if row_idx in boundaries]
        blank_rows = [
            row_idx
            for row_idx in gap_rows
            if int(context.non_blank_count[row_idx]) == 0
        ]
        if boundary_in_gap:
            # The ruled title belongs to the table below it.
            split_starts.append(boundary_in_gap[0])
        else:
            split_starts.append((blank_rows[-1] + 1) if blank_rows else previous_end + 1)

    split_regions: list[TableRegion] = []
    for idx, local_start in enumerate(split_starts):
        local_end = split_starts[idx + 1] if idx + 1 < len(split_starts) else len(region_df)
        candidate = TableRegion(
            start_row=region.start_row + local_start,
            end_row=region.start_row + local_end,
            start_col=region.start_col,
            end_col=region.end_col,
        )
        candidate_df = raw_df.iloc[
            candidate.start_row:candidate.end_row,
            candidate.start_col:candidate.end_col,
        ]
        if _region_has_table_signal(candidate_df):
            split_regions.append(candidate)
    return split_regions or [region]


def _row_is_repeated_title(context: SheetScanContext, row_idx: int) -> bool:
    tokens = [
        str(token).strip()
        for token in context.text[row_idx].tolist()
        if str(token).strip()
    ]
    return bool(tokens) and len(set(tokens)) == 1


def _looks_like_row_label_header(header: str) -> bool:
    """Identify columns that describe rows rather than contain observations."""
    leaf = str(header or "").split("|")[-1].strip().lower()
    leaf = re.sub(r"[^a-z0-9]+", " ", leaf).strip()
    return leaf in {
        "account",
        "book",
        "bucket",
        "category",
        "counterparty",
        "description",
        "entity",
        "exposure",
        "hierarchy",
        "issuer",
        "item",
        "limit",
        "line item",
        "metric",
        "portfolio",
        "section",
        "statement",
    }


def _looks_like_table_title_text(value: str) -> bool:
    text = re.sub(r"\s+", " ", str(value or "")).strip().lower()
    if not text:
        return False
    if re.match(r"^t\d+\b", text):
        return True
    return any(
        phrase in text
        for phrase in (
            "balance sheet",
            "income statement",
            "cash flow",
            "capital adequacy",
            "credit risk",
            "liquidity",
            "watchlist",
            "ratings",
            "financial position",
        )
    )


def _composite_headers_for_region(
        raw_df: pd.DataFrame,
        header_rows: list[int],
        value_cols: list[int],
) -> tuple[dict[int, str], dict[int, str]]:
    """Build vertical headers such as `Midcorp | 2024` for each value column."""
    context = _sheet_scan_context(raw_df)
    row_tokens: dict[int, list[str]] = {}
    first_value_col = min(value_cols)
    for row_idx in header_rows:
        carried: list[str] = [""] * raw_df.shape[1]
        value_tokens = [
            str(context.text[row_idx, col_idx]).strip()
            for col_idx in value_cols
            if str(context.text[row_idx, col_idx]).strip()
        ]
        if len(set(value_tokens)) == 1 and value_tokens:
            sole_token = value_tokens[0]
            for col_idx in value_cols:
                carried[col_idx] = sole_token
            row_tokens[row_idx] = carried
            continue

        current = ""
        for col_idx in range(first_value_col, raw_df.shape[1]):
            token = str(context.text[row_idx, col_idx]).strip()
            if token:
                current = token
            carried[col_idx] = current
        row_tokens[row_idx] = carried

    base_headers: dict[int, str] = {}
    column_groups: dict[int, str] = {}
    for col_idx in value_cols:
        parts: list[str] = []
        for row_idx in header_rows:
            token = row_tokens[row_idx][col_idx]
            if not token:
                continue
            if token.lower() in {"line item", "risk metric", "metric", "description"}:
                continue
            if parts and parts[-1] == token:
                continue
            parts.append(token)
        parts = _compact_summary_header_tokens(parts, max_tokens=5)
        base_headers[col_idx] = " | ".join(parts) if parts else f"Column {_excel_col_letter(col_idx + 1)}"
        column_groups[col_idx] = parts[0] if len(parts) > 1 else ""

    duplicate_counts: dict[str, int] = {}
    for header in base_headers.values():
        duplicate_counts[header] = duplicate_counts.get(header, 0) + 1

    headers: dict[int, str] = {}
    for col_idx, header in base_headers.items():
        if duplicate_counts.get(header, 0) > 1:
            headers[col_idx] = f"{header} | Column {_excel_col_letter(col_idx + 1)}"
        else:
            headers[col_idx] = header
    return headers, column_groups


def auto_flatten_rectangular_table(raw_df: pd.DataFrame) -> pd.DataFrame:
    """Flatten one bounded table and preserve multi-row column headers."""
    if raw_df is None or raw_df.empty or raw_df.shape[0] < 3 or raw_df.shape[1] < 2:
        return pd.DataFrame()

    candidates = _candidate_data_rows_for_region(raw_df)
    cluster = _largest_candidate_row_cluster(candidates, max_row_gap=2)
    if not cluster:
        return pd.DataFrame()

    data_start = int(cluster[0][0])
    first_value_cols = [int(value_col) for _, value_col in cluster]
    data_col_start = int(pd.Series(first_value_cols).mode().iloc[0])
    if data_col_start <= 0 or data_col_start >= raw_df.shape[1]:
        return pd.DataFrame()

    values = raw_df.to_numpy(dtype=object, copy=False)
    data_rows: list[int] = []
    blank_run = 0
    for row_idx in range(data_start, len(raw_df)):
        label_present = any(
            cell_text(values[row_idx, col_idx])
            for col_idx in range(data_col_start)
        )
        payload_present = any(
            not is_blank(values[row_idx, col_idx])
            for col_idx in range(data_col_start, raw_df.shape[1])
        )
        if label_present and payload_present:
            data_rows.append(row_idx)
            blank_run = 0
            continue
        if not label_present and not payload_present:
            blank_run += 1
            if blank_run >= 2:
                break
            continue
        blank_run = 0

    if not data_rows:
        return pd.DataFrame()

    value_cols = [
        col_idx
        for col_idx in range(data_col_start, raw_df.shape[1])
        if any(not is_blank(values[row_idx, col_idx]) for row_idx in data_rows)
    ]
    if not value_cols:
        return pd.DataFrame()

    context = _sheet_scan_context(raw_df)
    title_row = -1
    title = ""
    title_candidates: list[tuple[int, str, bool]] = []
    lookback_start = max(0, data_start - 6)
    for row_idx in range(lookback_start, data_start):
        row_non_blank = int(context.non_blank_count[row_idx])
        if row_non_blank == 0:
            continue
        payload_non_blank = int((~context.blank_mask[row_idx, data_col_start:]).sum())
        first_col = int(context.first_non_blank_col[row_idx])
        first_token = str(context.text[row_idx, first_col]).strip() if first_col >= 0 else ""
        repeated_title = _row_is_repeated_title(context, row_idx)
        if (
            first_token
            and first_col < data_col_start
            and (
                payload_non_blank == 0
                or repeated_title
                or _looks_like_table_title_text(first_token)
            )
        ):
            title_candidates.append(
                (row_idx, first_token, bool(context.row_bold_flags[row_idx]))
            )

    bold_titles = [candidate for candidate in title_candidates if candidate[2]]
    selected_title = bold_titles[-1] if bold_titles else (title_candidates[-1] if title_candidates else None)
    if selected_title is not None:
        title_row, title, _ = selected_title

    subtitle_parts: list[str] = []
    for candidate_row, candidate_text, _ in title_candidates:
        if candidate_row == title_row:
            continue
        normalized_candidate = candidate_text.strip()
        if (
            not normalized_candidate
            or is_unit_label(normalized_candidate)
            or _looks_like_date_or_period_text(normalized_candidate)
            or normalized_candidate.lower() == title.strip().lower()
            or normalized_candidate.lower() in {
                existing.lower() for existing in subtitle_parts
            }
        ):
            continue
        subtitle_parts.append(normalized_candidate)
    table_subtitle = " | ".join(subtitle_parts)

    title_candidate_rows = {candidate[0] for candidate in title_candidates}
    header_scan_start = max(0, title_row + 1)
    if (
        title_row >= 0
        and bool((~context.blank_mask[title_row, data_col_start:]).any())
    ):
        # A title row can also contain sibling table/group headers to the
        # right. Preserve it in the composite header stack instead of
        # discarding the entire row after taking the left-most title.
        header_scan_start = title_row
        title_candidate_rows.discard(title_row)
    header_rows = [
        row_idx
        for row_idx in range(header_scan_start, data_start)
        if row_idx not in title_candidate_rows
        if int(context.non_blank_count[row_idx]) > 0
        and bool((~context.blank_mask[row_idx, data_col_start:]).any())
    ]
    if not header_rows and data_start > 0:
        header_rows = [data_start - 1]
    header_rows = header_rows[-4:]

    headers, column_groups = _composite_headers_for_region(
        raw_df,
        header_rows,
        value_cols,
    )

    if not title:
        for row_idx in range(max(0, data_start - 5), data_start):
            first_col = int(context.first_non_blank_col[row_idx])
            if first_col < 0:
                continue
            candidate = str(context.text[row_idx, first_col]).strip()
            if candidate and not is_unit_label(candidate):
                title = candidate
                break
    title = title or "Table"

    unit = ""
    for row_idx in reversed(header_rows):
        for col_idx in range(min(data_col_start + 1, raw_df.shape[1])):
            candidate = cell_text(values[row_idx, col_idx])
            if is_unit_label(candidate) or _is_col0_unit_anchor(candidate):
                unit = candidate
                break
        if unit:
            break

    source_row_offset = int(raw_df.attrs.get("_source_row_offset", 0) or 0)
    source_col_offset = int(raw_df.attrs.get("_source_col_offset", 0) or 0)
    source_range = str(raw_df.attrs.get("_source_range", ""))

    records: list[dict] = []
    current_section = ""
    exact_section_labels = {
        "assets",
        "liabilities",
        "liabilities and equity",
        "equity",
        "income statement",
        "balance sheet",
        "cash flow statement",
    }
    for row_idx in data_rows:
        label_tokens = [
            cell_text(values[row_idx, col_idx])
            for col_idx in range(data_col_start)
            if cell_text(values[row_idx, col_idx])
        ]
        if not label_tokens:
            continue

        line_item = label_tokens[-1]
        parent_line_item = label_tokens[-2] if len(label_tokens) >= 2 else ""
        line_item_path = " > ".join(label_tokens)
        normalized_label = re.sub(r"\s+", " ", line_item).strip().lower()
        if normalized_label in exact_section_labels:
            current_section = line_item

        nested_payload_label_cols: set[int] = set()
        for future_col in value_cols:
            future_text = cell_text(values[row_idx, future_col]).lower()
            if (
                _parse_numeric_value(values[row_idx, future_col]) is None
                and future_text not in {"n.m.", "n.m", "nm"}
            ):
                continue
            for previous_col in range(future_col - 1, data_col_start - 1, -1):
                previous_text = cell_text(values[row_idx, previous_col])
                if not previous_text:
                    continue
                if (
                    _parse_numeric_value(values[row_idx, previous_col]) is None
                    and not is_unit_label(previous_text)
                    and not _looks_like_date_or_period_text(previous_text)
                ):
                    nested_payload_label_cols.add(previous_col)
                break
        for possible_label_col in value_cols:
            possible_label = cell_text(values[row_idx, possible_label_col])
            if (
                not possible_label
                or _parse_numeric_value(values[row_idx, possible_label_col]) is not None
                or possible_label.lower() in {"n.m.", "n.m", "nm"}
                or not _looks_like_row_label_header(headers.get(possible_label_col, ""))
            ):
                continue
            if any(
                not is_blank(values[row_idx, later_col])
                for later_col in value_cols
                if later_col > possible_label_col
            ):
                nested_payload_label_cols.add(possible_label_col)

        for col_idx in value_cols:
            value = values[row_idx, col_idx]
            if is_blank(value):
                continue
            value_text = cell_text(value).lower()
            if (
                _parse_numeric_value(value) is None
                and value_text not in {"n.m.", "n.m", "nm"}
                and col_idx in nested_payload_label_cols
            ):
                continue

            record_line_item = line_item
            record_parent_line_item = parent_line_item
            record_line_item_path = line_item_path
            local_label_cols = [
                label_col
                for label_col in nested_payload_label_cols
                if label_col < col_idx
                and cell_text(values[row_idx, label_col])
            ]
            if local_label_cols:
                local_label_col = max(local_label_cols)
                record_line_item = cell_text(values[row_idx, local_label_col])
                record_parent_line_item = ""
                record_line_item_path = record_line_item

            metric = headers.get(col_idx, f"Column {_excel_col_letter(col_idx + 1)}")
            metric_leaf = metric.split("|")[-1].strip()
            source_row = source_row_offset + row_idx + 1
            source_col_idx = source_col_offset + col_idx + 1
            source_column = _excel_col_letter(source_col_idx)
            record = {
                "table_name": title,
                "table_subtitle": table_subtitle,
                "section": current_section,
                "column_group": column_groups.get(col_idx, ""),
                "unit": unit,
                "parent_line_item": record_parent_line_item,
                "line_item_path": record_line_item_path,
                "line_item": record_line_item,
                "metric": metric,
                "value": value,
                "source_range": source_range,
                "source_row": source_row,
                "source_column": source_column,
                "source_column_index": source_col_idx,
                "source_cell": f"{source_column}{source_row}",
                "block_key": source_range or f"region:{source_row_offset}:{source_col_offset}",
                "block_start_column": source_col_offset,
            }
            record.update(metric_context(metric_leaf))
            records.append(record)

    result = pd.DataFrame(records)
    result.attrs = {}
    return result


def _rescue_row_payload_groups(
    raw_df: pd.DataFrame,
) -> dict[int, list[dict]]:
    """Find credible value cells independently of table-region detection.

    The normal region parser is intentionally strict. This pass is deliberately
    narrower in scope: it only identifies numeric/n.m. payload cells that have a
    genuine text label somewhere to their left on the same row.
    """
    if raw_df is None or raw_df.empty:
        return {}

    values = raw_df.to_numpy(dtype=object, copy=False)
    row_groups: dict[int, list[dict]] = {}
    for row_idx, row in enumerate(values):
        texts = [cell_text(value) for value in row]
        parsed_cols: list[int] = []
        parsed_values: dict[int, float | None] = {}
        for col_idx, value in enumerate(row):
            text = texts[col_idx]
            parsed = _parse_numeric_value(value)
            if parsed is None and text.lower() not in {"n.m.", "n.m", "nm"}:
                continue
            parsed_cols.append(col_idx)
            parsed_values[col_idx] = parsed

        if not parsed_cols:
            continue

        finite_values = [
            value for value in parsed_values.values()
            if value is not None
        ]
        year_only = bool(finite_values) and all(
            float(value).is_integer() and 1900 <= int(value) <= 2100
            for value in finite_values
        )

        nearest_label: dict[int, int] = {}
        current_label_col = -1
        for col_idx, text in enumerate(texts):
            is_payload = col_idx in parsed_values
            if (
                text
                and not is_payload
                and not is_unit_label(text)
                and not _looks_like_date_or_period_text(text)
            ):
                current_label_col = col_idx
            if is_payload and current_label_col >= 0:
                nearest_label[col_idx] = current_label_col

        path_tokens_by_label: dict[int, list[str]] = {}
        for label_col in set(nearest_label.values()):
            previous_payload = max(
                (col_idx for col_idx in parsed_cols if col_idx < label_col),
                default=-1,
            )
            path_tokens_by_label[label_col] = [
                texts[col_idx]
                for col_idx in range(previous_payload + 1, label_col + 1)
                if texts[col_idx]
                and col_idx not in parsed_values
                and not is_unit_label(texts[col_idx])
                and not _looks_like_date_or_period_text(texts[col_idx])
            ]

        cell_groups: list[dict] = []
        for value_col in parsed_cols:
            label_col = nearest_label.get(value_col, -1)
            if label_col < 0:
                continue

            label = texts[label_col]
            if not label:
                continue

            # Period header rows commonly look like "EUR mn | 2023 | 2024".
            left_tokens = [
                texts[col_idx]
                for col_idx in range(value_col)
                if texts[col_idx]
            ]
            header_label = label.strip().lower()
            if year_only and (
                len(finite_values) >= 2
                or any(is_unit_label(token) for token in left_tokens)
                or header_label in {
                    "metric",
                    "period",
                    "year",
                    "date",
                    "line item",
                    "line_item",
                    "description",
                }
            ):
                continue

            path_tokens = path_tokens_by_label.get(label_col, [])
            if not path_tokens:
                path_tokens = [label]

            cell_groups.append(
                {
                    "row_idx": int(row_idx),
                    "value_col": int(value_col),
                    "label_col": int(label_col),
                    "line_item": label,
                    "path_tokens": path_tokens,
                    "value": row[value_col],
                }
            )

        if cell_groups:
            row_groups[row_idx] = cell_groups
    return row_groups


def _rescue_financial_value_cells(raw_df: pd.DataFrame) -> pd.DataFrame:
    """Recover credible value cells omitted or misassigned by region parsers."""
    row_groups = _rescue_row_payload_groups(raw_df)
    if not row_groups:
        return pd.DataFrame()

    candidate_rows = sorted(row_groups)
    row_clusters = _candidate_row_clusters(
        [(row_idx, min(item["value_col"] for item in row_groups[row_idx]))
         for row_idx in candidate_rows],
        max_row_gap=2,
    )
    row_to_cluster: dict[int, int] = {}
    cluster_rows: dict[int, list[int]] = {}
    for cluster_idx, cluster in enumerate(row_clusters, start=1):
        rows = [int(row_idx) for row_idx, _ in cluster]
        cluster_rows[cluster_idx] = rows
        for row_idx in rows:
            row_to_cluster[row_idx] = cluster_idx

    values = raw_df.to_numpy(dtype=object, copy=False)
    context = _sheet_scan_context(raw_df)
    source_row_offset = int(raw_df.attrs.get("_source_row_offset", 0) or 0)
    source_col_offset = int(raw_df.attrs.get("_source_col_offset", 0) or 0)
    source_range = str(raw_df.attrs.get("_source_range", ""))

    all_data_rows = {
        int(row_idx)
        for rows in cluster_rows.values()
        for row_idx in rows
    }

    cluster_headers: dict[int, tuple[dict[int, str], dict[int, str]]] = {}
    cluster_titles: dict[int, str] = {}
    cluster_units: dict[int, str] = {}
    for cluster_idx, rows in cluster_rows.items():
        first_data_row = min(rows)
        previous_cluster_end = max(
            (
                max(other_rows)
                for other_idx, other_rows in cluster_rows.items()
                if other_idx < cluster_idx
            ),
            default=-1,
        )
        header_start = max(previous_cluster_end + 1, first_data_row - 8, 0)
        header_rows = [
            row_idx
            for row_idx in range(header_start, first_data_row)
            if int(context.non_blank_count[row_idx]) > 0
        ][-4:]
        value_cols = sorted(
            {
                int(item["value_col"])
                for row_idx in rows
                for item in row_groups[row_idx]
            }
        )
        if header_rows and value_cols:
            headers_map, groups_map = _composite_headers_for_region(
                raw_df,
                header_rows,
                value_cols,
            )
        else:
            headers_map, groups_map = {}, {}

        # Later value bands under one shared header (e.g. an "in %" ratio block
        # below an "EUR mn" block) have no header rows of their own between
        # them and the previous band, so the clamped window resolves nothing.
        # For columns that still carry only a positional fallback, look further
        # up, skipping other bands' data rows, to inherit the shared
        # group/period header band.
        extended_rows: list[int] = []
        unresolved = [
            col
            for col in value_cols
            if not headers_map.get(col) or headers_map[col].startswith("Column ")
        ]
        if unresolved:
            extended_rows = [
                row_idx
                for row_idx in range(max(0, first_data_row - 8), first_data_row)
                if row_idx not in all_data_rows
                and int(context.non_blank_count[row_idx]) > 0
            ][-4:]
            if extended_rows and extended_rows != header_rows:
                ext_headers, ext_groups = _composite_headers_for_region(
                    raw_df,
                    extended_rows,
                    value_cols,
                )
                for col in unresolved:
                    candidate = ext_headers.get(col, "")
                    if candidate and not candidate.startswith("Column "):
                        headers_map[col] = candidate
                        if not groups_map.get(col):
                            groups_map[col] = ext_groups.get(col, "")
        cluster_headers[cluster_idx] = (headers_map, groups_map)

        unit = ""
        title = ""
        title_windows = [header_rows]
        if extended_rows and extended_rows != header_rows:
            title_windows.append(extended_rows)
        for window in title_windows:
            for row_idx in reversed(window):
                for col_idx in range(raw_df.shape[1]):
                    token = cell_text(values[row_idx, col_idx])
                    if not token:
                        continue
                    if not unit and (is_unit_label(token) or _is_col0_unit_anchor(token)):
                        unit = token
                    if (
                        not title
                        and not is_unit_label(token)
                        and not _looks_like_date_or_period_text(token)
                        and _parse_numeric_value(values[row_idx, col_idx]) is None
                    ):
                        title = token
                if unit and title:
                    break
            if unit and title:
                break
        cluster_units[cluster_idx] = unit
        cluster_titles[cluster_idx] = title or "Table"

    # A one-cell narrative row containing a number is not a table. Keep a
    # single-value row only when neighboring rows share its local structure or
    # a period header sits directly above the value column.
    credible_keys: set[tuple[int, int]] = set()
    for row_idx, groups in row_groups.items():
        cluster_idx = row_to_cluster[row_idx]
        rows = cluster_rows[cluster_idx]
        for item in groups:
            value_col = int(item["value_col"])
            label_col = int(item["label_col"])
            same_structure = any(
                other_row != row_idx
                and abs(other_row - row_idx) <= 2
                and any(
                    abs(int(other["label_col"]) - label_col) <= 1
                    and abs(int(other["value_col"]) - value_col) <= 1
                    for other in row_groups[other_row]
                )
                for other_row in rows
            )
            multi_value_row = len(groups) >= 2
            nearby_period_header = any(
                _looks_like_date_or_period_text(cell_text(values[header_row, value_col]))
                for header_row in range(max(0, row_idx - 3), row_idx)
            )
            if same_structure or multi_value_row or nearby_period_header:
                credible_keys.add((row_idx, value_col))

    records: list[dict] = []
    for row_idx, groups in row_groups.items():
        cluster_idx = row_to_cluster[row_idx]
        headers, column_groups = cluster_headers[cluster_idx]
        for item in groups:
            value_col = int(item["value_col"])
            if (row_idx, value_col) not in credible_keys:
                continue

            label_col = int(item["label_col"])
            path_tokens = list(item["path_tokens"])
            line_item = str(item["line_item"])
            metric = headers.get(
                value_col,
                f"Column {_excel_col_letter(source_col_offset + value_col + 1)}",
            )
            source_row = source_row_offset + row_idx + 1
            source_col_idx = source_col_offset + value_col + 1
            source_column = _excel_col_letter(source_col_idx)
            record = {
                "table_id": f"rescue_{cluster_idx}_{label_col}",
                "table_name": cluster_titles[cluster_idx],
                "source_range": source_range,
                "source_cell": f"{source_column}{source_row}",
                "source_row": source_row,
                "source_column": source_column,
                "source_column_index": source_col_idx,
                "section": "",
                "column_group": column_groups.get(value_col, ""),
                "unit": cluster_units[cluster_idx],
                "parent_line_item": path_tokens[-2] if len(path_tokens) >= 2 else "",
                "line_item_path": " > ".join(path_tokens),
                "line_item": line_item,
                "metric": metric,
                "value": item["value"],
                "block_key": f"rescue:{cluster_idx}:{label_col}",
                "block_start_column": source_col_offset + label_col,
                "extraction_method": "coverage_rescue",
            }
            record.update(metric_context(metric.split("|")[-1].strip()))
            records.append(record)

    result = pd.DataFrame(records)
    result.attrs = {}
    return result


def _merge_extraction_with_coverage_rescue(
    extracted: pd.DataFrame,
    rescued: pd.DataFrame,
) -> pd.DataFrame:
    """Add missing source cells and replace rows assigned to the wrong label."""
    if rescued is None or rescued.empty:
        return extracted
    if extracted is None or extracted.empty:
        return rescued
    if "source_cell" not in extracted.columns:
        # Without source coordinates on the primary extraction, cell-level
        # reconciliation is impossible. Concatenating both passes duplicates
        # every value they both captured (the rescue pass sweeps ALL credible
        # cells, not just missed ones), so keep exactly one side: the rescue
        # frame when it covers at least as many cells AND carries strictly
        # richer dimensional context, otherwise the primary extraction.
        context_columns = (
            "section",
            "column_group",
            "shared_header",
            "unit",
            "rate_type",
            "period",
            "valuation_date",
            "tenor",
            "contract_type",
            "currency",
            "parent_line_item",
            "line_item_path",
            "metric_type",
            "metric_date",
        )

        def _context_richness(frame: pd.DataFrame) -> int:
            score = 0
            for column in context_columns:
                if column in frame.columns and _series_cell_text(frame[column]).ne("").any():
                    score += 1
            return score

        primary_values = (
            extracted["value"] if "value" in extracted.columns else pd.Series(dtype="object")
        )
        primary_mostly_numeric = bool(
            len(primary_values)
            and primary_values.map(
                lambda item: _parse_numeric_value(item) is not None
                or cell_text(item).lower() in {"n.m.", "n.m", "nm"}
            ).mean() >= 0.9
        )
        rescue_covers_primary = len(rescued) >= int(0.9 * len(extracted))
        if (
            rescue_covers_primary
            and primary_mostly_numeric
            and _context_richness(rescued) > _context_richness(extracted)
        ):
            return rescued

        # Keeping the primary: still append rescue rows for cells the primary
        # demonstrably missed. Without coordinates the best available identity
        # is the (line_item, value) pair as a MULTISET — a rescue occurrence
        # beyond the primary's count for the same pair is an uncovered cell,
        # not a duplicate.
        def _pair_key(label, value) -> tuple:
            return (cell_text(label), cell_text(value))

        if {"line_item", "value"}.issubset(extracted.columns) and {"line_item", "value"}.issubset(rescued.columns):
            primary_counts: dict = {}
            for label, value in zip(extracted["line_item"].tolist(), extracted["value"].tolist()):
                key = _pair_key(label, value)
                primary_counts[key] = primary_counts.get(key, 0) + 1

            take_positions: list[int] = []
            running: dict = {}
            for position, (label, value) in enumerate(
                zip(rescued["line_item"].tolist(), rescued["value"].tolist())
            ):
                key = _pair_key(label, value)
                running[key] = running.get(key, 0) + 1
                if running[key] > primary_counts.get(key, 0):
                    take_positions.append(position)
            if take_positions:
                return pd.concat(
                    [extracted, rescued.iloc[take_positions]],
                    ignore_index=True,
                    sort=False,
                )
        return extracted

    existing = extracted.copy()
    rescued = rescued.copy()

    # Coverage-rescue rows are discovered independently from table regions.
    # Reattach them to an existing logical table when their row span, metrics,
    # units, and nearby columns clearly identify the parent region. This keeps
    # complete extraction from fragmenting one business table into rescue-only
    # pseudo tables in the SQL semantic catalog.
    required_match_cols = {
        "table_id",
        "source_row",
        "source_column_index",
        "metric",
    }
    if required_match_cols.issubset(existing.columns) and required_match_cols.issubset(rescued.columns):
        existing_profiles: list[dict] = []
        for table_id, table_df in existing.groupby("table_id", dropna=False, sort=False):
            rows = pd.to_numeric(table_df["source_row"], errors="coerce").dropna()
            cols = pd.to_numeric(table_df["source_column_index"], errors="coerce").dropna()
            if rows.empty or cols.empty:
                continue
            metrics = {
                cell_text(value)
                for value in table_df["metric"].tolist()
                if cell_text(value)
            }
            units = {
                cell_text(value)
                for value in table_df.get("unit", pd.Series(dtype="object")).tolist()
                if cell_text(value)
            }
            existing_profiles.append(
                {
                    "table_id": table_id,
                    "table_name": (
                        cell_text(table_df["table_name"].iloc[0])
                        if "table_name" in table_df.columns
                        else ""
                    ),
                    "source_range": (
                        cell_text(table_df["source_range"].iloc[0])
                        if "source_range" in table_df.columns
                        else ""
                    ),
                    "row_min": float(rows.min()),
                    "row_max": float(rows.max()),
                    "col_min": float(cols.min()),
                    "col_max": float(cols.max()),
                    "metrics": metrics,
                    "units": units,
                }
            )

        for rescue_id, rescue_df in rescued.groupby("table_id", dropna=False, sort=False):
            rescue_rows = pd.to_numeric(rescue_df["source_row"], errors="coerce").dropna()
            rescue_cols = pd.to_numeric(rescue_df["source_column_index"], errors="coerce").dropna()
            if rescue_rows.empty or rescue_cols.empty:
                continue
            rescue_metrics = {
                cell_text(value)
                for value in rescue_df["metric"].tolist()
                if cell_text(value)
            }
            rescue_units = {
                cell_text(value)
                for value in rescue_df.get("unit", pd.Series(dtype="object")).tolist()
                if cell_text(value)
            }
            rescue_row_min = float(rescue_rows.min())
            rescue_row_max = float(rescue_rows.max())
            rescue_col_mid = float(rescue_cols.median())

            best_profile = None
            best_score = float("-inf")
            for profile in existing_profiles:
                row_intersection = max(
                    0.0,
                    min(rescue_row_max, profile["row_max"])
                    - max(rescue_row_min, profile["row_min"])
                    + 1.0,
                )
                row_span = max(rescue_row_max - rescue_row_min + 1.0, 1.0)
                row_overlap = row_intersection / row_span
                metric_union = rescue_metrics | profile["metrics"]
                metric_overlap = (
                    len(rescue_metrics & profile["metrics"]) / len(metric_union)
                    if metric_union
                    else 0.0
                )
                unit_overlap = bool(rescue_units & profile["units"])
                if rescue_col_mid < profile["col_min"]:
                    col_distance = profile["col_min"] - rescue_col_mid
                elif rescue_col_mid > profile["col_max"]:
                    col_distance = rescue_col_mid - profile["col_max"]
                else:
                    col_distance = 0.0

                score = (
                    metric_overlap * 100.0
                    + row_overlap * 20.0
                    + (10.0 if unit_overlap else 0.0)
                    - min(col_distance, 30.0)
                )
                # A cluster separated from the table by a 2+ blank-row gap is
                # a stacked sibling table that merely shares the header band
                # (dashboards stack mini-tables under one header row), not a
                # missed piece of the same table — mirror the whitespace
                # threshold used by region detection.
                vertical_gap = max(
                    rescue_row_min - profile["row_max"],
                    profile["row_min"] - rescue_row_max,
                    0.0,
                )
                credible = vertical_gap < 3.0 and (
                    metric_overlap >= 0.20
                    or (row_overlap >= 0.80 and col_distance <= 4)
                    or (row_overlap >= 0.80 and unit_overlap and col_distance <= 24)
                )
                if credible and score > best_score:
                    best_profile = profile
                    best_score = score

            if best_profile is not None:
                mask = rescued["table_id"].eq(rescue_id)
                rescued.loc[mask, "table_id"] = best_profile["table_id"]
                if "table_name" in rescued.columns and best_profile["table_name"]:
                    rescued.loc[mask, "table_name"] = best_profile["table_name"]
                if "source_range" in rescued.columns and best_profile["source_range"]:
                    rescued.loc[mask, "source_range"] = best_profile["source_range"]

    rescue_by_cell = rescued.drop_duplicates(subset=["source_cell"], keep="last").set_index("source_cell")
    existing_cells = existing["source_cell"].astype(str)
    replace_mask = np.zeros(len(existing), dtype=bool)
    if "line_item" in existing.columns:
        existing_labels = existing["line_item"].map(cell_text)
        for position, (cell, label) in enumerate(zip(existing_cells.tolist(), existing_labels.tolist())):
            if cell not in rescue_by_cell.index:
                continue
            rescued_label = cell_text(rescue_by_cell.at[cell, "line_item"])
            if rescued_label and rescued_label != label:
                replace_mask[position] = True

    if bool(replace_mask.any()):
        existing = existing.loc[~replace_mask].copy()

    remaining_cells = set(existing["source_cell"].astype(str))
    additions = rescued.loc[
        ~rescued["source_cell"].astype(str).isin(remaining_cells)
    ]
    # Concatenating against a genuinely empty frame (common: nothing left to
    # rescue) triggers pandas' empty/all-NA concat dtype-inference FutureWarning
    # and is wasted work, so skip pd.concat entirely in that case.
    if additions.empty:
        combined = existing.copy()
    elif existing.empty:
        combined = additions.copy()
    else:
        combined = pd.concat([existing, additions], ignore_index=True, sort=False)
    combined.attrs = {}
    return combined


def _auto_flatten_single_region_legacy(raw_df: pd.DataFrame, extraction_profile: str = "auto") -> pd.DataFrame:
    """General extractor for visually formatted report sheets.

    This is the single user-facing auto mode. Internally it tries a few
    structural strategies and chooses the richest useful output. That keeps the
    UI general without pretending every visual report has one physical shape.

    KEY FIX: If ALL strategies fail (return empty), fall back to detect and flatten
    matrix-style data where both rows and columns contain meaningful identifiers.
    """
    profile = (extraction_profile or "auto").strip().lower()
    if profile not in {"auto", "general", "matrix"}:
        profile = "auto"

    raw_df = trim_effective_row_window(raw_df, tail_buffer=3)

    # Keep the original grid for sparse wide layouts (e.g., horizontal balance sheet)
    # where top rows are intentionally sparse and would be removed by error-density skipping.
    original_df = raw_df
    raw_df = _skip_error_heavy_rows(raw_df, max_check=50)
    raw_df = trim_effective_row_window(raw_df, tail_buffer=2)

    first_col_texts = [cell_text(value).lower() for value in raw_df.iloc[:, 0].tolist()]
    looks_like_rates_sheet = any(
        text.startswith("exchange rates") or text.startswith("valuation rates")
        for text in first_col_texts
    )

    top_values = [
        str(v).strip().upper()
        for row in raw_df.head(12).values
        for v in row
        if not is_blank(v)
    ]
    has_matrix_signals = any(v in {"#REF!", "#DIV/0!", "#N/A"} for v in top_values)
    is_heavy_layout = raw_df.shape[0] >= 500 and raw_df.shape[1] >= 80

    if is_heavy_layout:
        fast_df = auto_flatten_financial_summary_layout(raw_df)
        if fast_df is None or fast_df.empty:
            fast_df = auto_flatten_heavy_summary_grid(raw_df)
        if fast_df is not None and not fast_df.empty:
            best_df = fast_df.copy()
            normalized = best_df
            # Continue with the shared normalization path below.
        else:
            normalized = pd.DataFrame()
    else:
        normalized = pd.DataFrame()

    strategy_order = []
    if normalized.empty and (profile == "matrix" or (profile == "auto" and has_matrix_signals and not looks_like_rates_sheet)):
        strategy_order = [
            ("fallback_matrix", _fallback_matrix_flatten),
            ("horizontal_balance_sheet", auto_flatten_horizontal_balance_sheet),
            ("financial_summary_layout", auto_flatten_financial_summary_layout),
        ]
    elif normalized.empty and looks_like_rates_sheet:
        strategy_order = [
            ("financial_summary_layout", auto_flatten_financial_summary_layout),
            ("sectioned_tables", auto_flatten_sectioned_financial_sheet),
            ("grouped_metric_blocks", auto_flatten_grouped_metric_blocks),
            ("side_by_side_blocks", auto_flatten_report_blocks),
            ("stacked_tables", auto_flatten_stacked_tables),
            ("fallback_matrix", _fallback_matrix_flatten),
            ("horizontal_balance_sheet", auto_flatten_horizontal_balance_sheet),
        ]
    elif normalized.empty:
        strategy_order = [
            ("horizontal_balance_sheet", auto_flatten_horizontal_balance_sheet),
            ("financial_summary_layout", auto_flatten_financial_summary_layout),
            ("heavy_summary_grid", auto_flatten_heavy_summary_grid),
            ("side_by_side_blocks", auto_flatten_report_blocks),
            ("stacked_tables", auto_flatten_stacked_tables),
            ("sectioned_tables", auto_flatten_sectioned_financial_sheet),
            ("fallback_matrix", _fallback_matrix_flatten),
        ]
        if is_heavy_layout:
            strategy_order = [
                ("horizontal_balance_sheet", auto_flatten_horizontal_balance_sheet),
                ("financial_summary_layout", auto_flatten_financial_summary_layout),
                ("heavy_summary_grid", auto_flatten_heavy_summary_grid),
                ("side_by_side_blocks", auto_flatten_report_blocks),
            ]

    best_name = ""
    best_df = normalized.copy()
    best_score = len(best_df) if not best_df.empty else 0
    for name, extractor in strategy_order:
        source_df = original_df if name == "horizontal_balance_sheet" else raw_df
        df = extractor(source_df)
        if df is None or df.empty:
            continue

        if name == "horizontal_balance_sheet" and len(df) >= 40:
            best_name = name
            best_df = df.copy()
            best_score = len(df)
            break

        # If user explicitly selected matrix mode, trust matrix extractor output.
        if profile == "matrix" and name == "fallback_matrix":
            best_name = name
            best_df = df.copy()
            best_score = len(df)
            break

        context_cols = [
            col for col in df.columns
            if col not in {"value", "block_start_column"}
        ]
        score = len(df) * max(len(context_cols), 1)
        if looks_like_rates_sheet and name == "sectioned_tables":
            # Preserve Spot/Average/tenor context on market-data style sheets.
            score *= 10
        if score > best_score:
            best_name = name
            best_df = df.copy()
            best_score = score

        # Early exit for obvious wins to reduce work on large sheets.
        if name == "fallback_matrix" and len(df) >= 500:
            break
        if looks_like_rates_sheet and name == "sectioned_tables" and len(df) >= 80:
            break

    if best_df.empty:
        return best_df

    normalized = best_df.copy()

    hierarchy_df = propagate_section_context(raw_df)
    if "line_item" in normalized.columns and not hierarchy_df.empty and "line_item" in hierarchy_df.columns:
        lineage_cols = [col for col in hierarchy_df.columns if isinstance(col, str) and col.startswith("_section_L")]
        if lineage_cols:
            hierarchy_map = hierarchy_df.loc[:, ["line_item", *lineage_cols]].copy()
            hierarchy_map["line_item"] = _series_cell_text(hierarchy_map["line_item"])
            hierarchy_map = hierarchy_map[hierarchy_map["line_item"].ne("")]
            hierarchy_map = hierarchy_map.drop_duplicates(subset=["line_item"], keep="first")
            normalized["line_item"] = _series_cell_text(normalized["line_item"])
            normalized = normalized.merge(hierarchy_map, on="line_item", how="left")

    if "block_id" not in normalized.columns:
        if "block_key" in normalized.columns:
            starts = normalized["block_key"].fillna("-1")
            normalized = normalized.drop(columns=["block_key"])
            normalized.insert(
                1,
                "block_id",
                pd.factorize(starts.astype(str))[0] + 1,
            )
        elif "block_start_column" in normalized.columns:
            starts = normalized["block_start_column"].fillna(-1)
            normalized.insert(
                1,
                "block_id",
                pd.factorize(starts.astype(str))[0] + 1,
            )
        else:
            normalized.insert(1, "block_id", 1)

    if "block_start_column" in normalized.columns:
        normalized = normalized.drop(columns=["block_start_column"])

    if "line_item" not in normalized.columns:
        if "currency" in normalized.columns:
            # currency IS the line item for rates sheets; don't create a duplicate column
            normalized["line_item"] = normalized["currency"]
            normalized = normalized.drop(columns=["currency"])
        else:
            normalized["line_item"] = ""

    # Drop currency if it's now a duplicate of line_item
    if (
            "currency" in normalized.columns
            and "line_item" in normalized.columns
            and columns_semantically_equal(normalized["currency"], normalized["line_item"])
    ):
        normalized = normalized.drop(columns=["currency"])

    if "metric" not in normalized.columns:
        metric_parts = []
        for _, row in normalized.iterrows():
            parts = [
                cell_text(row.get("period", "")),
                cell_text(row.get("valuation_date", "")),
                cell_text(row.get("tenor", "")),
                cell_text(row.get("rate_type", "")),
            ]
            metric_parts.append(" | ".join([part for part in parts if part]) or "value")
        normalized["metric"] = metric_parts

    # For rates sheets, populate metric_date/year/quarter from period or valuation_date
    # when those context columns are missing or empty.
    date_source_col = None
    if "period" in normalized.columns and normalized["period"].astype(str).str.strip().ne("").any():
        date_source_col = "period"
    elif "valuation_date" in normalized.columns and normalized["valuation_date"].astype(str).str.strip().ne("").any():
        date_source_col = "valuation_date"

    if date_source_col is not None:
        context_cols_to_fill = [
            "metric_type",
            "metric_date",
            "metric_quarter",
            "comparison_year",
            "comparison_date",
        ]
        # Only fill if all context columns are missing or entirely empty
        needs_fill = all(
            col not in normalized.columns or normalized[col].astype(str).str.strip().eq("").all()
            for col in context_cols_to_fill
        )
        if needs_fill:
            parsed = [metric_context(str(v)) for v in normalized[date_source_col].tolist()]
            for ctx_col in context_cols_to_fill:
                normalized[ctx_col] = [item.get(ctx_col, "") for item in parsed]

    if "table_name" not in normalized.columns:
        normalized["table_name"] = normalized["section"] if "section" in normalized.columns else ""

    # Drop table_name if it's semantically identical to section
    if (
            "table_name" in normalized.columns
            and "section" in normalized.columns
            and columns_semantically_equal(normalized["table_name"], normalized["section"])
    ):
        normalized = normalized.drop(columns=["table_name"])

    if "unit" not in normalized.columns:
        normalized["unit"] = ""

    for col in normalized.columns:
        if col != "value":
            normalized[col] = normalized[col].fillna("")

    normalized = _assign_schema_group_id(normalized)
    normalized = _reassign_block_id_from_schema_group(normalized)

    normalized = normalized.drop(
        columns=[
            col for col in (
                "block_id",
                "schema_group_id",
                "source_block_id",
                "source_schema_group_id",
                "logical_block_key",
                "block_key",
                "block_start_column",
            )
            if col in normalized.columns
        ],
        errors="ignore",
    )

    ordered_columns = [
        "tab_name",
        "table_id",
        "table_name",
        "table_subtitle",
        "source_range",
        "source_cell",
        "source_row",
        "source_column",
        "schema_signature",
        "section",
        "column_group",
        "shared_header",
        "unit",
        "rate_type",
        "period",
        "valuation_date",
        "tenor",
        "contract_type",
        "currency",
        "parent_line_item",
        "line_item_path",
        "line_item",
        "metric",
        "metric_detail",
        "metric_type",
        "metric_date",
        "comparison_date",
        "metric_quarter",
        "comparison_year",
        "value",
    ]
    existing_ordered = [col for col in ordered_columns if col in normalized.columns]
    remaining = [col for col in normalized.columns if col not in existing_ordered]
    result = drop_all_blank_columns(normalized[existing_ordered + remaining])
    result.attrs = {}
    return result


RE_FALLBACK_COLUMN_HEADER = re.compile(r"(?:^|\|\s*)Column\s+[A-Z]+$")


def _quarantine_isolated_columns(combined: pd.DataFrame) -> pd.DataFrame:
    """Flag observations from stray value columns far outside their table's
    value block — typically interim/helper numbers parked beside a report.

    Nothing is deleted: flagged rows get is_isolated_cell=1 and a table_id
    suffix so the SQL catalog and guardrails keep them out of aggregates while
    the values stay inspectable. A column cluster is quarantined only when it
    is separated from the main block by a wide gap (>= 3 columns, beyond the
    single spacer columns striped layouts use), carries a small share of the
    table's observations, and is either nearly empty or headerless (positional
    "Column X" fallback headers).
    """
    required = {"table_id", "source_row", "source_column_index"}
    if combined is None or combined.empty or not required.issubset(combined.columns):
        return combined

    col_numbers = pd.to_numeric(combined["source_column_index"], errors="coerce")
    row_numbers = pd.to_numeric(combined["source_row"], errors="coerce")
    table_ids = combined["table_id"].map(cell_text)
    metrics_text = (
        _series_cell_text(combined["metric"])
        if "metric" in combined.columns
        else pd.Series("", index=combined.index, dtype="string")
    )
    isolated_mask = np.zeros(len(combined), dtype=bool)

    for table_id in table_ids.unique():
        if not table_id:
            continue
        member_mask = table_ids.eq(table_id) & col_numbers.notna() & row_numbers.notna()
        member_total = int(member_mask.sum())
        if member_total < 6:
            continue
        member_cols = col_numbers[member_mask]
        member_rows = row_numbers[member_mask]
        total_rows = int(member_rows.nunique())
        if total_rows < 3:
            continue

        # Cluster the table's columns on gaps wider than striped spacers.
        distinct_cols = sorted(int(value) for value in member_cols.unique())
        clusters: list[list[int]] = [[distinct_cols[0]]]
        for column in distinct_cols[1:]:
            if column - clusters[-1][-1] >= 3:
                clusters.append([column])
            else:
                clusters[-1].append(column)
        if len(clusters) <= 1:
            continue

        cluster_sizes = [
            int(member_cols.isin(cluster).sum()) for cluster in clusters
        ]
        main_cluster = clusters[cluster_sizes.index(max(cluster_sizes))]
        for cluster, cluster_obs in zip(clusters, cluster_sizes):
            if cluster is main_cluster:
                continue
            if cluster_obs > 0.25 * member_total:
                continue
            cluster_member_mask = member_mask & col_numbers.isin(cluster)
            cluster_rows = row_numbers[cluster_member_mask]
            sparse_fill = int(cluster_rows.nunique()) <= max(1, int(0.25 * total_rows))
            cluster_metrics = metrics_text[cluster_member_mask]
            headerless = (
                bool(len(cluster_metrics))
                and float(
                    cluster_metrics.map(
                        lambda m: bool(RE_FALLBACK_COLUMN_HEADER.search(str(m)))
                    ).mean()
                ) >= 0.5
            )
            if sparse_fill or headerless:
                isolated_mask |= cluster_member_mask.to_numpy()

    if not bool(isolated_mask.any()):
        return combined
    combined = combined.copy()
    combined["is_isolated_cell"] = isolated_mask.astype(int)
    combined.loc[isolated_mask, "table_id"] = (
        combined.loc[isolated_mask, "table_id"].map(cell_text) + "::isolated"
    )
    return combined


def auto_flatten_report_tables(raw_df: pd.DataFrame, extraction_profile: str = "auto") -> pd.DataFrame:
    """Extract every detected table region instead of selecting one winner per tab."""
    if raw_df is None or raw_df.empty:
        return pd.DataFrame()

    profile = (extraction_profile or "auto").strip().lower()
    if profile not in {"auto", "general", "matrix"}:
        profile = "auto"

    regions = detect_table_regions(raw_df)
    parent_report_titles = _parent_report_titles_for_regions(raw_df, regions)
    extracted_regions: list[pd.DataFrame] = []

    for region_idx, region in enumerate(regions, start=1):
        region_df = _slice_table_region(raw_df, region)
        extracted = auto_flatten_rectangular_table(region_df)
        extraction_method = "region_grid"

        if extracted is None or extracted.empty:
            extracted = _auto_flatten_single_region_legacy(
                region_df,
                extraction_profile=profile,
            )
            extraction_method = "legacy_region_fallback"

        if extracted is None or extracted.empty:
            continue

        extracted = extracted.copy()
        extracted.attrs = {}
        if "table_id" not in extracted.columns:
            extracted.insert(0, "table_id", f"table_{region_idx}")
        if "source_range" not in extracted.columns:
            extracted["source_range"] = region.source_range
        parent_report_name = parent_report_titles.get(region_idx, "")
        if parent_report_name:
            extracted["parent_report_name"] = parent_report_name
        extracted["source_region_order"] = region_idx
        extracted["extraction_method"] = extraction_method
        extracted_regions.append(extracted)

    if extracted_regions:
        combined = pd.concat(extracted_regions, ignore_index=True, sort=False)
        combined.attrs = {}
        if profile != "matrix":
            rescued = _rescue_financial_value_cells(raw_df)
            combined = _merge_extraction_with_coverage_rescue(combined, rescued)
        combined = _apply_sheet_level_context(combined, raw_df)
        combined = _quarantine_isolated_columns(combined)
        if "parent_report_name" in combined.columns and "table_id" in combined.columns:
            parent_by_table = (
                combined.assign(
                    _parent_text=combined["parent_report_name"].map(cell_text)
                )
                .loc[lambda frame: frame["_parent_text"].ne("")]
                .drop_duplicates("table_id")
                .set_index("table_id")["_parent_text"]
                .to_dict()
            )
            combined["parent_report_name"] = [
                cell_text(value) or parent_by_table.get(table_id, "")
                for value, table_id in zip(
                    combined["parent_report_name"].tolist(),
                    combined["table_id"].tolist(),
                )
            ]

        source_keys = [
            col
            for col in ("source_cell",)
            if col in combined.columns
        ]
        if source_keys:
            combined = combined.drop_duplicates(subset=source_keys, keep="first")

        sort_cols = [
            col
            for col in (
                "source_region_order",
                "source_row",
                "source_column_index",
                "table_id",
            )
            if col in combined.columns
        ]
        if sort_cols:
            combined = combined.sort_values(sort_cols, kind="stable")
        return combined.reset_index(drop=True)

    fallback = _auto_flatten_single_region_legacy(
        raw_df,
        extraction_profile=profile,
    )
    if fallback is not None and not fallback.empty and "table_id" not in fallback.columns:
        fallback.insert(0, "table_id", "table_1")
    if profile != "matrix":
        rescued = _rescue_financial_value_cells(raw_df)
        fallback = _merge_extraction_with_coverage_rescue(fallback, rescued)
    fallback = _apply_sheet_level_context(fallback, raw_df)
    fallback = _quarantine_isolated_columns(fallback)
    if fallback is not None:
        fallback.attrs = {}
    return fallback


def auto_flatten_stacked_tables(raw_df: pd.DataFrame) -> pd.DataFrame:
    """Flatten common stacked mini-table layouts into metric/value rows.

    Looks for header rows where column 0 is a unit label (for example EUR mn)
    and later nonblank cells are metrics/periods. Rows below become line items
    until a run of blank rows or the next header.
    """
    records = []
    if raw_df is None or raw_df.empty:
        return pd.DataFrame()

    unit_prefixes = ("eur ", "usd ", "gbp ", "in %")
    i = 0
    while i < len(raw_df):
        row = raw_df.iloc[i]
        first = cell_text(row.iloc[0])
        first_lower = first.lower()
        metric_cols = nonblank_col_indexes(row, start_col=1)

        if first_lower.startswith(unit_prefixes) and metric_cols:
            unit = first
            table_name, column_group = "", ""
            if _is_col0_unit_anchor(row.iloc[0]):
                table_name, column_group = _resolve_block_header_from_unit_anchor(raw_df, i)
            if not table_name:
                table_name = nearest_title_above(raw_df, i)
            if not column_group:
                column_group = nearest_group_above(raw_df, i)
            metric_headers = {
                col_idx: cell_text(row.iloc[col_idx])
                for col_idx in metric_cols
            }
            j = i + 1
            blank_label_rows = 0
            while j < len(raw_df):
                data_row = raw_df.iloc[j]
                item = cell_text(data_row.iloc[0])
                item_lower = item.lower()
                if not item:
                    blank_label_rows += 1
                    if blank_label_rows >= 5:
                        break
                    j += 1
                    continue
                blank_label_rows = 0
                if item_lower.startswith(unit_prefixes):
                    break
                if not row_has_values(data_row, start_col=1):
                    j += 1
                    continue

                row_context = indented_row_context(raw_df, j, 0)
                for col_idx, metric in metric_headers.items():
                    value = data_row.iloc[col_idx]
                    if not is_blank(value):
                        record = {
                            "table_name": table_name,
                            "column_group": column_group,
                            "unit": unit,
                            "line_item": item,
                            "metric": metric,
                            "value": value,
                        }
                        record.update(row_context)
                        record.update(metric_context(metric))
                        records.append(record)
                j += 1
            i = j
            continue

        i += 1

    return pd.DataFrame(records)


# Function-level cache for when Streamlit session_state isn't available
_QUICK_READ_CACHE_MAX = 32
_quick_read_memory_cache: "OrderedDict[str, pd.DataFrame]" = OrderedDict()

# Shared pd.ExcelFile handles so per-sheet pandas reads of the same workbook
# parse the archive metadata (zip directory, shared strings, styles) only once.
_EXCEL_FILE_CACHE: OrderedDict = OrderedDict()
_EXCEL_FILE_CACHE_MAX = 2


def _preferred_excel_engine() -> str | None:
    """Pick the fastest available engine for VALUE-ONLY workbook reads.

    python-calamine (a Rust reader, 10-30x faster than openpyxl) is used when
    installed and pandas is new enough to support it. Style-aware reads
    (_open_display_workbook) always stay on openpyxl — calamine carries no
    style metadata. Returning None keeps pandas' default engine.
    """
    try:
        import python_calamine  # noqa: F401
    except Exception:
        return None
    try:
        major, minor = (int(part) for part in pd.__version__.split(".")[:2])
    except Exception:
        return None
    return "calamine" if (major, minor) >= (2, 2) else None


_EXCEL_VALUE_ENGINE = _preferred_excel_engine()


def _shared_excel_file(file_bytes: bytes, digest: str | None = None) -> pd.ExcelFile:
    key = digest or hashlib.sha256(file_bytes).hexdigest()
    cached = _EXCEL_FILE_CACHE.get(key)
    if cached is not None:
        _EXCEL_FILE_CACHE.move_to_end(key)
        return cached
    handle = pd.ExcelFile(io.BytesIO(file_bytes), engine=_EXCEL_VALUE_ENGINE)
    _EXCEL_FILE_CACHE[key] = handle
    while len(_EXCEL_FILE_CACHE) > _EXCEL_FILE_CACHE_MAX:
        _, evicted = _EXCEL_FILE_CACHE.popitem(last=False)
        try:
            evicted.close()
        except Exception:
            pass
    return handle


def _quick_read_cached(file_bytes: bytes, sheet_name: str) -> pd.DataFrame:
    """Fast pandas read for structure detection, cached when possible."""
    digest = hashlib.sha256(file_bytes).hexdigest()
    cache_key = f"_quick_{digest}_{sheet_name}"

    # Try Streamlit session_state first (best caching in production)
    try:
        if cache_key not in st.session_state:
            st.session_state[cache_key] = pd.read_excel(
                _shared_excel_file(file_bytes, digest), sheet_name=sheet_name, header=None
            )
        return st.session_state[cache_key]
    except Exception:
        # Fallback: bounded LRU when session_state is unavailable (outside
        # Streamlit, in tests, or offline). Bounded so a long-lived process that
        # reads many workbooks does not grow without limit.
        cache = _quick_read_memory_cache
        if cache_key in cache:
            cache.move_to_end(cache_key)
            return cache[cache_key]
        frame = pd.read_excel(
            _shared_excel_file(file_bytes, digest), sheet_name=sheet_name, header=None
        )
        cache[cache_key] = frame
        while len(cache) > _QUICK_READ_CACHE_MAX:
            cache.popitem(last=False)
        return frame


def infer_report_layout_from_quick_df(quick_df: pd.DataFrame) -> tuple[str, str, bool, bool]:
    """Infer whether a sheet looks like a general report block layout or matrix layout.

    Not cached: it only samples the top rows, so recomputing is cheaper than the
    whole-DataFrame hash @st.cache_data needs to build a cache key (measured
    ~2x). Callers that would otherwise invoke it twice for the same frame should
    compute it once and pass the result into infer_flat_table_structure().
    """
    if quick_df is None or quick_df.empty:
        return "general", "Sheet is mostly empty in quick preview.", False, False

    has_report_signals = (
            quick_df.shape[1] > 0
            and any(
        str(v).lower().startswith(
            ("exchange rates", "valuation rates", "eur ", "usd ", "gbp ", "in %")
        )
        for v in quick_df.iloc[:, 0].tolist()
        if not is_blank(v)
    )
    )

    top_values = [
        str(v).strip()
        for row in quick_df.head(20).values
        for v in row
        if not is_blank(v)
    ]
    # Check for error tokens more thoroughly (up to 20 rows, not just 12)
    error_tokens_list = {"#REF!", "#DIV/0!", "#N/A", "#VALUE!", "#NULL!", "#NUM!", "#ERROR!", "#NAME?"}
    has_matrix_signals = any(v.upper() in error_tokens_list for v in top_values)

    first_col_texts = [cell_text(v).lower() for v in quick_df.iloc[:, 0].head(40).tolist()]
    top_left_texts = [
        cell_text(v).lower()
        for row in quick_df.head(12).iloc[:, : min(12, quick_df.shape[1])].values
        for v in row
    ]
    if any("matrix" in text for text in first_col_texts if text) or any(
            "matrix" in text for text in top_left_texts if text):
        has_matrix_signals = True

    # Matrix sheets are often very wide and have many value-heavy rows with sparse label columns.
    row_sample = quick_df.iloc[: min(len(quick_df), 120)].to_numpy(
        dtype=object,
        copy=False,
    )
    first_col_nonblank_top = sum(1 for value in quick_df.iloc[:, 0].head(20).tolist() if cell_text(value))
    wide_sheet_signal = quick_df.shape[1] >= 40
    value_heavy_rows = 0
    for values in row_sample:
        label_left = cell_text(values[0]) if len(values) else ""
        metric_left = cell_text(values[1]) if len(values) > 1 else ""
        numeric_hits = sum(1 for value in values[2:] if looks_like_data_value(value))
        if numeric_hits >= 8 and (label_left or metric_left):
            value_heavy_rows += 1
    structural_matrix_signal = (
            wide_sheet_signal
            and value_heavy_rows >= 5
            and first_col_nonblank_top <= 5  # Relaxed from 3 to account for error tokens being skipped
    )
    if structural_matrix_signal:
        has_matrix_signals = True

    # Run expensive general header detection only when matrix signals are absent.
    if not has_report_signals and not has_matrix_signals:
        has_report_signals = sheet_looks_general_report_tables(quick_df)

    # Formula errors alone are not enough to classify a narrow financial
    # schedule as a matrix. Published reports often contain a few broken
    # workbook links while still having ordinary report-block structure.
    if (
            has_matrix_signals
            and has_report_signals
            and quick_df.shape[1] < 20
    ):
        return "general", "Narrow report structure overrides isolated matrix/error signals.", True, False
    if has_matrix_signals:
        return "matrix", "Detected matrix/error-token style header signals.", has_report_signals, has_matrix_signals
    if has_report_signals:
        return "general", "Detected general report block signals.", has_report_signals, has_matrix_signals
    return "general", "No strong report signals detected.", has_report_signals, has_matrix_signals


def infer_flat_table_structure(
        quick_df: pd.DataFrame,
        max_header_scan: int = 10,
        report_layout: tuple[str, str, bool, bool] | None = None,
) -> dict:
    """Positively identify a conventional header-row + record-row table.

    A sheet is considered flat only when it demonstrates stable tabular
    structure. Merely lacking report signals is not enough.

    report_layout lets a caller pass a precomputed
    infer_report_layout_from_quick_df() result to avoid recomputing it.
    """
    result = {
        "is_flat": False,
        "header_row": 0,
        "score": 0.0,
        "reason": "No stable flat-table structure detected.",
    }
    if quick_df is None or quick_df.empty:
        result["reason"] = "Sheet is empty."
        return result
    if quick_df.shape[0] < 6 or quick_df.shape[1] < 2:
        result["reason"] = "Too few rows or columns for a record-oriented flat table."
        return result

    _, report_reason, has_report_signals, has_matrix_signals = (
        report_layout
        if report_layout is not None
        else infer_report_layout_from_quick_df(quick_df)
    )
    if has_matrix_signals:
        result["reason"] = f"Matrix safeguard: {report_reason}"
        return result

    values = quick_df.to_numpy(dtype=object, copy=False)
    row_count, col_count = values.shape
    scan_limit = min(max(int(max_header_scan), 1), row_count - 4)
    best_candidate: dict | None = None

    for header_row in range(scan_limit):
        header_text = [cell_text(value) for value in values[header_row]]
        non_blank_header_cols = [
            col_idx for col_idx, text in enumerate(header_text) if text
        ]
        if len(non_blank_header_cols) < 2:
            continue

        first_col = min(non_blank_header_cols)
        last_col = max(non_blank_header_cols)
        active_width = last_col - first_col + 1
        if active_width < 2:
            continue
        header_slice = header_text[first_col:last_col + 1]
        header_non_blank = [text for text in header_slice if text]
        header_fill = len(header_non_blank) / active_width
        header_unique = len({text.lower() for text in header_non_blank}) / max(
            len(header_non_blank), 1
        )
        header_text_ratio = sum(
            1
            for text in header_non_blank
            if RE_ALPHA.search(text)
            and _parse_numeric_value(text) is None
        ) / max(len(header_non_blank), 1)
        if header_fill < 0.75 or header_unique < 0.80 or header_text_ratio < 0.60:
            continue

        sample_end = min(row_count, header_row + 201)
        body = values[header_row + 1:sample_end, first_col:last_col + 1]
        if body.shape[0] < 4:
            continue
        body_blank = np.zeros(body.shape, dtype=bool)
        for row_idx in range(body.shape[0]):
            for col_idx in range(body.shape[1]):
                body_blank[row_idx, col_idx] = is_blank(body[row_idx, col_idx])
        row_fill = 1.0 - body_blank.mean(axis=1)
        nonempty_rows = row_fill > 0
        if int(nonempty_rows.sum()) < 4:
            continue
        populated_fill = row_fill[nonempty_rows]
        median_row_fill = float(np.median(populated_fill))
        low_density_rows = float((populated_fill < 0.50).mean())
        blank_separator_ratio = float((~nonempty_rows).mean())
        column_fill = 1.0 - body_blank[nonempty_rows].mean(axis=0)
        stable_columns = float((column_fill >= 0.60).mean())

        # Flat files normally have one header followed immediately by records.
        leading_gap = 0
        for fill in row_fill:
            if fill > 0:
                break
            leading_gap += 1

        score = (
            25.0 * header_fill
            + 20.0 * header_unique
            + 15.0 * header_text_ratio
            + 20.0 * median_row_fill
            + 15.0 * stable_columns
            - 20.0 * low_density_rows
            - 25.0 * blank_separator_ratio
            - 5.0 * min(leading_gap, 3)
        )
        if has_report_signals:
            # The general report heuristic intentionally has high recall and can
            # flag ordinary transaction tables. Require stronger positive flat
            # evidence instead of treating that heuristic as an absolute veto.
            score -= 8.0

        # A very wide table with mostly numeric payload and just one or two
        # label columns is a matrix, even if its header row is clean.
        numeric_heavy_rows = 0
        if active_width >= 35:
            for row in body[: min(len(body), 80)]:
                numeric_hits = sum(
                    1 for value in row[2:] if _parse_numeric_value(value) is not None
                )
                if numeric_hits >= max(8, int((active_width - 2) * 0.60)):
                    numeric_heavy_rows += 1
        if active_width >= 35 and numeric_heavy_rows >= 5:
            continue

        candidate = {
            "is_flat": score >= (72.0 if has_report_signals else 65.0)
            and median_row_fill >= 0.65
            and stable_columns >= 0.70
            and blank_separator_ratio <= 0.10,
            "header_row": int(header_row),
            "score": round(score, 2),
            "reason": (
                f"Flat-table score {score:.1f}: header row {header_row + 1}, "
                f"{active_width} populated columns, median row fill "
                f"{median_row_fill:.0%}, stable columns {stable_columns:.0%}"
                f"{'; passed a stricter report-signal threshold' if has_report_signals else ''}."
            ),
        }
        if best_candidate is None or candidate["score"] > best_candidate["score"]:
            best_candidate = candidate

    return best_candidate or result


def _read_flat_excel_sheet(
        file_bytes: bytes,
        sheet_name: str,
        header_row: int = 0,
) -> pd.DataFrame:
    """Read a positively detected flat sheet without report extraction."""
    frame = pd.DataFrame(
        pd.read_excel(
            _shared_excel_file(file_bytes),
            sheet_name=sheet_name,
            header=max(int(header_row), 0),
        )
    )
    frame = frame.dropna(axis=0, how="all").dropna(axis=1, how="all")
    frame.columns = dedupe_columns(
        [
            column
            if not str(column).lower().startswith("unnamed:")
            else f"column_{idx + 1}"
            for idx, column in enumerate(frame.columns)
        ]
    )
    return drop_empty_or_zero_columns(frame.reset_index(drop=True))


@st.cache_data(show_spinner=False)
def _extract_report_cached(file_bytes: bytes, sheet_name: str, extraction_profile: str = "auto") -> pd.DataFrame:
    # Always use openpyxl for proper merged cell handling; rely on extraction strategy optimization instead
    raw_df = _read_display_sheet_cached(file_bytes, sheet_name)
    return auto_flatten_report_tables(raw_df, extraction_profile=extraction_profile)


MAX_WORKBOOK_ARCHIVE_MEMBERS = 10_000
MAX_WORKBOOK_UNCOMPRESSED_BYTES = 1 << 30  # 1 GiB


def _validate_workbook_archive(file_bytes: bytes) -> None:
    """Reject archives whose declared contents are far larger than any real
    spreadsheet, before openpyxl/pandas decompress them into memory."""
    with zipfile.ZipFile(io.BytesIO(file_bytes)) as archive:
        members = archive.infolist()
        if len(members) > MAX_WORKBOOK_ARCHIVE_MEMBERS:
            raise ValueError(
                "Workbook archive has too many entries to be a real spreadsheet."
            )
        declared_size = sum(member.file_size for member in members)
        if declared_size > MAX_WORKBOOK_UNCOMPRESSED_BYTES:
            raise ValueError(
                "Workbook contents exceed the supported decompressed size limit."
            )


def _workbook_has_merged_cells(file_bytes: bytes) -> bool:
    """Check workbook XML once so openpyxl can choose the cheapest safe mode."""
    has_merged_cells = False
    tag = b"mergeCell"
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as archive:
            for member in archive.namelist():
                if not member.startswith("xl/worksheets/") or not member.endswith(".xml"):
                    continue
                # Some writers namespace worksheet tags (for example
                # `<x:mergeCell>`), so search for the tag name rather than an
                # exact unprefixed opening tag. Stream in chunks so one
                # worksheet member never has to fit in memory whole.
                with archive.open(member) as handle:
                    carry = b""
                    while not has_merged_cells:
                        chunk = handle.read(1 << 20)
                        if not chunk:
                            break
                        if tag in carry + chunk:
                            has_merged_cells = True
                        carry = chunk[-(len(tag) - 1):]
                if has_merged_cells:
                    break
    except Exception:
        has_merged_cells = False
    return has_merged_cells


_DISPLAY_WORKBOOK_CACHE: OrderedDict = OrderedDict()
_DISPLAY_WORKBOOK_CACHE_MAX = 2


def _open_display_workbook(file_bytes: bytes):
    """Open one values-only workbook while retaining merged-cell metadata.

    The parsed workbook is kept in a small LRU keyed by content hash: single-
    sheet callers otherwise re-parse the entire archive once per sheet, which
    dominates extraction time on multi-sheet files. Callers must NOT close the
    returned workbook — eviction from the cache closes it.
    """
    key = hashlib.sha256(file_bytes).hexdigest()
    cached = _DISPLAY_WORKBOOK_CACHE.get(key)
    if cached is not None:
        _DISPLAY_WORKBOOK_CACHE.move_to_end(key)
        return cached

    _validate_workbook_archive(file_bytes)
    workbook = openpyxl.load_workbook(
        io.BytesIO(file_bytes),
        read_only=not _workbook_has_merged_cells(file_bytes),
        data_only=True,
        keep_vba=False,
        rich_text=False,
    )
    _DISPLAY_WORKBOOK_CACHE[key] = workbook
    while len(_DISPLAY_WORKBOOK_CACHE) > _DISPLAY_WORKBOOK_CACHE_MAX:
        _, evicted = _DISPLAY_WORKBOOK_CACHE.popitem(last=False)
        try:
            evicted.close()
        except Exception:
            pass
    return workbook


def _worksheet_to_display_df(
        workbook,
        worksheet,
        include_style_metadata: bool = True,
        style_columns: int | None = None,
) -> pd.DataFrame:
    """Convert one already-open worksheet to the extractor's display grid.

    Internal blank columns remain in this structural grid because report
    parsers use their original spacing to separate adjacent financial blocks.
    Safe column pruning happens after the sheet profile is known.

    style_columns bounds where the per-cell INDENT lookup runs on wide sheets,
    where it is the dominant cost. None reads indent on every column (full
    fidelity). A positive K reads indent on the first K columns AND on any text
    cell beyond them — numeric value cells carry no indent, so they are skipped,
    while every label/header is styled wherever it sits. That text rule keeps
    multiple side-by-side tables working: each table's labels are text, so they
    are styled regardless of x-position. BOLD is read on every cell (with a
    per-row short-circuit) regardless of K, so bold section/total rows are never
    missed even when only a numeric cell is bold. Per-row outline levels are
    always read in full (O(rows), not O(cells)).
    """

    # Optimize: only read the actual used range, not the entire worksheet
    dims = ""
    try:
        dims = worksheet.calculate_dimension() or ""
    except Exception:
        dims = getattr(worksheet, "dimensions", "") or ""
    if dims:
        # Parse dimensions string like "A1:F100"
        if dims and ":" in dims:
            start_cell, end_cell = dims.split(":")
            # Extract row/col from end_cell
            match = RE_DIMENSIONS.match(end_cell)
            if match:
                max_col_letter, max_row = match.groups()
                max_row = int(max_row)
                # Convert column letter to number
                max_col = 0
                for char in max_col_letter:
                    max_col = max_col * 26 + (ord(char) - ord('A') + 1)
            else:
                max_row = worksheet.max_row
                max_col = worksheet.max_column
        else:
            max_row = worksheet.max_row
            max_col = worksheet.max_column
    else:
        max_row = worksheet.max_row
        max_col = worksheet.max_column

    row_merge_index: dict[int, list[tuple[int, int, object]]] = {}
    merged_range_metadata: list[dict] = []
    worksheet_merged = getattr(getattr(worksheet, "merged_cells", None), "ranges", None)
    if worksheet_merged and not getattr(workbook, "read_only", False):
        for merged_range in worksheet_merged:
            source_cell = worksheet.cell(merged_range.min_row, merged_range.min_col)
            merged_range_metadata.append(
                {
                    "range": str(merged_range),
                    "min_row": int(merged_range.min_row - 1),
                    "max_row": int(merged_range.max_row),
                    "min_col": int(merged_range.min_col - 1),
                    "max_col": int(merged_range.max_col),
                    "text": formatted_excel_value(source_cell),
                }
            )
            for row_idx in range(merged_range.min_row, merged_range.max_row + 1):
                row_merge_index.setdefault(row_idx, []).append(
                    (merged_range.min_col, merged_range.max_col, source_cell)
                )
        for row_idx in row_merge_index:
            row_merge_index[row_idx].sort(key=lambda item: item[0])

    resolved_cell_cache: dict[tuple[int, int], object] = {}

    def _resolve_actual_cell(cell):
        if not hasattr(cell, "row") or not hasattr(cell, "column"):
            return cell
        key = (cell.row, cell.column)
        if key in resolved_cell_cache:
            return resolved_cell_cache[key]

        intervals = row_merge_index.get(cell.row)
        if intervals:
            col_idx = cell.column
            for min_col, max_col, source_cell in intervals:
                if min_col <= col_idx <= max_col:
                    resolved_cell_cache[key] = source_cell
                    return source_cell
                if col_idx < min_col:
                    break

        resolved_cell_cache[key] = cell
        return cell

    rows = []
    indents = []
    row_outline_levels: list[int] = []
    row_bold_flags: list[bool] = []
    row_merged_flags: list[bool] = []
    row_indent_levels: list[int] = []
    # Horizontal border rules ("lines") mark header underlines, ruled section
    # titles, and total separators. A row counts as ruled when a long run of
    # cells carries that border edge — single boxed cells don't qualify. The
    # scan window is bounded like the other style reads; rules start at the
    # left of a table, so the first columns carry the signal.
    border_scan_cols = min(max_col, 40)
    rule_threshold = max(3, int(0.3 * border_scan_cols))
    row_top_rule_flags: list[bool] = []
    row_bottom_rule_flags: list[bool] = []
    # Bounded iteration avoids parsing full XML rows outside the used range.
    for row_idx, row in enumerate(
            worksheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col),
            start=1,
    ):
        row_data = []
        row_indents = []
        row_has_bold = False
        row_has_merged = False
        top_border_count = 0
        bottom_border_count = 0
        for col_idx, cell in enumerate(row):
            actual_cell = _resolve_actual_cell(cell)
            raw_value = getattr(actual_cell, "value", None)
            row_data.append(formatted_excel_value(actual_cell))
            if include_style_metadata and col_idx < border_scan_cols:
                # Borders come from the raw cell, not the merge-resolved one:
                # a vertically merged anchor would bleed its border into every
                # spanned row and fake a rule where none is drawn.
                border = getattr(cell, "border", None)
                if border is not None:
                    top_side = getattr(border, "top", None)
                    if top_side is not None and getattr(top_side, "style", None):
                        top_border_count += 1
                    bottom_side = getattr(border, "bottom", None)
                    if bottom_side is not None and getattr(bottom_side, "style", None):
                        bottom_border_count += 1
            if include_style_metadata:
                # Indent only lives on label/header (text) cells and the leading
                # floor; numeric value cells carry none, so skip the alignment
                # lookup for them on wide sheets. The text test also keeps
                # side-by-side tables working — each table's labels are text,
                # wherever they sit on the x-axis.
                if (
                    style_columns is None
                    or col_idx < style_columns
                    or isinstance(raw_value, str)
                ):
                    alignment = getattr(actual_cell, "alignment", None)
                    indent_value = (
                        float(getattr(alignment, "indent", 0) or 0)
                        if alignment else 0.0
                    )
                    # Exports from other systems often bake hierarchy into the
                    # label text as leading spaces (or NBSPs) instead of Excel
                    # indent — fold them into the same indent scale (~2 chars
                    # per level) so every downstream consumer sees them.
                    if isinstance(raw_value, str) and raw_value[:1] in (" ", "\u00a0"):
                        stripped = raw_value.replace("\u00a0", " ")
                        leading_spaces = len(stripped) - len(stripped.lstrip(" "))
                        if leading_spaces >= 2:
                            indent_value = max(indent_value, float(leading_spaces // 2))
                    row_indents.append(indent_value)
                else:
                    row_indents.append(0.0)
                # Bold marks section/total rows and can sit on a numeric cell, so
                # check every cell until the row is known bold, then short-circuit.
                if not row_has_bold:
                    font = getattr(actual_cell, "font", None)
                    if bool(getattr(font, "bold", False)):
                        row_has_bold = True
            else:
                row_indents.append(0.0)
            if actual_cell is not cell:
                row_has_merged = True
        rows.append(row_data)
        indents.append(row_indents)
        row_bold_flags.append(row_has_bold)
        row_merged_flags.append(row_has_merged)
        row_indent_levels.append(int(max(row_indents)) if row_indents else 0)
        row_top_rule_flags.append(top_border_count >= rule_threshold)
        row_bottom_rule_flags.append(bottom_border_count >= rule_threshold)

        outline_level = 0
        if include_style_metadata:
            try:
                outline_level = int(getattr(worksheet.row_dimensions.get(row_idx), "outlineLevel", 0) or 0)
            except Exception:
                outline_level = 0
        row_outline_levels.append(outline_level)

    df = pd.DataFrame(rows)
    df.attrs["excel_indents"] = indents
    df.attrs["excel_row_outline_levels"] = row_outline_levels
    df.attrs["excel_row_bold_flags"] = row_bold_flags
    df.attrs["excel_row_merged_flags"] = row_merged_flags
    df.attrs["excel_row_indent_levels"] = row_indent_levels
    df.attrs["excel_row_top_rule_flags"] = row_top_rule_flags
    df.attrs["excel_row_bottom_rule_flags"] = row_bottom_rule_flags
    df.attrs["excel_col_map"] = list(range(df.shape[1]))
    df.attrs["excel_merged_ranges"] = merged_range_metadata
    return _freeze_attrs(df)


# Leading-column floor for per-cell style reads on wide sheets. Styles are also
# read on any text cell beyond it (see _worksheet_to_display_df), so this floor
# only needs to cover bold-on-numeric cells in the primary left label block;
# side-by-side tables further right are covered by the text rule.
_STYLE_COLUMN_LIMIT = 24


def _style_read_plan(worksheet, sheet_name: str) -> tuple[bool, int | None]:
    """Decide the per-cell style budget for one worksheet.

    Returns (include_style_metadata, style_columns):
      - matrix sheets:      (False, None) — hierarchy styles are not used.
      - large report sheets:(True, _STYLE_COLUMN_LIMIT) — styles on the leading
        label block plus any text cell, so subtitle/hierarchy detection still
        works on big sheets (the previous all-or-nothing skip broke it) while
        skipping the numeric value cells that dominate a wide sheet's cost.
      - everything else:    (True, None) — full-fidelity styles.
    """
    if "matrix" in sheet_name.lower():
        return False, None
    cells = (worksheet.max_row or 0) * (worksheet.max_column or 0)
    if cells >= 10_000:
        return True, _STYLE_COLUMN_LIMIT
    return True, None


@st.cache_data(show_spinner=False)
def _read_display_sheet_cached(file_bytes: bytes, sheet_name: str) -> pd.DataFrame:
    """Read one sheet with full style metadata. Batch callers should use
    read_display_sheets(). Styles are always read here because subtitle and
    hierarchy detection depend on them; this path is the accuracy-preserving
    fallback for a single sheet."""
    workbook = _open_display_workbook(file_bytes)
    # The workbook is owned by the module-level LRU; do not close it here.
    return _worksheet_to_display_df(workbook, workbook[sheet_name])


def read_display_sheets(
        file_bytes: bytes,
        sheet_names: list[str] | tuple[str, ...],
) -> dict[str, pd.DataFrame]:
    """Read many worksheets with a single workbook parse."""
    requested = [str(name) for name in sheet_names]
    if not requested:
        return {}
    workbook = _open_display_workbook(file_bytes)
    # The workbook is owned by the module-level LRU; do not close it here.
    result: dict[str, pd.DataFrame] = {}
    for sheet_name in requested:
        if sheet_name not in workbook.sheetnames:
            continue
        worksheet = workbook[sheet_name]
        include_style, style_cols = _style_read_plan(worksheet, sheet_name)
        result[sheet_name] = _worksheet_to_display_df(
            workbook,
            worksheet,
            include_style_metadata=include_style,
            style_columns=style_cols,
        )
    return result


# Keywords that mark a row as a sub-item (child) of the nearest preceding root row.
_HIERARCHY_CHILD_PREFIXES: tuple[str, ...] = (
    "thereof",
    "of which",
    "whereof",
    "includes",
    "including",
    "excl.",
    "excluding",
    "hereof",
)


def _enrich_hierarchy_from_text_patterns(df: pd.DataFrame) -> pd.DataFrame:
    """Fill parent_line_item, line_item_path, and indent_level using text patterns.

    Handles melted/matrix data by resolving hierarchy from ordered unique labels per
    block and applying the result back to all rows.
    """
    if df is None or df.empty:
        return pd.DataFrame() if df is None else df

    label_candidates = [c for c in ("line_item", "metric", "metric_detail") if c in df.columns]
    if not label_candidates:
        return df

    working = df.copy()

    for col in ("parent_line_item", "line_item_path"):
        if col not in working.columns:
            working[col] = ""
    if "indent_level" not in working.columns:
        working["indent_level"] = 0

    def _is_child_keyword(text: str) -> bool:
        return text.strip().lower().startswith(_HIERARCHY_CHILD_PREFIXES)

    def _resolve_hierarchy_for_sequence(unique_items: list[str]) -> dict[str, dict]:
        result: dict[str, dict] = {}
        stack: list[tuple[int, str]] = []

        def _terminates_colon_group(text: str) -> bool:
            lowered = text.strip().lower()
            if lowered.endswith(":"):
                return True
            if lowered.startswith(("total", "sum ", "net income", "net result")):
                return True
            letters = [ch for ch in text if ch.isalpha()]
            return len(letters) >= 4 and all(ch.isupper() for ch in letters)

        colon_parent: str | None = None
        for item in unique_items:
            if not item:
                continue

            if colon_parent is not None and not _is_child_keyword(item):
                if _terminates_colon_group(item):
                    colon_parent = None
                else:
                    result[item] = {
                        "parent": colon_parent,
                        "path": f"{colon_parent} > {item}",
                        "depth": 1,
                    }
                    stack = [(0, colon_parent)]
                    continue

            if item.strip().endswith(":") and len(item.strip()) > 3:
                colon_parent = item
                result[item] = {"parent": "", "path": "", "depth": 0}
                stack = [(0, item)]
                continue

            if _is_child_keyword(item):
                if stack and stack[-1][0] >= 1:
                    depth = stack[-1][0]
                else:
                    depth = 1

                while stack and stack[-1][0] >= depth:
                    stack.pop()

                parent_text = stack[-1][1] if stack else ""
                ancestors = [entry[1] for entry in stack]
                path = " > ".join([*ancestors, item]) if ancestors else item
                result[item] = {"parent": parent_text, "path": path, "depth": depth}
                stack.append((depth, item))

            else:
                if item in result and result[item]["depth"] == 0:
                    while stack and stack[-1][0] > 0:
                        stack.pop()
                    if not stack or stack[-1][1] != item:
                        stack = [(0, item)]
                else:
                    stack = [(0, item)]
                result[item] = {"parent": "", "path": "", "depth": 0}

        return result

    group_col = "block_id" if "block_id" in working.columns else None

    def _process_group(group_df: pd.DataFrame):
        # Select the label column that best represents hierarchical line labels.
        label_col = label_candidates[0]
        best_score = (-1, -1)
        for cand in label_candidates:
            values = [cell_text(v) for v in group_df[cand].tolist()]
            non_blank = [v for v in values if v]
            child_hits = sum(1 for v in non_blank if _is_child_keyword(v))
            score = (child_hits, len(set(non_blank)))
            if score > best_score:
                best_score = score
                label_col = cand

        # Column-wise pass instead of per-cell .at reads/writes: pull every
        # needed column once, compute locally, write back in three batched
        # assignments. Pass-1 indent updates are carried through new_indent so
        # pass 2 sees them exactly like the original row-by-row flow did.
        idx_list = list(group_df.index)
        labels_text = _series_cell_text(group_df[label_col]).tolist()
        paths_text = _series_cell_text(group_df["line_item_path"]).tolist()
        parents_text = _series_cell_text(group_df["parent_line_item"]).tolist()
        raw_parents = group_df["parent_line_item"].tolist()
        raw_paths = group_df["line_item_path"].tolist()
        new_indent = group_df["indent_level"].tolist()
        new_parent = list(raw_parents)
        new_path = list(raw_paths)

        seen: dict[str, int] = {}
        for i, text in enumerate(labels_text):
            if text and text not in seen:
                seen[text] = i
        ordered_unique = sorted(seen.keys(), key=lambda k: seen[k])

        excel_hierarchy: dict[str, dict] = {}
        for pos, existing_path in enumerate(paths_text):
            if existing_path and ">" in existing_path:
                depth = existing_path.count(">")
                excel_hierarchy[labels_text[pos]] = {
                    "parent": parents_text[pos],
                    "path": existing_path,
                    "depth": depth,
                }
                new_indent[pos] = depth

        hierarchy = _resolve_hierarchy_for_sequence(ordered_unique)

        for pos, item in enumerate(labels_text):
            if not item:
                new_indent[pos] = 0
                continue

            if item in excel_hierarchy:
                info = excel_hierarchy[item]
            elif item in hierarchy:
                info = hierarchy[item]
            else:
                info = {"parent": "", "path": "", "depth": 0}

            if info["depth"] > 0:
                if is_blank(new_parent[pos]):
                    new_parent[pos] = info["parent"]
                if is_blank(new_path[pos]):
                    new_path[pos] = info["path"]

            current_level = int(new_indent[pos]) if not is_blank(new_indent[pos]) else 0
            new_indent[pos] = max(current_level, info["depth"])

        working.loc[idx_list, "parent_line_item"] = new_parent
        working.loc[idx_list, "line_item_path"] = new_path
        working.loc[idx_list, "indent_level"] = new_indent

    if group_col and group_col in working.columns:
        for _, grp in working.groupby(group_col, dropna=False, sort=False):
            _process_group(grp)
    else:
        _process_group(working)

    working["indent_level"] = pd.to_numeric(working["indent_level"], errors="coerce").fillna(0).astype(int)

    hier_cols = ["parent_line_item", "line_item_path", "indent_level"]
    present_hier = [c for c in hier_cols if c in working.columns]
    if present_hier:
        other_cols = [c for c in working.columns if c not in present_hier]
        anchor = next((c for c in ("line_item", "metric", "value") if c in other_cols), None)
        if anchor:
            pos = other_cols.index(anchor)
            new_order = other_cols[:pos] + present_hier + other_cols[pos:]
        else:
            new_order = other_cols + present_hier
        working = working[new_order]

    return working


def _enrich_hierarchy_from_totals(df: pd.DataFrame) -> pd.DataFrame:
    """Infer parent/child links arithmetically: a row whose value equals the
    sum of an adjacent run of rows is that run's parent (a total/subtotal).

    Each value series (one metric/unit/group combination, ordered by source
    row) is rebuilt from the long format and every row is tested as a
    total-below (children directly above) and a total-above (children directly
    below). A candidate link is accepted only when at least two independent
    series agree — or every series containing that label agrees — so a
    coincidental sum in one column cannot invent a hierarchy. Existing
    hierarchy (indentation, outline levels, keywords) always wins: only blank
    parent/path fields are filled.
    """
    if df is None or df.empty or "line_item" not in df.columns:
        return df
    value_source = "value_numeric" if "value_numeric" in df.columns else "value"
    if value_source not in df.columns:
        return df

    series_key_columns = [
        column for column in (
            "tab_name", "table_id", "block_id", "section", "column_group",
            "shared_header", "unit", "rate_type", "period", "valuation_date",
            "tenor", "metric",
        )
        if column in df.columns
    ]
    if not series_key_columns:
        return df

    values_all = pd.to_numeric(df[value_source], errors="coerce")
    labels_all = _series_cell_text(df["line_item"])
    positions = pd.Series(np.arange(len(df), dtype=float), index=df.index)
    if "source_row" in df.columns:
        source_rows = pd.to_numeric(df["source_row"], errors="coerce")
        order_key = source_rows.fillna(positions + 1e9)
    else:
        order_key = positions

    MAX_RUN = 25
    MAX_SERIES_ROWS = 400

    def _series_step(series_values: list) -> float:
        """Rounding granularity of a series (10^-max_decimals, capped)."""
        max_decimals = 0
        for item in series_values:
            if item is None:
                continue
            text = f"{item:.6f}".rstrip("0")
            decimals = len(text.split(".")[1]) if "." in text else 0
            max_decimals = max(max_decimals, min(decimals, 6))
            if max_decimals >= 6:
                break
        return 10.0 ** -max_decimals

    def _tolerance(total: float, run_length: int, step: float) -> float:
        # True totals hold to the file's own rounding granularity; anything
        # looser admits correlated-series coincidences (FX rates, AuM shares).
        return max(0.002 * abs(total), 0.6 * step * run_length)

    # candidate -> set of series ids supporting it
    support: dict[tuple, set[int]] = {}
    parent_series_seen: dict[str, set[int]] = {}

    grouping = df.groupby(
        [df[column].astype(str) for column in series_key_columns],
        sort=False,
        dropna=False,
    )
    for series_id, (_, series_index) in enumerate(grouping.indices.items()):
        series_idx = df.index[series_index]
        if len(series_idx) < 3 or len(series_idx) > MAX_SERIES_ROWS:
            continue
        ordered = sorted(series_idx, key=lambda idx: float(order_key.loc[idx]))
        vals = [float(values_all.loc[idx]) if pd.notna(values_all.loc[idx]) else None for idx in ordered]
        labs = [labels_all.loc[idx] for idx in ordered]
        n = len(ordered)
        step = _series_step(vals)

        for label, value in zip(labs, vals):
            if label and value is not None:
                parent_series_seen.setdefault(label, set()).add(series_id)

        for t in range(n):
            total = vals[t]
            parent_label = labs[t]
            if total is None or total == 0 or not parent_label:
                continue
            # totals-below: children are the contiguous run just above
            acc = 0.0
            for k in range(1, min(t, MAX_RUN) + 1):
                child_value = vals[t - k]
                if child_value is None:
                    break
                acc += child_value
                if k >= 2 and abs(acc - total) <= _tolerance(total, k, step):
                    children = tuple(labs[t - k:t])
                    if all(children) and parent_label not in children and len(set(children)) == len(children):
                        key = (parent_label, children)
                        support.setdefault(key, set()).add(series_id)
            # totals-above: children are the contiguous run just below
            acc = 0.0
            for k in range(1, min(n - 1 - t, MAX_RUN) + 1):
                child_value = vals[t + k]
                if child_value is None:
                    break
                acc += child_value
                if k >= 2 and abs(acc - total) <= _tolerance(total, k, step):
                    children = tuple(labs[t + 1:t + k + 1])
                    if all(children) and parent_label not in children and len(set(children)) == len(children):
                        key = (parent_label, children)
                        support.setdefault(key, set()).add(series_id)

    if not support:
        return df

    accepted: list[tuple[int, int, str, tuple]] = []
    for (parent_label, children), series_ids in support.items():
        # A "thereof"-style label is by definition a child; arithmetic must
        # never promote it to a parent (that inverts the relationship).
        if parent_label.strip().lower().startswith(_HIERARCHY_CHILD_PREFIXES):
            continue
        series_with_parent = len(parent_series_seen.get(parent_label, set()))
        required = (
            1
            if series_with_parent <= 1
            else max(2, -(-6 * series_with_parent // 10))  # ceil(0.6 * seen)
        )
        if len(series_ids) >= required:
            accepted.append((-len(series_ids), len(children), parent_label, children))
    if not accepted:
        return df

    # Shorter runs bind tighter (immediate parent beats a grand total that
    # happens to match a longer span); higher support wins first.
    accepted.sort()
    child_to_parent: dict[str, str] = {}
    for _, _, parent_label, children in accepted:
        for child in children:
            child_to_parent.setdefault(child, parent_label)
    # Never let two labels claim each other.
    child_to_parent = {
        child: parent
        for child, parent in child_to_parent.items()
        if child_to_parent.get(parent) != child
    }
    if not child_to_parent:
        return df

    working = df.copy()
    for column in ("parent_line_item", "line_item_path"):
        if column not in working.columns:
            working[column] = ""
    if "indent_level" not in working.columns:
        working["indent_level"] = 0

    existing_parent = _series_cell_text(working["parent_line_item"]).tolist()
    existing_path = _series_cell_text(working["line_item_path"]).tolist()
    indent_raw = working["indent_level"].tolist()
    labels_list = labels_all.tolist()

    new_parent = list(working["parent_line_item"].tolist())
    new_path = list(working["line_item_path"].tolist())
    new_indent = list(indent_raw)
    touched = False
    for position, label in enumerate(labels_list):
        inferred = child_to_parent.get(label, "")
        if not inferred:
            continue
        if not existing_parent[position]:
            new_parent[position] = inferred
            touched = True
        # A path holding just the label itself is a placeholder, not real
        # hierarchy — extractors pre-fill it that way for root rows.
        if not existing_path[position] or existing_path[position] == label:
            new_path[position] = f"{inferred} > {label}"
            touched = True
        current_level = 0
        if not is_blank(indent_raw[position]):
            try:
                current_level = int(indent_raw[position])
            except (TypeError, ValueError):
                current_level = 0
        if current_level < 1:
            new_indent[position] = 1
            touched = True
    if not touched:
        return df
    working["parent_line_item"] = new_parent
    working["line_item_path"] = new_path
    working["indent_level"] = new_indent
    return working


def _assign_schema_group_id(df: pd.DataFrame) -> pd.DataFrame:
    """Split mixed schemas within the same block into stable schema_group_id values."""
    if df is None or df.empty:
        return pd.DataFrame() if df is None else df

    working = df.copy()
    if "schema_group_id" not in working.columns:
        working["schema_group_id"] = ""
    if "schema_signature" not in working.columns:
        working["schema_signature"] = ""

    essential_presence_cols = [
        "column_group",
        "unit",
        "rate_type",
        "period",
        "valuation_date",
        "tenor",
        "contract_type",
        "metric_type",
        "metric_detail",
    ]
    valuation_cols = ["valuation_date", "tenor", "contract_type"]
    period_cols = ["period", "rate_type"]

    _AXIS_INHERITANCE: dict[str, list[str]] = {
        "period_like": [],
        "valuation_like": ["period_like"],
        "hybrid": ["period_like", "valuation_like"],
        "generic": [],
    }
    value_shape_cache: dict[str, str] = {}

    def _value_shape(value) -> str:
        text = cell_text(value)
        if not text:
            return "blank"
        if text in value_shape_cache:
            return value_shape_cache[text]

        if _parse_numeric_value(text) is not None:
            value_shape_cache[text] = "numeric"
            return "numeric"

        lowered = text.lower()

        if RE_DURATION_TOKEN.fullmatch(text.strip()):
            value_shape_cache[text] = "duration"
            return "duration"

        if RE_PERIOD_TOKEN.search(lowered):
            value_shape_cache[text] = "period_token"
            return "period_token"
        if RE_DATE_DMY.search(text):
            value_shape_cache[text] = "date"
            return "date"
        if RE_DATE_YEARISH.search(text):
            value_shape_cache[text] = "date"
            return "date"
        if (
            any(char.isdigit() for char in text)
            and (
                RE_MONTH_YEAR_TOKEN.search(lowered)
                or bool(re.search(r"\d[./-]\d", text))
            )
        ):
            try:
                parsed = pd.to_datetime(text, errors="coerce", dayfirst=True)
                if pd.notna(parsed):
                    value_shape_cache[text] = "date"
                    return "date"
            except Exception:
                pass

        if RE_UPPER_CODE.fullmatch(text):
            value_shape_cache[text] = "code"
            return "code"
        if "|" in text:
            value_shape_cache[text] = "composite"
            return "composite"
        if RE_ALPHA.search(text):
            value_shape_cache[text] = "text"
            return "text"
        value_shape_cache[text] = "other"
        return "other"

    def _block_row_signatures(block_df: pd.DataFrame) -> list[tuple]:
        """Build per-row structural signatures for a whole block column-wise.

        Avoids materializing one pandas Series per row; on wide extracts the
        per-row .loc lookups dominated schema-group assignment.
        """
        row_total = len(block_df)
        texts: dict[str, list[str]] = {}
        for col in essential_presence_cols:
            if col in block_df.columns:
                col_series = block_df[col]
                if isinstance(col_series, pd.DataFrame):
                    col_series = col_series.iloc[:, 0]
                texts[col] = _series_cell_text(col_series).tolist()
            else:
                texts[col] = [""] * row_total

        essential_texts = [texts[col] for col in essential_presence_cols]
        shape_columns = [
            [(_value_shape(text) if text else "blank") for text in texts[col]]
            for col in essential_presence_cols
        ]
        valuation_texts = [texts[col] for col in valuation_cols]
        period_texts = [texts[col] for col in period_cols]

        signatures: list[tuple] = []
        for i in range(row_total):
            presence = tuple(bool(column[i]) for column in essential_texts)
            shape_profile = tuple(column[i] for column in shape_columns)
            valuation_hits = sum(1 for column in valuation_texts if column[i])
            period_hits = sum(1 for column in period_texts if column[i])

            if valuation_hits >= 1 and period_hits == 0:
                axis_kind = "valuation_like"
            elif period_hits >= 1 and valuation_hits == 0:
                axis_kind = "period_like"
            elif valuation_hits >= 1 and period_hits >= 1:
                axis_kind = "hybrid"
            else:
                axis_kind = "generic"

            signatures.append((axis_kind, sum(presence), presence, shape_profile))
        return signatures

    def _context_anchor(row: pd.Series | None) -> dict:
        if row is None:
            return {}
        result: dict = {}
        for col in ["section", "table_name", "column_group", "unit", "rate_type", "contract_type"]:
            raw = cell_text(row.get(col, ""))
            if raw:
                result[col] = re.sub(r"\s+", " ", raw).strip()[:64]
        return result

    def _signature_label(signature: tuple, row: pd.Series | None = None) -> str:
        axis_kind = signature[0]
        presence_bits = signature[2]
        shape_profile = signature[3]

        bases = _AXIS_INHERITANCE.get(axis_kind, [])

        ctx = _context_anchor(row)
        section = ctx.get("section") or ctx.get("table_name") or ""
        qualifier = ctx.get("rate_type") or ctx.get("contract_type") or ctx.get("column_group") or ""
        unit = ctx.get("unit") or ""

        present_cols = [
            col for col, present in zip(essential_presence_cols, presence_bits)
            if bool(present)
        ]

        note_shapes = {"date", "period_token", "duration", "code"}
        shape_hints = {
            col: shape
            for col, present, shape in zip(essential_presence_cols, presence_bits, shape_profile)
            if bool(present) and shape in note_shapes
        }

        payload = {
            "schema_version": "v2",
            "axis": {
                "kind": axis_kind,
                "inherits": bases,
            },
            "context": {
                "section": section,
                "qualifier": qualifier,
                "unit": unit,
            },
            "fields_present": present_cols,
            "shape_hints": shape_hints,
        }
        return json.dumps(payload, ensure_ascii=True, sort_keys=True, separators=(",", ":"))

    if "block_id" in working.columns:
        block_groups = working.groupby("block_id", dropna=False, sort=False)
    else:
        block_groups = [("default", working)]

    for block_key, block_df in block_groups:
        if block_df.empty:
            continue

        base = "default" if pd.isna(block_key) else str(block_key)
        signatures = pd.Series(
            _block_row_signatures(block_df),
            index=block_df.index,
            dtype="object",
        )
        if signatures.empty:
            working.loc[block_df.index, "schema_group_id"] = base
            continue

        counts = signatures.value_counts()
        if len(counts) <= 1:
            working.loc[block_df.index, "schema_group_id"] = base
            signature = counts.index[0]
            sample_row = working.loc[block_df.index[0]] if len(block_df.index) > 0 else None
            working.loc[block_df.index, "schema_signature"] = _signature_label(signature, sample_row)
            continue

        # Trend logic: split only when alternate row-types have enough mass, otherwise smooth as noise.
        dominant_signature = counts.index[0]
        dominant_count = int(counts.iloc[0])
        total_count = len(signatures)
        dominant_ratio = dominant_count / max(total_count, 1)
        outlier_threshold = max(1, int(total_count * 0.10))
        significance_threshold = max(2, int(total_count * 0.15))

        signature_map: dict[tuple, tuple] = {}
        if dominant_ratio >= 0.70:
            dominant_axis = dominant_signature[0]
            for signature, count in counts.items():
                axis = signature[0]
                count_int = int(count)
                if count_int <= outlier_threshold:
                    signature_map[signature] = dominant_signature
                elif axis != dominant_axis and count_int < significance_threshold:
                    signature_map[signature] = dominant_signature
                else:
                    signature_map[signature] = signature
        else:
            for signature in counts.index.tolist():
                signature_map[signature] = signature

        normalized_signatures = pd.Series(
            [signature_map[sig] for sig in signatures.tolist()],
            index=signatures.index,
            dtype="object",
        )
        unique_signatures = list(dict.fromkeys(normalized_signatures.tolist()))

        if len(unique_signatures) <= 1:
            working.loc[block_df.index, "schema_group_id"] = base
            sample_row = working.loc[block_df.index[0]] if len(block_df.index) > 0 else None
            label = _signature_label(unique_signatures[0], sample_row) if unique_signatures else "[generic] unknown"
            working.loc[block_df.index, "schema_signature"] = label
            continue

        signature_to_suffix = {
            signature: idx + 1
            for idx, signature in enumerate(sorted(unique_signatures, key=lambda sig: str(sig)))
        }
        # Column-wise context + per-(signature, context) label cache instead of
        # materializing one pandas Series per row: _context_anchor only reads
        # six columns via row.get(), which a plain dict serves identically.
        anchor_columns = ["section", "table_name", "column_group", "unit", "rate_type", "contract_type"]
        block_context: dict[str, list[str]] = {}
        for col in anchor_columns:
            if col in working.columns:
                col_series = working.loc[block_df.index, col]
                if isinstance(col_series, pd.DataFrame):
                    col_series = col_series.iloc[:, 0]
                block_context[col] = _series_cell_text(col_series).tolist()

        group_ids: list[str] = []
        labels: list[str] = []
        label_cache: dict[tuple, str] = {}
        for position, signature in enumerate(normalized_signatures.tolist()):
            group_ids.append(f"{base}:{signature_to_suffix[signature]}")
            row_ctx = {
                col: values[position]
                for col, values in block_context.items()
                if values[position]
            }
            cache_key = (signature, tuple(sorted(row_ctx.items())))
            label = label_cache.get(cache_key)
            if label is None:
                label = _signature_label(signature, row_ctx)
                label_cache[cache_key] = label
            labels.append(label)
        working.loc[normalized_signatures.index, "schema_group_id"] = group_ids
        working.loc[normalized_signatures.index, "schema_signature"] = labels

    return working


def _reassign_block_id_from_schema_group(df: pd.DataFrame) -> pd.DataFrame:
    """Assign distinct numeric block_id values when one source block splits into multiple schema groups."""
    if df is None or df.empty:
        return pd.DataFrame() if df is None else df
    if "schema_group_id" not in df.columns:
        return df

    working = df.copy()
    # Use schema_signature as canonical grouping key without exposing helper columns.
    if "schema_signature" in working.columns:
        canonical_key = working["schema_signature"].astype(str)
    else:
        canonical_key = working["schema_group_id"].astype(str)

    logical_keys = sorted(canonical_key.unique().tolist())
    key_to_id = {key: idx + 1 for idx, key in enumerate(logical_keys)}
    working["block_id"] = canonical_key.map(key_to_id).astype(int)

    # Canonical schema groups are logical (cross-source), not extractor-order based.
    key_to_schema_group = {key: f"schema_{idx + 1}" for idx, key in enumerate(logical_keys)}
    working["schema_group_id"] = canonical_key.map(key_to_schema_group)

    # Remove intermediate helper columns if they already exist in older dataframes.
    drop_cols = [c for c in ("source_block_id", "source_schema_group_id", "logical_block_key") if c in working.columns]
    if drop_cols:
        working = working.drop(columns=drop_cols)
    return working


@st.cache_data(show_spinner=False)
def is_sheet_already_flat(file_bytes: bytes, sheet_name: str) -> bool:
    """
    Pre-flight flat-file detection — FAST check (no extraction).

    Returns True only if the preview positively matches a stable flat table.
    Returns False if extraction might be beneficial.

    This function only reads a preview, making it ~100x faster than full extraction.
    """
    try:
        # Quick preview read only
        quick_df = _quick_read_cached(file_bytes, sheet_name)

        return bool(infer_flat_table_structure(quick_df)["is_flat"])
    except Exception as e:
        # On error, default to False (do extraction to be safe)
        print(f"[is_sheet_already_flat] Error checking sheet '{sheet_name}': {e}")
        return False


INTERNAL_EXTRACTION_COLUMNS = {
    "source_range",
    "source_cell",
    "source_row",
    "source_column",
    "source_column_index",
    "source_region_order",
    "extraction_method",
    "schema_signature",
    "schema_group_id",
    "block_id",
    "source_block_id",
    "source_schema_group_id",
    "logical_block_key",
    "block_key",
    "block_start_column",
    "_agent_table_name",
}


def prepare_sql_agent_table(df: pd.DataFrame) -> pd.DataFrame:
    """Return the compact, typed table intended for users and SQL agents.

    Cell addresses and extractor bookkeeping remain available inside the
    extraction pipeline, but are deliberately excluded from public outputs.
    """
    if df is None or df.empty:
        return pd.DataFrame() if df is None else df.copy()

    public = df.drop(
        columns=[col for col in INTERNAL_EXTRACTION_COLUMNS if col in df.columns],
        errors="ignore",
    ).copy()

    looks_extracted = (
        "value" in public.columns
        and ("table_id" in public.columns or "table_name" in public.columns)
        and ("line_item" in public.columns or "metric" in public.columns)
    )
    if looks_extracted:
        generated_parts = [
            col
            for col in public.columns
            if re.search(r"_(?:part|dash_part)_\d+$", str(col))
        ]
        if generated_parts:
            public = public.drop(columns=generated_parts)

    if "value" in public.columns:
        parsed_values: list[float | None] = []
        value_kinds: list[str] = []
        for value in public["value"].tolist():
            if is_blank(value):
                parsed_values.append(None)
                value_kinds.append("blank")
                continue
            parsed = _parse_numeric_value(value)
            if parsed is not None:
                parsed_values.append(parsed)
                value_kinds.append("numeric")
                continue
            parsed_values.append(None)
            lowered = cell_text(value).lower()
            value_kinds.append(
                "not_meaningful"
                if lowered in {"n.m.", "n.m", "nm", "n/a", "na"}
                else "text"
            )

        insert_at = public.columns.get_loc("value") + 1
        if "value_numeric" not in public.columns:
            public.insert(
                insert_at,
                "value_numeric",
                pd.Series(parsed_values, index=public.index, dtype="Float64"),
            )
            insert_at += 1
        if "value_kind" not in public.columns:
            public.insert(insert_at, "value_kind", value_kinds)

    preferred_order = [
        "tab_name",
        "table_id",
        "table_name",
        "parent_report_name",
        "table_subtitle",
        "section",
        "column_group",
        "shared_header",
        "unit",
        "rate_type",
        "period",
        "valuation_date",
        "tenor",
        "contract_type",
        "currency",
        "business_line",
        "parent_line_item",
        "line_item_path",
        "indent_level",
        "line_item",
        "metric",
        "metric_detail",
        "metric_type",
        "metric_date",
        "comparison_date",
        "metric_quarter",
        "comparison_year",
        "value",
        "value_numeric",
        "value_kind",
    ]
    ordered = [col for col in preferred_order if col in public.columns]
    remaining = [col for col in public.columns if col not in ordered]
    public = public[ordered + remaining]
    public.attrs = {}
    return drop_all_blank_columns(public).reset_index(drop=True)


SQL_AGENT_SCHEMA_VERSION = "4.0"
SQL_AGENT_DIALECT = "sqlite"
SQL_AGENT_COLUMN_DESCRIPTIONS = {
    "workbook_id": "Stable identifier derived from the uploaded file contents.",
    "sheet_id": "Stable identifier for one worksheet within a workbook.",
    "logical_table_id": "Stable identifier for one detected business table within a worksheet.",
    "observation_id": "Stable identifier for one extracted Excel observation.",
    "file_name": "Original uploaded file name.",
    "sheet_name": "Original Excel worksheet or CSV table name.",
    "table_name": "Human-readable detected table title.",
    "parent_report_name": "Shared parent report or dashboard heading spanning multiple logical tables.",
    "table_subtitle": "Narrative subtitle or qualifier attached to the logical table.",
    "table_type": "Inferred business table category such as balance_sheet, income_statement, ratio_table, or generic.",
    "section": "Financial statement section or subsection.",
    "column_group": "Higher-level header such as company, segment, geography, or scenario.",
    "shared_header": "Header inherited from a shared band governing multiple child tables, commonly a date or period.",
    "unit": "Currency, percentage, or scale attached to the observation.",
    "rate_type": "Rate category such as spot, average, or closing.",
    "period": "Reporting period label from the workbook.",
    "valuation_date": "Date on which the observation was valued.",
    "tenor": "Contract, rate, or instrument maturity.",
    "contract_type": "Contract category.",
    "currency": "Currency dimension.",
    "business_line": "Business unit or operating segment.",
    "parent_line_item": "Immediate parent of the financial line item.",
    "line_item_path": "Full hierarchy path for the financial line item.",
    "indent_level": "Hierarchy depth of the line item.",
    "is_isolated_cell": (
        "1 when the value came from a stray column far outside its table's value "
        "block — usually an interim/helper computation. Exclude these rows from "
        "aggregates unless explicitly requested (filter is_isolated_cell = 0)."
    ),
    "line_item": "Row concept being measured, for example Revenue, Cash, or CET1 ratio.",
    "metric": "Column concept, period, scenario, or measure attached to the observation.",
    "metric_detail": "Additional qualifier parsed from the metric header.",
    "metric_type": "Semantic category inferred from the metric.",
    "metric_date": "Primary date parsed from the metric.",
    "comparison_date": "Comparison date parsed from a variance metric.",
    "metric_quarter": "Quarter associated with the metric.",
    "comparison_year": "Comparison year associated with the metric.",
    "value_text": "Original displayed Excel value represented as text.",
    "value_numeric": "Safely parsed numeric value. Use this column for arithmetic.",
    "value_kind": "Classification of the original value: numeric, text, blank, or not_meaningful.",
}

SQL_AGENT_COLUMN_SYNONYMS = {
    "file_name": ["file", "workbook", "document", "source file"],
    "sheet_name": ["sheet", "tab", "worksheet"],
    "table_name": ["table", "report", "statement", "schedule"],
    "parent_report_name": ["parent report", "dashboard", "report group", "umbrella heading"],
    "table_subtitle": ["subtitle", "table qualifier", "report qualifier"],
    "section": ["category", "subsection", "statement section"],
    "column_group": ["segment", "entity", "company", "scenario", "geography"],
    "shared_header": ["shared date header", "inherited header", "common period"],
    "unit": ["currency unit", "scale", "denomination"],
    "line_item": ["account", "row", "financial item", "balance sheet item"],
    "line_item_path": ["hierarchy", "account path", "financial hierarchy"],
    "metric": ["column", "period", "measure header", "scenario"],
    "metric_date": ["date", "as of date", "reporting date"],
    "period": ["year", "quarter", "reporting period"],
    "value_numeric": ["amount", "figure", "balance", "numeric value"],
    "value_text": ["display value", "raw value", "text value"],
}


def _stable_agent_id(prefix: str, *parts, length: int = 16) -> str:
    payload = json.dumps(parts, ensure_ascii=False, sort_keys=True, default=str, separators=(",", ":"))
    digest = hashlib.sha256(payload.encode("utf-8")).hexdigest()[:length]
    return f"{prefix}_{digest}"


def _safe_sql_name(value: str, fallback: str = "object") -> str:
    cleaned = clean_column_name(value or fallback)
    return cleaned[:48] or fallback


def _sql_text(value) -> str:
    if value is None or value is pd.NA:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    if isinstance(value, (date, datetime, pd.Timestamp)):
        return value.isoformat()
    return cell_text(value)


def _sqlite_value(value):
    if value is None or value is pd.NA:
        return None
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    if isinstance(value, np.generic):
        value = value.item()
    if isinstance(value, (date, datetime, pd.Timestamp)):
        return value.isoformat()
    if isinstance(value, bool):
        return int(value)
    return value


def _is_sensitive_agent_column(column: str) -> bool:
    normalized = clean_column_name(column)
    sensitive_tokens = {
        "email",
        "phone",
        "mobile",
        "address",
        "ssn",
        "social_security",
        "passport",
        "account_number",
        "iban",
        "swift",
        "customer_name",
        "employee_name",
    }
    return any(token in normalized for token in sensitive_tokens)


def _infer_agent_table_type(table_name: str, unit_values: list[str]) -> str:
    text = str(table_name or "").lower()
    if "balance sheet" in text or "financial position" in text:
        return "balance_sheet"
    if "income statement" in text or "profit and loss" in text or re.search(r"\bp&l\b", text):
        return "income_statement"
    if "cash flow" in text:
        return "cash_flow"
    if "rating" in text:
        return "categorical_table"
    if any(token in text for token in ("ratio", "adequacy", "margin", "yield", "rate")):
        return "ratio_table"
    if any("%" in unit.lower() or "percent" in unit.lower() for unit in unit_values):
        return "ratio_table"
    if any(token in text for token in ("risk", "exposure", "liquidity", "capital")):
        return "risk_schedule"
    return "generic_financial_table"


def _agent_aggregation_policy(table_type: str, unit_values: list[str]) -> dict:
    time_dimensions = ["period", "metric_date", "valuation_date", "metric_quarter", "comparison_year"]
    if table_type == "ratio_table":
        return {
            "default_aggregation": "AVG",
            "additivity": "non_additive",
            "non_additive_dimensions": time_dimensions,
            "requires_unit_filter": True,
            "reason": "Ratios, rates, and percentages should generally be averaged or inspected directly, not summed.",
        }
    if table_type in {"income_statement", "cash_flow"}:
        return {
            "default_aggregation": "SUM",
            "additivity": "semi_additive",
            "non_additive_dimensions": time_dimensions,
            "requires_unit_filter": True,
            "reason": "Flow values can often be summed across compatible entities but should not be summed across overlapping periods.",
        }
    if table_type == "balance_sheet":
        return {
            "default_aggregation": "NONE",
            "additivity": "semi_additive",
            "non_additive_dimensions": time_dimensions,
            "requires_unit_filter": True,
            "reason": "Point-in-time balances and subtotals can double-count when summed; aggregate only with explicit scope.",
        }
    return {
        "default_aggregation": "NONE",
        "additivity": "unknown",
        "non_additive_dimensions": time_dimensions,
        "requires_unit_filter": bool(unit_values),
        "reason": "The source is unstructured, so aggregation requires an explicit business question and compatible units.",
    }



def _schema_column_sql_type(column: str, series: pd.Series) -> str:
    if column in {"indent_level", "comparison_year"}:
        return "INTEGER"
    if pd.api.types.is_bool_dtype(series.dtype):
        return "BOOLEAN"
    if pd.api.types.is_integer_dtype(series.dtype):
        return "INTEGER"
    if pd.api.types.is_numeric_dtype(series.dtype):
        return "REAL"
    if pd.api.types.is_datetime64_any_dtype(series.dtype):
        return "TIMESTAMP"
    return "TEXT"


def _schema_column_role(column: str, series: pd.Series) -> str:
    if column == "value_numeric" or pd.api.types.is_numeric_dtype(series.dtype):
        return "measure"
    if column in {"period", "valuation_date", "metric_date", "comparison_date", "metric_quarter", "comparison_year"}:
        return "time_dimension"
    if column in {"parent_line_item", "line_item_path", "indent_level", "line_item"}:
        return "hierarchy"
    return "dimension"


def build_excel_schema_package(
        sources: list[dict],
        max_sample_values: int = 10,
) -> dict:
    """Profile physical flat tables without materializing observation records."""
    workbook_rows: dict[str, dict] = {}
    sheet_rows: list[dict] = []
    logical_rows: list[dict] = []
    column_rows: list[dict] = []
    metric_rows: list[dict] = []

    for source_index, source in enumerate(sources):
        frame = source.get("frame")
        if frame is None or frame.empty:
            continue
        frame = pd.DataFrame(frame).copy()
        file_name = str(source.get("file_name") or "uploaded_file.xlsx")
        file_bytes = source.get("file_bytes") or b""
        content_hash = (
            hashlib.sha256(file_bytes).hexdigest()
            if file_bytes
            else hashlib.sha256(file_name.encode("utf-8")).hexdigest()
        )
        workbook_id = _stable_agent_id("wb", content_hash)
        workbook_rows[workbook_id] = {
            "workbook_id": workbook_id,
            "file_name": file_name,
            "file_type": Path(file_name).suffix.lower().lstrip(".") or "table",
            "content_sha256": content_hash,
            "byte_size": len(file_bytes),
        }

        sheet_name = str(source.get("sheet_name") or f"table_{source_index + 1}")
        physical_table = str(
            source.get("physical_table_name")
            or clean_column_name(sheet_name)
        )
        sheet_id = _stable_agent_id("sh", workbook_id, source_index, sheet_name)
        logical_table_id = _stable_agent_id(
            "lt",
            sheet_id,
            physical_table,
        )
        numeric_columns = [
            str(column)
            for column in frame.columns
            if pd.api.types.is_numeric_dtype(frame[column])
            or str(column) == "value_numeric"
        ]
        units = (
            list(dict.fromkeys(
                cell_text(value)
                for value in frame["unit"].tolist()
                if cell_text(value)
            ))[:20]
            if "unit" in frame.columns
            else []
        )
        table_type = _infer_agent_table_type(sheet_name, units)
        policy = _agent_aggregation_policy(table_type, units)
        logical_name = (
            f"{_safe_sql_name(Path(file_name).stem)}__"
            f"{_safe_sql_name(physical_table)}"
        )
        search_terms = list(dict.fromkeys(
            [
                file_name,
                sheet_name,
                physical_table,
                *[str(column) for column in frame.columns],
            ]
        ))

        sheet_rows.append(
            {
                "sheet_id": sheet_id,
                "workbook_id": workbook_id,
                "sheet_name": sheet_name,
                "sheet_index": int(source.get("sheet_index") or source_index),
                "parsing_mode": str(source.get("parsing_mode") or "flattened"),
                "source_row_count": int((source.get("source_shape") or frame.shape)[0]),
                "source_column_count": int((source.get("source_shape") or frame.shape)[1]),
                "logical_table_count": 1,
                "observation_count": int(len(frame)),
            }
        )
        logical_rows.append(
            {
                "logical_table_id": logical_table_id,
                "sheet_id": sheet_id,
                "parent_report_id": None,
                "logical_name": logical_name,
                "display_name": sheet_name,
                "table_type": table_type,
                "description": (
                    f'Physical SQLite table "{physical_table}" loaded from '
                    f"'{file_name}' / '{sheet_name}'."
                ),
                "grain": "One row in the flattened physical table.",
                "physical_table": physical_table,
                "filter_sql": "",
                "row_count": int(len(frame)),
                "primary_measure": (
                    "value_numeric"
                    if "value_numeric" in frame.columns
                    else (numeric_columns[0] if numeric_columns else None)
                ),
                "default_aggregation": policy["default_aggregation"],
                "additivity": policy["additivity"],
                "non_additive_dimensions_json": json.dumps(
                    policy["non_additive_dimensions"]
                ),
                "requires_unit_filter": int(policy["requires_unit_filter"]),
                "known_units_json": json.dumps(units, ensure_ascii=False),
                "known_sections_json": "[]",
                "search_terms_json": json.dumps(
                    search_terms[:80],
                    ensure_ascii=False,
                ),
            }
        )

        for column in frame.columns:
            column_name = str(column)
            series = frame[column]
            non_blank = [
                value for value in series.tolist()
                if not is_blank(value)
            ]
            is_sensitive_column = _is_sensitive_agent_column(column_name)
            samples = (
                []
                if is_sensitive_column
                else list(dict.fromkeys(
                    cell_text(value) for value in non_blank
                    if cell_text(value)
                ))[:max_sample_values]
            )
            distinct_count = int(series.nunique(dropna=True))
            role = _schema_column_role(column_name, series)
            is_numeric = pd.api.types.is_numeric_dtype(series)
            allowed_aggregations = (
                ["SUM", "AVG", "MIN", "MAX", "COUNT"]
                if is_numeric
                else ["COUNT", "COUNT_DISTINCT"]
            )
            column_rows.append(
                {
                    "column_id": _stable_agent_id(
                        "col",
                        logical_table_id,
                        column_name,
                    ),
                    "logical_table_id": logical_table_id,
                    "column_name": column_name,
                    "sql_type": _schema_column_sql_type(column_name, series),
                    "semantic_role": role,
                    "description": SQL_AGENT_COLUMN_DESCRIPTIONS.get(
                        column_name,
                        f"Column from the flattened source table: {column_name}.",
                    ),
                    "synonyms_json": json.dumps(
                        SQL_AGENT_COLUMN_SYNONYMS.get(column_name, []),
                        ensure_ascii=False,
                    ),
                    "sample_values_json": json.dumps(
                        samples,
                        ensure_ascii=False,
                    ),
                    "sample_values_redacted": int(is_sensitive_column),
                    "nullable": int(series.isna().any()),
                    "is_unique": int(
                        len(series) > 0
                        and distinct_count == len(series)
                    ),
                    "is_enum": int(0 < distinct_count <= 30),
                    "is_filterable": 1,
                    "is_hidden": 0,
                    "fill_rate": round(
                        len(non_blank) / max(len(series), 1),
                        4,
                    ),
                    "distinct_count": distinct_count,
                    "allowed_aggregations_json": json.dumps(
                        allowed_aggregations
                    ),
                    "non_additive_dimensions_json": json.dumps(
                        policy["non_additive_dimensions"]
                        if is_numeric
                        else []
                    ),
                }
            )

        for column_name in numeric_columns:
            metric_rows.append(
                {
                    "metric_id": _stable_agent_id(
                        "metric",
                        logical_table_id,
                        column_name,
                    ),
                    "logical_table_id": logical_table_id,
                    "metric_name": column_name,
                    "description": f'Numeric column "{column_name}".',
                    # Escape embedded quotes so a hostile column header cannot
                    # break out of the quoted identifier in downstream SQL.
                    "expression_sql": '"{}"'.format(column_name.replace('"', '""')),
                    "aggregation": policy["default_aggregation"],
                    "unit_column": "unit" if "unit" in frame.columns else None,
                    "non_additive_dimensions_json": json.dumps(
                        policy["non_additive_dimensions"]
                    ),
                    "requires_unit_filter": int(policy["requires_unit_filter"]),
                }
            )

    relationship_rows = [
        {
            "relationship_id": _stable_agent_id(
                "rel",
                "logical_tables_to_sheets",
            ),
            "relationship_name": "logical_tables_to_sheets",
            "from_table": "agent_logical_tables",
            "from_column": "sheet_id",
            "to_table": "agent_sheets",
            "to_column": "sheet_id",
            "cardinality": "many_to_one",
            "join_type": "inner",
        }
    ]
    frames = {
        "agent_workbooks": pd.DataFrame(workbook_rows.values()),
        "agent_sheets": pd.DataFrame(sheet_rows),
        "agent_parent_reports": pd.DataFrame(),
        "agent_logical_tables": pd.DataFrame(logical_rows),
        "excel_observations": pd.DataFrame(),
        "observation_provenance": pd.DataFrame(),
        "agent_columns": pd.DataFrame(column_rows),
        "agent_metrics": pd.DataFrame(metric_rows),
        "agent_relationships": pd.DataFrame(relationship_rows),
        "agent_verified_queries": pd.DataFrame(),
    }
    catalog = {
        "schema_version": SQL_AGENT_SCHEMA_VERSION,
        "dialect": SQL_AGENT_DIALECT,
        "purpose": "Natural-language SQL analysis of physical flattened tables.",
        "logical_tables": [
            {
                **row,
                "known_units": json.loads(row["known_units_json"]),
                "known_sections": [],
                "search_terms": json.loads(row["search_terms_json"]),
                "non_additive_dimensions": json.loads(
                    row["non_additive_dimensions_json"]
                ),
            }
            for row in logical_rows
        ],
        "parent_reports": [],
        "relationships": relationship_rows,
        "metrics": metric_rows,
        "verified_queries": [],
        "statistics": {
            "workbook_count": len(workbook_rows),
            "sheet_count": len(sheet_rows),
            "parent_report_count": 0,
            "logical_table_count": len(logical_rows),
            "observation_count": sum(
                int(row["observation_count"]) for row in sheet_rows
            ),
            "verified_query_count": 0,
        },
    }
    return {
        "catalog": catalog,
        "ddl_sql": "",
        "frames": frames,
    }


def build_embedded_schema_frame(package: dict) -> pd.DataFrame:
    """Flatten SQL-agent metadata into one Excel-friendly Schema worksheet."""
    frames = package.get("frames", {}) if package else {}
    workbooks = frames.get("agent_workbooks", pd.DataFrame()).copy()
    sheets = frames.get("agent_sheets", pd.DataFrame()).copy()
    parent_reports = frames.get("agent_parent_reports", pd.DataFrame()).copy()
    logical_tables = frames.get("agent_logical_tables", pd.DataFrame()).copy()
    columns = frames.get("agent_columns", pd.DataFrame()).copy()
    metrics = frames.get("agent_metrics", pd.DataFrame()).copy()
    relationships = frames.get("agent_relationships", pd.DataFrame()).copy()
    verified_queries = frames.get("agent_verified_queries", pd.DataFrame()).copy()

    workbook_lookup = (
        workbooks.set_index("workbook_id").to_dict("index")
        if not workbooks.empty and "workbook_id" in workbooks.columns
        else {}
    )
    sheet_lookup = (
        sheets.set_index("sheet_id").to_dict("index")
        if not sheets.empty and "sheet_id" in sheets.columns
        else {}
    )
    logical_lookup = (
        logical_tables.set_index("logical_table_id").to_dict("index")
        if not logical_tables.empty and "logical_table_id" in logical_tables.columns
        else {}
    )
    parent_report_lookup = (
        parent_reports.set_index("parent_report_id").to_dict("index")
        if not parent_reports.empty and "parent_report_id" in parent_reports.columns
        else {}
    )

    schema_columns = [
        "record_type",
        "workbook_id",
        "file_name",
        "sheet_id",
        "sheet_name",
        "parent_report_id",
        "parent_report_name",
        "logical_table_id",
        "logical_name",
        "display_name",
        "table_type",
        "object_name",
        "sql_type",
        "semantic_role",
        "description",
        "synonyms",
        "sample_values",
        "row_count",
        "grain",
        "physical_table",
        "filter_sql",
        "primary_measure",
        "metric_expression",
        "aggregation",
        "additivity",
        "unit_column",
        "requires_unit_filter",
        "non_additive_dimensions",
        "known_units",
        "known_sections",
        "search_terms",
        "nullable",
        "is_unique",
        "is_enum",
        "is_filterable",
        "is_hidden",
        "fill_rate",
        "distinct_count",
        "allowed_aggregations",
        "relationship_from",
        "relationship_to",
        "cardinality",
        "join_type",
        "question",
        "sql_query",
        "validation_status",
    ]

    def _base_row(
            workbook_id: str = "",
            sheet_id: str = "",
            logical_table_id: str = "",
    ) -> dict:
        logical = logical_lookup.get(logical_table_id, {})
        if not sheet_id:
            sheet_id = _sql_text(logical.get("sheet_id", ""))
        sheet = sheet_lookup.get(sheet_id, {})
        if not workbook_id:
            workbook_id = _sql_text(sheet.get("workbook_id", ""))
        workbook = workbook_lookup.get(workbook_id, {})
        parent_report_id = _sql_text(logical.get("parent_report_id", ""))
        parent_report = parent_report_lookup.get(parent_report_id, {})
        return {
            "workbook_id": workbook_id,
            "file_name": _sql_text(workbook.get("file_name", "")),
            "sheet_id": sheet_id,
            "sheet_name": _sql_text(sheet.get("sheet_name", "")),
            "parent_report_id": parent_report_id,
            "parent_report_name": _sql_text(parent_report.get("report_name", "")),
            "logical_table_id": logical_table_id,
            "logical_name": _sql_text(logical.get("logical_name", "")),
            "display_name": _sql_text(logical.get("display_name", "")),
            "table_type": _sql_text(logical.get("table_type", "")),
        }

    rows: list[dict] = []
    for _, record in workbooks.iterrows():
        row = _base_row(workbook_id=_sql_text(record.get("workbook_id", "")))
        row.update(
            {
                "record_type": "workbook",
                "object_name": row["file_name"],
                "description": (
                    f"Uploaded {record.get('file_type', 'file')} containing "
                    f"{int(record.get('byte_size', 0) or 0):,} bytes."
                ),
            }
        )
        rows.append(row)

    for _, record in sheets.iterrows():
        row = _base_row(
            workbook_id=_sql_text(record.get("workbook_id", "")),
            sheet_id=_sql_text(record.get("sheet_id", "")),
        )
        row.update(
            {
                "record_type": "sheet",
                "object_name": row["sheet_name"],
                "description": (
                    f"Worksheet parsed with mode '{_sql_text(record.get('parsing_mode', ''))}'."
                ),
                "row_count": _sqlite_value(record.get("observation_count", None)),
            }
        )
        rows.append(row)

    for _, record in logical_tables.iterrows():
        logical_table_id = _sql_text(record.get("logical_table_id", ""))
        row = _base_row(
            sheet_id=_sql_text(record.get("sheet_id", "")),
            logical_table_id=logical_table_id,
        )
        row.update(
            {
                "record_type": "logical_table",
                "object_name": row["logical_name"],
                "description": _sql_text(record.get("description", "")),
                "row_count": _sqlite_value(record.get("row_count", None)),
                "grain": _sql_text(record.get("grain", "")),
                "physical_table": _sql_text(record.get("physical_table", "")),
                "filter_sql": _sql_text(record.get("filter_sql", "")),
                "primary_measure": _sql_text(record.get("primary_measure", "")),
                "aggregation": _sql_text(record.get("default_aggregation", "")),
                "additivity": _sql_text(record.get("additivity", "")),
                "requires_unit_filter": _sqlite_value(record.get("requires_unit_filter", None)),
                "non_additive_dimensions": _sql_text(
                    record.get("non_additive_dimensions_json", "")
                ),
                "known_units": _sql_text(record.get("known_units_json", "")),
                "known_sections": _sql_text(record.get("known_sections_json", "")),
                "search_terms": _sql_text(record.get("search_terms_json", "")),
            }
        )
        rows.append(row)

    for _, record in parent_reports.iterrows():
        parent_report_id = _sql_text(record.get("parent_report_id", ""))
        sheet_id = _sql_text(record.get("sheet_id", ""))
        sheet = sheet_lookup.get(sheet_id, {})
        workbook_id = _sql_text(sheet.get("workbook_id", ""))
        row = _base_row(workbook_id=workbook_id, sheet_id=sheet_id)
        row.update(
            {
                "record_type": "parent_report",
                "parent_report_id": parent_report_id,
                "parent_report_name": _sql_text(record.get("report_name", "")),
                "object_name": _sql_text(record.get("report_name", "")),
                "description": _sql_text(record.get("description", "")),
                "row_count": _sqlite_value(record.get("child_table_count", None)),
                "search_terms": _sql_text(record.get("search_terms_json", "")),
            }
        )
        rows.append(row)

    for _, record in columns.iterrows():
        logical_table_id = _sql_text(record.get("logical_table_id", ""))
        row = _base_row(logical_table_id=logical_table_id)
        row.update(
            {
                "record_type": "column",
                "object_name": _sql_text(record.get("column_name", "")),
                "sql_type": _sql_text(record.get("sql_type", "")),
                "semantic_role": _sql_text(record.get("semantic_role", "")),
                "description": _sql_text(record.get("description", "")),
                "synonyms": _sql_text(record.get("synonyms_json", "")),
                "sample_values": _sql_text(record.get("sample_values_json", "")),
                "nullable": _sqlite_value(record.get("nullable", None)),
                "is_unique": _sqlite_value(record.get("is_unique", None)),
                "is_enum": _sqlite_value(record.get("is_enum", None)),
                "is_filterable": _sqlite_value(record.get("is_filterable", None)),
                "is_hidden": _sqlite_value(record.get("is_hidden", None)),
                "fill_rate": _sqlite_value(record.get("fill_rate", None)),
                "distinct_count": _sqlite_value(record.get("distinct_count", None)),
                "allowed_aggregations": _sql_text(
                    record.get("allowed_aggregations_json", "")
                ),
                "non_additive_dimensions": _sql_text(
                    record.get("non_additive_dimensions_json", "")
                ),
            }
        )
        rows.append(row)

    for _, record in metrics.iterrows():
        logical_table_id = _sql_text(record.get("logical_table_id", ""))
        row = _base_row(logical_table_id=logical_table_id)
        row.update(
            {
                "record_type": "metric",
                "object_name": _sql_text(record.get("metric_name", "")),
                "description": _sql_text(record.get("description", "")),
                "metric_expression": _sql_text(record.get("expression_sql", "")),
                "aggregation": _sql_text(record.get("aggregation", "")),
                "unit_column": _sql_text(record.get("unit_column", "")),
                "requires_unit_filter": _sqlite_value(
                    record.get("requires_unit_filter", None)
                ),
                "non_additive_dimensions": _sql_text(
                    record.get("non_additive_dimensions_json", "")
                ),
            }
        )
        rows.append(row)

    for _, record in relationships.iterrows():
        row = {
            "record_type": "relationship",
            "object_name": _sql_text(record.get("relationship_name", "")),
            "description": "Join relationship between SQL-agent metadata tables.",
            "relationship_from": (
                f"{_sql_text(record.get('from_table', ''))}."
                f"{_sql_text(record.get('from_column', ''))}"
            ),
            "relationship_to": (
                f"{_sql_text(record.get('to_table', ''))}."
                f"{_sql_text(record.get('to_column', ''))}"
            ),
            "cardinality": _sql_text(record.get("cardinality", "")),
            "join_type": _sql_text(record.get("join_type", "")),
        }
        rows.append(row)

    for _, record in verified_queries.iterrows():
        logical_table_id = _sql_text(record.get("logical_table_id", ""))
        row = _base_row(logical_table_id=logical_table_id)
        row.update(
            {
                "record_type": "verified_query",
                "object_name": _sql_text(record.get("verified_query_id", "")),
                "description": _sql_text(record.get("description", "")),
                "question": _sql_text(record.get("question", "")),
                "sql_query": _sql_text(record.get("sql_query", "")),
                "validation_status": _sql_text(record.get("validation_status", "")),
            }
        )
        rows.append(row)

    schema_df = pd.DataFrame(rows)
    for column in schema_columns:
        if column not in schema_df.columns:
            schema_df[column] = ""
    schema_df = schema_df[schema_columns]
    if not schema_df.empty:
        schema_df = schema_df.sort_values(
            ["file_name", "sheet_name", "logical_name", "record_type", "object_name"],
            kind="stable",
        ).reset_index(drop=True)
    return schema_df



def finalize_extracted_sheet(df: pd.DataFrame, strip_text: bool = True, split_hierarchy: bool = True) -> pd.DataFrame:
    """Apply final post-processing transformations to an extracted/flattened sheet.

    Includes:
    - Trimming whitespace in text columns
    - Optionally deriving hierarchy helper columns while preserving originals
    - Cleaning up blank and extractor-only columns
    """
    if df is None or df.empty:
        return df

    cleaned = df.copy()
    # Worksheet scan metadata contains large NumPy masks that are invalid once
    # extraction changes the dataframe shape and very expensive for pandas to copy.
    cleaned.attrs = {}

    # Trim whitespace in text columns
    if strip_text:
        for col in cleaned.columns:
            if not (pd.api.types.is_object_dtype(cleaned[col].dtype) or pd.api.types.is_string_dtype(
                    cleaned[col].dtype)):
                continue
            cleaned[col] = cleaned[col].astype(str).str.strip()

    # Apply hierarchy column splitting (| and dash delimiters)
    if split_hierarchy:
        cleaned = _auto_split_hierarchy_columns(cleaned, dash_split_mode="spaced")

    has_schema_signature = (
        "schema_signature" in cleaned.columns
        and cleaned["schema_signature"].astype("string").fillna("").str.strip().ne("").any()
    )
    if not has_schema_signature:
        cleaned = _assign_schema_group_id(cleaned)
        cleaned = _reassign_block_id_from_schema_group(cleaned)

    cleaned = _enrich_hierarchy_from_text_patterns(cleaned)
    cleaned = _enrich_hierarchy_from_totals(cleaned)

    # Final cleanup: remove only entirely blank columns. Identical populated
    # columns remain because they can still be semantically distinct.
    cleaned = drop_all_blank_columns(cleaned).reset_index(drop=True)

    hierarchy_defaults: list[tuple[str, object]] = [
        ("parent_line_item", ""),
        ("line_item_path", ""),
        ("indent_level", 0),
    ]
    missing_hier = [(c, d) for c, d in hierarchy_defaults if c not in cleaned.columns]
    if missing_hier:
        for hcol, hdefault in missing_hier:
            cleaned[hcol] = hdefault
        hier_present = [c for c, _ in hierarchy_defaults if c in cleaned.columns]
        other_cols = [c for c in cleaned.columns if c not in hier_present]
        anchor = next((c for c in ("line_item", "metric", "value") if c in other_cols), None)
        if anchor:
            pos = other_cols.index(anchor)
            cleaned = cleaned[other_cols[:pos] + hier_present + other_cols[pos:]]

    cleaned = cleaned.drop(
        columns=[
            col for col in (
                "block_id",
                "schema_group_id",
                "source_block_id",
                "source_schema_group_id",
                "logical_block_key",
                "block_key",
                "block_start_column",
            )
            if col in cleaned.columns
        ],
        errors="ignore",
    )

    return prepare_sql_agent_table(cleaned)


def extract_and_flatten_sheet(file_bytes: bytes, sheet_name: str, preferred_profile: str = "auto") -> pd.DataFrame:
    """Extract and flatten complex Excel sheets, or return quickly for already-flat sheets.

    For flat sheets: Returns immediately with just a pd.read_excel() call (< 1 second).
    For complex sheets: Runs the full extraction pipeline with UI feedback.
    """
    quick_df = _quick_read_cached(file_bytes, sheet_name)
    hinted_profile = (preferred_profile or "auto").strip().lower()
    if hinted_profile not in {"auto", "general", "matrix"}:
        hinted_profile = "auto"

    report_layout = infer_report_layout_from_quick_df(quick_df)
    flat_detection = infer_flat_table_structure(quick_df, report_layout=report_layout)
    if hinted_profile == "auto" and flat_detection["is_flat"]:
        result = _read_flat_excel_sheet(
            file_bytes,
            sheet_name,
            header_row=flat_detection["header_row"],
        )
        if result is not None and not result.empty and "tab_name" not in result.columns:
            result.insert(0, "tab_name", sheet_name)
        return result

    auto_profile = report_layout[0]
    if "matrix" in sheet_name.lower():
        auto_profile = "matrix"
    extraction_profile = hinted_profile if hinted_profile != "auto" else auto_profile

    if extraction_profile == "matrix":
        raw_df = drop_empty_or_zero_columns(
            _read_display_sheet_cached(file_bytes, sheet_name)
        )
        result = _finalize_matrix_extract(
            _fallback_matrix_flatten(raw_df),
        )
        if result is not None and not result.empty and "tab_name" not in result.columns:
            result.insert(0, "tab_name", sheet_name)
        return result

    extracted_df = _extract_report_cached(file_bytes, sheet_name, extraction_profile=extraction_profile)
    if extracted_df is None or extracted_df.empty:
        fallback_df = pd.read_excel(_shared_excel_file(file_bytes), sheet_name=sheet_name)
        result = finalize_extracted_sheet(pd.DataFrame(fallback_df), strip_text=True, split_hierarchy=False)
    else:
        result = finalize_extracted_sheet(pd.DataFrame(extracted_df), strip_text=True, split_hierarchy=False)
    if result is not None and not result.empty and "tab_name" not in result.columns:
        result.insert(0, "tab_name", sheet_name)
    return result


def _flat_frame_from_raw(raw_df: pd.DataFrame, header_row: int) -> pd.DataFrame:
    """Build a conventional flat table directly from an already-read grid."""
    if raw_df is None or raw_df.empty or header_row >= len(raw_df):
        return pd.DataFrame()
    headers = [
        cell_text(value) or f"column_{idx + 1}"
        for idx, value in enumerate(raw_df.iloc[header_row].tolist())
    ]
    frame = raw_df.iloc[header_row + 1:].copy()
    frame.attrs = {}
    frame.columns = dedupe_columns(headers)
    return (
        drop_empty_or_zero_columns(frame)
        .dropna(axis=0, how="all")
        .dropna(axis=1, how="all")
        .reset_index(drop=True)
    )


def _finalize_matrix_extract(extracted: pd.DataFrame) -> pd.DataFrame:
    """Keep matrix finalization linear and avoid report-only schema passes."""
    if extracted is None or extracted.empty:
        return pd.DataFrame()
    result = pd.DataFrame(extracted).copy()
    result.attrs = {}
    if "table_id" not in result.columns:
        result.insert(0, "table_id", "table_1")
    for column, default in (
        ("parent_line_item", ""),
        ("line_item_path", ""),
        ("indent_level", 0),
    ):
        if column not in result.columns:
            result[column] = default
    return prepare_sql_agent_table(result)


def _extract_from_raw_sheet(
        raw_df: pd.DataFrame,
        sheet_name: str,
        preferred_profile: str = "auto",
) -> pd.DataFrame:
    """Extract one preloaded worksheet without reopening the workbook."""
    hinted_profile = (preferred_profile or "auto").strip().lower()
    if hinted_profile not in {"auto", "general", "matrix"}:
        hinted_profile = "auto"

    report_layout = infer_report_layout_from_quick_df(raw_df)
    flat_detection = infer_flat_table_structure(raw_df, report_layout=report_layout)
    if hinted_profile == "auto" and flat_detection["is_flat"]:
        result = _flat_frame_from_raw(
            raw_df,
            int(flat_detection["header_row"]),
        )
    else:
        auto_profile = report_layout[0]
        if "matrix" in sheet_name.lower():
            auto_profile = "matrix"
        profile = hinted_profile if hinted_profile != "auto" else auto_profile
        profile_df = (
            drop_empty_or_zero_columns(raw_df)
            if profile == "matrix"
            else raw_df
        )
        extracted = (
            _fallback_matrix_flatten(profile_df)
            if profile == "matrix"
            else auto_flatten_report_tables(
                profile_df,
                extraction_profile=profile,
            )
        )
        result = (
            (
                _finalize_matrix_extract(extracted)
                if profile == "matrix"
                else finalize_extracted_sheet(
                    pd.DataFrame(extracted),
                    strip_text=True,
                    split_hierarchy=False,
                )
            )
            if extracted is not None and not extracted.empty
            else _flat_frame_from_raw(raw_df, 0)
        )

    if result is not None and not result.empty and "tab_name" not in result.columns:
        result.insert(0, "tab_name", sheet_name)
    if (
            result is not None
            and not result.empty
            and "value" not in result.columns
            and sheet_name.strip().lower() in {
                "cover",
                "contents",
                "index",
                "readme",
                "instructions",
            }
    ):
        return pd.DataFrame()
    return result


def extract_workbook_sheets(
        file_bytes: bytes,
        sheet_names: list[str] | tuple[str, ...],
        preferred_profile: str = "auto",
        progress_callback=None,
) -> dict[str, pd.DataFrame]:
    """Extract selected sheets after parsing the workbook only once.

    progress_callback, when provided, is invoked as
    ``callback(done, total, label)`` — once with done=0 after the workbook
    parse starts, then after every requested tab finishes flattening — so a
    caller can render "N/total tabs (NN%)" style progress.
    """
    requested = [str(name) for name in sheet_names]
    total = len(requested)

    def _report(done: int, label: str) -> None:
        if progress_callback is None or not total:
            return
        try:
            progress_callback(done, total, label)
        except Exception:
            pass  # progress display must never break extraction

    _report(0, "reading workbook")
    raw_sheets = read_display_sheets(file_bytes, requested)
    extracted: dict[str, pd.DataFrame] = {}
    for done, sheet_name in enumerate(requested, start=1):
        if sheet_name in raw_sheets:
            extracted[sheet_name] = _extract_from_raw_sheet(
                raw_sheets[sheet_name],
                sheet_name,
                preferred_profile=preferred_profile,
            )
        _report(done, sheet_name)
    return extracted


def _auto_split_hierarchy_columns(
        df: pd.DataFrame,
        max_parts: int = 12,
        dash_split_mode: str = "spaced",
) -> pd.DataFrame:
    """Expand common hierarchy delimiters into dedicated `<column>_*_part_n` columns."""
    if df is None or df.empty or max_parts < 2:
        return pd.DataFrame() if df is None else df

    expanded = df.copy()

    def _unique_name(base_name: str) -> str:
        if base_name not in expanded.columns:
            return base_name
        idx = 2
        while f"{base_name}_{idx}" in expanded.columns:
            idx += 1
        return f"{base_name}_{idx}"

    # (label, "contains" pattern, "split" pattern). The pipe delimiter is always
    # active; the dash delimiter depends on dash_split_mode.
    delimiter_specs: list[tuple[str, str, str]] = [
        ("part", r"\|", r"\s*\|\s*"),
    ]
    mode = (dash_split_mode or "spaced").strip().lower()
    if mode == "spaced":
        delimiter_specs.append(("dash_part", r"\s+[-–—]\s+", r"\s+[-–—]\s+"))
    elif mode == "any":
        delimiter_specs.append(("dash_part", r"[-–—]", r"\s*[-–—]\s*"))

    for source_col in list(df.columns):
        text_values = _series_cell_text(expanded[source_col]).astype(str)

        for label, contains_pattern, split_pattern in delimiter_specs:
            has_delim = text_values.str.contains(contains_pattern, regex=True, na=False)
            if not has_delim.any():
                continue

            split_df = (
                text_values.where(has_delim, "")
                .str.split(split_pattern, expand=True, regex=True)
                .fillna("")
                .replace(r"^\s+|\s+$", "", regex=True)
            )
            col_has_values = split_df.ne("").any(axis=0).to_numpy()
            if int(col_has_values.sum()) < 2:
                continue

            # Use the position of the last populated part, not the count, so an
            # empty interior part (e.g. "A |  | C") does not drop trailing parts.
            populated = np.flatnonzero(col_has_values)
            observed_max = int(populated[-1]) + 1
            part_count = min(observed_max, max_parts)
            for idx in range(part_count):
                base_name = clean_column_name(f"{source_col}_{label}_{idx + 1}")
                new_col = _unique_name(base_name)
                expanded[new_col] = split_df.iloc[:, idx] if idx < split_df.shape[1] else ""

    return expanded


def apply_pandas_cleanup(
        df: pd.DataFrame,
        drop_columns,
        rename_map,
        split_config,
        drop_blank_columns,
        type_conversions,
        strip_text,
        clean_names,
        dash_split_mode,
) -> pd.DataFrame:
    cleaned = df.copy()
    if cleaned.empty:
        return cleaned

    if drop_columns:
        cleaned = cleaned.drop(columns=[c for c in drop_columns if c in cleaned.columns])

    if rename_map:
        cleaned = cleaned.rename(columns={
            old: new for old, new in rename_map.items()
            if old in cleaned.columns and new
        })

    if strip_text:
        for col in cleaned.select_dtypes(include="object").columns:
            series = cleaned[col]
            cleaned[col] = series.where(series.notna(), "").astype(str).str.strip()

    if (dash_split_mode or "off").strip().lower() != "off":
        cleaned = _auto_split_hierarchy_columns(cleaned, dash_split_mode=dash_split_mode)

    if split_config and split_config.get("column") in cleaned.columns:
        source_col = split_config["column"]
        delimiter = split_config.get("delimiter") or " "
        max_parts = int(split_config.get("max_parts") or 2)
        prefix = split_config.get("prefix") or source_col
        keep_original = split_config.get("keep_original", True)
        source_series = cleaned[source_col]
        parts = (
            source_series.where(source_series.notna(), "")
            .astype(str)
            .str.split(delimiter, n=max_parts - 1, expand=True, regex=False)
            .fillna("")
        )
        for idx in range(max_parts):
            new_col = clean_column_name(f"{prefix}_{idx + 1}")
            cleaned[new_col] = parts[idx] if idx in parts.columns else ""
        if not keep_original:
            cleaned = cleaned.drop(columns=[source_col])

    if drop_blank_columns:
        present = [c for c in drop_blank_columns if c in cleaned.columns]
        if present:
            candidate = cleaned[present]
            flat = pd.Series(candidate.to_numpy(dtype=object, copy=False).ravel(), dtype="object")
            blank_flat = _series_cell_text(flat).eq("").to_numpy(dtype=bool)
            blank_matrix = blank_flat.reshape(candidate.shape)
            mask = pd.Series(blank_matrix.all(axis=1), index=candidate.index)
            cleaned = cleaned.loc[~mask].copy()

    for col, target_type in (type_conversions or {}).items():
        if col not in cleaned.columns or target_type == "keep":
            continue
        if target_type == "numeric":
            cleaned[col] = pd.to_numeric(cleaned[col], errors="coerce")
        elif target_type == "datetime":
            cleaned[col] = pd.to_datetime(cleaned[col], errors="coerce")
        elif target_type == "text":
            series = cleaned[col]
            cleaned[col] = series.where(series.notna(), "").astype(str)

    if clean_names:
        cleaned.columns = dedupe_columns(cleaned.columns)

    # Columns that happen to contain the same values can still have different
    # analytical meanings (for example value and value_numeric). Do not remove
    # them solely because their current contents match.
    return drop_all_blank_columns(cleaned).reset_index(drop=True)


def _format_schema_excel_sheet(worksheet, row_count: int, column_count: int) -> None:
    """Apply compact metadata-table formatting to an openpyxl Schema sheet."""
    if worksheet is None or row_count <= 0 or column_count <= 0:
        return
    from openpyxl.styles import Alignment, Font, PatternFill

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.sheet_view.showGridLines = False

    compact_widths = {
        "A": 18,
        "B": 22,
        "C": 28,
        "D": 22,
        "E": 28,
        "F": 24,
        "G": 42,
        "H": 34,
        "I": 22,
        "J": 28,
        "K": 16,
        "L": 20,
        "M": 52,
        "N": 34,
        "O": 42,
        "P": 14,
        "Q": 38,
        "R": 28,
        "S": 32,
        "T": 22,
        "U": 30,
        "V": 22,
        "W": 18,
        "X": 18,
        "Y": 18,
        "Z": 18,
    }
    for column_letter, width in compact_widths.items():
        worksheet.column_dimensions[column_letter].width = width

    for row in worksheet.iter_rows(min_row=2, max_row=row_count + 1):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=False)


def _neutralize_formula_text(df: pd.DataFrame) -> pd.DataFrame:
    """openpyxl writes any string cell beginning with '=' as a live formula,
    so text copied from an untrusted upload could smuggle executable formulas
    into the output workbook. Quote those strings to keep them inert."""

    def _sanitize(value):
        if isinstance(value, str) and value.startswith("="):
            return "'" + value
        return value

    sanitized = df.copy()
    for position, dtype in enumerate(sanitized.dtypes):
        if dtype == object or pd.api.types.is_string_dtype(dtype):
            sanitized.isetitem(position, sanitized.iloc[:, position].map(_sanitize))
    sanitized.columns = [_sanitize(column) for column in sanitized.columns]
    return sanitized


def to_excel_bytes(df, schema_df: pd.DataFrame | None = None):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        _neutralize_formula_text(df).to_excel(writer, sheet_name="flat_table", index=False)
        if schema_df is not None and not schema_df.empty:
            _neutralize_formula_text(schema_df).to_excel(writer, sheet_name="Schema", index=False)
            _format_schema_excel_sheet(
                writer.sheets.get("Schema"),
                row_count=len(schema_df),
                column_count=len(schema_df.columns),
            )
    return output.getvalue()


def to_multisheet_excel_bytes(
        frames: dict[str, pd.DataFrame],
        schema_df: pd.DataFrame | None = None,
) -> bytes:
    """Write multiple DataFrames into one Excel workbook, one sheet per entry.

    Sheet names are truncated to 31 characters (Excel limit) and de-duplicated.
    """
    output = io.BytesIO()
    used_names: list[str] = ["Schema"]

    def _safe_sheet_name(name: str) -> str:
        # Excel sheet names: max 31 chars, no [ ] : * ? / \
        cleaned = re.sub(r"[\\/:*?\[\]]", "_", str(name))[:31]
        base = cleaned
        counter = 2
        while cleaned in used_names:
            suffix = f"_{counter}"
            cleaned = base[: 31 - len(suffix)] + suffix
            counter += 1
        used_names.append(cleaned)
        return cleaned

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, df in frames.items():
            if df is None or df.empty:
                continue
            requested_name = "Schema_Data" if str(sheet_name).strip().lower() == "schema" else sheet_name
            _neutralize_formula_text(df).to_excel(
                writer, sheet_name=_safe_sheet_name(requested_name), index=False
            )
        if schema_df is not None and not schema_df.empty:
            _neutralize_formula_text(schema_df).to_excel(writer, sheet_name="Schema", index=False)
            _format_schema_excel_sheet(
                writer.sheets.get("Schema"),
                row_count=len(schema_df),
                column_count=len(schema_df.columns),
            )
    return output.getvalue()


def generate_structural_schema(
        df: pd.DataFrame,
        max_unique_values: int = 100,
        group_by: tuple[str, ...] = ("tab_name", "table_id", "table_name"),
        source_context: dict | None = None,
) -> dict:
    """Return the lightweight schema catalog used by the integrated app."""
    context = dict(source_context or {})
    context["frame"] = df
    context.setdefault("file_name", "uploaded_table.xlsx")
    context.setdefault("sheet_name", "flat_table")
    context.setdefault("sheet_index", 0)
    context.setdefault("parsing_mode", "flattened")
    context.setdefault("source_shape", df.shape if df is not None else (0, 0))
    package = build_excel_schema_package(
        [context],
        max_sample_values=min(max(int(max_unique_values), 1), 20),
    )
    return package["catalog"]


# ============================================================================
# STANDALONE STREAMLIT UI
# ============================================================================

def _standalone_excel_sheet_names(file_bytes: bytes) -> list[str]:
    """Return workbook sheet names without loading their cell data."""
    try:
        # Shared handle: owned by the module LRU, must not be closed here.
        return [str(name) for name in _shared_excel_file(file_bytes).sheet_names]
    except Exception:
        return []


def _standalone_flat_sheets(
        file_bytes: bytes,
        sheet_names: list[str],
) -> dict[str, pd.DataFrame]:
    """Read user-confirmed flat sheets in one workbook parse."""
    if not sheet_names:
        return {}
    loaded = pd.read_excel(_shared_excel_file(file_bytes), sheet_name=sheet_names)
    if isinstance(loaded, pd.DataFrame):
        loaded = {sheet_names[0]: loaded}
    return {
        str(name): drop_empty_or_zero_columns(pd.DataFrame(frame))
        for name, frame in loaded.items()
        if frame is not None and not frame.empty
    }


def _standalone_schema_and_workbook(
        file_name: str,
        frames: dict[str, pd.DataFrame],
        parsing_mode: str,
) -> tuple[pd.DataFrame, bytes]:
    """Build the shared Schema tab and common workbook for standalone output."""
    sources = [
        {
            "file_name": file_name,
            "sheet_name": sheet_name,
            "physical_table_name": sheet_name,
            "sheet_index": sheet_index,
            "parsing_mode": parsing_mode,
            "source_shape": frame.shape,
            "frame": frame,
        }
        for sheet_index, (sheet_name, frame) in enumerate(frames.items())
        if frame is not None and not frame.empty
    ]
    package = build_excel_schema_package(sources)
    schema_frame = build_embedded_schema_frame(package)
    return schema_frame, to_multisheet_excel_bytes(
        frames,
        schema_df=schema_frame,
    )


@st.cache_data(show_spinner=False)
def _standalone_process_excel(
        file_name: str,
        file_bytes: bytes,
        sheet_names: tuple[str, ...],
        already_flat: bool,
        _progress_callback=None,
) -> tuple[dict[str, pd.DataFrame], pd.DataFrame, bytes]:
    """Process one workbook through the same backend used by the main app.

    _progress_callback is underscore-prefixed so st.cache_data does not try to
    hash it; on a cache hit it simply never fires (the result is instant).
    """
    selected = list(sheet_names)
    frames = (
        _standalone_flat_sheets(file_bytes, selected)
        if already_flat
        else extract_workbook_sheets(
            file_bytes,
            selected,
            preferred_profile="auto",
            progress_callback=_progress_callback,
        )
    )
    frames = {
        name: drop_empty_or_zero_columns(pd.DataFrame(frame))
        for name, frame in frames.items()
        if frame is not None and not frame.empty
    }
    schema_frame, workbook_bytes = _standalone_schema_and_workbook(
        file_name,
        frames,
        "already_flat" if already_flat else "auto_extracted",
    )
    return frames, schema_frame, workbook_bytes


@st.cache_data(show_spinner=False)
def _standalone_process_csv(
        file_name: str,
        file_bytes: bytes,
) -> tuple[dict[str, pd.DataFrame], pd.DataFrame, bytes]:
    frame = drop_empty_or_zero_columns(
        pd.read_csv(io.BytesIO(file_bytes))
    )
    table_name = clean_column_name(Path(file_name).stem) or "flat_table"
    frames = {table_name: frame}
    schema_frame, workbook_bytes = _standalone_schema_and_workbook(
        file_name,
        frames,
        "already_flat",
    )
    return frames, schema_frame, workbook_bytes


def run_standalone_app() -> None:
    """Render the standalone flat-file builder Streamlit interface."""
    st.set_page_config(
        page_title="Excel Flat-File Builder",
        page_icon="📊",
        layout="wide",
    )
    st.title("Excel Flat-File Builder")
    st.caption(
        "Convert messy Excel reports, multi-table tabs, and matrices into "
        "SQL-friendly flat tables with an embedded Schema tab."
    )

    uploaded = st.file_uploader(
        "Upload an Excel workbook or CSV",
        type=["xlsx", "xlsm", "csv"],
        accept_multiple_files=False,
    )
    if uploaded is None:
        st.info("Upload a file to begin.")
        return

    file_name = str(uploaded.name)
    file_bytes = bytes(uploaded.getvalue())
    extension = Path(file_name).suffix.lower()
    selected_sheets: list[str] = []
    already_flat = extension == ".csv"

    if extension != ".csv":
        try:
            _validate_workbook_archive(file_bytes)
        except Exception as exc:
            st.error(f"The workbook was rejected: {exc}")
            return
        sheet_names = _standalone_excel_sheet_names(file_bytes)
        if not sheet_names:
            st.error("The workbook could not be opened or contains no sheets.")
            return
        selected_sheets = st.multiselect(
            "Sheets to process",
            options=sheet_names,
            default=sheet_names,
        )
        processing_mode = st.radio(
            "Processing mode",
            options=[
                "Auto-detect flat/report/matrix structure",
                "Already flat — skip report extraction",
            ],
            horizontal=True,
        )
        already_flat = processing_mode.startswith("Already flat")

    process_clicked = st.button(
        "Build Common Excel Spreadsheet",
        type="primary",
        disabled=extension != ".csv" and not selected_sheets,
    )
    result_key = "standalone_builder_result_" + hashlib.sha256(
        repr(
            (
                hashlib.sha256(file_bytes).hexdigest(),
                tuple(selected_sheets),
                already_flat,
            )
        ).encode()
    ).hexdigest()
    if process_clicked:
        with st.spinner("Extracting tables and building schema..."):
            try:
                if extension == ".csv":
                    result = _standalone_process_csv(file_name, file_bytes)
                else:
                    progress_text = st.empty()
                    progress_bar = st.progress(0.0)

                    def _report_progress(done: int, total: int, label: str) -> None:
                        fraction = min(max(done / max(total, 1), 0.0), 1.0)
                        progress_bar.progress(fraction)
                        message = f"Processed {done}/{total} tabs ({fraction:.0%})"
                        if label:
                            message += f" — {label}"
                        progress_text.caption(message)

                    result = _standalone_process_excel(
                        file_name,
                        file_bytes,
                        tuple(selected_sheets),
                        already_flat,
                        _progress_callback=_report_progress,
                    )
                    progress_bar.progress(1.0)
                    progress_text.caption(
                        f"Processed {len(selected_sheets)}/{len(selected_sheets)} tabs (100%)"
                    )
                st.session_state[result_key] = result
            except Exception as exc:
                st.error(f"Processing failed: {exc}")
                return

    result = st.session_state.get(result_key)
    if result is None:
        return
    frames, schema_frame, workbook_bytes = result
    if not frames:
        st.warning("No extractable data was found in the selected sheets.")
        return

    total_rows = sum(len(frame) for frame in frames.values())
    metric_a, metric_b, metric_c = st.columns(3)
    metric_a.metric("Output tables", len(frames))
    metric_b.metric("Output rows", f"{total_rows:,}")
    metric_c.metric("Schema records", f"{len(schema_frame):,}")

    st.download_button(
        "Download Common Excel Spreadsheet",
        data=workbook_bytes,
        file_name=f"{Path(file_name).stem}_common_flat_file.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    preview_tabs = st.tabs([*frames.keys(), "Schema"])
    for tab, (table_name, frame) in zip(
            preview_tabs[:-1],
            frames.items(),
    ):
        with tab:
            st.caption(
                f"{len(frame):,} rows × {len(frame.columns):,} columns"
            )
            st.dataframe(frame.head(500), use_container_width=True)
    with preview_tabs[-1]:
        st.dataframe(schema_frame, use_container_width=True)

if __name__ == "__main__":
    run_standalone_app()

#----------------------------
#new
#----------------------------

import streamlit as st
import os
import html
from pathlib import Path
import io
from datetime import datetime
from dotenv import load_dotenv
import hashlib
from typing import Any

# ── Streamlit version compatibility ─────────────────────────────────────────
# The app must also run on older Streamlit releases (e.g. 1.23 installed via
# Homebrew). Backfill APIs that newer code paths use unconditionally:
#   - st.rerun was added in 1.27 (previously st.experimental_rerun)
#   - st.toggle was added in 1.26 (st.checkbox is the functional equivalent)
if not hasattr(st, "rerun") and hasattr(st, "experimental_rerun"):
    st.rerun = st.experimental_rerun
if not hasattr(st, "toggle"):
    st.toggle = st.checkbox

# Load .env from the same directory as THIS file — must happen before any
# azure.identity import so env vars are in the process environment in time.
_ENV_PATH = Path(__file__).resolve().parent / ".env"
load_dotenv(_ENV_PATH, override=True)

from sql_agent_core import (
    FilesDatabaseManager,
    SQLAgentOrchestrator,
    initialize_azure_client,
    load_dataframe_into_files_db,
    refresh_flat_file_schema_artifacts,
    route_schema_for_question as core_route_schema_for_question,
    summarize_loaded_tables,
)
import pandas as pd
from flat_file_builder import (
    drop_empty_or_zero_columns,
    extract_and_flatten_sheet,
    extract_workbook_sheets,
    finalize_extracted_sheet,
    apply_pandas_cleanup,
    clean_column_name,
    to_excel_bytes,
    _validate_workbook_archive,
    _shared_excel_file,
)

# ============================================================================
# CONFIGURATION
# ============================================================================

# Azure OpenAI Configuration — read from env if set, otherwise keep hardcoded defaults.
GPT_ENDPOINT   = os.getenv("GPT_ENDPOINT",   "https://cog-bnl-0001-prp-ext002-oai.openai.azure.com/")
GPT_DEPLOYMENT = os.getenv("GPT_DEPLOYMENT", "gpt-5.1-Finance")
GPT_API_VERSION = os.getenv("GPT_API_VERSION", "2024-12-01-preview")


# ── Config debug logging (True/False only — no secrets printed) ────────────
print(f"[app] .env path        : {_ENV_PATH}")
print(f"[app] .env exists       : {_ENV_PATH.exists()}")
print(f"[app] GPT_ENDPOINT set  : {bool(GPT_ENDPOINT)}")
print(f"[app] GPT_DEPLOYMENT set: {bool(GPT_DEPLOYMENT)}")
print(f"[app] GPT_API_VERSION   : {bool(GPT_API_VERSION)}")


# Custom CSS for better styling
CUSTOM_CSS = """
<style>
    /* Main container styling */
    .main {
        padding: 2rem;
    }
   
    /* Header styling */
    h1 {
        color: #1e3a8a;
        font-weight: 700;
        margin-bottom: 0.5rem;
    }
   
    h2 {
        color: #2563eb;
        font-weight: 600;
    }
   
    h3 {
        color: #3b82f6;
    }
   
    /* Card-like containers */
    .stExpander {
        border: 1px solid #e5e7eb;
        border-radius: 10px;
        margin-bottom: 1rem;
        box-shadow: 0 1px 3px rgba(0, 0, 0, 0.1);
    }
   
    /* Dataframe styling */
    .stDataFrame {
        border-radius: 10px;
        overflow: hidden;
    }
   
    /* Button styling */
    .stButton > button {
        border-radius: 8px;
        font-weight: 600;
        transition: all 0.3s ease;
    }
   
    .stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
    }
   
    /* Text input styling */
    .stTextInput > div > div > input {
        border-radius: 8px;
        border: 2px solid #e5e7eb;
        padding: 0.75rem;
        font-size: 1rem;
    }
   
    .stTextInput > div > div > input:focus {
        border-color: #3b82f6;
        box-shadow: 0 0 0 3px rgba(59, 130, 246, 0.1);
    }
   
    /* Success/Error message styling */
    .stSuccess, .stError, .stWarning, .stInfo {
        border-radius: 8px;
        padding: 1rem;
    }
   
    /* Code block styling */
    .stCodeBlock {
        border-radius: 8px;
        border: 1px solid #e5e7eb;
    }
   
    /* Tab styling */
    .stTabs [data-baseweb="tab-list"] {
        gap: 1rem;
    }
   
    .stTabs [data-baseweb="tab"] {
        border-radius: 8px 8px 0 0;
        padding: 0.75rem 1.5rem;
        font-weight: 600;
    }
   
    /* Sidebar styling */
    [data-testid="stSidebar"] {
        background-color: #f8fafc;
    }
   
    /* Metrics container */
    .metric-container {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 1.5rem;
        border-radius: 12px;
        color: white;
        margin: 1rem 0;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
    }
   
    .metric-value {
        font-size: 2.5rem;
        font-weight: 700;
        margin: 0;
    }
   
    .metric-label {
        font-size: 1rem;
        opacity: 0.9;
        margin: 0;
    }
</style>
"""


# ============================================================================
# BENELUX-AWARE EXTRACTION HELPER
# ============================================================================

def _extract_sheet_safely(
    file_bytes: bytes,
    sheet_name: str,
    preferred_profile: str = "auto",
) -> pd.DataFrame:
    """Extract one sheet, falling back to a direct flat read on failure."""
    try:
        df = extract_and_flatten_sheet(file_bytes, sheet_name, preferred_profile=preferred_profile)
        if df is not None and not df.empty:
            return df
    except Exception as e:
        print(f"[extract_sheet] Structured extraction failed: {e}")

    try:
        df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name)
        if df is not None and not df.empty:
            return finalize_extracted_sheet(df, strip_text=True, split_hierarchy=True)
    except Exception as e:
        print(f"[extract_sheet] Flat fallback failed: {e}")

    return pd.DataFrame()


# ============================================================================
# SESSION STATE
# ============================================================================


def initialize_session_state():
    """Initialize Streamlit session state"""
    if 'initialized' not in st.session_state:
        st.session_state.initialized = False
        st.session_state.azure_client = None
        st.session_state.azure_config = None
        st.session_state.agent_orchestrator = None
        st.session_state.files_db = None
        st.session_state.sql_agent = None
        st.session_state.query_history = []
        st.session_state.graph_memory = []
        st.session_state.use_graph_orchestration = True
        st.session_state.graph_thread_id = hashlib.sha256(
            str(datetime.now()).encode()
        ).hexdigest()[:12]
        st.session_state.uploaded_files = []
        st.session_state.files_loaded = False
        st.session_state.file_sheets = {}  # Maps file names to their sheets
        st.session_state.selected_sheets = {}  # Maps file names to selected sheets
        st.session_state.flat_file_overrides = {}  # Maps Excel file names to 'already flat' boolean
        st.session_state.show_sheet_selector = False

    # Backfill keys for existing sessions after app updates.
    if 'flat_file_overrides' not in st.session_state:
        st.session_state.flat_file_overrides = {}
    if 'agent_orchestrator' not in st.session_state:
        st.session_state.agent_orchestrator = None
    if 'graph_memory' not in st.session_state:
        st.session_state.graph_memory = []
    if 'use_graph_orchestration' not in st.session_state:
        st.session_state.use_graph_orchestration = True
    if 'graph_thread_id' not in st.session_state:
        st.session_state.graph_thread_id = hashlib.sha256(
            str(datetime.now()).encode()
        ).hexdigest()[:12]
    if 'pending_clarified_question' not in st.session_state:
        st.session_state.pending_clarified_question = ""


# ============================================================================
# INITIALIZATION
# ============================================================================


def get_excel_sheets_from_bytes(file_bytes: bytes):
    """Extract sheet names directly from Excel bytes (no temp file needed)."""
    try:
        # Shared handle from flat_file_builder's LRU (calamine-accelerated
        # when available); owned by that cache — do not close it here.
        return list(_shared_excel_file(file_bytes).sheet_names)
    except Exception:
        return []


def _read_excel_sheets_from_bytes(file_bytes: bytes, sheet_names: list[str]) -> dict[str, pd.DataFrame]:
    """Read multiple sheets in one workbook parse to reduce repeated IO overhead."""
    if not sheet_names:
        return {}

    try:
        result = pd.read_excel(_shared_excel_file(file_bytes), sheet_name=sheet_names)
    except Exception:
        # Keep behavior resilient: if batch read fails, fall back per sheet.
        result = {}
        for sheet_name in sheet_names:
            try:
                result[sheet_name] = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name)
            except Exception:
                result[sheet_name] = pd.DataFrame()

    if isinstance(result, pd.DataFrame):
        return {sheet_names[0]: pd.DataFrame(result)}

    return {str(name): pd.DataFrame(df) for name, df in result.items()}


def _uploaded_file_name(uploaded_file: Any) -> str:
    name = getattr(uploaded_file, "name", None)
    return str(name) if isinstance(name, str) and name else "uploaded_file"


def _uploaded_file_bytes(uploaded_file: Any) -> bytes:
    if hasattr(uploaded_file, "getbuffer"):
        try:
            return bytes(uploaded_file.getbuffer())
        except Exception:
            pass
    if hasattr(uploaded_file, "getvalue"):
        data = uploaded_file.getvalue()
        if isinstance(data, (bytes, bytearray)):
            return bytes(data)
    if hasattr(uploaded_file, "read"):
        data = uploaded_file.read()
        if isinstance(data, (bytes, bytearray)):
            return bytes(data)
    raise ValueError("Unsupported uploaded file object: expected Streamlit UploadedFile-like input")


def analyze_uploaded_files(uploaded_files: list[Any] | None):
    """Analyze uploaded files and extract sheet information"""
    if not uploaded_files:
        return {}

    file_sheets = {}
   
    for uploaded_file in uploaded_files:
        file_name = _uploaded_file_name(uploaded_file)
        file_extension = os.path.splitext(file_name)[1].lower()

        if file_extension in ['.xlsx', '.xls']:
            file_bytes = _uploaded_file_bytes(uploaded_file)
            try:
                _validate_workbook_archive(file_bytes)
            except ValueError as exc:
                st.error(f"'{file_name}' was rejected: {exc}")
                continue
            except Exception:
                pass  # non-zip (e.g. legacy .xls); let pandas decide below
            sheets = get_excel_sheets_from_bytes(file_bytes)
            if sheets:
                file_sheets[file_name] = {
                    'sheets': sheets,
                    'path': None,
                    'type': 'excel'
                }
        elif file_extension == '.csv':
            file_sheets[file_name] = {
                'sheets': None,
                'path': None,
                'type': 'csv'
            }

    return file_sheets


def _apply_default_pandas_cleanup(df: pd.DataFrame) -> pd.DataFrame:
    """Apply deterministic pandas cleanup defaults used by app loading paths."""
    if df is None or df.empty:
        return pd.DataFrame() if df is None else df
    return pd.DataFrame(
        apply_pandas_cleanup(
            df,
            drop_columns=[],
            rename_map={},
            split_config={},
            drop_blank_columns=[],
            type_conversions={},
            strip_text=True,
            clean_names=False,
            dash_split_mode="spaced",
        )
    )


def _apply_flat_file_cleanup(df: pd.DataFrame) -> pd.DataFrame:
    """Lightweight cleanup for user-confirmed flat files.

    Keeps processing minimal and removes columns containing only blank/zero
    values before SQLite and schema generation see them.
    """
    if df is None or df.empty:
        return pd.DataFrame() if df is None else df

    cleaned = drop_empty_or_zero_columns(pd.DataFrame(df).copy())
    text_cols = [
        col for col in cleaned.columns
        if pd.api.types.is_object_dtype(cleaned[col].dtype) or pd.api.types.is_string_dtype(cleaned[col].dtype)
    ]
    if text_cols:
        cleaned[text_cols] = cleaned[text_cols].apply(lambda s: s.astype(str).str.strip())

    return pd.DataFrame(cleaned)


def _table_overviews_signature() -> tuple:
    """Signature of the loaded tables, so cached summaries invalidate when the
    set of tables (or their row counts) changes."""
    files_db = st.session_state.get("files_db")
    if files_db is None or not files_db.tables_info:
        return ()
    return tuple(
        (name, int(info.get("row_count", 0) or 0), len(info.get("columns", [])))
        for name, info in files_db.tables_info.items()
    )


def _build_table_overviews(use_ai: bool) -> dict[str, dict]:
    """Compute per-table overviews, using the LLM when requested and available."""
    files_db = st.session_state.get("files_db")
    if files_db is None or not files_db.tables_info:
        return {}
    agent = st.session_state.get("sql_agent")
    if use_ai and agent is not None and st.session_state.get("azure_client") is not None:
        try:
            return agent.summarize_tables(use_ai=True)
        except Exception:
            pass  # fall back to the deterministic summaries below
    return summarize_loaded_tables(files_db)


def _get_table_overviews() -> dict[str, dict]:
    """Return cached overviews, rebuilding the heuristic ones when the loaded
    tables change. AI summaries are produced only on explicit request and kept
    until the table set changes."""
    signature = _table_overviews_signature()
    cached = st.session_state.get("table_overviews")
    if cached is not None and st.session_state.get("table_overviews_sig") == signature:
        return cached
    overviews = _build_table_overviews(use_ai=False)
    st.session_state["table_overviews"] = overviews
    st.session_state["table_overviews_sig"] = signature
    return overviews


def _display_schema_route_preview(files_db: FilesDatabaseManager, user_question: str) -> None:
    """Show the local table/column route that will be used for a free-text query."""
    if not user_question.strip():
        return
    route = core_route_schema_for_question(files_db, user_question)
    selected = route.get("selected", [])
    if not selected:
        return

    with st.expander("🧭 Schema route preview", expanded=False):
        st.caption(
            f"{len(selected)} table(s) selected from "
            f"{route.get('available_table_count', 0)} available"
            + (" · ambiguous match" if route.get("ambiguous") else "")
        )
        for item in selected:
            table = item.get("table", {})
            table_name = str(table.get("table_name", ""))
            st.markdown(
                f"**{table_name}** · {int(table.get('row_count', 0) or 0):,} rows"
            )
            columns = []
            for col in item.get("columns", []):
                samples = col.get("samples") or []
                columns.append({
                    "column": col.get("name", ""),
                    "dtype": col.get("dtype", ""),
                    "role": col.get("role", "dimension"),
                    "examples": ", ".join(str(v) for v in samples[:3]),
                })
            if columns:
                st.dataframe(pd.DataFrame(columns), use_container_width=True, hide_index=True)


def _quote_ident(name: str) -> str:
    return '"' + str(name).replace('"', '""') + '"'


def _warning_is_financial_guardrail(warning: str) -> bool:
    text = str(warning or "").lower()
    return any(
        phrase in text
        for phrase in (
            "multiple unit",
            "multiple currency",
            "multiple value_kind",
            "spans multiple",
            "subtotal/total",
            "pre-aggregated totals",
            "isolated helper cells",
        )
    )


def _refined_question_from_clarification(result: dict, option: str, label: str = "") -> str:
    base_question = str(result.get("question") or "").strip()
    clarification = str(result.get("clarification_question") or "").strip()
    chosen = f"{label}: {option}" if label else option
    parts = [base_question]
    if clarification:
        parts.append(f"Clarification requested: {clarification}")
    parts.append(f"Use this interpretation: {chosen}")
    return "\n\n".join(part for part in parts if part)


def load_data(
    uploaded_files: list[Any] | None,
    selected_sheets=None,
    existing_files_db: FilesDatabaseManager | None = None,
    flat_file_overrides: dict[str, bool] | None = None,
):
    """Load uploaded files into the database"""
    if not uploaded_files:
        st.error("❌ No files were provided for loading.")
        return None

    try:
        with st.spinner("Loading files..."):
            # Reuse existing in-memory DB so newly added files do not wipe previous ones.
            files_db = existing_files_db if existing_files_db is not None else FilesDatabaseManager()
            loaded_count = 0
           
            for uploaded_file in uploaded_files:
                file_name = _uploaded_file_name(uploaded_file)
                file_bytes = _uploaded_file_bytes(uploaded_file)
                file_extension = os.path.splitext(file_name)[1].lower()
                force_flat = bool(flat_file_overrides and flat_file_overrides.get(file_name, False))

                if file_extension in ['.xlsx', '.xls']:
                    try:
                        _validate_workbook_archive(file_bytes)
                    except ValueError as exc:
                        st.error(f"❌ '{file_name}' was rejected: {exc}")
                        continue
                    except Exception:
                        pass  # non-zip (e.g. legacy .xls); let pandas decide below

                # Explicit user override: treat selected Excel files as already flat.
                if file_extension in ['.xlsx', '.xls'] and force_flat:
                    if selected_sheets and file_name in selected_sheets:
                        sheets_to_process = selected_sheets[file_name]
                    else:
                        sheets_to_process = get_excel_sheets_from_bytes(file_bytes)

                    if not sheets_to_process:
                        st.warning(f"⚠️ No sheets selected/found for: {file_name}")
                        continue

                    preloaded_flat_sheets = _read_excel_sheets_from_bytes(file_bytes, sheets_to_process)

                    loaded_sheet_mappings = []
                    for sheet_name in sheets_to_process:
                        try:
                            flat_df = preloaded_flat_sheets.get(sheet_name, pd.DataFrame())
                            if flat_df is None or flat_df.empty:
                                flat_df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name)
                            if flat_df is None or flat_df.empty:
                                continue
                            flat_df = _apply_flat_file_cleanup(pd.DataFrame(flat_df))
                            table_name = load_dataframe_into_files_db(
                                files_db,
                                flat_df,
                                file_name,
                                sheet_name,
                            )
                            if table_name:
                                loaded_sheet_mappings.append(f"{sheet_name} -> {table_name}")
                        except Exception as e:
                            st.warning(f"⚠️ Failed to load flat sheet '{sheet_name}' from '{file_name}': {e}")

                    if loaded_sheet_mappings:
                        files_db.loaded_files.append({
                            'file_path': file_name,
                            'type': 'Excel',
                            'sheets': loaded_sheet_mappings,
                        })
                        loaded_count += 1
                        st.success(f"✅ Loaded flat file: {file_name} ({len(loaded_sheet_mappings)} sheet(s))")
                    else:
                        st.warning(f"⚠️ Failed to load: {file_name}")
                    continue

                # Fast path: process selected Excel sheets directly in-memory and load into DB.
                # This avoids flatten->write temp workbook->read workbook roundtrip.
                if file_extension in ['.xlsx', '.xls'] and not force_flat:
                    if selected_sheets and file_name in selected_sheets:
                        sheets_to_process = selected_sheets[file_name]
                    else:
                        sheets_to_process = get_excel_sheets_from_bytes(file_bytes)

                    if not sheets_to_process:
                        st.warning(f"⚠️ No sheets selected/found for: {file_name}")
                        continue

                    progress_text = st.empty()
                    progress_bar = st.progress(0.0)

                    def _report_extract_progress(
                        done: int,
                        total: int,
                        label: str,
                        _bar=progress_bar,
                        _text=progress_text,
                        _file=file_name,
                    ) -> None:
                        fraction = min(max(done / max(total, 1), 0.0), 1.0)
                        _bar.progress(fraction)
                        message = f"🔄 {_file}: {done}/{total} tabs processed ({fraction:.0%})"
                        if label:
                            message += f" — {label}"
                        _text.caption(message)

                    try:
                        extracted_sheets = extract_workbook_sheets(
                            file_bytes,
                            sheets_to_process,
                            preferred_profile="auto",
                            progress_callback=_report_extract_progress,
                        )
                    except Exception as exc:
                        st.warning(
                            "⚠️ Batch extraction failed; retrying selected sheets "
                            f"individually. ({exc})"
                        )
                        extracted_sheets = {}

                    prepared_sheets: dict[str, pd.DataFrame] = {}
                    failed_sheet_names: list[str] = []
                    fallback_sheets: dict[str, pd.DataFrame] | None = None

                    for sheet_name in sheets_to_process:
                        progress_text.caption(f"🔄 {file_name}: finalizing {sheet_name}")
                        structured_df = extracted_sheets.get(
                            sheet_name,
                            pd.DataFrame(),
                        )
                        if structured_df is None or structured_df.empty:
                            structured_df = _extract_sheet_safely(
                                file_bytes,
                                sheet_name,
                                preferred_profile="auto",
                            )

                        if structured_df is None or structured_df.empty:
                            if fallback_sheets is None:
                                fallback_sheets = _read_excel_sheets_from_bytes(
                                    file_bytes,
                                    sheets_to_process,
                                )
                            fallback_df = fallback_sheets.get(sheet_name, pd.DataFrame())
                            if fallback_df is None or fallback_df.empty:
                                fallback_df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name)
                            if fallback_df is None or fallback_df.empty:
                                failed_sheet_names.append(sheet_name)
                                st.warning(f"⚠️ Sheet '{sheet_name}' is empty, skipping...")
                                continue
                            structured_df = pd.DataFrame(fallback_df)
                            structured_df = finalize_extracted_sheet(structured_df, strip_text=True, split_hierarchy=True)

                        structured_df = drop_empty_or_zero_columns(
                            _apply_default_pandas_cleanup(
                                pd.DataFrame(structured_df)
                            )
                        )
                        if structured_df.empty:
                            failed_sheet_names.append(sheet_name)
                            st.warning(f"⚠️ Sheet '{sheet_name}' produced no rows after cleanup, skipping...")
                            continue

                        prepared_sheets[sheet_name] = structured_df

                    progress_bar.progress(1.0)
                    progress_text.caption(
                        f"✅ {file_name}: {len(sheets_to_process)}/{len(sheets_to_process)} tabs processed (100%)"
                    )
                    if prepared_sheets:
                        st.success(
                            f"✅ Prepared {len(prepared_sheets)} sheet(s) for DB load"
                            + (f" | skipped: {len(failed_sheet_names)}" if failed_sheet_names else "")
                        )
                    elif failed_sheet_names:
                        st.warning(f"⚠️ No sheets prepared from '{file_name}'. Skipped: {len(failed_sheet_names)}")

                    # Phase 2: load prepared sheets into DB.
                    loaded_sheet_mappings = []
                    for sheet_name, structured_df in prepared_sheets.items():
                        try:
                            table_name = load_dataframe_into_files_db(
                                files_db,
                                structured_df,
                                file_name,
                                sheet_name,
                            )
                            if table_name:
                                loaded_sheet_mappings.append(f"{sheet_name} -> {table_name}")
                        except Exception as e:
                            st.warning(f"⚠️ Failed to load prepared sheet '{sheet_name}' into DB: {e}")

                    if loaded_sheet_mappings:
                        files_db.loaded_files.append({
                            'file_path': file_name,
                            'type': 'Excel',
                            'sheets': loaded_sheet_mappings,
                        })
                        loaded_count += 1
                        st.success(f"✅ Loaded: {file_name} ({len(loaded_sheet_mappings)} sheet(s))")
                    else:
                        st.warning(f"⚠️ Failed to load: {file_name}")
                    continue

                if file_extension == '.csv':
                    try:
                        csv_df = pd.read_csv(io.BytesIO(file_bytes))
                        csv_df = _apply_flat_file_cleanup(csv_df)
                        table_name = load_dataframe_into_files_db(
                            files_db,
                            csv_df,
                            file_name,
                            Path(file_name).stem or "flat_table",
                        )
                    except Exception as exc:
                        table_name = ""
                        st.warning(f"⚠️ Failed to load CSV '{file_name}': {exc}")
                    if table_name:
                        files_db.loaded_files.append(
                            {
                                "file_path": file_name,
                                "type": "CSV",
                                "sheets": [f"{file_name} -> {table_name}"],
                            }
                        )
                        loaded_count += 1
                        st.success(f"✅ Loaded: {file_name}")
                    else:
                        st.warning(f"⚠️ Failed to load: {file_name}")
                    continue

                st.warning(f"⚠️ Unsupported file type: {file_name}")

            if loaded_count > 0:
                try:
                    refresh_flat_file_schema_artifacts(files_db)
                except Exception as exc:
                    st.warning(
                        "⚠️ Data loaded, but the AI schema context/common workbook "
                        f"could not be refreshed: {exc}"
                    )
                return files_db
            if existing_files_db is not None and files_db.tables_info:
                st.info("ℹ️ No new tables were added, keeping existing loaded files.")
                return files_db
            st.error("❌ No files were loaded successfully.")
            return None
    except Exception as e:
        st.error(f"Error loading files: {e}")
        return None


def initialize_azure():
    """Initialize Azure OpenAI client only."""
    if not st.session_state.initialized:
        with st.spinner("Initializing Azure OpenAI..."):
            # .env is already loaded at module level via __file__.
            # No second load_dotenv() call needed here.
            try:
                azure_client, azure_config = initialize_azure_client(
                    GPT_ENDPOINT,
                    GPT_DEPLOYMENT,
                    GPT_API_VERSION,
                )
                st.session_state.azure_client = azure_client
                st.session_state.azure_config = azure_config
                st.session_state.agent_orchestrator = SQLAgentOrchestrator(
                    azure_client,
                    azure_config.deployment_name,
                )
                st.session_state.agent_orchestrator.set_memory(
                    st.session_state.graph_memory
                )
                st.session_state.initialized = True
                return True
            except Exception as e:
                st.error(f"Failed to initialize Azure OpenAI client: {e}")
                return False

    if st.session_state.agent_orchestrator is None and st.session_state.azure_client is not None:
        st.session_state.agent_orchestrator = SQLAgentOrchestrator(
            st.session_state.azure_client,
            st.session_state.azure_config.deployment_name,
        )
        st.session_state.agent_orchestrator.set_memory(
            st.session_state.graph_memory
        )
    return True


def load_uploaded_files(uploaded_files, selected_sheets=None, flat_file_overrides=None):
    """Load uploaded files and initialize SQL agent"""
    if uploaded_files:
        existing_db = st.session_state.files_db if st.session_state.files_loaded else None
        files_db = load_data(
            uploaded_files,
            selected_sheets,
            existing_files_db=existing_db,
            flat_file_overrides=flat_file_overrides,
        )

        if files_db is None or not files_db.tables_info:
            st.error("No data was loaded. Please check your files.")
            return False

        orchestrator = st.session_state.agent_orchestrator
        if orchestrator is None:
            orchestrator = SQLAgentOrchestrator(
                st.session_state.azure_client,
                st.session_state.azure_config.deployment_name,
            )
            orchestrator.set_memory(st.session_state.graph_memory)
            st.session_state.agent_orchestrator = orchestrator

        sql_agent = orchestrator.build_agent(
            files_db,
            previous_agent=st.session_state.sql_agent,
        )

        st.session_state.files_db = files_db
        st.session_state.sql_agent = sql_agent

        # Keep previously uploaded files listed and add any new ones.
        previous_uploaded = st.session_state.uploaded_files or []
        merged_uploaded = list(previous_uploaded)
        seen_uploaded_names = {
            _uploaded_file_name(item)
            for item in merged_uploaded
        }
        for item in uploaded_files:
            item_name = _uploaded_file_name(item)
            if item_name not in seen_uploaded_names:
                merged_uploaded.append(item)
                seen_uploaded_names.add(item_name)
        st.session_state.uploaded_files = merged_uploaded

        st.session_state.files_loaded = True
        st.session_state.show_sheet_selector = False

        return True

    return False


# ============================================================================
# UI COMPONENTS
# ============================================================================


def display_tables_info():
    """Display information about loaded tables"""
    if st.session_state.files_db and st.session_state.files_db.tables_info:
        common_workbook_bytes = getattr(
            st.session_state.files_db,
            "common_workbook_bytes",
            None,
        )
        if common_workbook_bytes:
            st.download_button(
                "Download Common Flat-File Workbook",
                data=common_workbook_bytes,
                file_name="loaded_files_common_flat_file.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="download_loaded_common_workbook",
            )
            st.caption(
                "Contains one tab per loaded table and a generated Schema tab "
                "used by the SQL agent."
            )

        # Display overview metrics
        col1, col2, col3 = st.columns(3)
       
        total_tables = len(st.session_state.files_db.tables_info)
        total_rows = sum(info['row_count'] for info in st.session_state.files_db.tables_info.values())
        total_columns = sum(len(info['columns']) for info in st.session_state.files_db.tables_info.values())
       
        with col1:
            st.metric(
                label="📊 Tables Loaded",
                value=total_tables,
                delta=None
            )
       
        with col2:
            st.metric(
                label="📝 Total Rows",
                value=f"{total_rows:,}",
                delta=None
            )
       
        with col3:
            st.metric(
                label="🔢 Total Columns",
                value=total_columns,
                delta=None
            )
       
        st.markdown("---")
        overviews = _get_table_overviews()
        st.subheader("🧭 What each tab contains")
        ai_ready = (
            st.session_state.get("agent_orchestrator") is not None
            and st.session_state.get("azure_client") is not None
        )
        any_ai = any(v.get("source") == "ai" for v in overviews.values())
        head_col, btn_col = st.columns([3, 1])
        with head_col:
            st.caption(
                "AI-written one-line summary of each tab."
                if any_ai
                else "Quick auto-generated summary of each tab — click to enrich with AI."
            )
        with btn_col:
            if st.button(
                "✨ Enhance with AI" if not any_ai else "🔄 Refresh AI",
                use_container_width=True,
                disabled=not ai_ready,
                help=None if ai_ready else "Azure OpenAI is still initializing.",
                key="generate_tab_overviews",
            ):
                with st.spinner("Summarizing tabs..."):
                    st.session_state["table_overviews"] = _build_table_overviews(use_ai=True)
                    st.session_state["table_overviews_sig"] = _table_overviews_signature()
                st.rerun()
        for table_name, overview in overviews.items():
            badge = "🤖" if overview.get("source") == "ai" else "•"
            st.markdown(f"{badge} **{table_name}** — {overview.get('summary', '')}")

        st.markdown("---")
        st.subheader("📋 Table Summary")

        summary_df = st.session_state.files_db.get_tables_summary()
        st.dataframe(
            summary_df,
            use_container_width=True,
            hide_index=True,
            column_config={
                "Table": st.column_config.TextColumn("Table Name", width="medium"),
                "Source": st.column_config.TextColumn("Source File", width="large"),
                "Rows": st.column_config.NumberColumn("Rows", format="%d"),
                "Columns": st.column_config.NumberColumn("Columns", format="%d"),
            }
        )
       
        st.markdown("---")
       
        # Show expandable details for each table
        for table_name, info in st.session_state.files_db.tables_info.items():
            with st.expander(f"📄 **{table_name}** - {info['row_count']:,} rows"):
                table_overview = overviews.get(table_name, {})
                if table_overview.get("summary"):
                    badge = "🤖 AI summary" if table_overview.get("source") == "ai" else "📝 Summary"
                    st.caption(f"{badge}: {table_overview['summary']}")
                col_a, col_b = st.columns([1, 2])
               
                with col_a:
                    st.markdown("**📂 Source Information**")
                    st.write(f"**File:** {info['source_file']}")
                    if 'source_sheet' in info:
                        st.write(f"**Sheet:** {info['source_sheet']}")
                    st.write(f"**Type:** {info.get('source_type', 'Excel')}")
               
                with col_b:
                    st.markdown("**📊 Column Information**")
                    st.write(f"**Total Columns:** {len(info['columns'])}")
                    st.write(f"**Column Names:** {', '.join(info['columns'][:10])}")
                    if len(info['columns']) > 10:
                        st.write(f"*... and {len(info['columns']) - 10} more columns*")
               
                # Show data preview
                try:
                    preview_query = f"SELECT * FROM {_quote_ident(table_name)} LIMIT 5"
                    preview_df = st.session_state.files_db.execute_query(preview_query)
                    st.markdown("**🔍 Data Preview (first 5 rows):**")
                    st.dataframe(preview_df, use_container_width=True, hide_index=True)
                except Exception as e:
                    st.warning(f"Could not load preview: {e}")

                st.markdown("---")
                st.caption("Use advanced cleanup to rename columns, split hierarchy values, and re-apply cleaned data to this in-memory table.")
                show_cleanup = st.checkbox(
                    "Open advanced pandas cleanup",
                    key=f"adv_cleanup_toggle_{table_name}",
                    value=False,
                )
                if show_cleanup:
                    try:
                        full_df = st.session_state.files_db.execute_query(
                            f"SELECT * FROM {_quote_ident(table_name)}"
                        )
                        final_df = render_advanced_table_preview(
                            full_df,
                            file_stem=table_name,
                            key_prefix=f"adv_cleanup_{table_name}",
                        )
                        if final_df is not None and not final_df.empty:
                            apply_col, info_col = st.columns([1, 2])
                            with apply_col:
                                if st.button(
                                    "Apply cleaned table",
                                    key=f"apply_cleaned_{table_name}",
                                    type="primary",
                                    use_container_width=True,
                                ):
                                    orchestrator = st.session_state.agent_orchestrator
                                    if orchestrator is None:
                                        orchestrator = SQLAgentOrchestrator(
                                            st.session_state.azure_client,
                                            st.session_state.azure_config.deployment_name,
                                        )
                                        orchestrator.set_memory(
                                            st.session_state.graph_memory
                                        )
                                        st.session_state.agent_orchestrator = orchestrator
                                    replaced_table = orchestrator.replace_table(
                                        st.session_state.files_db,
                                        table_name,
                                        final_df,
                                        str(info.get('source_file', 'unknown')),
                                        str(info.get('source_sheet', table_name)),
                                        sql_agent=st.session_state.sql_agent,
                                    )
                                    if replaced_table:
                                        st.success(f"✅ Updated table: {replaced_table} ({len(final_df)} rows)")
                                        st.rerun()
                                    else:
                                        st.error("❌ Failed to update table.")
                            with info_col:
                                st.info("This updates the active in-memory table used for querying in this session.")
                    except Exception as e:
                        st.warning(f"Could not open advanced cleanup for '{table_name}': {e}")


def display_query_interface():
    """Display the query interface"""
    # Welcome message
    st.markdown("""
    <div style='background-color: #f8fafc;
                padding: 1.5rem;
                border-radius: 10px;
                border-left: 4px solid #3b82f6;
                margin-bottom: 2rem;'>
        <h3 style='margin: 0; color: #1e3a8a;'>💬 Ask Your Question</h3>
        <p style='margin: 0.5rem 0 0 0; color: #64748b;'>
            Type your question in natural language and let AI generate the SQL query for you
        </p>
    </div>
    """, unsafe_allow_html=True)
   
    # Example questions in a nicer format
    with st.expander("💡 Need inspiration? Click here for example questions"):
        col1, col2 = st.columns(2)
       
        with col1:
            st.markdown("**📊 Data Exploration:**")
            st.markdown("""
            - How many rows are in each table?
            - Show me all unique values in [column_name]
            - Display the first 20 records
            """)
       
        with col2:
            st.markdown("**📈 Analysis & Aggregation:**")
            st.markdown("""
            - What is the total sum of [column_name]?
            - Calculate the average [column_name] by [group_column]
            - Show me the top 10 records by [column_name]
            """)
   
    files_db = st.session_state.files_db
    # A clarification option was clicked on the previous run. Consume the flag
    # now, but run the refined query further down so the input box and the
    # persistent result area both still render this pass.
    pending_clarified = str(st.session_state.get("pending_clarified_question", "") or "").strip()
    if pending_clarified:
        st.session_state.pending_clarified_question = ""

    # Graph-level reasoning (planning, sanity checks, cross-question memory)
    # is embedded by default; the opt-out lives in the sidebar settings.
    langgraph_available, _ = SQLAgentOrchestrator.langgraph_status()
    graph_engine = "LangGraph" if langgraph_available else "Local graph"
    memory_count = len(st.session_state.graph_memory)
    if st.session_state.get("use_graph_orchestration", True):
        st.caption(f"🧠 Graph reasoning on · {graph_engine} · {memory_count} memory record(s)")
    else:
        st.caption("🧠 Graph reasoning off — re-enable it in the sidebar to reuse verified results across questions.")

    submit_button = False
    user_question = ""

    # Free-text natural-language question (the only query mode).
    user_question = st.text_area(
        "Your Question:",
        placeholder="e.g., What is the total cost broken down by department?",
        height=100,
        key="user_question",
        help="Type your question in natural language. Be as specific as possible for better results."
    )
    if files_db is not None:
        _display_schema_route_preview(files_db, user_question)

    # Action buttons with better layout
    col1, col2, col3 = st.columns([2, 2, 6])
    with col1:
        submit_button = st.button("🔍 Generate", type="primary", use_container_width=True,)
    with col2:
        clear_button = st.button("🗑️ Clear History", use_container_width=True,)
   
    if clear_button:
        st.session_state.query_history = []
        st.session_state.graph_memory = []
        st.session_state.sql_agent.conversation_history = []
        if st.session_state.agent_orchestrator is not None:
            st.session_state.agent_orchestrator.clear_memory()
        st.success("✨ History cleared successfully!")
        st.rerun()
   
    if pending_clarified:
        execute_query(pending_clarified)
    elif submit_button and user_question:
        execute_query(user_question)

    # Persistently render the most recent result so its interactive elements
    # (notably the clarification option buttons) stay live across reruns. A
    # Streamlit button only reports its click on the run where it is
    # re-instantiated, so the result must be drawn here on every pass — not
    # only as a one-shot side effect of execute_query.
    latest_history = st.session_state.get("query_history") or []
    if latest_history:
        display_query_result(latest_history[0])


def execute_query(user_question: str):
    """Execute a query and display results"""
    with st.spinner("Generating SQL and executing query..."):
        orchestrator = st.session_state.agent_orchestrator
        if orchestrator is None:
            orchestrator = SQLAgentOrchestrator(
                st.session_state.azure_client,
                st.session_state.azure_config.deployment_name,
            )
            orchestrator.set_memory(st.session_state.graph_memory)
            st.session_state.agent_orchestrator = orchestrator
        result = orchestrator.run_free_text_query(
            st.session_state.sql_agent,
            st.session_state.files_db,
            user_question,
            graph_memory=st.session_state.graph_memory,
            use_langgraph=True,
            enable_memory=bool(st.session_state.use_graph_orchestration),
            thread_id=st.session_state.graph_thread_id,
        )
        if st.session_state.use_graph_orchestration:
            st.session_state.graph_memory = list(result.get("graph_memory") or [])

        # Store the result; rendering happens in display_query_interface so it
        # (and any clarification buttons) persists across reruns.
        st.session_state.query_history.insert(0, result)


def display_graph_diagnostics(result: dict):
    """Show compact graph/memory context metadata without exposing hidden reasoning."""
    graph_mode = result.get("graph_mode")
    prompt_tokens = int(result.get("prompt_context_est_tokens") or 0)
    schema_tokens = int(result.get("schema_context_est_tokens") or 0)
    memory_tokens = int(result.get("memory_context_est_tokens") or 0)
    examples_tokens = int(result.get("verified_examples_context_est_tokens") or 0)
    memory_used = result.get("memory_used") or []
    examples_used = result.get("verified_examples_used") or []
    selected_tables = result.get("selected_schema_tables") or []
    schema_confidence = result.get("schema_confidence")
    schema_expanded = bool(result.get("schema_expanded"))
    result_sanity = result.get("result_sanity") or {}

    if graph_mode or prompt_tokens or schema_tokens or memory_tokens or examples_tokens:
        parts = []
        if graph_mode:
            parts.append(f"Graph: {graph_mode}")
        if schema_confidence:
            parts.append(f"schema confidence: {schema_confidence}")
        if schema_expanded:
            parts.append("expanded route")
        if prompt_tokens:
            parts.append(f"prompt context ~{prompt_tokens} tokens")
        if schema_tokens:
            parts.append(f"schema ~{schema_tokens}")
        if memory_tokens:
            parts.append(f"memory ~{memory_tokens}")
        if examples_tokens:
            parts.append(f"examples ~{examples_tokens}")
        if memory_used:
            parts.append(f"{len(memory_used)} memory record(s) used")
        if examples_used:
            parts.append(f"{len(examples_used)} verified example(s)")
        if parts:
            st.caption(" · ".join(parts))

    graph_trace = result.get("graph_trace") or []
    if graph_trace or selected_tables or result_sanity:
        with st.expander("🧭 Graph trace", expanded=False):
            if selected_tables:
                st.caption("Selected schema tables: " + ", ".join(selected_tables))
            if result_sanity:
                flags = result_sanity.get("flags") or []
                st.caption(
                    f"Result sanity: {result_sanity.get('status', 'unknown')}"
                    + (f" ({', '.join(flags)})" if flags else "")
                )
            for line in graph_trace:
                st.text(line)
            if memory_used:
                st.markdown("**Memory used:**")
                for item in memory_used:
                    st.caption(str(item.get("question", "")))
            if examples_used:
                st.markdown("**Verified examples used:**")
                for item in examples_used:
                    st.caption(str(item.get("question", "")))


def display_query_result(result: dict):
    """Display the result of a query (v3: clarification | success | failure)."""
    result_container = st.container()

    with result_container:
        st.markdown("---")

        # ── Clarification needed ───────────────────────────────────────────
        if result.get('needs_clarification'):
            st.markdown("**Your Question:**")
            st.info(result['question'])
            display_graph_diagnostics(result)

            st.markdown(
                "<div style='background:#fffbeb; border-left:4px solid #f59e0b; "
                "padding:1.25rem 1.5rem; border-radius:6px; margin:1rem 0;'>"
                "<div style='font-size:1.05rem; font-weight:700; color:#92400e; "
                "margin-bottom:0.6rem;'>🤔 Clarification needed</div>"
                "<div style='color:#78350f; font-size:1rem;'>"
                + html.escape(str(result.get('clarification_question', 'Could you clarify your question?')))
                + "</div></div>",
                unsafe_allow_html=True,
            )

            options = result.get('clarification_options', [])
            labels  = result.get('option_labels', [])
            if options:
                st.markdown("**Suggested interpretations:**")
                for i, opt in enumerate(options, 1):
                    label = labels[i - 1] if i - 1 < len(labels) and labels[i - 1] else None
                    if label:
                        st.markdown(f"&nbsp;&nbsp;**{i}. {label}:** {opt}")
                    else:
                        st.markdown(f"&nbsp;&nbsp;**{i}.** {opt}")
                    button_label = f"Use option {i}" + (f": {label}" if label else "")
                    button_key = "clarify_" + hashlib.sha256(
                        f"{result.get('question', '')}|{i}|{opt}".encode()
                    ).hexdigest()[:12]
                    if st.button(button_label, key=button_key, use_container_width=True):
                        st.session_state.pending_clarified_question = _refined_question_from_clarification(
                            result,
                            str(opt),
                            str(label or ""),
                        )
                        st.rerun()

            secondary_note = result.get('secondary_note', '')
            if secondary_note:
                st.info(f"💡 {secondary_note}")

            reason = result.get('clarification_reason', '')
            if reason:
                st.caption(f"ℹ️ {reason}")

            st.caption("Or edit the question manually above and run it again.")
            return

        # ── Success ───────────────────────────────────────────────────────
        if result['success']:
            warnings_list = result.get('warnings', [])
            has_financial_guardrail = any(
                _warning_is_financial_guardrail(w) for w in warnings_list
            )
            if has_financial_guardrail:
                st.warning("⚠️ Query ran with financial guardrail warnings.")
            else:
                st.success("✅ Query executed successfully!")


            # Question
            st.markdown("**Your Question:**")
            st.info(result['question'])
            display_graph_diagnostics(result)

            if warnings_list:
                if has_financial_guardrail:
                    st.error("Review financial guardrail warnings before trusting this result.")
                with st.expander(
                    f"⚠️ {len(warnings_list)} warning(s)",
                    expanded=has_financial_guardrail,
                ):
                    for w in warnings_list:
                        st.warning(w)

            # ── Answer summary (new — most prominent) ─────────────────────
            answer_summary = result.get('answer_summary', '')
            if answer_summary:
                safe_answer_summary = html.escape(str(answer_summary)).replace("\n", "<br>")
                st.markdown("**💡 Answer:**")
                st.markdown(
                    f"<div style='background:#f0fdf4; border-left:4px solid #22c55e; "
                    f"padding:1rem; border-radius:6px; margin-bottom:1rem;'>"
                    f"{safe_answer_summary}</div>",
                    unsafe_allow_html=True,
                )

            # ── Repair notice ──────────────────────────────────────────────
            repair_attempts = result.get('repair_attempts', 0)
            if repair_attempts:
                st.info(f"🔧 SQL was repaired automatically ({repair_attempts} attempt(s)).")


            # ── Generated SQL ──────────────────────────────────────────────
            with st.expander("🔍 View Generated SQL", expanded=False):
                st.code(result.get('sql_query', ''), language='sql')


            # ── Debug trace ───────────────────────────────────────────────
            trace = result.get('trace', [])
            if trace:
                with st.expander("🔎 Pipeline trace (debug)", expanded=False):
                    for line in trace:
                        st.text(line)


            # ── Results table ─────────────────────────────────────────────
            st.markdown("---")
            st.subheader("📊 Results")
            results_df = result['results']


            if results_df is not None and len(results_df) > 0:
                # Single-value result → show as metric
                if len(results_df) == 1 and len(results_df.columns) == 1:
                    value = results_df.iloc[0, 0]
                    col_name = results_df.columns[0]
                    try:
                        if isinstance(value, float):
                            formatted = f"{value:,.2f}"
                        elif isinstance(value, int):
                            formatted = f"{value:,}"
                        else:
                            formatted = str(value)
                    except Exception:
                        formatted = str(value)
                    st.markdown(f"**{col_name}:** {formatted}")
                elif len(results_df) <= 20:
                    st.dataframe(results_df, use_container_width=True, hide_index=True)
                else:
                    st.dataframe(
                        results_df,
                        use_container_width=True,
                        hide_index=True,
                        height=400,
                    )


                # Download
                st.markdown("---")
                csv = results_df.to_csv(index=False)
                col_dl, _ = st.columns([1, 3])
                with col_dl:
                    st.download_button(
                        label="📥 Download CSV",
                        data=csv,
                        file_name=f"query_results_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.csv",
                        mime="text/csv",
                        use_container_width=True,
                    )
            else:
                st.warning("⚠️ No results found for this query.")


        else:
            # ── Failure path ───────────────────────────────────────────────
            st.error("❌ Query execution failed")


            st.markdown("**Your Question:**")
            st.info(result['question'])
            display_graph_diagnostics(result)


            # Error message
            with st.expander("🔍 Error Details", expanded=True):
                st.error(result.get('error', 'Unknown error'))


            # SQL that was attempted (if any)
            if result.get('sql_query'):
                with st.expander("🔍 Last SQL Attempted", expanded=False):
                    st.code(result['sql_query'], language='sql')


            # Warnings
            warnings_list = result.get('warnings', [])
            if warnings_list:
                has_financial_guardrail = any(
                    _warning_is_financial_guardrail(w) for w in warnings_list
                )
                with st.expander("⚠️ Warnings", expanded=has_financial_guardrail):
                    for w in warnings_list:
                        st.warning(w)


            # Trace
            trace = result.get('trace', [])
            if trace:
                with st.expander("🔎 Pipeline trace (debug)", expanded=False):
                    for line in trace:
                        st.text(line)


def display_query_history():
    """Display query history (v3: handles clarification, success, and failure entries)."""
    if st.session_state.query_history:
        st.markdown(f"### 📜 Query History ({len(st.session_state.query_history)} queries)")
        st.markdown("---")

        for idx, result in enumerate(st.session_state.query_history):
            # Choose icon based on result type
            if result.get('needs_clarification'):
                status_icon = "🤔"
            elif result['success']:
                status_icon = "✅"
            else:
                status_icon = "❌"

            question_preview = result['question'][:60]
            suffix           = "..." if len(result['question']) > 60 else ""

            with st.expander(
                f"{status_icon} Query {idx + 1}: {question_preview}{suffix}",
                expanded=False,
            ):
                st.markdown(f"**Full Question:** {result['question']}")

                # ── Clarification entry ─────────────────────────────────
                if result.get('needs_clarification'):
                    st.markdown(
                        "<div style='background:#fffbeb; border-left:4px solid #f59e0b; "
                        "padding:0.75rem 1rem; border-radius:6px; margin:0.5rem 0;'>"
                        "<b>🤔 Clarification was requested:</b><br>"
                        + result.get('clarification_question', '') + "</div>",
                        unsafe_allow_html=True,
                    )
                    options = result.get('clarification_options', [])
                    labels  = result.get('option_labels', [])
                    if options:
                        parts = []
                        for i, opt in enumerate(options, 1):
                            label = labels[i - 1] if i - 1 < len(labels) and labels[i - 1] else None
                            parts.append(f"**{i}. {label}:** {opt}" if label else f"**{i}.** {opt}")
                        st.markdown("**Options offered:** " + " · ".join(parts))
                    secondary_note = result.get('secondary_note', '')
                    if secondary_note:
                        st.caption(f"💡 {secondary_note}")

                # ── Success entry ───────────────────────────────────────
                elif result['success']:
                    answer_summary = result.get('answer_summary', '')
                    if answer_summary:
                        st.markdown("**💡 Answer:**")
                        st.markdown(answer_summary)

                    st.markdown("**Generated SQL:**")
                    st.code(result.get('sql_query', ''), language='sql')

                    if result.get('repair_attempts', 0):
                        st.info(f"🔧 SQL was repaired ({result['repair_attempts']} attempt(s)).")

                    results_df = result.get('results')
                    if results_df is not None and len(results_df) > 0:
                        st.markdown("**Results:**")
                        if len(results_df) <= 5:
                            st.dataframe(
                                results_df,
                                use_container_width=True,
                                hide_index=True,
                            )
                        else:
                            st.dataframe(
                                results_df,
                                use_container_width=True,
                                hide_index=True,
                                height=300,
                            )

                        st.caption(f"📊 {len(results_df)} rows returned")

                    warnings_list = result.get('warnings', [])
                    if warnings_list:
                        with st.expander("⚠️ Warnings", expanded=False):
                            for w in warnings_list:
                                st.warning(w)

                # ── Failure entry ───────────────────────────────────────
                else:
                    st.error(f"**Error:** {result.get('error', 'Unknown error')}")


    else:
        st.markdown(
            """
            <div style='text-align:center; padding:3rem; background-color:#f9fafb;
                        border-radius:15px; border:2px dashed #d1d5db;'>
                <h3 style='color:#6b7280;'>📜 No queries yet</h3>
                <p style='color:#9ca3af;'>Your query history will appear here
                after you run your first query</p>
            </div>
            """,
            unsafe_allow_html=True,
        )


# ============================================================================
# ADVANCED TABLE RENDERING
# ============================================================================

def render_advanced_table_preview(df: pd.DataFrame, file_stem: str, key_prefix: str):
    """Render advanced table preview with full pandas cleanup and data operations.

    This function provides all the functionality from flat_file_builder.py for:
    - Column deletion
    - Column renaming
    - Column splitting
    - Type conversions
    - Hierarchy splitting on | and - delimiters
    - Whitespace trimming
    - Column name cleaning
    - Duplicate column removal
    - Note row filtering
    - Excel download export
    """
    # Use a key that includes the file_stem so all widget state
    # automatically resets when the user switches to a different file/sheet.
    scope = hashlib.sha256(f"{key_prefix}_{file_stem}".encode()).hexdigest()[:8]
    kp = f"{key_prefix}_{scope}"

    st.markdown("### Advanced Table Operations")
    with st.expander("Edit the table with pandas-style operations", expanded=False):
        cleanup_cols = list(df.columns)
        drop_columns = st.multiselect(
            "Delete columns",
            cleanup_cols,
            help="Remove columns from the final output.",
            key=f"{kp}_drop_columns",
        )

        remaining_for_rename = [c for c in cleanup_cols if c not in drop_columns]
        rename_df = pd.DataFrame({
            "Column": remaining_for_rename,
            "New Name": remaining_for_rename,
        })
        edited_rename_df = st.data_editor(
            rename_df,
            hide_index=True,
            use_container_width=True,
            num_rows="fixed",
            key=f"{kp}_rename_columns_editor",
        )
        rename_map = {
            row["Column"]: row["New Name"]
            for _, row in edited_rename_df.iterrows()
            if row["Column"] != row["New Name"]
        }

        st.markdown("**Split a column**")
        split_enabled = st.checkbox(
            "Split one column into multiple columns",
            key=f"{kp}_split_enabled",
        )
        split_config = {}
        if split_enabled and remaining_for_rename:
            col_split_a, col_split_b, col_split_c = st.columns(3)
            with col_split_a:
                split_column = st.selectbox(
                    "Column to split",
                    remaining_for_rename,
                    key=f"{kp}_split_column",
                )
            with col_split_b:
                delimiter = st.text_input(
                    "Delimiter",
                    value=" | ",
                    key=f"{kp}_split_delimiter",
                )
            with col_split_c:
                max_parts = st.number_input(
                    "Number of output columns",
                    min_value=2,
                    max_value=12,
                    value=2,
                    key=f"{kp}_split_max_parts",
                )
            col_prefix, col_keep = st.columns([2, 1])
            with col_prefix:
                split_prefix = st.text_input(
                    "Output column prefix",
                    value=clean_column_name(split_column),
                    key=f"{kp}_split_prefix",
                )
            with col_keep:
                keep_original = st.checkbox(
                    "Keep original",
                    value=True,
                    key=f"{kp}_split_keep_original",
                )
            split_config = {
                "column": split_column,
                "delimiter": delimiter,
                "max_parts": max_parts,
                "prefix": split_prefix,
                "keep_original": keep_original,
            }

        drop_blank_columns = []
        type_conversions = {}

        col_strip, col_clean, col_dash = st.columns(3)
        with col_strip:
            strip_text = st.checkbox(
                "Trim whitespace in text columns",
                value=True,
                key=f"{kp}_strip_text",
            )
        with col_clean:
            clean_names = st.checkbox(
                "Clean final column names",
                value=False,
                key=f"{kp}_clean_names",
            )
        with col_dash:
            dash_split_mode = st.selectbox(
                "Auto-split '-' mode",
                options=["Off", "Spaced only ( - )", "Any dash (-)"],
                index=1,
                key=f"{kp}_dash_split_mode",
                help="Controls automatic dash splitting in hierarchy columns.",
            )
        dash_split_mode_map = {
            "Off": "off",
            "Spaced only ( - )": "spaced",
            "Any dash (-)": "any",
        }
        dash_split_mode_value = dash_split_mode_map[dash_split_mode]

    # Cache cleanup result in session_state keyed on inputs so it only
    # recomputes when the user actually changes a cleanup setting.
    cleanup_cache_version = "v2_hierarchy_split_regex"
    cleanup_sig = hashlib.sha256(
        str((
            cleanup_cache_version,
            list(df.columns), len(df),
            sorted(drop_columns),
            sorted(rename_map.items()),
            str(split_config),
            strip_text,
            clean_names,
            dash_split_mode_value,
        )).encode()
    ).hexdigest()
    cleanup_cache_key = f"{kp}_final_df_{cleanup_sig}"

    if cleanup_cache_key not in st.session_state:
        st.session_state[cleanup_cache_key] = apply_pandas_cleanup(
            df,
            drop_columns=drop_columns,
            rename_map=rename_map,
            split_config=split_config,
            drop_blank_columns=drop_blank_columns,
            type_conversions=type_conversions,
            strip_text=strip_text,
            clean_names=clean_names,
            dash_split_mode=dash_split_mode_value,
        )
    final_df = st.session_state[cleanup_cache_key]

    st.markdown("### Final Table Preview")
    col_rows, col_cols = st.columns(2)
    col_rows.metric("Rows", f"{len(final_df):,}")
    col_cols.metric("Columns", f"{len(final_df.columns):,}")
    st.dataframe(final_df.head(200), use_container_width=True)

    # Build Excel bytes only when the user explicitly requests the download.
    # Using a callback keeps the heavy serialization out of every re-render.
    dl_key = f"{kp}_download"
    dl_state_key = f"{kp}_excel_bytes_{cleanup_sig}"  # keyed on cleanup sig so stale bytes are never served

    def _prepare_download():
        st.session_state[dl_state_key] = to_excel_bytes(final_df)

    if dl_state_key not in st.session_state:
        st.button(
            "Prepare Download",
            on_click=_prepare_download,
            use_container_width=True,
            key=f"{kp}_prepare_btn",
        )
    else:
        st.download_button(
            "⬇️ Download Excel",
            data=st.session_state[dl_state_key],
            file_name=f"{file_stem}_processed.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key=dl_key,
        )

    return final_df


# ============================================================================
# MAIN APP
# ============================================================================


def main():
    """Main application"""
    st.set_page_config(
        page_title="SQL Query Agent - Natural Language to SQL",
        page_icon="🔍",
        layout="wide",
        initial_sidebar_state="expanded"
    )
   
    # Apply custom CSS
    st.markdown(CUSTOM_CSS, unsafe_allow_html=True)
   
    # Header block
    st.markdown("""
    <div style='background: linear-gradient(to right, #3b82f6, #2563eb);
                padding: 2rem;
                border-radius: 12px;
                margin-bottom: 2rem;
                box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);'>
        <h1 style='color: white; margin: 0; font-size: 2.5rem;'>🔍 SQL Query Agent</h1>
        <p style='color: white; opacity: 0.95; margin: 0.5rem 0 0 0; font-size: 1.1rem;'>
            Transform your questions into insights with AI-powered SQL generation
        </p>
    </div>
    """, unsafe_allow_html=True)
   
    # Initialize session state
    initialize_session_state()
   
    # Initialize Azure OpenAI
    if not initialize_azure():
        st.stop()
   
    # Sidebar with enhanced styling
    with st.sidebar:
        st.markdown("""
        <div style='text-align: center; padding: 1rem; margin-bottom: 1rem;'>
            <h2 style='color: #1e3a8a; margin: 0;'>📤 Upload Files</h2>
        </div>
        """, unsafe_allow_html=True)
       
        # File uploader
        uploaded_files = st.file_uploader(
            "Upload Excel or CSV files",
            type=['xlsx', 'xls', 'csv'],
            accept_multiple_files=True,
            help="Upload one or more Excel (.xlsx, .xls) or CSV (.csv) files to query",
            key="file_uploader"
        )
        uploaded_file_list = uploaded_files if isinstance(uploaded_files, list) else ([uploaded_files] if uploaded_files else [])
       
        # Analyze files and show sheet selection
        if uploaded_file_list:
            # Analyze files to get sheets
            if st.button("📋 Analyze Files", type="secondary", use_container_width=True):
                with st.spinner("Analyzing files..."):
                    file_sheets = analyze_uploaded_files(uploaded_file_list)
                    st.session_state.file_sheets = file_sheets
                    # Initialize selected sheets with all sheets (preserve existing selections when possible)
                    existing_selected = dict(st.session_state.selected_sheets)
                    existing_overrides = dict(st.session_state.flat_file_overrides)
                    st.session_state.selected_sheets = {}
                    st.session_state.flat_file_overrides = {}
                    for file_name, info in file_sheets.items():
                        if info['type'] == 'excel' and info['sheets']:
                            st.session_state.selected_sheets[file_name] = existing_selected.get(file_name, info['sheets'].copy())
                            st.session_state.flat_file_overrides[file_name] = bool(existing_overrides.get(file_name, False))
                    st.session_state.show_sheet_selector = True
                    st.success("✅ Files analyzed! Select sheets below.")
                    st.rerun()
           
            # Show uploaded file names
            st.markdown("**📁 Selected Files:**")
            for uploaded in uploaded_file_list:
                uploaded_name = _uploaded_file_name(uploaded)
                uploaded_bytes = _uploaded_file_bytes(uploaded)
                file_size = len(uploaded_bytes) / (1024 * 1024)  # Convert to MB
                file_extension = os.path.splitext(uploaded_name)[1].lower()
                file_type = "Excel" if file_extension in ['.xlsx', '.xls'] else "CSV"
                st.text(f"• {uploaded_name} ({file_size:.2f} MB, {file_type})")

            # Sheet selection interface
            if st.session_state.show_sheet_selector and st.session_state.file_sheets:
                st.markdown("---")
                st.markdown("**📊 Select Sheets to Load:**")
               
                # Show sheet selection for each Excel file
                has_excel = False
                for file_name, info in st.session_state.file_sheets.items():
                    if info['type'] == 'excel' and info['sheets']:
                        has_excel = True
                        with st.expander(f"📁 {file_name}", expanded=True):
                            st.markdown(f"*Found {len(info['sheets'])} sheet(s)*")
                           
                            # Select all/none buttons
                            col1, col2 = st.columns(2)
                            with col1:
                                if st.button(f"✅ Select All", key=f"select_all_{file_name}", use_container_width=True,):
                                    st.session_state.selected_sheets[file_name] = info['sheets'].copy()
                                    # Force widget state update by clearing checkbox keys
                                    for sheet in info['sheets']:
                                        widget_key = f"sheet_{file_name}_{sheet}"
                                        if widget_key in st.session_state:
                                            st.session_state[widget_key] = True
                                    st.rerun()
                            with col2:
                                if st.button(f"❌ Clear All", key=f"clear_all_{file_name}", use_container_width=True,):
                                    st.session_state.selected_sheets[file_name] = []
                                    # Force widget state update by clearing checkbox keys
                                    for sheet in info['sheets']:
                                        widget_key = f"sheet_{file_name}_{sheet}"
                                        if widget_key in st.session_state:
                                            st.session_state[widget_key] = False
                                    st.rerun()
                           
                            # Checkboxes for each sheet with callback
                            def update_sheet_selection(file_name, sheet):
                                """Callback to update sheet selection"""
                                widget_key = f"sheet_{file_name}_{sheet}"
                                is_checked = st.session_state.get(widget_key, False)
                               
                                if file_name not in st.session_state.selected_sheets:
                                    st.session_state.selected_sheets[file_name] = []
                               
                                if is_checked and sheet not in st.session_state.selected_sheets[file_name]:
                                    st.session_state.selected_sheets[file_name].append(sheet)
                                elif not is_checked and sheet in st.session_state.selected_sheets[file_name]:
                                    st.session_state.selected_sheets[file_name].remove(sheet)
                           
                            for sheet in info['sheets']:
                                is_selected = sheet in st.session_state.selected_sheets.get(file_name, [])
                                st.checkbox(
                                    sheet,
                                    value=is_selected,
                                    key=f"sheet_{file_name}_{sheet}",
                                    on_change=update_sheet_selection,
                                    args=(file_name, sheet)
                                )

                            st.markdown("---")
                            st.caption("Workbook processing mode")
                            mode_key = f"processing_mode_{file_name}"
                            current_override = bool(st.session_state.flat_file_overrides.get(file_name, False))
                            mode_options = [
                                "Auto extract report/matrix",
                                "Already flat (skip extraction)",
                            ]
                            default_idx = 1 if current_override else 0
                            selected_mode = st.radio(
                                "How should this workbook be processed?",
                                options=mode_options,
                                index=default_idx,
                                key=mode_key,
                                help="This is separate from sheet selection. Choose 'Already flat' only when tabs are clean table layouts.",
                            )
                            st.session_state.flat_file_overrides[file_name] = (selected_mode == mode_options[1])

                if not has_excel:
                    st.info("ℹ️ No Excel files to configure. CSV files will be loaded automatically.")
               
                # Flat file conversion is always enabled
                st.markdown("---")
                st.info("ℹ️ Excel sheets are auto-flattened by default. You can mark specific files as already flat to skip extraction.")

                # Load button
                st.markdown("---")
                total_selected = sum(len(sheets) for sheets in st.session_state.selected_sheets.values())
                action_label = "Add / Load Files" if st.session_state.files_loaded else "Load Files"
                if st.button(f"🔄 {action_label} ({total_selected} sheet(s) selected)", type="primary", use_container_width=True):
                    with st.spinner("Loading files..."):
                        if load_uploaded_files(
                            uploaded_file_list,
                            st.session_state.selected_sheets,
                            st.session_state.flat_file_overrides,
                        ):
                            st.success("✅ Files loaded successfully!")
                            st.rerun()

        # Clear files button (only show if files are loaded)
        if st.session_state.files_loaded:
            st.markdown("---")
            if st.button("🗑️ Clear All Files", use_container_width=True):
                st.session_state.files_db = None
                st.session_state.sql_agent = None
                st.session_state.query_history = []
                st.session_state.graph_memory = []
                if st.session_state.agent_orchestrator is not None:
                    st.session_state.agent_orchestrator.clear_memory()
                st.session_state.uploaded_files = []
                st.session_state.files_loaded = False
                st.session_state.file_sheets = {}
                st.session_state.selected_sheets = {}
                st.session_state.flat_file_overrides = {}
                st.session_state.show_sheet_selector = False
                st.success("✨ Files cleared! Upload new files to continue.")
                st.rerun()
       
        st.markdown("---")
       
        st.markdown("""
        <div style='text-align: center; padding: 1rem; margin-bottom: 1rem;'>
            <h2 style='color: #1e3a8a; margin: 0;'>ℹ️ About</h2>
        </div>
        """, unsafe_allow_html=True)
       
        st.markdown("""
        <div style='background-color: white;
                    padding: 1.5rem;
                    border-radius: 10px;
                    border: 1px solid #e5e7eb;
                    box-shadow: 0 1px 3px rgba(0, 0, 0, 0.1);'>
            <p style='margin: 0;'>
                Query Excel and CSV files using natural language - no SQL knowledge required!
            </p>
        </div>
        """, unsafe_allow_html=True)
       
        st.markdown("<br>", unsafe_allow_html=True)
       
        st.markdown("**✨ Key Features:**")
        st.markdown("""
        - 📊 Upload your files & select specific sheets
        - 📋 Support for Excel & CSV
        - 💬 Natural language queries
        - 🔍 AI-powered SQL generation
        - 📥 Export results to CSV
        - 📜 Query history tracking
        """)
       
        st.markdown("---")
       
        # Status section with metrics
        st.markdown("**📊 System Status:**")
        if st.session_state.initialized:
            st.success("✅ Azure OpenAI Ready")
        else:
            st.warning("⚠️ Initializing...")
        langgraph_available, _ = SQLAgentOrchestrator.langgraph_status()
        if langgraph_available:
            st.success("✅ LangGraph Ready")
        else:
            st.info("ℹ️ Local graph fallback active")
        st.toggle(
            "Graph reasoning & memory",
            key="use_graph_orchestration",
            help=(
                "Embedded by default: plans queries as a graph, sanity-checks "
                "results, and reuses verified answers across questions. "
                "Disable only to run each question in isolation."
            ),
        )
       
        if st.session_state.files_loaded and st.session_state.files_db and st.session_state.files_db.tables_info:
            st.success("✅ Files Loaded")
           
            # Display quick stats
            total_tables = len(st.session_state.files_db.tables_info)
            total_queries = len(st.session_state.query_history)
            files_loaded = len(st.session_state.files_db.loaded_files)
           
            st.info(f"📁 Active: {files_loaded} file(s)")
           
            col1, col2 = st.columns(2)
            with col1:
                st.metric("Tables", total_tables)
            with col2:
                st.metric("Queries", total_queries)
        else:
            st.info("ℹ️ No files loaded yet")
       
        st.markdown("---")
        st.caption("💡 Tip: Be specific in your questions for better results!")
   
    # Main content with tabs
    if not st.session_state.files_loaded:
        # Show welcome message when no files are loaded
        st.markdown("""
        <div style='text-align: center;
                    padding: 4rem 2rem;
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    border-radius: 20px;
                    margin: 2rem 0;
                    box-shadow: 0 10px 40px rgba(0, 0, 0, 0.2);'>
            <h2 style='color: white; font-size: 2.5rem; margin: 0 0 1rem 0;'>👋 Welcome to SQL Query Agent</h2>
            <p style='color: white; font-size: 1.3rem; opacity: 0.95; margin: 0;'>
                Get started by uploading your Excel or CSV files using the sidebar
            </p>
        </div>
        """, unsafe_allow_html=True)
       
        # Instructions
        col1, col2, col3 = st.columns(3)
       
        with col1:
            st.markdown("""
            <div style='text-align: center; padding: 2rem; background-color: #f8fafc; border-radius: 15px; height: 100%;'>
                <div style='font-size: 3rem; margin-bottom: 1rem;'>📤</div>
                <h3 style='color: #1e3a8a;'>1. Upload Files</h3>
                <p style='color: #64748b;'>Select Excel (.xlsx, .xls) or CSV files from your computer</p>
            </div>
            """, unsafe_allow_html=True)
       
        with col2:
            st.markdown("""
            <div style='text-align: center; padding: 2rem; background-color: #f8fafc; border-radius: 15px; height: 100%;'>
                <div style='font-size: 3rem; margin-bottom: 1rem;'>📊</div>
                <h3 style='color: #1e3a8a;'>2. Select Sheets</h3>
                <p style='color: #64748b;'>Choose which Excel sheets to process</p>
            </div>
            """, unsafe_allow_html=True)
       
        with col3:
            st.markdown("""
            <div style='text-align: center; padding: 2rem; background-color: #f8fafc; border-radius: 15px; height: 100%;'>
                <div style='font-size: 3rem; margin-bottom: 1rem;'>💬</div>
                <h3 style='color: #1e3a8a;'>3. Ask & Analyze</h3>
                <p style='color: #64748b;'>Get instant insights with natural language</p>
            </div>
            """, unsafe_allow_html=True)
       
        st.markdown("<br>", unsafe_allow_html=True)
       
        # Additional info
        with st.expander("📖 How to use this application", expanded=False):
            st.markdown("""
            ### Getting Started
           
            1. **Upload Your Data Files**
               - Click the file uploader in the sidebar
               - Select one or more Excel or CSV files
               - Click "Analyze Files" to scan them
           
            2. **Select Sheets (for Excel files)**
               - View all available sheets from your Excel files
               - Select or deselect specific sheets to load
               - Use "Select All" / "Clear All" for quick selection
               - CSV files are automatically included
           
            3. **Choose Processing Options**
               - Selected sheets are automatically checked for flat, report-block, or matrix structure
               - Already-flat sheets bypass expensive extraction
               - Complex sheets are flattened while remaining separate logical tables

            4. **Load Your Data**
               - Click "Load Files" with your selected sheets
               - Generated schema metadata is supplied directly to the AI
               - A common Excel workbook containing all data tables and the Schema tab becomes available
               - Wait for the files to be processed
           
            5. **Ask Questions in Natural Language**
               - Go to the "Ask Questions" tab
               - Type your question like: "What is the total sales by region?"
               - Click "Generate" to execute
           
            6. **View and Export Results**
               - Results appear as tables
               - Download results as CSV
               - View query history anytime
           
            ### Flat File Conversion
            The integrated flat-file builder automatically handles:
            - Complex layouts or merged cells
            - Matrix formats (data in both rows and columns)
            - Unstructured data that's hard to query
            - Existing flat tables, which are loaded without report extraction
           
            ### Supported File Types
            - Excel files (.xlsx, .xls) - choose specific sheets
            - CSV files (.csv) - loaded automatically
           
            ### Example Questions
            - "Show me the first 10 rows"
            - "What is the total revenue by product?"
            - "Calculate the average price by category"
            - "List all unique customer names"
            """)
    else:
        # Show main tabs when files are loaded
        tab1, tab2, tab3 = st.tabs([
            "🔍 Ask Questions",
            "📊 View Tables",
            "📜 Query History"
        ])
       
        with tab1:
            display_query_interface()
       
        with tab2:
            display_tables_info()
       
        with tab3:
            display_query_history()


if __name__ == "__main__":
    main()
